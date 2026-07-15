from flask import Flask, render_template, request, Response, send_file, redirect, url_for, session, flash, abort, g
from datetime import date, timedelta, datetime
from psycopg2 import sql
from psycopg2.extras import RealDictCursor
from apscheduler.schedulers.background import BackgroundScheduler
from q_launch_registry import Q_LAUNCH_ITEMS
from services.coach import build_coach, build_action_cards
import os
from decimal import Decimal, InvalidOperation
import time
import csv
import io
import imaplib
import email
from email.header import decode_header
from openpyxl import Workbook
from dotenv import load_dotenv
from openpyxl.styles import Font
from db import get_db_connection
output = io.StringIO()
file_data = io.BytesIO()
load_dotenv()
from services.sms_service import send_sms_telnyx


app = Flask(__name__)

# --------------------------------------------------
# Application Branding
# --------------------------------------------------

app.config["PRODUCT_NAME"] = "Peach Suite Pro"
app.config["COMPANY_NAME"] = "Just Peachy Data LLC"
app.config["APP_VERSION"] = "1.0.0"
app.config["TAGLINE"] = "Helping Your Business Bear Fruit™"


# Mailgun config
MAILGUN_API_KEY = os.getenv("MAILGUN_API_KEY")
MAILGUN_DOMAIN = os.getenv("MAILGUN_DOMAIN")
MAILGUN_FROM = os.getenv("MAILGUN_FROM")

SMS_ENABLED = os.getenv("SMS_ENABLED", "false").lower() == "true"


app.secret_key = os.environ.get("SECRET_KEY", "local-dev-key")


from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps



print("APP.PY LOADED - SWITCH TEST VERSION", flush=True)

print("APP FILE LOADED")


print("GENERAL EMAIL SEND HIT", flush=True)
print("MAILGUN DOMAIN:", MAILGUN_DOMAIN, flush=True)
print("MAILGUN FROM:", MAILGUN_FROM, flush=True)
print("MAILGUN KEY STARTS:", MAILGUN_API_KEY[:4] if MAILGUN_API_KEY else None, flush=True)




# ==========================================================
# PEACH SUITE PRO APPLICATION SETTINGS
# ==========================================================

def build_q_launch_items():
    items = []

    for item in Q_LAUNCH_ITEMS:
        if not item.get("active", True):
            continue

        try:
            item_copy = item.copy()
            item_copy["url"] = url_for(
                item["endpoint"],
                **item.get("endpoint_args", {})
            )
            items.append(item_copy)

        except Exception as e:
            print(
                f"[Q Launch] Skipping {item.get('title')} "
                f"({item.get('endpoint')}): {e}",
                flush=True
            )

    return items












@app.context_processor
def inject_global_context():
    from datetime import datetime

    # These values should always be available
    context = {
        "product_name": app.config.get("PRODUCT_NAME"),
        "company_name": app.config.get("COMPANY_NAME"),
        "app_version": app.config.get("APP_VERSION"),
        "tagline": app.config.get("TAGLINE"),
        "current_year": datetime.now().year,
        "godaddy_unreviewed_count": 0
    }

    context["q_launch_items"] = build_q_launch_items()

    if "user_id" not in session or "spa_id" not in session:
        return context

    spa_id = current_spa_id()

    if not spa_id:
        return context

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE spa_id = %s
          AND external_source = 'godaddy'
          AND COALESCE(import_reviewed, FALSE) = FALSE
    """, (spa_id,))

    context["godaddy_unreviewed_count"] = cur.fetchone()[0]

    cur.close()
    conn.close()

    return context


# =====================================================
# LOGGING HELPERS
# =====================================================



def log_sms(message):
    print(f"[SMS] {message}", flush=True)

def log_email(message):
    print(f"[EMAIL] {message}", flush=True)

def log_godaddy(message):
    print(f"[GODADDY] {message}", flush=True)

def log_scheduler(message):
    print(f"[SCHEDULER] {message}", flush=True)

def log_reminder(message):
    print(f"[REMINDER] {message}", flush=True)

def log_ai(message):
    print(f"[AI] {message}", flush=True)






###################################
#
#
###################################



def log_event(category, message, severity="INFO", spa_id=None, related_type=None, related_id=None, created_by=None):
    print(f"[{category}] {severity}: {message}", flush=True)

    save_system_activity(
        category=category,
        severity=severity,
        message=message,
        spa_id=spa_id,
        related_type=related_type,
        related_id=related_id,
        created_by=created_by
    )

def log_sms(message, severity="INFO", spa_id=None, related_type=None, related_id=None, created_by=None):
    log_event("SMS", message, severity, spa_id, related_type, related_id, created_by)


def log_email(message, severity="INFO", spa_id=None, related_type=None, related_id=None, created_by=None):
    log_event("EMAIL", message, severity, spa_id, related_type, related_id, created_by)


def log_godaddy(message, severity="INFO", spa_id=None, related_type=None, related_id=None, created_by=None):
    log_event("GODADDY", message, severity, spa_id, related_type, related_id, created_by)


def log_scheduler(message, severity="INFO", spa_id=None, related_type=None, related_id=None, created_by=None):
    log_event("SCHEDULER", message, severity, spa_id, related_type, related_id, created_by)


def log_reminder(message, severity="INFO", spa_id=None, related_type=None, related_id=None, created_by=None):
    log_event("REMINDER", message, severity, spa_id, related_type, related_id, created_by)


def log_ai(message, severity="INFO", spa_id=None, related_type=None, related_id=None, created_by=None):
    log_event("AI", message, severity, spa_id, related_type, related_id, created_by)




#######################################
#  Save SYSTEM LOGGING
#####################################



def save_system_activity(
    category,
    message,
    severity="INFO",
    spa_id=None,
    related_type=None,
    related_id=None,
    created_by=None
):
    try:
        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO system_logs (
                spa_id,
                category,
                severity,
                message,
                related_type,
                related_id,
                created_by
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            spa_id,
            category,
            severity,
            message,
            related_type,
            related_id,
            created_by
        ))

        conn.commit()
        cur.close()
        conn.close()

    except Exception as e:
        print(f"[SYSTEM LOGGING ERROR] {e}", flush=True)








########################################
#LOG EVENT
#LOGGING HELPER FUNCTION
#################################




def log_event(category, message, severity="INFO", spa_id=None, related_type=None, related_id=None, created_by=None):
    print(f"[{category}] {severity}: {message}", flush=True)

    save_system_activity(
        category=category,
        message=message,
        severity=severity,
        spa_id=spa_id,
        related_type=related_type,
        related_id=related_id,
        created_by=created_by
    )




#  ---------------------
#        HELPERS
#  --------------------








######################################
#
#   LOG GODADDY MESSAGE
#
######################################


def log_godaddy(message):
    print(f"[GODADDY IMPORT] {message}", flush=True)


















#   ---------------------
#
#  MASTER ADMIN ACCESS 
#
#
#   --------------------


def is_master_admin():
    return session.get("role") == "master_admin"


def current_spa_filter():
    if is_master_admin():
        return "", ()
    return " AND spa_id = %s ", (current_spa_id(),)


def current_spa_id():
    from flask import g, session

    if "user_id" not in session:
        return None

    if session.get("role") == "master_admin":
        return None

    return getattr(g, "spa_id", None)





# ############################################
#
#  Messaging Compliance Center - Onboarding 
#
#  
###############################################


def get_messaging_onboarding(spa_id):
    ...



def save_messaging_onboarding(spa_id, form_data):
    ...


def get_messaging_onboarding(spa_id):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute("""
        SELECT *
        FROM messaging_onboarding
        WHERE spa_id = %s
    """, (spa_id,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    return row



##################################################
#
#   SMS  & EMAIL OPT OUT FOOTER FROM MASTER ADMIN
#
###################################################

def get_messaging_footer(footer_type):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT footer_text
        FROM messaging_footer_settings
        WHERE footer_type = %s
          AND is_active = TRUE
        LIMIT 1
    """, (footer_type,))
    row = cur.fetchone()
    cur.close()
    conn.close()

    return row[0] if row else ""









    ####################################
    #
    #   LOG BOOKING IMPORT
    #
    #####################################


def log_booking_import(
    spa_id,
    source,
    status,
    email_id=None,
    external_order_id=None,
    email_subject=None,
    appointment_id=None,
    client_id=None,
    parser_version=None,
    error_message=None,
    processing_time_ms=None
):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            INSERT INTO booking_import_logs (
                spa_id,
                source,
                email_id,
                external_order_id,
                email_subject,
                status,
                appointment_id,
                client_id,
                parser_version,
                error_message,
                processing_time_ms
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (spa_id, source, email_id, status)
            DO NOTHING
        """, (
            spa_id,
            source,
            email_id,
            external_order_id,
            email_subject,
            status,
            appointment_id,
            client_id,
            parser_version,
            error_message,
            processing_time_ms
        ))

        conn.commit()

    except Exception as e:
        conn.rollback()
        log_godaddy(f"Failed to write booking import log: {e}")

    finally:
        cur.close()
        conn.close()







###################################################
###############################################
#############################################
##########################################
#
#   COMMUNICATIONS ENGINE
#
#
#
#
#
######################################










##########################################
#
#   GET REQUEST LANGUAGE
######################################

def get_request_language():
    """
    Official messaging route language resolver.

    Priority:
    1. URL query string
    2. Submitted form
    3. Platform default language
    """
    return normalize_language_code(
        request.args.get("language_code")
        or request.form.get("language_code")
        or get_current_language()
        or get_default_language()
    )










##########################################
#
#   SMS EMAIL Communications Group
######################################

######### number  1 first
def append_sms_footer(message):

    footer = get_messaging_footer("sms_opt_out")

    message = (message or "").strip()

    if footer and footer.lower() in message.lower():
        return message

    if footer:
        return f"{message}\n\n{footer}"

    return message



#################################
#
#   APPEND EMAIL FOOTER
#
############## number 2

def append_email_footer(message):
    footer = get_messaging_footer(
        "email_unsubscribe"
    )

    message = (message or "").strip()

    if footer and footer.lower() in message.lower():
        return message

    if footer:
        return f"{message}\n\n{footer}"

    return message



##################################
#
#   GET APPROVED SMS TEMPLATE
#       6-30-26
#################################
############# number 3




def get_approved_sms_template(
    spa_id,
    template_type,
    language_code=None,
    template_id=None
):
    language_code = normalize_language_code(language_code or get_default_language())
    default_language = normalize_language_code(get_default_language())

    conn = get_db_connection()
    cur = conn.cursor()

    if template_id:
        cur.execute("""
            SELECT
                template_id,
                message_text
            FROM messaging_templates
            WHERE template_id = %s
              AND spa_id = %s
              AND channel = 'sms'
              AND is_active = TRUE
              AND approved_for_use = TRUE
              AND COALESCE(is_archived, FALSE) = FALSE
            LIMIT 1
        """, (template_id, spa_id))

        row = cur.fetchone()

        cur.close()
        conn.close()

        return row

    cur.execute("""
        SELECT
            template_id,
            message_text
        FROM messaging_templates
        WHERE spa_id = %s
          AND channel = 'sms'
          AND template_type = %s
          AND language_code = %s
          AND is_active = TRUE
          AND approved_for_use = TRUE
          AND COALESCE(is_archived, FALSE) = FALSE
        ORDER BY updated_at DESC, template_id DESC
        LIMIT 1
    """, (spa_id, template_type, language_code))

    row = cur.fetchone()

    if not row and language_code != default_language:
        cur.execute("""
            SELECT
                template_id,
                message_text
            FROM messaging_templates
            WHERE spa_id = %s
              AND channel = 'sms'
              AND template_type = %s
              AND language_code = %s
              AND is_active = TRUE
              AND approved_for_use = TRUE
              AND COALESCE(is_archived, FALSE) = FALSE
            ORDER BY updated_at DESC, template_id DESC
            LIMIT 1
        """, (spa_id, template_type, default_language))

        row = cur.fetchone()

    cur.close()
    conn.close()

    return row







############################################################
#       GET APPROVED EMAIL TEMPLATE
#
#
#       6-30-26
############################################################
############. number 4



def get_approved_email_template(
    spa_id,
    template_type,
    language_code=None,
    template_id=None
):
    language_code = normalize_language_code(language_code or get_default_language())
    default_language = normalize_language_code(get_default_language())

    conn = get_db_connection()
    cur = conn.cursor()

    if template_id:
        cur.execute("""
            SELECT
                template_id,
                subject_text,
                message_text
            FROM messaging_templates
            WHERE template_id = %s
              AND spa_id = %s
              AND channel = 'email'
              AND approved_for_use = TRUE
              AND is_active = TRUE
              AND COALESCE(is_archived, FALSE) = FALSE
            LIMIT 1
        """, (template_id, spa_id))

        row = cur.fetchone()

        cur.close()
        conn.close()

        if not row:
            return None

        return {
            "template_id": row[0],
            "subject": row[1] or "",
            "body": row[2] or ""
        }

    cur.execute("""
        SELECT
            template_id,
            subject_text,
            message_text
        FROM messaging_templates
        WHERE spa_id = %s
          AND template_type = %s
          AND channel = 'email'
          AND language_code = %s
          AND approved_for_use = TRUE
          AND is_active = TRUE
          AND COALESCE(is_archived, FALSE) = FALSE
        ORDER BY updated_at DESC, template_id DESC
        LIMIT 1
    """, (
        spa_id,
        template_type,
        language_code
    ))

    row = cur.fetchone()

    if not row and language_code != default_language:
        cur.execute("""
            SELECT
                template_id,
                subject_text,
                message_text
            FROM messaging_templates
            WHERE spa_id = %s
              AND template_type = %s
              AND channel = 'email'
              AND language_code = %s
              AND approved_for_use = TRUE
              AND is_active = TRUE
              AND COALESCE(is_archived, FALSE) = FALSE
            ORDER BY updated_at DESC, template_id DESC
            LIMIT 1
        """, (
            spa_id,
            template_type,
            default_language
        ))

        row = cur.fetchone()

    cur.close()
    conn.close()

    if not row:
        return None

    return {
        "template_id": row[0],
        "subject": row[1] or "",
        "body": row[2] or ""
    }











#####################################
#
#. APPLY MERGE FIELDS
###################################
#### number 5



def apply_merge_fields(text, merge_data):
    text = text or ""

    for field, value in (merge_data or {}).items():
        value = str(value or "")

        text = text.replace(f"{{{{{field}}}}}", value)
        text = text.replace(f"{{{{ {field} }}}}", value)
        text = text.replace(f"{{{field}}}", value)

    return text








#####################################
#   BUILD SMS MESSAGE
#      COMMUNICATION ENGINE PART 7
#
#
#   6-30-26
###################################
############## 



def build_sms_message(
    spa_id,
    template_type,
    merge_data,
    language_code=None,
    template_id=None
):
    merge_data = merge_data or {}

    client_id = merge_data.get("client_id")

    language_code = normalize_language_code(
        language_code
        or merge_data.get("language_code")
        or get_default_language()
    )

    merge_data["language_code"] = language_code

    template = get_approved_sms_template(
        spa_id=spa_id,
        template_type=template_type,
        language_code=language_code,
        template_id=template_id
    )

    if not template:
        return {
            "success": False,
            "error": f"No approved active SMS template found for: {template_type}",
            "message_body": None,
            "template_id": None,
            "language_code": language_code
        }

    template_id = template[0]
    template_text = template[1]

    merge_data = enrich_sms_merge_data(
        spa_id=spa_id,
        template_type=template_type,
        merge_data=merge_data
    )

    rendered_message = render_sms_template(
        template_text,
        merge_data
    )

    rendered_message = append_sms_footer(rendered_message)

    return {
        "success": True,
        "error": None,
        "message_body": rendered_message,
        "template_id": template_id,
        "language_code": language_code
    }












#####################################
#
#       BUILD EMAIL MESSAGE
#
#
#       6-30-26
###################################
##############. Number 7.  




def build_email_message(
    spa_id,
    template_type,
    merge_data,
    language_code=None,
    template_id=None
):
    merge_data = merge_data or {}

    language_code = normalize_language_code(
        language_code
        or merge_data.get("language_code")
        or get_default_language()
    )

    merge_data["language_code"] = language_code

    template = get_approved_email_template(
        spa_id=spa_id,
        template_type=template_type,
        language_code=language_code,
        template_id=template_id
    )

    if not template:
        return None

    subject = apply_merge_fields(
        template["subject"],
        merge_data
    )

    body = apply_merge_fields(
        template["body"],
        merge_data
    )

    body = append_email_footer(body)

    return {
        "subject": subject,
        "body": body,
        "template_id": template["template_id"],
        "language_code": language_code
    }











#####################################
#
#       BUILD COMMUNICATION 
#
#   6-30-26
#
###################################
############number 8 - - last



def build_communication(
    spa_id,
    channel,
    template_type,
    merge_data,
    language_code=None,
    template_id=None
):
    channel = (channel or "").lower().strip()
    language_code = normalize_language_code(language_code or get_default_language())

    if channel == "sms":
        sms = build_sms_message(
            spa_id=spa_id,
            template_type=template_type,
            merge_data=merge_data,
            language_code=language_code,
            template_id=template_id
        )

        if not sms or not sms.get("success"):
            return None

        return {
            "channel": "sms",
            "subject": None,
            "body": sms["message_body"],
            "template_id": sms["template_id"],
            "language_code": sms["language_code"]
        }

    if channel == "email":
        email = build_email_message(
            spa_id=spa_id,
            template_type=template_type,
            merge_data=merge_data,
            language_code=language_code,
            template_id=template_id
        )

        if not email:
            return None

        return {
            "channel": "email",
            "subject": email["subject"],
            "body": email["body"],
            "template_id": email.get("template_id"),
            "language_code": email.get("language_code", language_code)
        }

    return None








##################################
#
#     SEND COMMUNICATION
#   6-30-26
#################################


def send_communication(
    spa_id,
    channel,
    recipient,
    template_type,
    merge_data,
    client_id=None,
    message_type=None,
    language_code=None,
    template_id=None
):
    channel = (channel or "").lower().strip()
    language_code = normalize_language_code(language_code or get_default_language())

    if channel not in ("sms", "email"):
        raise ValueError(f"Unsupported communication channel: {channel}")

    merge_data = merge_data or {}
    merge_data["language_code"] = language_code

    if client_id and not merge_data.get("client_id"):
        merge_data["client_id"] = client_id

    communication = build_communication(
        spa_id=spa_id,
        channel=channel,
        template_type=template_type,
        merge_data=merge_data,
        language_code=language_code,
        template_id=template_id
    )

    if not communication:
        return {
            "success": False,
            "error": "Communication could not be built.",
            "channel": channel,
            "template_id": template_id,
            "template_type": template_type,
            "language_code": language_code
        }

    if channel == "sms":
        length_check = validate_sms_length(communication["body"])

        if not length_check["valid"]:
            return {
                "success": False,
                "error": length_check["error"],
                "channel": "sms",
                "message_body": communication["body"],
                "template_id": communication.get("template_id"),
                "template_type": template_type,
                "language_code": language_code
            }

        sms_result = send_compliant_sms(
            spa_id=spa_id,
            client_id=client_id,
            recipient_phone=recipient,
            message_body=communication["body"],
            message_type=message_type or template_type
        )

        sms_result["channel"] = "sms"
        sms_result["message_body"] = communication["body"]
        sms_result["template_id"] = communication.get("template_id")
        sms_result["template_type"] = template_type
        sms_result["language_code"] = language_code

        return sms_result

    if channel == "email":
        response = send_email(
            to=recipient,
            subject=communication["subject"],
            body=communication["body"]
        )

        if response.status_code == 200:
            return {
                "success": True,
                "status": "sent",
                "provider_status": response.status_code,
                "error": None,
                "channel": "email",
                "subject": communication["subject"],
                "body": communication["body"],
                "template_id": communication.get("template_id"),
                "template_type": template_type,
                "language_code": language_code
            }

        return {
            "success": False,
            "status": "failed",
            "provider_status": response.status_code,
            "error": response.text,
            "channel": "email",
            "subject": communication["subject"],
            "body": communication["body"],
            "template_id": communication.get("template_id"),
            "template_type": template_type,
            "language_code": language_code
        }







##################################
#
#   LOG SMS MESSAGE
#
##################################


def log_sms_message(
    spa_id,
    client_id,
    recipient_phone,
    message_body,
    message_type,
    direction,
    status,
    provider_message_id=None,
    provider_status=None,
    provider_error_code=None,
    provider_error_message=None
):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO sms_messages (
            spa_id,
            client_id,
            recipient_phone,
            message_body,
            message_type,
            direction,
            status,
            provider_message_id,
            provider_status,
            provider_error_code,
            provider_error_message,
            created_at
        )
        VALUES (
            %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW()
        )
    """, (
        spa_id,
        client_id,
        recipient_phone,
        message_body,
        message_type,
        direction,
        status,
        provider_message_id,
        provider_status,
        provider_error_code,
        provider_error_message
    ))

    conn.commit()
    cur.close()
    conn.close()





##################################
#
#   SEND EMAIL MESSAGE
#
##################################

def send_email_message(
    spa_id,
    client_id,
    recipient_email,
    subject,
    message_body,
    message_type
):
    result = send_email(
        to=recipient_email,
        subject=subject,
        body=message_body
    )

    # Later we can move this to a shared communication log table.
    log_email_message(
        spa_id=spa_id,
        client_id=client_id,
        recipient_email=recipient_email,
        subject=subject,
        message_body=message_body,
        message_type=message_type,
        status="sent"
    )

    return result




##################################
#
#      GET SMS ELIGIBLE CLIENTS
#
##################################


def get_sms_eligible_clients(
    spa_id,
    search="",
    client_status="",
    show_all=False
):
    clients = []

    if not search and not show_all:
        return clients

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            c.client_id,
            c.first_name,
            c.last_name,
            c.phone,
            cs.status_name
        FROM clients c
        LEFT JOIN client_statuses cs
            ON c.client_status = cs.status_name
        WHERE c.spa_id = %s
          AND c.sms_opt_in = TRUE
          AND COALESCE(c.sms_opt_out, FALSE) = FALSE
          AND c.active_client = TRUE
          AND c.phone IS NOT NULL
          AND TRIM(c.phone) <> ''
    """

    params = [spa_id]

    if search:
        query += """
          AND (
               LOWER(c.first_name) LIKE %s
               OR LOWER(c.last_name) LIKE %s
               OR c.phone LIKE %s
          )
        """
        params.extend([
            f"%{search.lower()}%",
            f"%{search.lower()}%",
            f"%{search}%"
        ])

    if client_status:
        query += " AND c.client_status = %s"
        params.append(client_status)

    query += " ORDER BY c.last_name, c.first_name"

    cur.execute(query, params)
    clients = cur.fetchall()

    cur.close()
    conn.close()

    return clients
    



##################################
#
#      GET CLIENT STATUSES
#
##################################


def get_client_statuses(spa_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            client_status_id,
            status_name
        FROM client_statuses
        WHERE spa_id = %s
        ORDER BY status_name
    """, (spa_id,))

    statuses = cur.fetchall()

    cur.close()
    conn.close()

    return statuses





##################################
#
#     GET EMAIL ELIGIBLE CLIENTS
#
##################################


def get_email_eligible_clients(
    spa_id,
    search="",
    client_status="",
    show_all=False
):
    clients = []

    # Allow a status selection to load clients even when
    # there is no search and Show All was not selected.
    if not search and not show_all and not client_status:
        return clients

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            c.client_id,
            c.first_name,
            c.last_name,
            c.email,
            cs.status_name
        FROM clients c
        LEFT JOIN client_statuses cs
            ON TRIM(c.client_status) = TRIM(cs.status_name)
           AND cs.spa_id = c.spa_id
        WHERE c.spa_id = %s
          AND c.email_opt_in = TRUE
          AND COALESCE(c.email_opt_out, FALSE) = FALSE
          AND c.email IS NOT NULL
          AND TRIM(c.email) <> ''
    """

    params = [spa_id]

    if search:
        query += """
          AND (
               LOWER(c.first_name) LIKE %s
               OR LOWER(c.last_name) LIKE %s
               OR LOWER(c.email) LIKE %s
          )
        """

        search_value = f"%{search.lower()}%"

        params.extend([
            search_value,
            search_value,
            search_value
        ])

    if client_status:
        query += """
          AND cs.client_status_id = %s
        """
        params.append(client_status)

    query += """
        ORDER BY
            c.last_name,
            c.first_name
    """

    cur.execute(query, params)
    clients = cur.fetchall()

    cur.close()
    conn.close()

    return clients


#################################
#
#      GET TEMPLATE BY ID
#
##################################


def get_template_by_id(
    spa_id,
    template_id,
    channel=None
):
    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            template_id,
            template_type,
            template_name,
            message_text,
            subject_text,
            language_code,
            approved_for_use,
            is_active,
            ai_score,
            ai_review,
            ai_risk_level,
            last_ai_reviewed_at,
            channel
        FROM messaging_templates
        WHERE spa_id = %s
          AND template_id = %s
    """

    params = [spa_id, template_id]

    if channel:
        query += " AND channel = %s"
        params.append(channel)

    query += " LIMIT 1"

    cur.execute(query, params)

    template = cur.fetchone()

    cur.close()
    conn.close()

    return template








##################################
#
#      Template name
#
##################################

def is_default_template_name(template_name):
    return (template_name or "").strip().lower() == "default"



##################################
#
#   GET ACTIVE MESSAGING TEMPLATES
#
##################################



def get_active_messaging_templates(spa_id, channel):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            template_id,
            template_type,
            template_name,
            message_text,
            subject_text,
            language_code
        FROM messaging_templates
        WHERE spa_id = %s
          AND channel = %s
          AND is_active = TRUE
          AND approved_for_use = TRUE
          AND COALESCE(is_archived, FALSE) = FALSE
        ORDER BY template_type, template_name
    """, (spa_id, channel))

    templates = cur.fetchall()

    cur.close()
    conn.close()

    return templates






##################################
#
#   GET SMS CLIENTS BY IDS
#
##################################


def get_sms_clients_by_ids(spa_id, client_ids):
    if not client_ids:
        return []

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            client_id,
            first_name,
            last_name,
            phone,
            sms_opt_in,
            sms_opt_out
        FROM clients
        WHERE spa_id = %s
          AND sms_opt_in = TRUE
          AND COALESCE(sms_opt_out, FALSE) = FALSE
          AND active_client = TRUE
          AND phone IS NOT NULL
          AND TRIM(phone) <> ''
          AND client_id = ANY(%s)
        ORDER BY last_name, first_name
    """, (spa_id, client_ids))

    clients = cur.fetchall()

    cur.close()
    conn.close()

    return clients




##################################
#
#   GET DEFAULT LANGUAGE
#
##################################


def get_default_language():
    languages = get_supported_languages(active_only=True)

    for lang in languages:
        if lang["is_default"]:
            return lang["language_code"]

    return "EN"


##################################
#
#   SUPPORTED LANGUAGES
#
##################################





from psycopg2.extras import RealDictCursor

def get_supported_languages(active_only=True):
    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    if active_only:
        cur.execute("""
            SELECT
                language_code,
                language_name,
                native_name,
                locale_code,
                is_default
            FROM supported_languages
            WHERE is_active = TRUE
            ORDER BY display_order, language_name
        """)
    else:
        cur.execute("""
            SELECT
                language_code,
                language_name,
                native_name,
                locale_code,
                is_active,
                is_default
            FROM supported_languages
            ORDER BY display_order, language_name
        """)

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows










##################################
#
#   NORMALIZE LANGUAGE CODE
#
##################################





def normalize_language_code(language_code):
    code = (language_code or "").strip().upper()

    supported_languages = get_supported_languages(active_only=True)

    supported_codes = {
        lang["language_code"]
        for lang in supported_languages
    }

    if code in supported_codes:
        return code

    return get_default_language()





##################################
#
#   GET LANGUAGE NAME
#
##################################

def get_language_name(language_code):
    code = normalize_language_code(language_code)

    languages = get_supported_languages(active_only=True)

    for lang in languages:
        if lang["language_code"] == code:
            return lang["native_name"]

    return code
























###################################################################
#       >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#   >>>>>. COMMUNICATIONS ENGING.  <<<<<<<<<<<<<
#
#####################################################################


##################################
#
#     RENDER SMS TEMPLATE. HELPER
#
#################################


import re

def render_sms_template(template_text, merge_data):
    if not template_text:
        return ""

    if merge_data is None:
        merge_data = {}

    def replace_field(match):
        field_name = match.group(1).strip()
        value = merge_data.get(field_name, "")
        return str(value or "")

    return re.sub(r"\{\{\s*([a-zA-Z0-9_]+)\s*\}\}", replace_field, template_text)








##################################
#
#     ENRICH SMS MERGE DATA
#
#################################

def enrich_sms_merge_data(spa_id, template_type, merge_data):
    merge_data = merge_data or {}

    if template_type == "appointment_reminder":
        return enrich_appointment_reminder_merge_data(spa_id, merge_data)

    return merge_data



def enrich_appointment_reminder_merge_data(spa_id, merge_data):
    # We will fill this in next.
    return merge_data






##################################
#
#     ENRICH APPOINTMENT REMINDER MERGE
#
#################################
def enrich_appointment_reminder_merge_data(spa_id, merge_data):
    reminder_id = merge_data.get("reminder_id")

    if not reminder_id:
        return merge_data

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            c.first_name,
            c.last_name,
            a.appointment_date,
            a.appointment_time,
            a.service_type,
            s.spa_name,
            s.owner_phone,
            NULL AS spa_website
        FROM reminder_queue rq
        JOIN clients c
          ON rq.client_id = c.client_id
         AND rq.spa_id = c.spa_id
        LEFT JOIN appointments a
          ON rq.appointment_id = a.appointment_id
         AND rq.spa_id = a.spa_id
        LEFT JOIN spas s
          ON rq.spa_id = s.spa_id
        WHERE rq.spa_id = %s
          AND rq.reminder_id = %s
        LIMIT 1
    """, (spa_id, reminder_id))

    row = cur.fetchone()

    cur.close()
    conn.close()

    if not row:
        return merge_data

    (
        first_name,
        last_name,
        appointment_date,
        appointment_time,
        service_name,
        spa_name,
        spa_phone,
        spa_website
    ) = row

    merge_data.update({
        "client_first_name": first_name or "",
        "client_last_name": last_name or "",
        "client_full_name": f"{first_name or ''} {last_name or ''}".strip(),
        "appointment_date": appointment_date.strftime("%m/%d/%Y") if appointment_date else "",
        "appointment_time": appointment_time.strftime("%I:%M %p") if appointment_time else "",
        "service_name": service_name or "",
        "spa_name": spa_name or "",
        "spa_phone": spa_phone or "",
        "spa_website": spa_website or ""
    })

    return merge_data





################################
#       SEND BIRTHDAY
##############################

def send_birthday_reminder_sms(reminder_id, spa_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            rq.reminder_id,
            rq.client_id,
            rq.recipient_phone,
            c.first_name,
            c.birth_date,
            s.spa_name
        FROM reminder_queue rq
        JOIN clients c
          ON rq.client_id = c.client_id
         AND rq.spa_id = c.spa_id
        JOIN spas s
          ON rq.spa_id = s.spa_id
        WHERE rq.reminder_id = %s
          AND rq.spa_id = %s
          AND rq.reminder_type = 'birthday'
          AND rq.send_method = 'sms'
    """, (reminder_id, spa_id))

    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return False, "Birthday reminder not found."

    (
        reminder_id,
        client_id,
        recipient_phone,
        first_name,
        birth_date,
        spa_name
    ) = row

    if not recipient_phone:
        cur.execute("""
            UPDATE reminder_queue
            SET status = 'skipped',
                error_message = %s
            WHERE reminder_id = %s
              AND spa_id = %s
        """, ("Missing phone number", reminder_id, spa_id))
        conn.commit()
        cur.close()
        conn.close()
        return False, "Missing phone number."

    merge_data = build_birthday_message_merge_data(
        client_id=client_id,
        first_name=first_name,
        birth_date=birth_date,
        spa_name=spa_name
    )

    result = send_communication(
        spa_id=spa_id,
        channel="sms",
        recipient=recipient_phone,
        template_type="birthday_message",
        merge_data=merge_data,
        client_id=client_id,
        message_type="birthday_message"
    )

    success = result.get("success", False)

    if success:
        cur.execute("""
            UPDATE reminder_queue
            SET status = 'sent',
                sent_at = NOW(),
                error_message = NULL
            WHERE reminder_id = %s
              AND spa_id = %s
        """, (reminder_id, spa_id))
    else:
        cur.execute("""
            UPDATE reminder_queue
            SET status = 'failed',
                error_message = %s
            WHERE reminder_id = %s
              AND spa_id = %s
        """, (str(result), reminder_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    return success, result


##################################
#
#     SEND TEMPLATE SMS
#
#################################



def send_template_sms(
    spa_id,
    client_id,
    recipient_phone,
    template_type,
    merge_data,
    message_type=None
):
    return send_communication(
        spa_id=spa_id,
        channel="sms",
        recipient=recipient_phone,
        template_type=template_type,
        merge_data=merge_data or {},
        client_id=client_id,
        message_type=message_type or template_type
    )






##################################
#
#     SEND appointment reminder sms
#
#################################


def send_appointment_reminder_sms(reminder_id, spa_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            rq.reminder_id,
            rq.client_id,
            rq.recipient_phone,
            a.appointment_date,
            a.appointment_time,
            a.service_type,
            s.spa_name,
            s.owner_phone,
            NULL AS spa_website
        FROM reminder_queue rq
        JOIN appointments a
          ON rq.appointment_id = a.appointment_id
         AND rq.spa_id = a.spa_id
        JOIN spas s
          ON rq.spa_id = s.spa_id
        WHERE rq.reminder_id = %s
          AND rq.spa_id = %s
          AND rq.reminder_type = 'appointment_reminder'
          AND rq.send_method = 'sms'
    """, (reminder_id, spa_id))

    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return False, "Reminder not found."

    (
        reminder_id,
        client_id,
        recipient_phone,
        appointment_date,
        appointment_time,
        service_name,
        spa_name,
        spa_phone,
        spa_website
    ) = row

    merge_data = build_appointment_reminder_merge_data(
        spa_id=spa_id,
        client_id=client_id,
        appointment_date=appointment_date,
        appointment_time=appointment_time,
        service_name=service_name,
        spa_name=spa_name,
        spa_phone=spa_phone,
        spa_website=spa_website
    )

    merge_data["reminder_id"] = reminder_id

    result = send_communication(
        spa_id=spa_id,
        channel="sms",
        recipient=recipient_phone,
        template_type="appointment_reminder",
        merge_data=merge_data,
        client_id=client_id,
        message_type="appointment_reminder"
    )

    success = result.get("success", False)

    if success:
        cur.execute("""
            UPDATE reminder_queue
            SET status = 'sent',
                sent_at = NOW(),
                error_message = NULL
            WHERE reminder_id = %s
              AND spa_id = %s
        """, (reminder_id, spa_id))
    else:
        cur.execute("""
            UPDATE reminder_queue
            SET status = 'failed',
                error_message = %s
            WHERE reminder_id = %s
              AND spa_id = %s
        """, (str(result), reminder_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    return success, result



##################################
#
#     BUILD APPOINTMENT REMINDER SMS
#
#################################

def build_appointment_reminder_merge_data(
    spa_id,
    client_id,
    appointment_date,
    appointment_time,
    service_name,
    spa_name,
    spa_phone=None,
    spa_website=None
):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            first_name,
            last_name
        FROM clients
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, spa_id))

    client = cur.fetchone()

    cur.close()
    conn.close()

    first_name = client[0] if client else ""
    last_name = client[1] if client else ""

    return {
        "client_id": client_id,
        "client_first_name": first_name or "",
        "client_last_name": last_name or "",
        "client_full_name": f"{first_name or ''} {last_name or ''}".strip(),
        "appointment_date": appointment_date.strftime("%m/%d/%Y") if appointment_date else "",
        "appointment_time": appointment_time.strftime("%I:%M %p") if appointment_time else "",
        "service_name": service_name or "",
        "spa_name": spa_name or "",
        "spa_phone": spa_phone or "",
        "spa_website": spa_website or ""
    }


##############################
#
#   BUILD COMMON MERGE DATA
############################

def build_common_merge_data(
    spa_id,
    client_id
):
    return {}




##############################
#
#   BUILD. BIRTHDAY MERGE DATA
############################

def build_birthday_message_merge_data(
    client_id,
    first_name,
    birth_date,
    spa_name
):
    return {
        "client_id": client_id,
        "client_first_name": first_name or "",
        "first_name": first_name or "",
        "birthday_month": birth_date.strftime("%B") if birth_date else "",
        "birthday_day": birth_date.strftime("%d") if birth_date else "",
        "spa_name": spa_name or ""
    }







##############################
#
#    REMINDER QUEUE AFTER APPOINTMENT
############################


def reminder_queue_after_appointment():
    return True, "Test"





####################################
#
#
#  SAVE MESSAGING ONBOARDING
#
######################################

def save_messaging_onboarding(spa_id, **fields):
    conn = get_db_connection()
    cur = conn.cursor()

    allowed_fields = [
        "legal_business_name",
        "dba_name",
        "ein",
        "business_type",
        "industry",
        "years_in_business",

        "owner_name",
        "owner_email",
        "owner_phone",
        "business_address1",
        "business_address2",
        "city",
        "state",
        "postal_code",
        "country",

        "website_url",
        "privacy_policy_url",
        "terms_url",
        "sms_policy_url",
        "opt_in_method",
        "consent_language",
        "sample_message_1",
        "sample_message_2",

        "opt_in_method_primary",
        "consent_language",
        "frequency_disclosed",
        "rates_disclosed",
        "stop_disclosed",
        "help_disclosed",
        "privacy_linked",
        "terms_linked",
        "double_opt_in",
        "consent_record_retention",
        "opt_in_screenshot",
        "consent_last_reviewed"
    ]

    clean_fields = {
        key: value
        for key, value in fields.items()
        if key in allowed_fields
    }

    cur.execute("""
        SELECT onboarding_id
        FROM messaging_onboarding
        WHERE spa_id = %s
    """, (spa_id,))

    existing = cur.fetchone()

    if existing:
        if clean_fields:
            set_clause = ", ".join(
                [f"{key} = %s" for key in clean_fields.keys()]
            )

            values = list(clean_fields.values())
            values.append(spa_id)

            cur.execute(f"""
                UPDATE messaging_onboarding
                SET {set_clause}
                WHERE spa_id = %s
            """, values)

    else:
        columns = ["spa_id"] + list(clean_fields.keys())
        placeholders = ["%s"] * len(columns)
        values = [spa_id] + list(clean_fields.values())

        cur.execute(f"""
            INSERT INTO messaging_onboarding
            ({", ".join(columns)})
            VALUES ({", ".join(placeholders)})
        """, values)

    conn.commit()
    cur.close()
    conn.close()


############################################################
# MESSAGING COMPLIANCE CHECK
############################################################

def run_messaging_compliance_check(onboarding):
    issues = []
    score = 100

    if not onboarding:
        return {
            "score": 0,
            "status": "Not Started",
            "issues": ["Messaging onboarding has not been started yet."]
        }

    website_url = onboarding.get("website_url")
    privacy_policy_url = onboarding.get("privacy_policy_url")
    terms_url = onboarding.get("terms_url")
    sms_policy_url = onboarding.get("sms_policy_url")
    consent_language = onboarding.get("consent_language") or ""
    sample_message_1 = onboarding.get("sample_message_1") or ""
    sample_message_2 = onboarding.get("sample_message_2") or ""

    if not website_url:
        issues.append("Website URL is missing.")
        score -= 15

    if not privacy_policy_url:
        issues.append("Privacy Policy URL is missing.")
        score -= 15

    if not terms_url:
        issues.append("Terms & Conditions URL is missing.")
        score -= 10

    if not sms_policy_url:
        issues.append("SMS consent policy URL is missing.")
        score -= 15

    required_consent_phrases = [
        "STOP",
        "HELP",
        "message and data rates",
        "consent is not a condition"
    ]

    for phrase in required_consent_phrases:
        if phrase.lower() not in consent_language.lower():
            issues.append(f"Consent language may be missing: {phrase}")
            score -= 10

    if not sample_message_1:
        issues.append("Sample message 1 is missing.")
        score -= 10

    if not sample_message_2:
        issues.append("Sample message 2 is missing.")
        score -= 10

    score = max(score, 0)

    if score >= 85:
        status = "Ready for Review"
    elif score >= 60:
        status = "Needs Review"
    else:
        status = "Incomplete"

    return {
        "score": score,
        "status": status,
        "issues": issues
    }
    



############################################################
#    MESSAGING COMPLIANCE CHECK - ONBOARDING
############################################################






################################################################
#.   COMMUNICATIONS HELPERS
###############################################################





##########################################################
#       AI TEMPLATE REVIEW
#
#   REVIEW TEMPLATE AI BASIC
#   6-30-26
##########################################################


def review_template_ai_basic(template_type, message_text, channel="sms"):

    message_text = message_text or ""

    score = 100
    notes = []

    recommended_fields = {
        "appointment_reminder": [
            "{client_first_name}",
            "{appointment_date}",
            "{appointment_time}"
        ],
        "appointment_confirmation": [
            "{client_first_name}",
            "{appointment_date}",
            "{appointment_time}"
        ],
        "appointment_rescheduled": [
            "{client_first_name}",
            "{appointment_date}",
            "{appointment_time}"
        ],
        "appointment_cancelled": [
            "{client_first_name}"
        ],
        "birthday_message": [
            "{client_first_name}"
        ],
        "follow_up": [
            "{client_first_name}"
        ],
        "review_request": [
            "{client_first_name}"
        ],
        "gift_certificate": [
            "{client_first_name}"
        ]
    }

    fields = recommended_fields.get(template_type, [])

    message_lower = message_text.lower()
    normalized_message = message_text.replace(" ", "")

    if len(message_text.strip()) < 10:
        score -= 50
        notes.append("Template appears to be empty.")





    if channel == "sms":

        opt_out_terms = [
            "stop",
            "unsubscribe",
            "opt out",
            "opt-out",
            "reply stop",
            "text stop",
            "{{opt_out}}"
        ]

        if any(term in message_lower for term in opt_out_terms):
            score -= 40
            notes.append(
                "Remove STOP, unsubscribe, or opt-out language. Peach Suite Pro automatically appends the required compliance footer."
            )
        else:
            notes.append(
                "System compliance footer will be automatically appended."
            )

        single_segment_warning = get_system_setting_int(
            "sms_best_practice_length",
            160
        )

        sms_max_characters = get_system_setting_int(
            "sms_max_length",
            320
        )

        if len(message_text) > sms_max_characters:
            score -= 40
            notes.append(
                f"Template exceeds the maximum allowed SMS length of {sms_max_characters} characters."
            )


        elif len(message_text) > single_segment_warning:
            score -= 10
            notes.append(
                f"Message exceeds the recommended SMS length of {single_segment_warning} characters and may be delivered as multiple text messages."
            )

    elif channel == "email":
        notes.append(
            "Email compliance footer will be automatically appended."
        )


    for field in fields:
        normalized_field = field.replace(" ", "")

        if normalized_field not in normalized_message:
            score -= 10
            notes.append(f"⚠️ Missing recommended merge field: {field}")


  
    if score >= 90:
        risk = "Low"
    elif score >= 70:
        risk = "Medium"
    else:
        risk = "High"

    return {
        "score": max(score, 0),
        "review": "\n".join(notes),
        "risk_level": risk
    }













############################################################
#    template preview
# # TODO: Replace sample_data with build_merge_data() when Communications Engine is complete.
############################################################

def merge_template_preview(message_text, merge_data):


    sample = {
        "{{ client_first_name }}": "Sarah",
        "{{ service_name }}": "90 Minute Facial",
        "{{ appointment_date }}": "Tuesday, June 16",
        "{{ appointment_time }}": "2:00 PM",
        "{{ spa_name }}": "Clear Skin Esthetics",
        "{{ opt_out }}": "Reply STOP to opt out."
    }

    for key, value in sample.items():
        message_text = message_text.replace(key, value)

    return message_text





############################################################
#   SEND COMPLIANT SMS
#
#
#
#
############################################################

# ==========================================================
# PSP_REFACTOR          TODO
#
# Central SMS Communications Pipeline.
#
# All application SMS routes should call this function.
#
# Future Communications Engine enhancements:
#   - Verify SMS consent
#   - Load template
#   - Merge template fields
#   - AI compliance validation
#   - Append compliance footer
#   - Queue/schedule message
#   - Send through provider
#   - Log delivery and analytics
#
# Do NOT call send_sms_message() directly from routes.
# ==========================================================





def send_compliant_sms(spa_id, client_id, recipient_phone, message_body, message_type="manual"):
    """
    Central SMS sending pipeline:
    1. Send already-built compliant SMS body
    2. Return provider result with final message body
    """

    final_message = message_body

    result = send_sms_message(
        recipient_phone,
        final_message
    )

    result["client_id"] = client_id
    result["message_type"] = message_type
    result["final_message_body"] = final_message

    return result
















#####################################################################
#   >>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#
#   END OF SMS COMPLIANCE - AI - MESSAGING COMPLIANCE
#
#   <<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
#####################################################################











#####################################
#
#   GET SYSTEM SETTING
#
#
###################################



def get_system_setting(key, default=None):
    ...








#####################################
#
#   GET SET SYSTEM  SETTING
#
#
###################################



def set_system_setting(key, value):
    ...








#####################################
#
#   GET PLATFORM SETTING
#
#
###################################



def get_system_setting_int(setting_key, default_value):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT setting_value
        FROM system_settings
        WHERE setting_key = %s
        LIMIT 1
    """, (setting_key,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    if not row:
        return default_value

    try:
        return int(row[0])
    except (TypeError, ValueError):
        return default_value







#####################################
#
#   VALIDATE SMS LENGTH
#
#
###################################



def validate_sms_length(message_body):
    message_body = message_body or ""

    max_characters = get_system_setting_int(
        "sms_max_characters",
        320
    )

    if len(message_body) > max_characters:
        return {
            "valid": False,
            "error": f"SMS message is {len(message_body)} characters. Maximum allowed is {max_characters}."
        }

    return {
        "valid": True,
        "error": None
    }




















############################################################
# EXECUTIVE DASHBOARD DATA
############################################################

# ============================================================
# 🍑 MASTER DASHBOARD DATA HELPER
#
# Used by:
#   - Morning Briefing
#   - Business Summary
#   - Future Home Page
#   - Future Reports Center
#
# Add new dashboard metrics here.
# =======================================================
        
def get_dashboard_data(spa_id, spa_now):
    if spa_now is None:
        raise ValueError("get_dashboard_data requires spa_now")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    dashboard = {}
    
    from datetime import timedelta

    today = spa_now.date()
    current_time = spa_now.time().replace(tzinfo=None)
    tomorrow = today + timedelta(days=1)

    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    # Month boundaries
    month_start = today.replace(day=1)

    if month_start.month == 12:
        next_month_start = month_start.replace(
            year=month_start.year + 1,
            month=1
        )
    else:
        next_month_start = month_start.replace(
            month=month_start.month + 1
       )    

    
    cur.execute("""
    SELECT COUNT(*)
    FROM appointments
    WHERE appointment_date=%s
    AND spa_id=%s
    AND status IN ('booked','completed')
    """,(today,spa_id))  
 
    dashboard["appointments_today"] = cur.fetchone()[0] or 0

        # Expected revenue today
    cur.execute("""
        SELECT COALESCE(SUM(price_at_booking), 0)
        FROM appointments
        WHERE spa_id = %s
        AND appointment_date = CURRENT_DATE
        AND status NOT IN ('Cancelled', 'No Show')
    """, (spa_id,))

    expected_income_today = cur.fetchone()[0] or 0



    cur.execute("""
        SELECT COALESCE(SUM(price_at_booking), 0)
        FROM appointments
        WHERE appointment_date = %s
          AND spa_id = %s
          AND status IN ('booked', 'completed')
    """, (today, spa_id))

    dashboard["expected_revenue"] = cur.fetchone()[0] or 0


    cur.execute("""
        SELECT
            COALESCE(status, 'Unknown') AS status,
            COUNT(*) AS count
        FROM appointments
        WHERE spa_id = %s
        AND appointment_date >= date_trunc('month', CURRENT_DATE)
        AND appointment_date < date_trunc('month', CURRENT_DATE) + INTERVAL '1 month'
        GROUP BY status
        ORDER BY status
    """, (spa_id,))

    appointment_status_chart = [
        {
            "status": row[0],
            "count": row[1]
        }
        for row in cur.fetchall()
    ]

    cur.execute("""
        SELECT COUNT(*)
        FROM clients
        WHERE spa_id = %s
    """, (spa_id,))

    dashboard["total_clients"]  = cur.fetchone()[0] or 0
    dashboard["appointment_status_chart"] = appointment_status_chart
    dashboard["expected_income_today"] = expected_income_today
















    #####################################################
    
    #   TODAY

    ####################################################


        #####################################################
    #
    #   TODAY AND TOMORROW
    #
    #####################################################

    # Total appointments scheduled today
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE appointment_date = %s
          AND spa_id = %s
          AND LOWER(status) IN ('booked', 'completed')
    """, (today, spa_id))

    dashboard["appointments_today"] = cur.fetchone()[0] or 0


    # Appointments remaining today
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE appointment_date = %s
          AND appointment_time >= %s
          AND spa_id = %s
          AND LOWER(status) = 'booked'
    """, (today, current_time, spa_id))

    dashboard["appointments_remaining_today"] = (
        cur.fetchone()[0] or 0
    )


    # Expected revenue for all valid appointments today
    cur.execute("""
        SELECT COALESCE(SUM(price_at_booking), 0)
        FROM appointments
        WHERE appointment_date = %s
          AND spa_id = %s
          AND LOWER(status) NOT IN ('cancelled', 'no show')
    """, (today, spa_id))

    expected_income_today = cur.fetchone()[0] or 0

    dashboard["expected_income_today"] = expected_income_today
    dashboard["expected_revenue"] = expected_income_today


    # Expected revenue remaining today
    cur.execute("""
        SELECT COALESCE(SUM(price_at_booking), 0)
        FROM appointments
        WHERE appointment_date = %s
          AND appointment_time >= %s
          AND spa_id = %s
          AND LOWER(status) = 'booked'
    """, (today, current_time, spa_id))

    dashboard["expected_income_remaining_today"] = (
        cur.fetchone()[0] or 0
    )


    # Appointments tomorrow
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE appointment_date = %s
          AND spa_id = %s
          AND LOWER(status) IN ('booked', 'completed')
    """, (tomorrow, spa_id))

    dashboard["appointments_tomorrow"] = (
        cur.fetchone()[0] or 0
    )


    # Expected revenue tomorrow
    cur.execute("""
        SELECT COALESCE(SUM(price_at_booking), 0)
        FROM appointments
        WHERE appointment_date = %s
          AND spa_id = %s
          AND LOWER(status) NOT IN ('cancelled', 'no show')
    """, (tomorrow, spa_id))

    dashboard["expected_income_tomorrow"] = (
        cur.fetchone()[0] or 0
    )


    # Next appointment remaining today
    cur.execute("""
        SELECT
            a.appointment_time,
            c.first_name,
            c.last_name
        FROM appointments a
        JOIN clients c
          ON a.client_id = c.client_id
         AND a.spa_id = c.spa_id
        WHERE a.appointment_date = %s
          AND a.appointment_time >= %s
          AND a.spa_id = %s
          AND LOWER(a.status) = 'booked'
        ORDER BY a.appointment_time
        LIMIT 1
    """, (today, current_time, spa_id))

    dashboard["next_appointment"] = cur.fetchone()


    # Birthdays today
    cur.execute("""
        SELECT COUNT(*)
        FROM clients
        WHERE spa_id = %s
        AND birth_date IS NOT NULL
        AND EXTRACT(MONTH FROM birth_date) = %s
        AND EXTRACT(DAY FROM birth_date) = %s
    """, (
        spa_id,
        today.month,
        today.day
    ))

    dashboard["birthdays_today"] = cur.fetchone()[0] or 0



    # Completed appointments today
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE appointment_date=%s
          AND spa_id=%s
          AND status='completed'
    """,(today,spa_id))

    dashboard["completed_today"] = cur.fetchone()[0] or 0



    
 

    # Cancelled appointments today
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE appointment_date=%s
          AND spa_id=%s
          AND status='cancelled'
    """,(today,spa_id))

    dashboard["cancelled_today"] = cur.fetchone()[0] or 0


    # No shows today
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE appointment_date=%s
          AND spa_id=%s
          AND status='no show'
    """,(today,spa_id))

    dashboard["no_shows_today"] = cur.fetchone()[0] or 0




    



     # Daily Revenue
    cur.execute("""
        SELECT COALESCE(SUM(price_at_booking), 0)
        FROM appointments
        WHERE appointment_date = %s
        AND status = 'completed'
        AND spa_id = %s
    """, (today, spa_id))

    dashboard["daily_revenue"] = cur.fetchone()[0] or 0   





   #####################################################
    
    #   WEEK
    
   #################################################### 

    # Weekly Appointment Totals
    cur.execute("""
        SELECT
            COUNT(*) AS total_appointments,
            COUNT(*) FILTER (WHERE status = 'booked') AS booked_count,
            COUNT(*) FILTER (WHERE status = 'completed') AS completed_count,
            COUNT(*) FILTER (WHERE status = 'cancelled') AS cancelled_count
        FROM appointments
        WHERE appointment_date BETWEEN %s AND %s
            AND spa_id = %s
    """, (week_start, week_end, spa_id))

    weekly_totals = cur.fetchone() or (0, 0, 0, 0)

    dashboard["weekly_totals"] = weekly_totals
    dashboard["weekly_total_appointments"] = weekly_totals[0]
    dashboard["weekly_booked_count"] = weekly_totals[1]
    dashboard["weekly_completed_count"] = weekly_totals[2]
    dashboard["weekly_cancelled_count"] = weekly_totals[3]



    # Weekly revenue
    cur.execute(f"""
        SELECT COALESCE(SUM(a.price_at_booking), 0)
        FROM appointments a
        WHERE a.appointment_date BETWEEN %s AND %s
          AND a.status = 'completed'
          AND a.spa_id = %s
    """, [week_start, week_end, spa_id])
    weekly_revenue = cur.fetchone()[0] or 0

    dashboard["weekly_revenue"] = weekly_revenue


    #####################################################
    
    #   MONTH
    
    #################################################### 

    # No Shows This Month
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE appointment_date >= %s
          AND appointment_date < %s
          AND status = 'no show'
          AND spa_id = %s
        """, (
        month_start,
        next_month_start,
        spa_id
    ))

    dashboard["no_shows_month"] = cur.fetchone()[0] or 0    



    # New Clients This Month
    cur.execute("""
        SELECT COUNT(*)
        FROM clients
        WHERE spa_id = %s
        AND created_at >= %s
        AND created_at < %s
    """, (spa_id, month_start, next_month_start))

    dashboard["new_clients_month"] = cur.fetchone()[0] or 0


    # Monthly revenue
    cur.execute(f"""
        SELECT COALESCE(SUM(a.price_at_booking), 0)
        FROM appointments a
        WHERE a.appointment_date >= %s
        AND a.appointment_date < %s
        AND a.status = 'completed'
        AND a.spa_id = %s
    """, (month_start, next_month_start, spa_id))

    monthly_revenue = cur.fetchone()[0] or 0

    dashboard["monthly_revenue"] = monthly_revenue



    # Average ticket this month
    cur.execute("""
        SELECT COALESCE(AVG(a.price_at_booking), 0)
        FROM appointments a
        WHERE a.appointment_date >= %s
        AND a.appointment_date < %s
        AND a.status = 'completed'
        AND a.price_at_booking IS NOT NULL
        AND a.spa_id = %s
    """, (month_start, next_month_start, spa_id))

    average_ticket = cur.fetchone()[0] or 0

    dashboard["average_ticket"] = average_ticket






    #####################################################
        
    #   ALL TIME.... NOT WEEK, MONTH OR YEAR
          
    ####################################################

 

    # Cancelled Appointments (All Time)
    cur.execute("""
    SELECT COUNT(*)
    FROM appointments
    WHERE status = 'cancelled'
      AND spa_id = %s
    """, (spa_id,))

    dashboard["cancelled_count"] = cur.fetchone()[0] or 0


 

    #####################################################
    
    #   ALERTS
    
    #################################################### 
 

    ############################################################
    # BUSINESS HEALTH
    ############################################################


    dashboard["business_health_score"] = 0
    dashboard["business_health_label"] = "Temp Data - Not Calculated"
    dashboard["business_health_class"] = "kpi-neutral"



    cur.close()
    conn.close()
    
    return dashboard
    
    






#  -------
#
# Get Statis
# spa_id good
#  -------

def get_status_id(status_name):
    from flask import g

    spa_id = g.spa_id

    conn = get_db_connection()
    cur = conn.cursor()
        
    cur.execute("""   
        SELECT gift_certificate_status_id
        FROM gift_certificate_statuses
        WHERE status_name = %s
          AND spa_id = %s
    """, (status_name, spa_id))
    
    result = cur.fetchone()
    
    cur.close()
    conn.close()
    
    return result[0] if result else None



#   ----------------
#
#
#   --------------


def parse_bool(value):
    if value is None:
        return None

    value = str(value).strip().lower()

    if value in ("true", "yes", "y", "1", "on"):
        return True
    if value in ("false", "no", "n", "0", "off"):
        return False
    if value in ("none", "", "null"):
        return None

    return None
#   -----------------
#
#   ----------------

@app.context_processor
def inject_spa_name():
    try:
        spa_id = current_spa_id()
        if not spa_id:
            return dict(spa_name="")

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT spa_name
            FROM spas
            WHERE spa_id = %s
        """, (spa_id,))

        spa = cur.fetchone()

        cur.close()
        conn.close()

        return dict(spa_name=spa[0] if spa else "")

    except:
        return dict(spa_name="")







#   --------------------------
#
#  GODADDY IMPORT ALERT
#
#   --------------------------

@app.context_processor
def inject_godaddy_import_alert():
    if "user_id" not in session or "spa_id" not in session:
        return {}

    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE spa_id = %s
          AND external_source = 'godaddy'
          AND COALESCE(import_reviewed, FALSE) = FALSE
    """, (spa_id,))

    godaddy_unreviewed_count = cur.fetchone()[0]

    cur.close()
    conn.close()

    return {
        "godaddy_unreviewed_count": godaddy_unreviewed_count
    }





####################################
###################################


#   -------------------------- 
#   
#    PROGRESS  PERCENT   
#
#   -------------------------- 


def progress_percent(actual, goal):
    actual = float(actual or 0)
    goal = float(goal or 0)

    if goal <= 0:
        return 0

    return min(round((actual / goal) * 100, 1), 150)














#   --------------------------
#
#
#
#   --------------------------

def split_client_name(full_name):
    if not full_name:
        return "", ""

    parts = full_name.strip().split()

    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])



#   ----------------------------
#
#   CURRENT SPA ID
#
#   -----------------------

def current_spa_id():
    from flask import g, session

    # Not logged in yet
    if "user_id" not in session:
        return None

    # Master admin is not tied to one spa
    if session.get("role") == "master_admin":
        return None

    return getattr(g, "spa_id", None)



def get_spa_today():
    return get_spa_now().date()


def get_spa_current_time():
    return get_spa_now().time()


#   -------------------------
#
#
#   -------------------------        



##########################################
#ENGLISH   EMAIL_UNSUBSCRIBE_FOOTER = """
#SPANISH
#########################################

EMAIL_UNSUBSCRIBE_FOOTER_EN = """

----------------------------------------

You are receiving emails from Peach Suite Pro.

To unsubscribe from future marketing emails,
reply UNSUBSCRIBE or contact our office directly.
"""

EMAIL_UNSUBSCRIBE_FOOTER_ES = """

----------------------------------------

Usted está recibiendo correos electrónicos de Peach Suite Pro.

Para cancelar la suscripción a futuros correos electrónicos de marketing,
responda UNSUBSCRIBE o comuníquese directamente con nuestra oficina.
"""



def add_email_footer(body, language="English"):

    body = (body or "").strip()

    if "unsubscribe" in body.lower():
        return body

    if language == "Spanish":
        return f"{body}{EMAIL_UNSUBSCRIBE_FOOTER_ES}"

    return f"{body}{EMAIL_UNSUBSCRIBE_FOOTER_EN}"







        
    
#   -------------------------
#   AUDIT HELPERS
#
#   -------------------------
    
def log_audit(
    cur,
    spa_id,
    user_id,
    action_type,
    table_name=None,
    record_id=None,
    old_value=None,
    new_value=None,
    notes=None
):
    cur.execute("""
        INSERT INTO audit_log (
            spa_id,
            user_id,
            action_type,
            table_name,
            record_id,
            old_value,
            new_value,
            notes
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        spa_id,
        user_id,
        action_type,
        table_name,
        record_id,
        old_value,
        new_value,
        notes
    ))


def log_appointment_history(
    cur,
    spa_id,
    appointment_id,
    client_id,
    user_id,
    action_type,
    old_date=None,
    old_time=None,
    new_date=None,
    new_time=None,
    old_status=None,
    new_status=None,
    notes=None
):
    cur.execute("""
        INSERT INTO appointment_history (
            spa_id,
            appointment_id,
            client_id,
            user_id,
            action_type,
            old_date,
            old_time,
            new_date,
            new_time,
            old_status,
            new_status,
            notes
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (
        spa_id,
        appointment_id,
        client_id,
        user_id,
        action_type,
        old_date,
        old_time,
        new_date,
        new_time,
        old_status,
        new_status,
        notes
    ))
        


      






  
#   -------------------------
#
#
#   -------------------------




def get_spa_name(spa_id):
    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT spa_name
            FROM spas
            WHERE spa_id = %s
        """, (spa_id,))
        
        spa_row = cur.fetchone()
        return spa_row[0] if spa_row else "Your Spa"

    finally:
        cur.close()
        conn.close()


#   ---------------------------
#
#     LOGIN HELPERS
#
#
#   ---------------------------


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in first.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


        
#   -------------------------
#
#
#   -------------------------




def master_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != "master_admin":
            flash("Access denied.", "error")
            return redirect(url_for("home"))
        return f(*args, **kwargs)
    return decorated_function









#   ---------------------------
#
#      ALLOWED USERS
#
#       
#   ---------------------------
        
ALLOWED_USER_ROLES = ["master_admin", "admin", "manager", "staff"]


def clean_user_role(role):
    if role not in ALLOWED_USER_ROLES:
        return "staff"
    return role


def require_master_admin():
    if session.get("role") != "master_admin":
        abort(403)


def require_admin_or_master():
    if session.get("role") not in ["admin", "master_admin"]:
        abort(403)


def current_user_role():
    return session.get("role")








#   ---------------------------
#
#      DEF  GET ACCOUNTING 
#
#
#   ---------------------------

def get_accounting_ytd_summary(spa_id):


    cur.execute("""
        SELECT COALESCE(SUM(total_amount), 0)
        FROM income
        WHERE spa_id = %s
          AND income_date BETWEEN %s AND %s
    """, (spa_id, year_start, today))

    ytd_income = cur.fetchone()[0]


    cur.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM expenses
        WHERE spa_id = %s
          AND expense_date BETWEEN %s AND %s
    """, (spa_id, year_start, today))

    ytd_expenses = cur.fetchone()[0]



    # Birthday alert count
    # Adjust this query if your birthday table/logic differs
    cur.execute("""
        SELECT COUNT(*)
        FROM client_birthday_offers cbo
        JOIN clients c 
            ON c.client_id = cbo.client_id
           AND c.spa_id = cbo.spa_id
        WHERE c.spa_id = %s
          AND cbo.offer_sent = FALSE
          AND cbo.birthday_year = %s
    """, (spa_id, today.year))
    birthday_alert_count = cur.fetchone()[0] or 0

    # Expiring gift certificate count
    cur.execute("""
        SELECT COUNT(*)
        FROM gift_certificates gc 
        JOIN gift_certificate_statuses gcs
        ON gc.gift_certificate_status_id = gcs.gift_certificate_status_id
        WHERE gc.spa_id = %s
        AND gcs.status_name = 'Active'
        AND gc.amount_paid > 0
        AND gc.is_redeemed = FALSE
        AND gc.remaining_balance > 0
        AND gc.expires_date BETWEEN %s AND (%s + INTERVAL '60 days')
    """, (spa_id, today, today))
    expiring_gc_count = cur.fetchone()[0] or 0

    return {
    
        "ytd_income": ytd_income,
        "ytd_expenses": ytd_expenses,
    }



#   ---------------------------
#
#      DEF GET EMPLOYEE COMPENSATION
#
#   ---------------------------


#  get_employee_compensation_ytd_summary(spa_id)


    cur.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM employee_compensation
        WHERE spa_id = %s
          AND compensation_date BETWEEN %s AND %s
    """, (spa_id, year_start, today))

    ytd_employee_compensation = cur.fetchone()[0]

    return {
        "ytd_employee_compensation": ytd_employee_compensation
    }






#   ---------------------------
#
#      SMS EMAIL CONSENT RECORD
#
#
#   ---------------------------

def add_consent_record(
    cur,
    spa_id,
    client_id,
    consent_type,
    consent_status,
    consent_source="Admin Updated",
    consent_note=None,
    updated_by=None
):
    cur.execute("""
        INSERT INTO consent_records (
            spa_id,
            client_id,
            consent_type,
            consent_status,
            consent_source,
            consent_note,
            updated_by
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        spa_id,
        client_id,
        consent_type,
        consent_status,
        consent_source,
        consent_note,
        updated_by
    ))








#  ---------------
#  LOAD SPA
#
#
#
#
#
#
#
#  --------------


@app.before_request
def load_spa():

    if request.endpoint in (
        "login",  
        "logout", 
        "static",
        "mailgun_godaddy_booking",
        "godaddy_booking_intake",
        "telnyx_sms_webhook"

     ):
        return

    if "user_id" not in session:
        return redirect(url_for("login"))

    if session.get("role") == "master_admin":
        g.spa_id = None
        return

    spa_id = session.get("spa_id")

    if not spa_id:
        session.clear()
        return redirect(url_for("login"))

    g.spa_id = spa_id





#   ----------------------
#
#    DEF SMS EMAIL TERMS ACCEPTED
#
#   ---------------------

def sms_email_terms_accepted(spa_id):

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT owner_agreed_sms_email_terms
        FROM spas
        WHERE spa_id = %s
    """, (spa_id,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    return bool(row and row[0])



        
#   ----------------------
#       
#    SMS PLACEHOLDERS
#    
#   ---------------------
    
def apply_sms_placeholders(message, data):
    if not message:
        return ""

    for key, value in data.items():
        message = message.replace(
            "{" + key + "}",
            str(value or "")
        )

    return message







#   ----------------------
#
#    DEF SPA REQUIRES
#
#   ---------------------


from functools import wraps

def spa_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):

        if session.get("role") == "master_admin":
            return f(*args, **kwargs)

        spa_id = current_spa_id()

        if not spa_id:
            flash("No spa is assigned to this user.", "error")
            return redirect(url_for("login"))

        return f(*args, **kwargs)

    return decorated_function




    
@app.context_processor
def inject_spa():
    return dict(spa_id=g.get("spa_id"))





@app.context_processor
def inject_current_spa():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT spa_name
            FROM spas
            WHERE spa_id = %s
        """, (spa_id,))
        row = cur.fetchone()
        spa_name = row[0] if row else "Your Spa"

    finally:
        cur.close()
        conn.close()

    return {
        "active_spa_id": spa_id,
        "active_spa_name": spa_name
    }









#  ---------------
#   ERROR HANDLERS
#  --------------


@app.errorhandler(404)
def page_not_found(error):
    return render_template("404.html"), 404




@app.errorhandler(500)
def internal_server_error(error):
    return render_template("500.html"), 500







#  -------------------------
#  
#     SAVE TIME ZONE SETTINGS
#  
#    spa_id good
#
#  -------------------------


@app.route("/save_time_settings", methods=["POST"])
@login_required
@spa_required  

def save_time_settings():
    spa_id = current_spa_id()
    
    timezone_name = request.form.get("timezone_name", "").strip()
    
    allowed_timezones = {
        "America/New_York",
        "America/Chicago",
        "America/Denver",
        "America/Los_Angeles",
        "America/Anchorage",
        "Pacific/Honolulu",
        "America/Toronto",
        "America/Vancouver",
        "America/Edmonton",
        "America/Halifax",
        "Europe/London",
        "Europe/Dublin",
        "Europe/Paris",
        "Europe/Berlin",
        "Europe/Madrid",
        "Europe/Rome",
        "Australia/Sydney",
        "Australia/Perth",
        "Asia/Tokyo",
        "Asia/Singapore",
        "Asia/Dubai"
    }

    if timezone_name not in allowed_timezones:
        flash("Invalid timezone selected.", "error")
        return redirect(url_for("admin"))

    conn = get_db_connection()   
    cur = conn.cursor()
    
    cur.execute("""
        UPDATE spas
        SET timezone_name = %s
        WHERE spa_id = %s
    """, (timezone_name, spa_id))
        
    conn.commit()
    cur.close()
    conn.close()
        
    flash("Time settings updated successfully.", "success")
    return redirect(url_for("admin"))




######################################
#
#   TIMEZONE HELPER
#
#######################################

def get_spa_timezone(spa_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT timezone_name
        FROM spas
        WHERE spa_id = %s
        LIMIT 1
    """, (spa_id,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    return row[0] if row and row[0] else "UTC"












###############################################
#
#       SYSTEM ACTIVITY LOG
#
#
################################################



@app.route("/admin/system-activity")
@login_required
@spa_required
def system_activity():
    page = request.args.get("page", 1, type=int)
    per_page = 100
    offset = (page - 1) * per_page

    category = request.args.get("category", "").strip()
    severity = request.args.get("severity", "").strip()
    date_filter = request.args.get("date_filter", "today").strip()
    search = request.args.get("search", "").strip()

    where_clauses = []
    params = []

    if category:
        where_clauses.append("category = %s")
        params.append(category)

    if severity:
        where_clauses.append("severity = %s")
        params.append(severity)

    if search:
        where_clauses.append("message ILIKE %s")
        params.append(f"%{search}%")

    if date_filter == "today":
        where_clauses.append("created_at::date = CURRENT_DATE")
    elif date_filter == "yesterday":
        where_clauses.append("created_at::date = CURRENT_DATE - INTERVAL '1 day'")
    elif date_filter == "7":
        where_clauses.append("created_at >= NOW() - INTERVAL '7 days'")
    elif date_filter == "30":
        where_clauses.append("created_at >= NOW() - INTERVAL '30 days'")

    where_sql = ""
    if where_clauses:
        where_sql = "WHERE " + " AND ".join(where_clauses)

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    cur.execute(f"""
        SELECT COUNT(*)
        FROM system_logs
        {where_sql}
    """, params)

    cur.execute(f"""
        SELECT COUNT(*) AS total
        FROM system_logs
        {where_sql}
    """, params)

    total_logs = cur.fetchone()["total"]
    total_pages = max((total_logs + per_page - 1) // per_page, 1)

    cur.execute(f"""
        SELECT
            log_id,
            category,
            severity,
            message,
            created_at
        FROM system_logs
        {where_sql}
        ORDER BY created_at DESC
        LIMIT %s OFFSET %s
    """, params + [per_page, offset])

    logs = cur.fetchall()

    cur.execute("""
        SELECT DISTINCT category
        FROM system_logs
        WHERE category IS NOT NULL
        ORDER BY category
    """)
    categories = [row["category"] for row in cur.fetchall()]

    cur.close()
    conn.close()

    return render_template(
        "system_activity.html",
        logs=logs,
        categories=categories,
        page=page,
        total_pages=total_pages,
        total_logs=total_logs,
        category=category,
        severity=severity,
        date_filter=date_filter,
        search=search
    )

























#   ----------------------------------
#
#     SWITCH ROUTE
#
#    this will switch users
#
#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#
#    DELETE  DELETE  DELETE AFTER TESTNG
#
#
#    DELETE  DELETE  DELETE AFTER TESTING
#   ---------------------------------

@app.route("/switch-spa/<int:spa_id>")
@login_required
@master_admin_required
def switch_spa(spa_id):

    if not spa_exists(spa_id):
        flash("Spa not found.", "warning")
        return redirect(url_for("home"))

    session["spa_id"] = spa_id

    flash(f"Development mode: switched to spa {spa_id}.", "info")

    return redirect(url_for("home"))

#  -------------------------
#
#   TELNYX WEBHOOK
#                  
#  -------------------------


@app.route("/webhooks/telnyx/sms", methods=["POST"])
def telnyx_sms_webhook():
    from services.telnyx_webhook import process_telnyx_webhook
    return process_telnyx_webhook()




#  -------------------------
#
#  SQUARE WEBHOOK
#
#
# Not production-safe until signature validation, idempotency,
# spa/location mapping, and duplicate protection are complete.
#  ------------------
#  -------------------------

@app.route("/square/webhook", methods=["POST"])
def square_webhook():

    return "", 204

    payload = request.get_data(as_text=True)
    signature = request.headers.get("x-square-hmacsha256-signature")

    # validate signature here

    event = request.get_json()

    # look for booking.created
    # parse booking id and details
    # insert into incoming_square_bookings if not already present

    return "", 200






#  -------------------
#   Square Incomming Bookings
# SQUARE_INTEGRATION_STAGING
#
#
# Not production-safe until signature validation, idempotency,
# spa/location mapping, and duplicate protection are complete.
#  ------------------

@app.route("/incoming_bookings")
def incoming_bookings():

    return "", 204

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            incoming_booking_id,
            square_booking_id,
            client_name,
            client_email,
            client_phone,
            appointment_date,
            appointment_time,
            service_name,
            status,
            created_at
        FROM incoming_square_bookings
        ORDER BY
            CASE WHEN status = 'new' THEN 0 ELSE 1 END,
            appointment_date,
            appointment_time,
            created_at DESC
    """)
    bookings = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("incoming_bookings.html", bookings=bookings)


#  -------------------------------------
#   Square Incoming Booking Review
#
#
# Not production-safe until signature validation, idempotency,
# spa/location mapping, and duplicate protection are complete.
#  ------------------
#  ------------------------------------

@app.route("/incoming_bookings/<int:incoming_booking_id>")
def review_incoming_booking(incoming_booking_id):
    
    return "", 204

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            incoming_booking_id,
            square_booking_id,
            client_name,
            client_email,
            client_phone,
            appointment_date,
            appointment_time,
            service_name,
            status,
            raw_payload,
            created_at
        FROM incoming_square_bookings
        WHERE incoming_booking_id = %s
    """, (incoming_booking_id,))
    booking = cur.fetchone()

    cur.close()
    conn.close()

    if not booking:
        flash("Incoming booking not found.", "error")
        return redirect(url_for("incoming_bookings"))

    return render_template("review_incoming_booking.html", booking=booking)




#  ----------------------------------
#
#  SQUARE Ignore incoming 
#
#
# Not production-safe until signature validation, idempotency,
# spa/location mapping, and duplicate protection are complete.
#  ------------------
#  ----------------------------------

@app.route("/incoming_bookings/<int:incoming_booking_id>/ignore", methods=["POST"])
def ignore_incoming_booking(incoming_booking_id):
    
    return "", 204

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE incoming_square_bookings
        SET status = 'ignored',
            reviewed_at = CURRENT_TIMESTAMP
        WHERE incoming_booking_id = %s
    """, (incoming_booking_id,))

    conn.commit()
    cur.close()
    conn.close()

    flash("Incoming booking marked as ignored.", "success")
    return redirect(url_for("incoming_bookings"))



#  ------------------------------
#
#  SQUARE ADD NEW CLIENT 
#
#
# Not production-safe until signature validation, idempotency,
# spa/location mapping, and duplicate protection are complete.
#  ------------------#  NOT SAFE    NOT  SAFE
#  -----------------------------
  

@app.route("/incoming_bookings/<int:incoming_booking_id>/add_new_client")
def add_new_client_from_booking(incoming_booking_id):
    
    return "", 204

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT *
            FROM incoming_bookings
            WHERE incoming_booking_id = %s
        """, (incoming_booking_id,))

        booking = cur.fetchone()

        if not booking:
            flash("Booking record not found.", "error")
            return redirect(url_for("home"))

        return render_template(
            "add_new_client.html",
            booking=booking
        )

    except Exception as e:
        flash(f"Error loading booking: {str(e)}", "error")
        return redirect(url_for("home"))

    finally:
        cur.close()
        conn.close()

    cur.execute("""
        SELECT
            incoming_booking_id,
            client_name,
            client_email,
            client_phone,
            appointment_date,
            appointment_time,
            service_name
        FROM incoming_square_bookings
        WHERE incoming_booking_id = %s
    """, (incoming_booking_id,))
    booking = cur.fetchone()

    cur.close()
    conn.close()

    if not booking:
        flash("Incoming booking not found.", "error")
        return redirect(url_for("incoming_bookings"))

    first_name, last_name = split_client_name(booking[1])

    session["incoming_booking_data"] = {
        "incoming_booking_id": booking[0],
        "first_name": first_name,
        "last_name": last_name,
        "email": booking[2] or "",
        "phone": booking[3] or "",
        "appointment_date": booking[4].strftime("%Y-%m-%d") if booking[4] else "",
        "appointment_time": booking[5].strftime("%H:%M:%S") if booking[5] else "",
        "service_name": booking[6] or ""
    }

    return redirect(url_for("add_new_client"))






#  ------------------
#
#  SQUARE MATCH EXISTING
#
#
# Not production-safe until signature validation, idempotency,
# spa/location mapping, and duplicate protection are complete.
#  ------------------#     NOT SAFE  NOT SAFE
#  ------------------



app.route("/incoming_bookings/<int:incoming_booking_id>/match_existing_client")
def match_existing_client_booking(incoming_booking_id):
    flash("Next step: choose existing client and create appointment.", "info")
    return redirect(url_for("review_incoming_booking", incoming_booking_id=incoming_booking_id))




#   --------------------------------------------------------------
#
#
#
#              MAILGUN   EMAIL 
#
#       this is good.... checked on 6/17/26
#   ----------------------------------------------------------


import os
import requests

def send_email(to, subject, body):

    final_body = add_email_footer(body)

    mailgun_api_key = os.environ.get("MAILGUN_API_KEY", "").strip()
    mailgun_domain = os.environ.get("MAILGUN_DOMAIN", "").strip()
    mailgun_from = os.environ.get("MAILGUN_FROM", "").strip()

    response = requests.post(
        f"https://api.mailgun.net/v3/{mailgun_domain}/messages",
        auth=("api", mailgun_api_key),
        data={
            "from": mailgun_from,
            "to": [to],
            "subject": subject,
            "text": final_body
        },
        timeout=20
    )

    return response



#  ----------------------
#
#   DATABASE HOME
#
#  ----------------------

@app.route("/")
@login_required
@spa_required  

def home():
    return render_template("home.html")









#   ---------------------------
#
#
#    DROP DOWN CONFIG
#
#
#   ---------------------------



DROPDOWN_CONFIG = {
    "skin_types": {
        "title": "Skin Types",
        "table": "skin_types",
        "pk": "skin_type_id",
        "value": "skin_type_name",
        "label": "Skin Type Name",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "skin_type_name"
    },

    "referral_sources": {
        "title": "Referral Sources",
        "table": "referral_sources",
        "pk": "referral_source_id",
        "value": "referral_source_name",
        "label": "Referral Source Name",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "referral_source_name"
    },

    "fitzpatrick_types": {
        "title": "Fitzpatrick Types",
        "table": "fitzpatrick_types",
        "pk": "fitzpatrick_id",
        "value": "fitzpatrick_level",
        "extra_value": "description",
        "label": "Fitzpatrick Level",
        "extra_label": "Description",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "fitzpatrick_level"
    },


    "appointment_status": {
        "title": "Appointment Status",
        "table": "appointment_status",
        "pk": "status_id",
        "value": "status_name",
        "label": "Status Name",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "status_name"
    },


    "client_form_names": {
        "title": "Client Form Names",
        "table": "client_form_names",
        "pk": "form_type",   # <-- fix this
        "value": "form_type_name",
        "label": "Form Type Name",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "form_type_name"
    },


#  ---------
#  clean up the following
#   --------


    "expense_categories": {
        "title": "Expense Categories",
        "table": "expense_categories",
        "pk": "expense_cat_id",
        "value": "expense_cat_name",
        "label": "Expense Category Name",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "expense_cat_name"
    },

    "income_types": {
        "title": "Income Types",
        "table": "income_types",
        "pk": "income_type_id",
        "value": "income_type_name",
        "label": "Income Type Name",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "income_type_name"
    },

    "payment_methods": {
        "title": "Payment Methods",
        "table": "payment_methods",
        "pk": "payment_method_id",
        "value": "payment_method",
        "label": "Payment Method",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "payment_method"
    },

    "service_name_types": {
        "title": "Service Name Types",
        "table": "service_name_types",
        "pk": "service_type_id",
        "value": "service_name",
        "label": "Service Name",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "service_name" 
    },

    "sex": {
        "title": "Sex",
        "table": "sex",
        "pk": "sex_type_id",
        "value": "sex_type",
        "label": "Sex Type",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "sex_type"
    },

    "treatment_rooms": {
        "title": "Treatment Rooms",
        "table": "treatment_rooms",
        "pk": "room_id",
        "value": "room_name",
        "label": "Room Name",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "room_name"
    },

    "vendor_name": {
        "title": "Vendor Name",
        "table": "vendor_name",
        "pk": "vendor_id",
        "value": "vendors_name",
        "label": "Vendor Name",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "vendors_name"
    },

    "compensation_types": {
        "title": "Compensation Types",
        "table": "compensation_types", 
        "pk": "compensation_type_id",
        "value": "compensation_type_name",
        "label": "Compensation Type",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "compensation_type_name"
    },

    "spa_locations": {
        "title": "Spa Locations",
        "table": "spa_locations",
        "pk": "spa_location_id",
        "value": "location_name",
        "label": "Spa Locations",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "location_name"
    },

    "client_statuses": {
        "title": "Client Statuses",     
        "table": "client_statuses",     
        "pk": "client_status_id",
        "value": "status_name",
        "label": "Client Statuses",    
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "status_name"
    },

    "preferred_contact_methods": {
        "title": "Preferred Contact Methods",
        "table": "preferred_contact_methods",
        "pk": "preferred_contact_method_id",
        "value": "method_name", 
        "label": "Preferred Contact Methods",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "method_name"
    },

    "preferred_languages": {
        "title": "Preferred Languages",
        "table": "preferred_languages",
        "pk": "preferred_language_id",
        "value": "language_name",   
        "label": "Preferred Languages",
        "spa_scoped": True,
        "active_column": "is_active",
        "order_by": "language_name"
    }



}


#   ------------------------------
#
#  Drop Down Helper
#
#
#
#   ------------------------------
   

def get_dropdown_options(config_key, spa_id):
    config = DROPDOWN_CONFIG[config_key]

    table = config["table"]
    pk = config["pk"]
    value = config["value"]
    order_by = config.get("order_by", value)

    query = f"""
        SELECT {pk}, {value}
        FROM {table}
    """

    conditions = []
    params = []

    if config.get("spa_scoped"):
        conditions.append("spa_id = %s")
        params.append(spa_id)

    if config.get("active_column"):
        conditions.append(f"{config['active_column']} = TRUE")

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    query += f" ORDER BY {order_by}"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(query, params)
    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows












#  -------------------------------
#           DROP DOWNS
#
#
#  APPOINTMENT STATUS
#  CLIENT FORM NAME  
#  EXPENSE CATEGORIES  
#  INCOME TYPES
#  PAYMENT METHODS
#  SERVICE NAME TYPES
#  SEX
#  TREATMENT ROOMS
#  VENDOR NAMES
#
#
#  --------------------------------







#   ----------------------------
#
#     ADMIN DROPDOWNS
#
#   spa_id good
#
#   -----------------------------



@app.route("/dropdowns/<dropdown_key>", methods=["GET", "POST"])
@login_required
@spa_required
def manage_dropdown(dropdown_key):
    spa_id = current_spa_id()

    config = DROPDOWN_CONFIG.get(dropdown_key)

    if not config:
        flash("Invalid dropdown selected.", "error")
        return redirect(url_for("admin"))

    table = config["table"]
    pk = config["pk"]
    value_col = config["value"]
    extra_col = config.get("extra_value")
    spa_scoped = config.get("spa_scoped", True)

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT dropdown_key, display_label
        FROM spa_dropdown_labels
        WHERE spa_id = %s
    """, (spa_id,))

    label_rows = cur.fetchall()
    dropdown_labels = {row[0]: row[1] for row in label_rows}
    current_label = dropdown_labels.get(dropdown_key, config["title"])

    if request.method == "POST":
        value = request.form.get("value", "").strip()
        extra_value = request.form.get("extra_value", "").strip() if extra_col else None

        if not value:
            flash(f"{config['label']} is required.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("manage_dropdown", dropdown_key=dropdown_key))

        if extra_col:
            cur.execute(f"""
                INSERT INTO {table} (spa_id, {value_col}, {extra_col})
                VALUES (%s, %s, %s)
            """, (spa_id, value, extra_value))
        else:
            cur.execute(f"""
                INSERT INTO {table} (spa_id, {value_col})
                VALUES (%s, %s)
            """, (spa_id, value))

        conn.commit()
        cur.close()
        conn.close()

        flash(f"{config['title']} item added.", "success")
        return redirect(url_for("manage_dropdown", dropdown_key=dropdown_key))

    if extra_col:
        cur.execute(f"""
            SELECT {pk}, {value_col}, {extra_col}
            FROM {table}
            WHERE spa_id = %s
            ORDER BY {value_col}
        """, (spa_id,))
    else:
        cur.execute(f"""
            SELECT {pk}, {value_col}
            FROM {table}
            WHERE spa_id = %s
            ORDER BY {value_col}
        """, (spa_id,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "manage_dropdown.html",
        dropdown_key=dropdown_key,
        config=config,
        rows=rows,
        dropdown_labels=dropdown_labels
    )








#  --------------------------------
#
#     DROP DOWNS  DELETE FUNCTION
#
# ROUTE:  admin/dropdowns
#           spa_id good
#  ------------------------------


@app.route("/dropdowns/<dropdown_key>/delete/<int:item_id>", methods=["POST"])
@login_required
@spa_required
def delete_dropdown_item(dropdown_key, item_id):
    spa_id = current_spa_id()

    config = DROPDOWN_CONFIG.get(dropdown_key)

    if not config:
        flash("Invalid dropdown selected.", "error")
        return redirect(url_for("admin"))

    table = config["table"]
    pk = config["pk"]

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute(f"""
        DELETE FROM {table}
        WHERE {pk} = %s
          AND spa_id = %s
    """, (item_id, spa_id))

    if cur.rowcount == 0:
        flash("Item not found or not authorized.", "error")
    else:
        flash("Item deleted.", "success")

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("manage_dropdown", dropdown_key=dropdown_key))





#  ---------------------
#     
#     UPDATE DROPDOWN LABELS
#  
#  -----------------


@app.route("/update_dropdown_labels", methods=["POST"])
@login_required
@spa_required
def update_dropdown_labels():
    spa_id = current_spa_id()

    skin_types = request.form.get("skin_types", "").strip()
    fitzpatrick_types = request.form.get("fitzpatrick_types", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    labels = [
        ("skin_types", skin_types),
        ("fitzpatrick_types", fitzpatrick_types)
    ]

    for key, label in labels:

        cur.execute("""
            INSERT INTO spa_dropdown_labels (
                spa_id,
                dropdown_key,
                display_label
            )
            VALUES (%s, %s, %s)

            ON CONFLICT (spa_id, dropdown_key)
            DO UPDATE SET
                display_label = EXCLUDED.display_label
        """, (spa_id, key, label))

    conn.commit()

    cur.close()
    conn.close()

    flash("Dropdown labels updated.", "success")

    return redirect(url_for("admin"))









#############################
#
#   SERVICE TYPES LIST
#
#
########################################

@app.route("/service-types")
@login_required
@spa_required
def service_types():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            service_type_id,
            service_name,
            default_duration_minutes,
            default_price,
            is_active
        FROM service_name_types
        WHERE spa_id = %s
        ORDER BY service_name
    """, (spa_id,))

    services = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "service_types.html",
        services=services
    )





#################################
#
#
#   ADD SERVICE TYPE 
#
#
####################################

@app.route("/service-types/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_service_type():
    spa_id = current_spa_id()

    if request.method == "POST":
        service_name = (
            request.form.get("service_name") or ""
        ).strip()

        duration_raw = (
            request.form.get("default_duration_minutes") or ""
        ).strip()

        price_raw = (
            request.form.get("default_price") or ""
        ).strip()

        is_active = request.form.get("is_active") == "1"

        if not service_name:
            flash("Service name is required.", "error")
            return redirect(url_for("add_service_type"))

        try:
            default_duration_minutes = int(duration_raw)
        except (TypeError, ValueError):
            flash("Default session length must be entered in minutes.", "error")
            return redirect(url_for("add_service_type"))

        if default_duration_minutes <= 0:
            flash("Default session length must be greater than zero.", "error")
            return redirect(url_for("add_service_type"))

        try:
            default_price = float(price_raw)
        except (TypeError, ValueError):
            flash("Default service price must be valid.", "error")
            return redirect(url_for("add_service_type"))

        if default_price < 0:
            flash("Default service price cannot be negative.", "error")
            return redirect(url_for("add_service_type"))

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO service_name_types (
                spa_id,
                service_name,
                default_duration_minutes,
                default_price,
                is_active
            )
            VALUES (%s, %s, %s, %s, %s)
        """, (
            spa_id,
            service_name,
            default_duration_minutes,
            default_price,
            is_active
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Service type added.", "success")
        return redirect(url_for("service_types"))

    return render_template(
        "service_type_form.html",
        service=None
    )








#################################
#
#
#   EDIT SERVICE TYPE
#
#
####################################



@app.route(
    "/service-types/<int:service_type_id>/edit",
    methods=["GET", "POST"]
)
@login_required
@spa_required
def edit_service_type(service_type_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        service_name = (
            request.form.get("service_name") or ""
        ).strip()

        duration_raw = (
            request.form.get("default_duration_minutes") or ""
        ).strip()

        price_raw = (
            request.form.get("default_price") or ""
        ).strip()

        is_active = request.form.get("is_active") == "1"

        if not service_name:
            flash("Service name is required.", "error")
            cur.close()
            conn.close()
            return redirect(url_for(
                "edit_service_type",
                service_type_id=service_type_id
            ))

        try:
            default_duration_minutes = int(duration_raw)
            default_price = float(price_raw)
        except (TypeError, ValueError):
            flash("Duration and price must be valid.", "error")
            cur.close()
            conn.close()
            return redirect(url_for(
                "edit_service_type",
                service_type_id=service_type_id
            ))

        if default_duration_minutes <= 0 or default_price < 0:
            flash("Duration and price values are invalid.", "error")
            cur.close()
            conn.close()
            return redirect(url_for(
                "edit_service_type",
                service_type_id=service_type_id
            ))

        cur.execute("""
            UPDATE service_name_types
            SET
                service_name = %s,
                default_duration_minutes = %s,
                default_price = %s,
                is_active = %s
            WHERE service_type_id = %s
              AND spa_id = %s
        """, (
            service_name,
            default_duration_minutes,
            default_price,
            is_active,
            service_type_id,
            spa_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Service type updated.", "success")
        return redirect(url_for("service_types"))

    cur.execute("""
        SELECT
            service_type_id,
            service_name,
            default_duration_minutes,
            default_price,
            is_active
        FROM service_name_types
        WHERE service_type_id = %s
          AND spa_id = %s
    """, (service_type_id, spa_id))

    service = cur.fetchone()

    cur.close()
    conn.close()

    if not service:
        flash("Service type not found.", "error")
        return redirect(url_for("service_types"))

    return render_template(
        "service_type_form.html",
        service=service
    )










#################################
#
#
#   
#
#
####################################

                 
#   -----------------------
#               
#     SPA MANAGEMENT PAGE
#               
#  ----------------------



@app.route("/spa_management")
@login_required
@spa_required
def spa_management():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT *
        FROM spas
        WHERE spa_id = %s
    """, (spa_id,))

    spa = cur.fetchone()

    cur.execute("""
        SELECT
            owner_agreed_sms_email_terms,
            owner_agreed_sms_email_terms_at,
            owner_agreed_sms_email_terms_version
        FROM spas
        WHERE spa_id = %s
    """, (spa_id,))

    terms_status = cur.fetchone()

    cur.close()
    conn.close()

    return render_template(
        "spa_management.html",
        spa=spa,
        terms_status=terms_status
    )





##############################################
#
#   USER SETTINGS 
#
#
###############################################




@app.route("/my-settings", methods=["GET", "POST"])
@login_required
@spa_required
def my_settings():
    user_id = session.get("user_id")

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        preferred_language = request.form.get("preferred_language", "EN").strip().upper()

        if preferred_language not in ("EN", "ES"):
            preferred_language = "EN"

        cur.execute("""
            UPDATE users
            SET preferred_language = %s
            WHERE user_id = %s
        """, (preferred_language, user_id))

        conn.commit()

        session["language_code"] = preferred_language

        flash("User settings updated.", "success")
        cur.close()
        conn.close()

        return redirect(url_for("my_settings"))

    cur.execute("""
        SELECT preferred_language
        FROM users
        WHERE user_id = %s
    """, (user_id,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    preferred_language = row[0] if row and row[0] else "EN"

    return render_template(
        "my_settings.html",
        preferred_language=preferred_language
    )




























####################################
#   COMMUNICATIONS HOME
###################################


@app.route("/communications")
@login_required
@spa_required  

def communications():
    return render_template("communications.html")
     








#   -----------------------
#
#     BUSINESS FINANCING PAGE
#   
#  ----------------------

@app.route("/business_financing_home")
@login_required
@spa_required  

def business_financing_home():
    return render_template("business_financing_home.html")
     



#   --------------------------- 
#
#   -------------------------




@app.route("/email_template_form1")
@login_required
@spa_required  

def email_template_form1():
    return render_template("email_template_form.html")


#   -----------------------
#       
#     ACCOUNTING HOME
#
#  ----------------------
        
@app.route("/accounting_home")
@login_required
@spa_required  

def accounting_home():
    return render_template("accounting_home.html")






#########################################
#   FINANCIALS HOME
#
#######################################


@app.route("/financials_home")
@login_required
@spa_required  

def financials_home():
    return render_template("financials_home.html")





#########################################
#   Reports
#
#######################################


@app.route("/financial_reports_home")
@login_required
@spa_required  

def financial_reports_home():
    return render_template("financial_reports_home.html")






@app.route("/reports_all")
@login_required
@spa_required  

def eports_all():
    return render_template("reports_all.html")





#########################################
#   MASTER ADMIN HOME
#
#######################################


@app.route("/master_admin_home")
@login_required
@spa_required  

def master_admin_home():
    return render_template("master_admin_home.html")




#########################################
#   MASTER ADMIN 
#
#   MASTER ADMIN SETTINGS
#
#######################################


@app.route("/master-admin/settings")
@login_required
@spa_required
def master_admin_settings():
    supported_languages = get_supported_languages(active_only=False)

    return render_template(
        "master_admin_settings.html",
        supported_languages=supported_languages
    )











#   -----------------------
#
#     TERMS & Conditions
#        PRIVACY
#  ----------------------


                
@app.route("/terms")
def terms():
    return redirect(url_for("view_help_page", page_key="terms"))

@app.route("/privacy")
def privacy():
    return redirect(url_for("view_help_page", page_key="privacy"))







#   -----------------------
#
#     Finance HOME    
#
#  ----------------------
            
@app.route("/finance_home")
@login_required
@spa_required  

def finance_home():
    return render_template("finance_home.html")



#   --------------------------- 
#
#   -------------------------



@app.route("/birthday_offers1_home")
@login_required
@spa_required  

def birthday_offers1_home():
    return render_template("birthday_offers1_home.html")




#   ------------------------------------------------
#
#
#     LOGIN
#
#
#
#
#   ---------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            SELECT user_id, spa_id, first_name, last_name, email, password_hash, role
            FROM users
            WHERE LOWER(email) = %s
              AND active = TRUE
        """, (email,))

        user = cur.fetchone()

        cur.close()
        conn.close()

        if not user:
            log_scheduler("Login failed: user account not found.")
            flash("Invalid email or password.", "error")
            return render_template("login.html")

        password_match = check_password_hash(user[5], password)

        if not password_match:
            log_scheduler("Login failed: invalid password.")
            flash("Invalid email or password.", "error")
            return render_template("login.html")

        role = user[6]

        session.clear()
        session["user_id"] = user[0]
        session["spa_id"] = user[1]
        session["first_name"] = user[2]
        session["last_name"] = user[3]
        session["email"] = user[4]
        session["role"] = role

        flash("Logged in successfully.", "success")


        return redirect(url_for("morning_briefing"))

    return render_template("login.html")










#   -------------------------
#
#
#         LOGOUT 
#
#
#
#   -------------------------



@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "success")
    return redirect(url_for("login"))




#   -------------------------
#  >>>>>>>>>>>>>>>>>>>>>>>>>>>>
#     
#        CLIENTS HOME
#      TEMP FIX
#   >>>>>>>>>>>>>>>>>>>>>>>>
#   
#   -------------------------

@app.route("/clients")
@login_required
@spa_required
def clients_home():
    return redirect(url_for("client_management"))





###############################################
#
#   REVIEW GODADDY IMPORTS
#
#
###############################################


@app.route("/godaddy-imports/review-calendar")
@login_required
@spa_required
def review_godaddy_imports_and_calendar():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE appointments
        SET import_reviewed = TRUE
        WHERE spa_id = %s
          AND external_source = 'godaddy'
          AND COALESCE(import_reviewed, FALSE) = FALSE
    """, (spa_id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("calendar_view"))





###############################################
#
#   REVIEW GODADDY IMPORTS DISMISS
#
#   THIS IS FOR THE IMPORT BANNER MESSAGE
###############################################





@app.route("/godaddy-imports/dismiss", methods=["POST"])
@login_required
@spa_required
def dismiss_godaddy_import_review():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE appointments
        SET import_reviewed = TRUE
        WHERE spa_id = %s
          AND external_source = 'godaddy'
          AND COALESCE(import_reviewed, FALSE) = FALSE
    """, (spa_id,))

    conn.commit()

    cur.close()
    conn.close()

    return redirect(url_for("calendar_view"))












# #################################
#
#
#       COMPLIANCE HOME
#
#
#
# #################################



# ==========================================================
# MESSAGING COMPLIANCE CENTER
# ==========================================================

@app.route('/admin/messaging-compliance')
@login_required
@spa_required
def messaging_compliance_dashboard():

    spa_id = session.get("spa_id")

    onboarding = get_messaging_onboarding(spa_id)

    if onboarding:
        current_step = onboarding.get("current_step") or 1
        last_completed_step = onboarding.get("last_completed_step") or 0
    else:
        current_step = 1
        last_completed_step = 0

    compliance_check = run_messaging_compliance_check(onboarding)

    progress_percent = int((last_completed_step / 7) * 100)

    return render_template(
        "admin/messaging_compliance/dashboard.html",
        onboarding=onboarding,
        current_step=current_step,
        last_completed_step=last_completed_step,
        progress_percent=progress_percent,
        compliance_check=compliance_check
    )










############################
#
#
#     COMPLIANCE ONBOARDING
#
#############################


@app.route(
    '/admin/messaging-compliance/onboarding',
    methods=['GET', 'POST']
)
@login_required
@spa_required
def messaging_compliance_onboarding():

    spa_id = session.get("spa_id")   # use your existing helper
    step = request.args.get("step", 1, type=int)


    if request.method == "POST":

        form_data = dict(request.form)
        form_data.pop("action", None)   # Don't save the button value

        save_messaging_onboarding(
            spa_id=spa_id,
            **form_data
        )



        action = request.form.get("action")

        flash("Onboarding information saved.", "success")

        if action == "save_exit":
            return redirect(url_for("messaging_compliance_dashboard"))

        next_step = min(step + 1, 7)

        return redirect(
            url_for("messaging_compliance_onboarding", step=next_step)
)

    onboarding = get_messaging_onboarding(spa_id)

    return render_template(
        "admin/messaging_compliance/onboarding_wizard.html",
        onboarding=onboarding,
        step=step
    )




############################
#      
#
#     COMPLIANCE OVERVIEW
#
#############################


@app.route('/admin/messaging-compliance/overview')
@login_required
@spa_required
def messaging_compliance_overview():
    return render_template("admin/messaging_compliance/compliance_overview.html")



############################
#      
#
#     COMPLIANCE BRAND
#
#############################


@app.route('/admin/messaging-compliance/brand')
@login_required
@spa_required
def messaging_compliance_brand():
    return render_template("admin/messaging_compliance/brand_registration.html")



############################
#      
#
#     COMPLIANCE CAMPAIGN
#
#############################


@app.route('/admin/messaging-compliance/campaign')
@login_required
@spa_required
def messaging_compliance_campaign():
    return render_template("admin/messaging_compliance/campaign_registration.html")





############################
#      TEMPLATE REVIEW
#
#     COMPLIANCE TEMPLATES
#
#############################



@app.route("/admin/messaging-compliance/template-library")
@login_required
@spa_required
def template_review_default():
    language_code = get_request_language()

    return redirect(url_for(
        "template_review",
        channel="sms",
        language_code=language_code
    ))


@app.route("/admin/messaging-compliance/template-library/<channel>")
@login_required
@spa_required
def template_review(channel):
    spa_id = current_spa_id()

    channel = (channel or "sms").lower().strip()
    language_code = get_request_language()
    show_archived = request.args.get("show_archived") == "1"

    if channel not in ("sms", "email"):
        flash("Invalid template library.", "warning")
        return redirect(url_for(
            "template_review",
            channel="sms",
            language_code=language_code
        ))

    page_title = (
        "SMS Messaging Template Library"
        if channel == "sms"
        else "Email Messaging Template Library"
    )

    switch_channel = "email" if channel == "sms" else "sms"

    switch_label = (
        "Go To: 📧 Email Templates"
        if channel == "sms"
        else "Go To: 💬 SMS Templates"
    )

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            t.template_type,
            t.display_name,
            t.template_category,
            m.template_id,
            m.template_name,
            m.is_active,
            m.approved_for_use,
            m.ai_score,
            m.updated_at,
            m.ai_risk_level,
            m.channel,
            COALESCE(m.is_archived, FALSE) AS is_archived,
            m.language_code
        FROM messaging_template_types t

        LEFT JOIN messaging_templates m
            ON t.template_type = m.template_type
           AND m.spa_id = %s
           AND m.channel = %s
           AND m.language_code = %s
           AND (
                %s = TRUE
                OR COALESCE(m.is_archived, FALSE) = FALSE
           )

        WHERE t.is_active = TRUE

        ORDER BY t.display_order
    """, (
        spa_id,
        channel,
        language_code,
        show_archived
    ))

    templates = cur.fetchall()

    completed = sum(1 for t in templates if t[3])
    total = len(templates)
    percent = round((completed / total) * 100) if total else 0

    supported_languages = get_supported_languages()

    cur.close()
    conn.close()

    print("SUPPORTED LANGUAGES:", supported_languages, flush=True)
    print("CURRENT LANGUAGE:", language_code, flush=True)

    return render_template(
        "admin/messaging_compliance/template_review.html",
        templates=templates,
        channel=channel,
        language_code=language_code,
        show_archived=show_archived,
        completed=completed,
        total=total,
        percent=percent,
        page_title=page_title,
        switch_channel=switch_channel,
        switch_label=switch_label,
        supported_languages=supported_languages
    )






###################################
#
#   EDIT MESSAGING TEMPLATE BY ID
#   6/30/26
###################################


@app.route("/admin/messaging-compliance/templates/edit/<int:template_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_messaging_template_by_id(template_id):
    spa_id = current_spa_id()
    language_code = get_request_language()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            m.template_id,              -- 0
            m.spa_id,                   -- 1
            m.channel,                  -- 2
            m.template_type,            -- 3
            m.template_name,            -- 4
            m.language_code,            -- 5
            m.subject_text,             -- 6
            m.message_text,             -- 7
            m.is_active,                -- 8
            m.approved_for_use,         -- 9
            m.ai_score,                 -- 10
            m.ai_review,                -- 11
            m.ai_risk_level,            -- 12
            m.last_ai_reviewed_at,      -- 13
            COALESCE(is_archived, FALSE), -- 14
            t.display_name            --15
        FROM messaging_templates m

        LEFT JOIN messaging_template_types t
            ON m.template_type = t.template_type

        WHERE m.template_id = %s
        AND m.spa_id = %s
                LIMIT 1
    """, (template_id, spa_id))

    template = cur.fetchone()

    if not template:
        cur.close()
        conn.close()
        flash("Template not found.", "warning")
        return redirect(url_for(
            "template_review",
            channel="sms",
            language_code=language_code
        ))

    channel = template[2]
    template_type = template[3]
    template_name = template[4] or "Default"
    display_name = template[15] or template_type.replace("_", " ").title()
    

    template_language_code = normalize_language_code(template[5] or language_code)

    language_name = get_language_name(template_language_code)

    is_default_template = template_name.strip().lower() == "default"

    if request.method == "POST":
        submitted_template_name = (
            request.form.get("template_name")
            or template_name
            or "Default"
        ).strip()

        final_template_name = "Default" if is_default_template else submitted_template_name

        form_language_code = normalize_language_code(
            request.form.get("language_code")
            or template_language_code
        )

        subject_text = (request.form.get("subject_text") or "").strip()
        message_text = (request.form.get("message_text") or "").strip()

        if channel == "sms":
            subject_text = None

        if not message_text:
            flash("Message text is required.", "warning")
            cur.close()
            conn.close()
            return redirect(url_for(
                "edit_messaging_template_by_id",
                template_id=template_id,
                language_code=form_language_code
            ))

        is_active = template[8]

        ai_result = review_template_ai_basic(
            template_type,
            message_text,
            channel=channel
        )

        approved_for_use = ai_result["score"] > 60

        
        ai_result = review_template_ai_basic(
            template_type,
            message_text,
            channel=channel
        )

        cur.execute("""
            UPDATE messaging_templates
            SET
                template_name = %s,
                language_code = %s,
                subject_text = %s,
                message_text = %s,
                is_active = %s,
                approved_for_use = %s,
                ai_score = %s,
                ai_review = %s,
                ai_risk_level = %s,
                last_ai_reviewed_at = NOW(),
                updated_at = NOW()
            WHERE template_id = %s
              AND spa_id = %s
        """, (
            final_template_name,
            form_language_code,
            subject_text,
            message_text,
            is_active,
            approved_for_use,
            ai_result["score"],
            ai_result["review"],
            ai_result["risk_level"],
            template_id,
            spa_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Template saved and AI reviewed.", "success")

        return redirect(url_for(
            "edit_messaging_template_by_id",
            template_id=template_id,
            language_code=form_language_code
        ))

    footer_key = "sms_opt_out" if channel == "sms" else "email_unsubscribe"
    compliance_footer = get_messaging_footer(footer_key)
    supported_languages = get_supported_languages()

    cur.close()
    conn.close()

    return render_template(
        "admin/messaging_compliance/edit_messaging_template.html",
        template=template,
        template_id=template_id,
        channel=channel,
        template_type=template_type,
        template_name=template_name,
        display_name=display_name,
        language_name=language_name,
        language_code=template_language_code,
        compliance_footer=compliance_footer,
        supported_languages=supported_languages,
        is_default_template=is_default_template,
        create_mode=False
    )











###################################
#
#   DUPLICATE MESSAGING TEMPLATE
#
###################################



@app.route(
    "/admin/messaging-compliance/templates/<int:template_id>/duplicate",
    methods=["POST"]
)
@login_required
@spa_required
def duplicate_messaging_template(template_id):

    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            channel,
            template_type,
            template_name,
            language_code,
            subject_text,
            message_text
        FROM messaging_templates
        WHERE template_id = %s
          AND spa_id = %s
          AND COALESCE(is_archived, FALSE) = FALSE
        LIMIT 1
    """, (template_id, spa_id))

    original = cur.fetchone()

    if not original:
        cur.close()
        conn.close()
        flash("Template not found or cannot be duplicated.", "warning")
        return redirect(url_for("template_review", channel="sms"))

    channel = original[0]
    template_type = original[1]
    template_name = original[2] or "Template"
    language_code = original[3] or "en"
    subject_text = original[4]
    message_text = original[5]

    copy_base = f"{template_name} Copy"

    cur.execute("""
        SELECT template_name
        FROM messaging_templates
        WHERE spa_id = %s
        AND channel = %s
        AND template_type = %s
        AND COALESCE(is_archived, FALSE) = FALSE
    """, (
        spa_id,
        channel,
        template_type
    ))

    existing_names = {
        row[0].strip()
        for row in cur.fetchall()
        if row[0]
    }

    if copy_base not in existing_names:
        new_template_name = copy_base
    else:
        counter = 2

        while f"{copy_base} {counter}" in existing_names:
            counter += 1

        new_template_name = f"{copy_base} {counter}"


    cur.execute("""
        INSERT INTO messaging_templates
        (
            spa_id,
            channel,
            template_type,
            template_name,
            language_code,
            subject_text,
            message_text,
            is_active,
            approved_for_use,
            is_archived,
            ai_score,
            ai_review,
            ai_risk_level,
            created_at,
            updated_at
        )
        VALUES
        (
            %s, %s, %s, %s, %s, %s, %s,
            FALSE, FALSE, FALSE,
            NULL, NULL, NULL,
            NOW(), NOW()
        )
        RETURNING template_id
    """, (
        spa_id,
        channel,
        template_type,
        new_template_name,
        language_code,
        subject_text,
        message_text
    ))

    new_template_id = cur.fetchone()[0]

    conn.commit()


    cur.close()
    conn.close()

    flash("Template duplicated. Review and save before using.", "success")

    return redirect(url_for(
        "edit_messaging_template_by_id",
        template_id=new_template_id
    ))































############################
#      
#         CREATE MESSAGING TEMPLATE
#     
#   6-30-26
#############################
##
#################



@app.route(
    "/admin/messaging-compliance/templates/<channel>/<template_type>",
    methods=["GET", "POST"]
)
@login_required
@spa_required
def create_messaging_template(channel, template_type):
    spa_id = current_spa_id()
    language_code = get_request_language()

    channel = (channel or "sms").lower().strip()

    if channel not in ("sms", "email"):
        flash("Invalid template channel.", "warning")
        return redirect(url_for(
            "template_review",
            channel="sms",
            language_code=language_code
        ))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT template_id
        FROM messaging_templates
        WHERE spa_id = %s
          AND channel = %s
          AND template_type = %s
          AND language_code = %s
          AND COALESCE(is_archived, FALSE) = FALSE
        ORDER BY updated_at DESC, template_id DESC
        LIMIT 1
    """, (
        spa_id,
        channel,
        template_type,
        language_code
    ))

    existing = cur.fetchone()

    if existing:
        cur.close()
        conn.close()
        return redirect(url_for(
            "edit_messaging_template_by_id",
            template_id=existing[0],
            language_code=language_code
        ))

    if request.method == "POST":
        template_name = (request.form.get("template_name") or "Default").strip()
        form_language_code = normalize_language_code(
            request.form.get("language_code") or language_code
        )

        subject_text = (request.form.get("subject_text") or "").strip()
        message_text = (request.form.get("message_text") or "").strip()

        if channel == "sms":
            subject_text = None

        if not message_text:
            flash("Message text is required.", "warning")
            cur.close()
            conn.close()
            return redirect(url_for(
                "create_messaging_template",
                channel=channel,
                template_type=template_type,
                language_code=form_language_code
            ))

        ai_result = review_template_ai_basic(
            template_type,
            message_text,
            channel=channel
        )

        is_active = ai_result["score"] > 60
        approved_for_use = ai_result["score"] > 60

        cur.execute("""
            INSERT INTO messaging_templates
            (
                spa_id,
                channel,
                template_name,
                template_type,
                language_code,
                subject_text,
                message_text,
                is_active,
                approved_for_use,
                ai_score,
                ai_review,
                ai_risk_level,
                last_ai_reviewed_at,
                created_at,
                updated_at
            )
            VALUES
            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW(),NOW())
            RETURNING template_id
        """, (
            spa_id,
            channel,
            template_name,
            template_type,
            form_language_code,
            subject_text,
            message_text,
            is_active,
            approved_for_use,
            ai_result["score"],
            ai_result["review"],
            ai_result["risk_level"]
        ))

        new_template_id = cur.fetchone()[0]

        conn.commit()
        cur.close()
        conn.close()

        flash("Template created and AI reviewed.", "success")

        return redirect(url_for(
            "edit_messaging_template_by_id",
            template_id=new_template_id,
            language_code=form_language_code
        ))

    footer_key = "sms_opt_out" if channel == "sms" else "email_unsubscribe"
    compliance_footer = get_messaging_footer(footer_key)
    supported_languages = get_supported_languages()

    cur.close()
    conn.close()

    return render_template(
        "admin/messaging_compliance/edit_messaging_template.html",
        template=None,
        template_id=None,
        channel=channel,
        template_type=template_type,
        template_name="Default",
        language_code=language_code,
        compliance_footer=compliance_footer,
        supported_languages=supported_languages,
        is_default_template=False,
        create_mode=True
    )











############################
#      
#
#     PREVIEW MESSAGING TEMPLATE BY ID
#   6-30-26
#############################


@app.route(
    "/admin/messaging-compliance/templates/preview/<int:template_id>",
    methods=["GET"]
)
@login_required
@spa_required
def preview_messaging_template_by_id(template_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            m.template_id,        -- 0
            m.channel,            -- 1
            m.template_type,      -- 2
            m.template_name,      -- 3
            m.language_code,      -- 4
            m.subject_text,       -- 5
            m.message_text,       -- 6
            t.display_name        -- 7
        FROM messaging_templates m
        LEFT JOIN messaging_template_types t
            ON m.template_type = t.template_type
        WHERE m.template_id = %s
          AND m.spa_id = %s
          AND COALESCE(m.is_archived, FALSE) = FALSE
        LIMIT 1
    """, (template_id, spa_id))

    template = cur.fetchone()

    cur.close()
    conn.close()

    if not template:
        flash("Template not found.", "warning")
        return redirect(url_for(
            "template_review",
            channel="sms",
            language_code=get_request_language()
        ))

    channel = template[1]
    template_type = template[2]
    template_name = template[3] or "Default"
    language_code = normalize_language_code(template[4] or get_default_language())
    display_name = template[7] or template_type.replace("_", " ").title()
    language_name = get_language_name(language_code)

    sample_data = {
        "client_id": None,
        "language_code": language_code,
        "preferred_language": language_code,
        "client_first_name": "Rick",
        "client_full_name": "Rick Conley",
        "appointment_date": "June 20, 2026",
        "appointment_time": "2:00 PM",
        "service_name": "Signature Facial",
        "spa_name": "Clear Skin Esthetics",
        "spa_phone": "(817) 555-1234",
        "spa_website": "https://peachsuitepro.com",
        "spa_address": "123 Main Street",
        "opt_out": "Reply STOP to opt out"
    }

    preview = build_communication(
        spa_id=spa_id,
        channel=channel,
        template_type=template_type,
        merge_data=sample_data,
        language_code=language_code,
        template_id=template_id
    )

    if not preview:
        flash("Unable to build preview.", "warning")
        return redirect(url_for(
            "edit_messaging_template_by_id",
            template_id=template_id,
            language_code=language_code
        ))

    return render_template(
        "admin/messaging_compliance/template_preview.html",
        template_id=template_id,
        display_name=display_name,
        template_name=template_name,
        channel=channel,
        template_type=template_type,
        preview=preview,
        language_code=language_code,
        language_name=language_name
    )







############################################
###########################################. TODO
#####test.  test.  test. remove rick conley from below

@app.route("/admin/test-template")
@login_required
@spa_required
def test_template():

    spa_id = session.get("spa_id")

    merge_data = {
        "client_first_name": "Rick",
        "client_last_name": "Conley",
        "appointment_date": "June 20, 2026",
        "appointment_time": "10:00 AM",
        "service_name": "Facial",
        "spa_name": "Clear Skin Esthetics"
    }

    result = build_sms_message(
        spa_id,
        "appointment_reminder",
        merge_data
    )

    return f"<pre>{result}</pre>"





############################
#      
#
#     COMPLIANCE DATA MIGRATION
#
#############################

@app.route('/admin/messaging-compliance/migration')
@login_required
@spa_required
def messaging_compliance_migration():
    return render_template("admin/messaging_compliance/data_migration.html")



############################
#      
#
#     COMPLIANCE DOCUMENTS
#
#############################


@app.route('/admin/messaging-compliance/documents')
@login_required
@spa_required
def messaging_compliance_documents():
    return render_template("admin/messaging_compliance/documents.html")





############################
#      
#
#     COMPLIANCE AUDIT LOG
#
#############################

@app.route('/admin/messaging-compliance/audit-log')
@login_required
@spa_required
def messaging_compliance_audit_log():
    return render_template("admin/messaging_compliance/audit_log.html")




############################
#
#
#     COMPLIANCE ONBOARDING
#
#############################


@app.route('/admin/messaging-compliance/campaign-registration')
@login_required
@spa_required
def messaging_compliance_campaign_registration():
    return render_template("admin/messaging_compliance/campaign_registration.html")







############################
#
#   COACH MEMORY HELPERS
#     
#
#############################


def get_or_create_coach_daily_session(
    cur,
    spa_id,
    user_id,
    session_date
):
    """
    Return today's Coach session.

    Creates the session on the first Daily Briefing visit.
    Updates the visit information on later visits.
    """

    cur.execute("""
        INSERT INTO coach_daily_sessions (
            spa_id,
            user_id,
            session_date
        )
        VALUES (%s, %s, %s)

        ON CONFLICT (spa_id, user_id, session_date)
        DO UPDATE SET
            last_opened_at = CURRENT_TIMESTAMP,
            open_count = coach_daily_sessions.open_count + 1

        RETURNING
            coach_session_id,
            started_at,
            last_opened_at,
            last_message_at,
            open_count,
            current_recommendation_key,
            session_status
    """, (
        spa_id,
        user_id,
        session_date
    ))

    row = cur.fetchone()

    return {
        "coach_session_id": row[0],
        "started_at": row[1],
        "last_opened_at": row[2],
        "last_message_at": row[3],
        "open_count": row[4],
        "current_recommendation_key": row[5],
        "session_status": row[6]
    }






############################
#
#   COACH MEMORY HELPERS
#     
#
#############################



def build_coach_recommendation_key(recommendation):
    """
    Create a stable key for Coach memory.

    The queue position cannot be used because recommendation_1
    may represent a different issue later in the day.
    """

    if not recommendation:
        return None

    category = str(
        recommendation.get("category", "recommendation")
    ).strip().lower()

    action_url = str(
        recommendation.get("action_url")
        or recommendation.get("url")
        or ""
    ).strip().lower()

    raw_key = f"{category}:{action_url}"

    cleaned_key = "".join(
        character
        if character.isalnum()
        else "_"
        for character in raw_key
    )

    while "__" in cleaned_key:
        cleaned_key = cleaned_key.replace("__", "_")

    return cleaned_key.strip("_")[:150]











############################
#
#   COACH MEMORY HELPERS
#     
#
#############################





def record_coach_interaction(
    cur,
    coach_session_id,
    spa_id,
    user_id,
    coach,
    message_type="briefing_open"
):
    """
    Save the Coach message shown during this page visit.
    """

    current_recommendation = coach.get(
        "current_recommendation"
    )

    recommendation_key = build_coach_recommendation_key(
        current_recommendation
    )

    category = None
    action_url = None

    if current_recommendation:
        category = current_recommendation.get("category")

        action_url = (
            current_recommendation.get("action_url")
            or current_recommendation.get("url")
        )

    cur.execute("""
        INSERT INTO coach_interactions (
            coach_session_id,
            spa_id,
            user_id,
            recommendation_key,
            category,
            message_type,
            message_text,
            action_url,
            interaction_status
        )
        VALUES (
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            %s,
            'presented'
        )
    """, (
        coach_session_id,
        spa_id,
        user_id,
        recommendation_key,
        category,
        message_type,
        coach.get("message"),
        action_url
    ))

    cur.execute("""
        UPDATE coach_daily_sessions
        SET
            last_message_at = CURRENT_TIMESTAMP,
            current_recommendation_key = %s
        WHERE coach_session_id = %s
    """, (
        recommendation_key,
        coach_session_id
    ))








##################################################
#
#   MASTER ADMIN FOOTERS
#
#
#
######################################################

@app.route("/master/messaging-footers", methods=["GET", "POST"])
@login_required
@master_admin_required
def master_messaging_footers():

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        sms_footer = request.form.get("sms_opt_out", "").strip()
        email_footer = request.form.get("email_unsubscribe", "").strip()

        cur.execute("""
            UPDATE messaging_footer_settings
            SET footer_text = %s,
                updated_at = CURRENT_TIMESTAMP,
                updated_by = %s
            WHERE footer_type = 'sms_opt_out'
        """, (sms_footer, session.get("user_id")))

        cur.execute("""
            UPDATE messaging_footer_settings
            SET footer_text = %s,
                updated_at = CURRENT_TIMESTAMP,
                updated_by = %s
            WHERE footer_type = 'email_unsubscribe'
        """, (email_footer, session.get("user_id")))

        conn.commit()
        flash("Messaging footer settings updated.", "success")
        return redirect(url_for("master_messaging_footers"))

    cur.execute("""
        SELECT footer_type, footer_text
        FROM messaging_footer_settings
        WHERE is_active = TRUE
    """)
    rows = cur.fetchall()

    footers = {row[0]: row[1] for row in rows}

    cur.close()
    conn.close()

    return render_template(
        "master_messaging_footers.html",
        footers=footers
    )






#   -------------------------
#
#
#       ADD SPA
#
#
#
#   -------------------------

@app.route("/add-spa", methods=["GET", "POST"])
@login_required
@spa_required  

def add_spa():
    require_master_admin()

    if request.method == "POST":
        spa_name = request.form.get("spa_name", "").strip()
        owner_first_name = request.form.get("owner_first_name", "").strip()
        owner_last_name = request.form.get("owner_last_name", "").strip()
        owner_email = request.form.get("owner_email", "").strip().lower()
        owner_phone = request.form.get("owner_phone", "").strip()
        timezone_name = request.form.get("timezone_name", "America/Chicago").strip()

        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "").strip()

        if not spa_name or not owner_email or not username or not password:
            flash("Spa name, owner email, username, and password are required.", "error")
            return redirect(url_for("add_spa"))

        password_hash = generate_password_hash(password)

        conn = get_db_connection()
        cur = conn.cursor()

        try:
            cur.execute("""
                INSERT INTO spas (
                    spa_name,
                    owner_first_name,
                    owner_last_name,
                    owner_email,
                    owner_phone,
                    timezone_name,
                    active,
                    sync_enabled,
                    subscription_status
                )
                VALUES (%s, %s, %s, %s, %s, %s, TRUE, FALSE, 'active')
                RETURNING spa_id
            """, (
                spa_name,
                owner_first_name,
                owner_last_name,
                owner_email,
                owner_phone,
                timezone_name
            ))

            spa_id = cur.fetchone()[0]

            cur.execute("""
                INSERT INTO users (
                    spa_id,
                    first_name,
                    last_name,
                    email,
                    username,
                    password_hash,
                    role,
                    active
                )
                VALUES (%s, %s, %s, %s, %s, %s, 'admin', TRUE)
            """, (
                spa_id,
                owner_first_name,
                owner_last_name,
                owner_email,
                username,
                password_hash
            ))

            conn.commit()
            flash("Spa and owner admin user created successfully.", "success")
            return redirect(url_for("home"))

        except Exception as e:
            conn.rollback()
            flash(f"Error creating spa: {e}", "error")
            return redirect(url_for("add_spa"))

        finally:
            cur.close()
            conn.close()

    return render_template("add_spa.html")




#   -------------------------
#
#       
#    ADD USER 
#
#
#
#   -------------------------


@app.route("/users/add", methods=["GET", "POST"])
@login_required
@spa_required

def add_user():
    require_admin_or_master()

    if request.method == "POST":
        spa_id = current_spa_id()

        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        role = clean_user_role(request.form.get("role", "staff"))

        if not email or not username or not password:
            flash("Email, username, and password are required.", "error")
            return redirect(url_for("add_user"))

        password_hash = generate_password_hash(password)

        conn = get_db_connection()
        cur = conn.cursor()

        try:
            cur.execute("""
                INSERT INTO users (
                    spa_id,
                    first_name,
                    last_name,
                    email,
                    username,
                    password_hash,
                    role,
                    active
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE)
            """, (
                spa_id,
                first_name,
                last_name,
                email,
                username,
                password_hash,
                role
            ))

            conn.commit()
            flash("User created successfully.", "success")
            return redirect(url_for("home"))

        except Exception as e:
            conn.rollback()
            flash(f"Error creating user: {e}", "error")
            return redirect(url_for("add_user"))

        finally:
            cur.close()
            conn.close()

    return render_template("add_user.html")






                    
                    
#   -------------------------
#
#
#        HELP - NEW & EDIT
#.       HELP NEW
#.       HELP EDIT
#
#   -------------------------
            
@app.route("/admin/help-pages/new", methods=["GET", "POST"])
@app.route("/admin/help-pages/edit/<page_key>", methods=["GET", "POST"])
@login_required
@master_admin_required
def edit_help_page(page_key=None):

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":

        if not page_key:
            page_key = request.form.get("page_key", "").strip()

        display_name = request.form.get("display_name", "").strip()
        title = request.form.get("title", "").strip()
        language_code = request.form.get("language_code", "EN").strip().upper()
        display_order = request.form.get("display_order") or None
        content = request.form.get("content", "").strip()
        is_active = request.form.get("is_active") == "on"

        cur.execute("""
            SELECT help_page_id
            FROM help_pages
            WHERE page_key = %s
              AND language_code = %s
        """, (page_key, language_code))

        existing = cur.fetchone()

        if existing:
            cur.execute("""
                UPDATE help_pages
                SET display_name = %s,
                    title = %s,
                    language_code = %s,
                    display_order = %s,
                    content = %s,
                    is_active = %s
                WHERE page_key = %s
                  AND language_code = %s
            """, (
                display_name,
                title,
                language_code,
                display_order,
                content,
                is_active,
                page_key,
                language_code
            ))
        else:
            cur.execute("""
                INSERT INTO help_pages
                    (
                        page_key,
                        display_name,
                        title,
                        language_code,
                        display_order,
                        content,
                        is_active
                    )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                page_key,
                display_name,
                title,
                language_code,
                display_order,
                content,
                is_active
            ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Help page saved.", "success")
        return redirect(
            url_for(
                "edit_help_page",
                page_key=page_key,
                lang=language_code
            )
        )

    language_code = request.args.get("lang", "EN").strip().upper()

    cur.execute("""
        SELECT
            page_key,
            display_name,
            title,
            language_code,
            display_order,
            content,
            is_active
        FROM help_pages
        WHERE page_key = %s
          AND language_code = %s
    """, (page_key, language_code))

    page = cur.fetchone()

    cur.close()
    conn.close()

    return render_template(
        "admin_edit_help_page.html",
        page=page,
        page_key=page_key,
        language_code=language_code,
        mode="edit" if page else "new"
    )



#   -------------------------
#  
#  
#       CURRENT LANGUAGE
#   
#   
#   
#   -------------------------







def get_current_language():
    user_id = session.get("user_id")

    if not user_id:
        return "EN"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT preferred_language
        FROM users
        WHERE user_id = %s
    """, (user_id,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    language = row[0] if row and row[0] else "EN"
    language = language.strip().upper()

    if language not in ("EN", "ES"):
        language = "EN"

    return language



#   -------------------------
#  
#  
#       GET CLIENT LANGUAGE
#   
#   
#   
#   -------------------------




def get_client_language(client_id):
    if not client_id:
        return "EN"

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT preferred_language
        FROM clients
        WHERE client_id = %s
    """, (client_id,))

    row = cur.fetchone()

    cur.close()
    conn.close()

    language = row[0] if row and row[0] else "EN"
    language = language.strip().upper()

    if language not in ("EN", "ES"):
        language = "EN"

    return language





#   -------------------------
#  
#  
#       VIEW HELP PAGE
#   
#   
#   
#   -------------------------




@app.route("/help/<page_key>")
@login_required
@spa_required
def view_help_page(page_key):

    
    conn = get_db_connection()
    cur = conn.cursor()

    # 1. Try requested language first
    session["language_code"] = "ES"

    requested_language = get_current_language()

    cur.execute("""
        SELECT title, content, language_code
        FROM help_pages
        WHERE page_key = %s
        AND language_code = %s
        AND is_active = TRUE
        ORDER BY help_page_id DESC
        LIMIT 1
    """, (page_key, requested_language))

    page = cur.fetchone()

    # 2. Fallback to English
    if not page and requested_language.upper() != "EN":
        cur.execute("""
            SELECT title, content, language_code
            FROM help_pages
            WHERE page_key = %s
              AND language_code = 'EN'
              AND is_active = TRUE
            ORDER BY help_page_id DESC
            LIMIT 1
        """, (page_key,))

        page = cur.fetchone()

    cur.close()
    conn.close()

    if not page:
        flash("Help page not found.", "warning")
        return redirect(url_for("help_center"))

    return render_template(
        "help_page.html",
        page=page,
        page_key=page_key,
        selected_language=page[2]
    )



#   -------------------------
#
#
#       HELP PAGE
#
#
#
#   -------------------------
        



@app.route("/page_help")
@login_required
@spa_required
def page_help():
    return redirect(url_for("view_help_page", page_key="calendar"))




@app.route("/admin/help-pages/new", methods=["GET"])
@login_required
@spa_required
def new_help_page():

    return render_template(
        "admin_edit_help_page.html",
        page=None,
        page_key=""
    )










    
#   -------------------------
#
#
#    HELP
#
#               
#
#   -------------------------
    

@app.route("/help")
@login_required
@spa_required
def help_center():

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()


    if search:
        search_term = f"%{search}%"

        cur.execute("""
            SELECT
                page_key,
                display_name,
                title,
                language_code,
                display_order
            FROM help_pages
            WHERE is_active = TRUE
            AND (
                    page_key ILIKE %s
                OR display_name ILIKE %s
                OR title ILIKE %s
                OR content ILIKE %s
            )
            ORDER BY display_order, page_key, language_code
        """, (
            search_term,
            search_term,
            search_term,
            search_term
        ))

    else:
        cur.execute("""
            SELECT
                page_key,
                display_name,
                title,
                language_code,
                display_order
            FROM help_pages
            WHERE is_active = TRUE
            ORDER BY display_order, display_name
        """)



    pages = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "help_center.html",
        pages=pages,
        search=search
    )







#   -------------------------
#  
#  
#    SMS Email TERMS
#   
#   
#   
#   -------------------------


@app.route("/sms-email-terms", methods=["GET", "POST"])
@login_required
@spa_required
def sms_email_terms():

    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    # Load help page content
    cur.execute("""
        SELECT title, content
        FROM help_pages
        WHERE page_key = %s
          AND spa_id = %s
          AND is_active = TRUE
    """, ("sms_email_terms", spa_id))

    page = cur.fetchone()

    if not page:
        cur.close()
        conn.close()

        flash("Terms page not found.", "warning")
        return redirect(url_for("home"))

    if request.method == "POST":

        agreed = "agree_terms" in request.form

        if not agreed:
            flash("You must agree to continue.", "warning")

            cur.close()
            conn.close()

            return render_template(
                "help_page.html",
                page=page,
                page_key="sms_email_terms"
            )

        cur.execute("""
            UPDATE spas
            SET owner_agreed_sms_email_terms = TRUE,
                owner_agreed_sms_email_terms_at = NOW(),
                owner_agreed_sms_email_terms_version = %s
            WHERE spa_id = %s
        """, ("v1.0", spa_id))

        conn.commit()

        cur.close()
        conn.close()

        flash("Terms accepted successfully.", "success")

        return redirect(url_for("home"))

    cur.close()
    conn.close()

    return render_template(
        "help_page.html",
        page=page,
        page_key="sms_email_terms"
    )







                    
#   ------------------------- 
#
#
#    DAILY BUILD LOG
#
#
#
#   -------------------------

@app.route('/log-it', methods=['GET', 'POST'])
@login_required
@master_admin_required
def log_it():
    if request.method == 'POST':
        log_title = request.form.get('log_title')
        log_notes = request.form.get('log_notes')

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO developer_change_log
                (log_title, log_notes)
            VALUES (%s, %s)
        """, (log_title, log_notes))
        conn.commit()
        cur.close()
        conn.close()

        flash("Daily change log saved successfully.", "success")
        return redirect(url_for('log_it_report'))

    return render_template('log_it.html')


@app.route('/log-it-report')
@login_required
@master_admin_required
def log_it_report():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT log_id, log_date, log_title, log_notes, created_at
        FROM developer_change_log
        ORDER BY log_date DESC, created_at DESC
    """)
    logs = cur.fetchall()
    cur.close()
    conn.close()

    return render_template('log_it_report.html', logs=logs)

















#   -------------------------
#
#
#  FEEDBACK FORM
#
#
#   this is checked 6-17-26
#   -------------------------


@app.route("/feedback", methods=["GET", "POST"])
@login_required
@spa_required
def feedback():
    spa_id = current_spa_id()

    if request.method == "POST":
        user_name = (request.form.get("user_name") or "").strip()
        user_email = (request.form.get("user_email") or "").strip()       
        page_name = (request.form.get("page_name") or "").strip()
        feedback_type = (request.form.get("feedback_type") or "").strip()
        message = (request.form.get("message") or "").strip()
        expected_behavior = (request.form.get("expected_behavior") or "").strip()
        severity = (request.form.get("severity") or "").strip()

        if not feedback_type:
            flash("Please select a feedback type.", "error")
            return redirect(url_for("feedback", page=page_name))

        if not message:
            flash("Please enter your feedback.", "error")
            return redirect(url_for("feedback", page=page_name))

        conn = get_db_connection()
        cur = conn.cursor()

        try:
            cur.execute("""
                INSERT INTO user_feedback (
                    spa_id,
                    user_name,
                    user_email,
                    page_name,
                    feedback_type,
                    message,
                    expected_behavior,
                    severity
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                spa_id,
                user_name if user_name else None,
                user_email if user_email else None,
                page_name if page_name else None,
                feedback_type,
                message,
                expected_behavior if expected_behavior else None,
                severity if severity else None
            ))

            conn.commit()

          
            if severity == "High":
                try:
                    response = send_email(
                        to="rickconleytx@gmail.com",
                        subject=f" SEV {severity} | {feedback_type} | {page_name}",
                        body=(
                            f"Spa ID: {spa_id}\n"
                            f"User Name: {user_name or 'Not provided'}\n"
                            f"User Email: {user_email or 'Not provided'}\n"
                            f"Page: {page_name or 'Not provided'}\n"
                            f"Type: {feedback_type}\n"
                            f"Severity: {severity or 'Not provided'}\n"
                            f"Expected Behavior: {expected_behavior or 'Not provided'}\n\n"
                            f"Message:\n{message}"
                        )
                    )

                    print("MAILGUN STATUS:", response.status_code)
                    print("MAILGUN BODY:", response.text)

                except Exception as email_error:
                    print("MAILGUN ERROR:", str(email_error))

            flash("Thank you. Your feedback has been submitted.", "success")
            return redirect(url_for("feedback", page=page_name))

        except Exception as db_error:
            conn.rollback()
            print("FEEDBACK DB ERROR:", str(db_error))
            flash("There was a problem saving your feedback.", "error")
            return redirect(url_for("feedback", page=page_name))

        finally:
            cur.close()
            conn.close()

    page_name = request.args.get("page") or request.referrer or request.url
    return render_template("feedback.html", page_name=page_name)













#   -------------------------
#       
#   
#  FEEDBACK ADMIN
#       
#   
#   
#   -------------------------



@app.route("/feedback-admin")
@login_required
@master_admin_required
def feedback_admin():
    feedback_type = (request.args.get("feedback_type") or "").strip()
    status = (request.args.get("status") or "open").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT COUNT(*)
        FROM user_feedback
        WHERE is_resolved IS FALSE OR is_resolved IS NULL
    """)
    open_count = cur.fetchone()[0]

    query = """
        SELECT
            uf.feedback_id,
            uf.user_name,
            uf.user_email,
            uf.page_name,
            uf.feedback_type,
            uf.severity,
            uf.expected_behavior,
            uf.message,
            uf.created_at,
            uf.is_resolved,
            uf.spa_id,
            s.spa_name,
            uf.action_taken_note
        FROM user_feedback uf
        LEFT JOIN spas s
            ON uf.spa_id = s.spa_id
        WHERE 1=1
    """

    params = []

    if feedback_type:
        query += " AND uf.feedback_type = %s"
        params.append(feedback_type)

    if status == "open":
        query += " AND (uf.is_resolved IS FALSE OR uf.is_resolved IS NULL)"
    elif status == "resolved":
        query += " AND uf.is_resolved IS TRUE"

    query += " ORDER BY uf.is_resolved ASC, uf.created_at DESC"

    cur.execute(query, tuple(params))
    feedback_items = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "feedback_admin.html",
        feedback_items=feedback_items,
        selected_type=feedback_type,
        selected_status=status,
        open_count=open_count
    )












#   -----------------------
#
#    FEEDBACK EDIT
#
#
#
#                           
#
#  ----------------------
            

@app.route("/feedback/edit/<int:feedback_id>", methods=["GET", "POST"])
@login_required
@master_admin_required
def edit_feedback(feedback_id):

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":

        is_resolved = request.form.get("is_resolved") == "on"
        action_taken_note = request.form.get(
            "action_taken_note", ""
        ).strip()

        cur.execute("""
            UPDATE user_feedback
            SET is_resolved = %s,
                action_taken_note = %s
            WHERE feedback_id = %s
        """, (
            is_resolved,
            action_taken_note,
            feedback_id
        ))

        conn.commit()

        cur.close()
        conn.close()

        flash("Feedback updated.", "success")
        return redirect(url_for("feedback_admin"))


    cur.execute("""
        SELECT
            feedback_id,
            user_name,
            user_email,
            page_name,
            feedback_type,
            severity,
            expected_behavior,
            message,
            created_at,
            is_resolved,
            spa_id,
            action_taken_note
        FROM user_feedback
        WHERE feedback_id = %s
    """, (feedback_id,))

    feedback = cur.fetchone()


    cur.close()
    conn.close()

    return render_template(
        "edit_feedback.html",
        feedback=feedback
    )




















#   -----------------------
#     
#    FEEDBACK RESOLVE
#     
#     
#     
#    
#  
#  ----------------------


@app.route("/feedback/resolve/<int:feedback_id>", methods=["POST"])
@login_required
def resolve_feedback(feedback_id):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE user_feedback
        SET is_resolved = TRUE
        WHERE feedback_id = %s
    """, (feedback_id,))

    conn.commit()
    cur.close()
    conn.close()

    flash("Feedback marked as resolved.", "success")
    return redirect(url_for("feedback_admin"))









        
#   -----------------------
#
#    FEEDBACK REOPEN 
#
#
#
#
#
#  ---------------------- 
       



@app.route("/feedback/reopen/<int:feedback_id>", methods=["POST"])
def reopen_feedback(feedback_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE user_feedback
        SET is_resolved = FALSE
        WHERE feedback_id = %s
          AND spa_id = %s
    """, (feedback_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Feedback reopened.", "success")
    return redirect(url_for("feedback_admin"))












#   -----------------------  
#
#    
#     
#
#
#     
#       
#  ----------------------     












#   -----------------------
#  
#     CREDIT   PROCESSORS
#
#
#
#     spa_id good
#
#  ----------------------

@app.route("/credit_processors")
@login_required
@spa_required  

def credit_processors():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            credit_processor_id,
            credit_processor_name,
            percentage_fee,
            flat_fee,
            additional_fee,
            is_active
        FROM credit_processors
        WHERE spa_id = %s
        ORDER BY credit_processor_name
    """, (spa_id,))

    processors = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "credit_processors.html",
        processors=processors
    )




#   -----------------------
#     
#      ADD CREDIT PROCESSORS
#
#       spa_id good  
#  ----------------------


@app.route("/credit_processors/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_credit_processor():
    spa_id = current_spa_id()
    
    if request.method == "POST":
        credit_processor_name = request.form.get("credit_processor_name", "").strip()
        percentage_fee = float(request.form.get("percentage_fee") or 0)
        flat_fee = float(request.form.get("flat_fee") or 0)
        additional_fee = float(request.form.get("additional_fee") or 0)
            
        if not credit_processor_name:
            flash("Processor name is required.", "error")
            return redirect(url_for("add_credit_processor"))
            
        conn = get_db_connection()
        cur = conn.cursor()
        
        try:
            cur.execute("""
                INSERT INTO credit_processors (
                    spa_id,
                    credit_processor_name,
                    percentage_fee,
                    flat_fee,
                    additional_fee,
                    is_active
                )
                VALUES (%s, %s, %s, %s, %s, TRUE)
            """, (
                spa_id,
                credit_processor_name,
                percentage_fee,
                flat_fee,
                additional_fee
            ))

            conn.commit()
            flash("Credit processor added successfully.", "success")

        except Exception as e:
            conn.rollback()
            flash(f"Error adding processor: {e}", "error")

        finally:
            cur.close()
            conn.close()

        return redirect(url_for("credit_processors"))

    return render_template("add_credit_processor.html")





#   ----------------------------
#     
#     EDIT CREDIT PROCESSOR
#
#     spa_id good  
#  ----------------------------



@app.route("/credit_processors/edit/<int:credit_processor_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_credit_processor(credit_processor_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        credit_processor_name = request.form.get("credit_processor_name", "").strip()
        percentage_fee = request.form.get("percentage_fee") or 0
        flat_fee = request.form.get("flat_fee") or 0
        additional_fee = request.form.get("additional_fee") or 0

        if not credit_processor_name:
            cur.close()
            conn.close()
            flash("Processor name is required.", "error")
            return redirect(url_for("edit_credit_processor", credit_processor_id=credit_processor_id))

        cur.execute("""
            UPDATE credit_processors
            SET credit_processor_name = %s,
                percentage_fee = %s,
                flat_fee = %s,
                additional_fee = %s
            WHERE credit_processor_id = %s
              AND spa_id = %s
        """, (
            credit_processor_name,
            percentage_fee,
            flat_fee,
            additional_fee,
            credit_processor_id,
            spa_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Credit processor updated successfully.", "success")
        return redirect(url_for("credit_processors"))

    cur.execute("""
        SELECT
            credit_processor_id,
            credit_processor_name,
            percentage_fee,
            flat_fee,
            additional_fee,
            is_active
        FROM credit_processors
        WHERE credit_processor_id = %s
          AND spa_id = %s
    """, (credit_processor_id, spa_id))

    processor = cur.fetchone()

    cur.close()
    conn.close()

    if not processor:
        flash("Credit processor not found.", "error")
        return redirect(url_for("credit_processors"))

    return render_template(
        "edit_credit_processor.html",
        processor=processor
    )




#   ------------------------------
#     TOGGLE ACTIVE/DEACTIVATE
#     CREDIT PROCESSORS
#  
#     spa_id good
#  ------------------------------


@app.route("/credit_processors/toggle/<int:credit_processor_id>", methods=["POST"])
@login_required
@spa_required
def toggle_credit_processor(credit_processor_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE credit_processors
        SET is_active = NOT is_active
        WHERE credit_processor_id = %s
          AND spa_id = %s
    """, (credit_processor_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Credit processor status updated.", "success")
    return redirect(url_for("credit_processors"))












#   -----------------------
#
#     FUTURE CALENDAR SYNC
#
#  ----------------------
    
    
@app.route("/admin/sync_calendar")
def sync_calendar():
    flash(
        "Google Calendar sync is not active yet. This feature will be added after deployment.",
        "info"
    )
    return redirect(url_for("admin"))





#   ------------------------------------------
#
#
#
#   ------------------------------------------


            
#   ------------------------------------------
#     MAILGUN INCOMING MAIL
#
#     MAILGUN GODADDY BOOKING
#   ------------------------------------------
    
    
@app.route("/mailgun/godaddy-booking", methods=["POST"])
def mailgun_godaddy_booking():

    print("MAILGUN GODADDY ROUTE HIT", flush=True)
    print("SUBJECT:", request.form.get("subject", ""), flush=True)
    print("FORM KEYS:", list(request.form.keys()), flush=True)

    subject = request.form.get("subject", "")

    body = (
        request.form.get("stripped-text")
        or request.form.get("body-plain")
        or request.form.get("body")
    )

    if not body:
        return {"error": "No email body received"}, 400

    spa_id = 1

    result = import_godaddy_booking(body, spa_id, subject)

    return result, 200








#   ----------------------------------------------
#     
#   
#    GoDaddy IMPORTS
#
#
#
#
#   --------------------------------------------
        

@app.route("/godaddy-imports")
@login_required
@spa_required
def godaddy_imports():
    spa_id = current_spa_id()

    review_filter = request.args.get("review", "").strip().lower()
    from_coach = request.args.get("from_coach", "").strip() == "1"    

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            a.appointment_id,
            a.appointment_date,
            a.appointment_time,
            c.first_name,
            c.last_name,
            c.phone,
            c.email,
            a.external_service_name,
            a.external_order_id,
            a.status,
            a.notes,
            COALESCE(a.import_reviewed, FALSE) AS import_reviewed
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
        AND a.spa_id = c.spa_id
        WHERE a.spa_id = %s
        AND a.external_source = 'godaddy'
    """

    params = [spa_id]

    if review_filter == "unreviewed":
        query += """
            AND COALESCE(a.import_reviewed, FALSE) = FALSE
        """

    query += """
        ORDER BY a.appointment_date DESC, a.appointment_time DESC
    """

    cur.execute(query, tuple(params))
    imports = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "godaddy_imports.html",
        imports=imports,
        review_filter=review_filter,
        import_count=len(imports),
        from_coach=from_coach
    )






        
            
           
        
#   ----------------------------------------------
#
#
#       GMAIL BOOKING POLL    MANUAL
#
#             
#
#
#   --------------------------------------------
    

@app.route("/test-gmail-booking-poll")
@login_required
@spa_required
def test_gmail_booking_poll():
    # later this will:
    # 1. connect to Gmail
    # 2. find GoDaddy booking emails
    # 3. read email body
    # 4. call import_godaddy_booking(body, spa_id, subject)
    return "Gmail booking poll test route ready."






###############################################
#
#   GODADDY PARSER HELPER
#
#
##############################################

from decimal import Decimal

def parse_money(value):
    if not value:
        return None

    value = value.replace("$", "").replace(",", "").strip()

    try:
        return Decimal(value)
    except Exception:
        return None






    
    
            
    
        
#   ----------------------------------------------   
#
#
#    GoDaddy WebHook Parser
#
#
#
#
#   --------------------------------------------
    



import re
from datetime import datetime


def parse_godaddy_booking_email(body):
    data = {}

    # ---------------------------------------------------------
    # Basic booking fields
    # ---------------------------------------------------------
    order_match = re.search(
        r"Order #\s+(.+)",
        body
    )
    if order_match:
        data["order_number"] = order_match.group(1).strip()

    name_match = re.search(
        r"Name:\s*(.+)",
        body
    )
    if name_match:
        data["customer_name"] = name_match.group(1).strip()

    phone_match = re.search(
        r"Phone:\s*(.+)",
        body
    )
    if phone_match:
        data["phone"] = phone_match.group(1).strip()

    email_match = re.search(
        r"Email:\s*(.+)",
        body
    )
    if email_match:
        data["email"] = email_match.group(1).strip()

    service_match = re.search(
        r"What:\s*(.+)",
        body
    )
    if service_match:
        data["service"] = service_match.group(1).strip()

    staff_match = re.search(
        r"Staff:\s*\r?\n\s*([^\r\n]+)",
        body,
        re.IGNORECASE
    )

    booking["provider_name_at_booking"] = (
        staff_match.group(1).strip()
        if staff_match
        else None
    )

    # ---------------------------------------------------------
    # Appointment date, time, and duration
    # ---------------------------------------------------------
    when_line_match = re.search(
        r"^When:\s*(.+)$",
        body,
        re.IGNORECASE | re.MULTILINE
    )

    if not when_line_match:
        raise ValueError(
            "GoDaddy booking email is missing the When field."
        )

    raw_when_line = when_line_match.group(1).strip()


    duration_match = re.search(
        r"\(\s*"
        r"(?:(\d+)\s*(hours?|hrs?))?"
        r"\s*"
        r"(?:(\d+)\s*(mins?|minutes?))?"
        r"\s*\)\s*$",
        raw_when_line,
        re.IGNORECASE
    )

    if not duration_match:
        raise ValueError(
            "Could not parse GoDaddy appointment duration: "
            f"{raw_when_line!r}"
        )

    hours_value = int(duration_match.group(1) or 0)
    minutes_value = int(duration_match.group(3) or 0)

    if hours_value == 0 and minutes_value == 0:
        raise ValueError(
            "GoDaddy appointment duration was empty: "
            f"{raw_when_line!r}"
        )

    duration_minutes = (
        hours_value * 60
        + minutes_value
    )

    raw_when = raw_when_line[
        :duration_match.start()
    ].strip()


    raw_when = re.sub(
        r"\s+",
        " ",
        raw_when
    ).strip()

    raw_when = re.sub(
        r"(\d)(AM|PM)\b",
        r"\1 \2",
        raw_when,
        flags=re.IGNORECASE
    )

    date_formats = [
        "%A, %B %d, %Y at %I:%M %p",
        "%A, %B %d, %Y at %I %p",
        "%A, %B %d, %Y at %H:%M",
        "%B %d, %Y at %I:%M %p",
        "%B %d, %Y at %I %p",
        "%m/%d/%Y at %I:%M %p",
        "%m/%d/%Y %I:%M %p",
    ]

    appointment_datetime = None

    for fmt in date_formats:
        try:
            appointment_datetime = datetime.strptime(
                raw_when,
                fmt
            )
            break
        except ValueError:
            continue

    if appointment_datetime is None:
        raise ValueError(
            "Could not parse GoDaddy appointment datetime: "
            f"{raw_when!r}. Full When value: "
            f"{raw_when_line!r}"
        )


    data["appointment_datetime"] = appointment_datetime
    data["duration_minutes"] = duration_minutes

    # ---------------------------------------------------------
    # Payment information
    # ---------------------------------------------------------
    payment_match = re.search(
        r"Payment status:\s*(.+)",
        body
    )
    if payment_match:
        data["payment_status"] = (
            payment_match.group(1).strip()
        )

    subtotal_match = re.search(
        r"Subtotal\s+\$([\d,]+\.\d{2})",
        body
    )
    if subtotal_match:
        data["subtotal"] = parse_money(
            subtotal_match.group(1)
        )

    order_total_match = re.search(
        r"Order Total\s+\$([\d,]+\.\d{2})",
        body
    )
    if order_total_match:
        data["order_total"] = parse_money(
            order_total_match.group(1)
        )

    paid_checkout_match = re.search(
        r"Paid at checkout\s+\$([\d,]+\.\d{2})",
        body
    )

    if paid_checkout_match:
        paid_amount = parse_money(
            paid_checkout_match.group(1)
        )

        data["paid_at_checkout_amount"] = paid_amount
        data["paid_at_checkout"] = paid_amount > 0
    else:
        data["paid_at_checkout"] = False
        data["paid_at_checkout_amount"] = 0

    data["raw_email_body"] = body

    return data























#   ------------------------------------------
#           
#   
#       
#   ------------------------------------------






#   ----------------------------------------------
#       
#       
#     SMS   HELPER
#   
#       
#       
#
#   --------------------------------------------



def get_sms_template(spa_id, template_type):
    result = build_communication(
        spa_id=spa_id,
        channel="sms",
        template_type=template_type,
        merge_data={}
    )

    if not result.get("success"):
        return None

    return result.get("message_body")




 




##################################################################
#
#   HELPER   SEND SMS MESSAGE
#
#################################################################


# ==========================================================
# Internal provider wrapper.   
#
# Called only by send_compliant_sms().
#
# This function abstracts the SMS provider (Telnyx today,
# future providers tomorrow) from the application.
# ==========================================================



def send_sms_message(to_phone, message_body):

    sms_enabled = os.getenv("SMS_ENABLED", "false").lower() == "true"

    final_message_body = message_body

    if not sms_enabled:
        return {
            "success": False,
            "status": "logged",
            "provider_message_id": None,
            "provider_status": None,
            "provider_error_code": None,
            "provider_error_message": "SMS sending disabled",
            "final_message_body": final_message_body
        }

    try:
        
        print("FINAL SMS BODY BEING SENT:", final_message_body, flush=True)
        result = send_sms_telnyx(to_phone, final_message_body)

        print("TELNYX RESULT:", result, flush=True)

        message_data = result.get("data", result)

        return {
            "success": True,
            "status": "sent",
            "provider_message_id": message_data.get("id"),
            "provider_status": message_data.get("record_type", "accepted"),
            "provider_error_code": None,
            "provider_error_message": None,
            "final_message_body": final_message_body
        }

    except Exception as e:
        return {
            "success": False,
            "status": "failed",
            "provider_message_id": None,
            "provider_status": None,
            "provider_error_code": None,
            "provider_error_message": str(e),
            "final_message_body": final_message_body
        }






#   ----------------------------------------------
#
#
#     BIRTHDAY HELPER
#
#
#
#
#   --------------------------------------------




def get_birthday_campaign_year(birth_date, today):
    current_year = today.year
    current_month = today.month
    today_day = today.day

    if current_month == 12:
        next_month = 1
        next_month_year = current_year + 1
    else:
        next_month = current_month + 1
        next_month_year = current_year

    birth_month = birth_date.month

    if birth_month == current_month:
        return current_year

    if today_day >= 15 and birth_month == next_month:
        return next_month_year

    return None







#   --------------------------------------------------
#
#
#              HELPER   QUERY
#
#              spa-id good
#  ----------------------------------------------------




from datetime import date


def get_loan_contribution_rows(spa_id, start_date=None, end_date=None):
    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            contribution_date AS activity_date,
            'Owner Contribution' AS activity_type,
            funding_source AS description,
            amount,
            NULL AS payment_method,
            NULL AS loan_name,
            notes
        FROM owner_contributions
        WHERE spa_id = %s
    """
    params = [spa_id]

    if start_date:
        query += " AND contribution_date >= %s"
        params.append(start_date)

    if end_date:
        query += " AND contribution_date <= %s"
        params.append(end_date)

    query += """

        UNION ALL

        SELECT
            reimbursement_date AS activity_date,
            'Owner Reimbursement' AS activity_type,
            NULL AS description,
            amount,
            payment_method,
            NULL AS loan_name,
            notes
        FROM owner_reimbursements
        WHERE spa_id = %s
    """
    params.append(spa_id)

    if start_date:
        query += " AND reimbursement_date >= %s"
        params.append(start_date)

    if end_date:
        query += " AND reimbursement_date <= %s"
        params.append(end_date)

    query += """

        UNION ALL

        SELECT
            lp.payment_date AS activity_date,
            'Loan Payment' AS activity_type,
            CONCAT('Principal: $', lp.principal_paid,
                   ' / Interest: $', lp.interest_paid) AS description,
            lp.total_payment AS amount,
            NULL AS payment_method,
            bl.loan_name,
            lp.notes
        FROM loan_payments lp
        LEFT JOIN business_loans bl
            ON lp.loan_id = bl.loan_id
           AND lp.spa_id = bl.spa_id
        WHERE lp.spa_id = %s
    """
    params.append(spa_id)

    if start_date:
        query += " AND lp.payment_date >= %s"
        params.append(start_date)

    if end_date:
        query += " AND lp.payment_date <= %s"
        params.append(end_date)

    query += " ORDER BY activity_date DESC"

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    cur.close()
    conn.close()
    return rows









            
#   ------------------------------------------------
#
#    >>>>>>>>>>>>  GO DADDY  <<<<<<<<<<<<<<<<<
#
#    GoDaddy Parser   
#
#    
#
#   -----------------------------------------------

@app.route("/test-godaddy-parser")
def test_godaddy_parser():

    with open("test_booking.txt", "r") as f:
        body = f.read()

    booking = parse_godaddy_booking_email(body)

    return f"""
    <pre>
    {booking}
    </pre>
    """






##########################################
# >>>>>>>>>>>  GO DADDY  <<<<<<<<<<<<<<<<<
#
#    GoDaddy EMAIL  Parser   
#
#    
#
#   -----------------------------------------------


import re
from datetime import datetime

def parse_godaddy_email_body(body):
    booking = {}

    order_match = re.search(r"Order #\s*(C-[A-Z0-9]+)", body)
    booking["external_order_id"] = order_match.group(1) if order_match else None

    subtotal_match = re.search(r"Subtotal\s*\$?([\d.]+)", body)
    booking["subtotal"] = float(subtotal_match.group(1)) if subtotal_match else None

    total_match = re.search(r"Order Total\s*\$?([\d.]+)", body)
    booking["order_total"] = float(total_match.group(1)) if total_match else None

    paid_match = re.search(r"Paid at checkout\s*\$?([\d.]+)", body)
    paid_amount = float(paid_match.group(1)) if paid_match else 0
    booking["paid_at_checkout"] = paid_amount > 0

    name_match = re.search(r"Name:\s*(.+)", body)
    booking["customer_name"] = name_match.group(1).strip() if name_match else None

    phone_match = re.search(r"Phone:\s*(.+)", body)
    booking["phone"] = phone_match.group(1).strip() if phone_match else None
    if booking["phone"] and len(booking["phone"]) == 10:
        booking["phone"] = (
            f"({booking['phone'][0:3]}) "
            f"{booking['phone'][3:6]}-"
            f"{booking['phone'][6:]}"
        )

    email_match = re.search(r"Email:\s*(.+)", body)
    booking["email"] = email_match.group(1).strip() if email_match else None

    what_match = re.search(r"What:\s*(.+)", body)
    booking["service_name"] = what_match.group(1).strip() if what_match else None


    what_match = re.search(r"What:\s*(.+)", body)
    booking["service_name"] = what_match.group(1).strip() if what_match else None

    staff_match = re.search(
        r"Staff:\s*\r?\n\s*([^\r\n]+)",
        body,
        re.IGNORECASE
    )

    booking["provider_name_at_booking"] = (
        staff_match.group(1).strip()
        if staff_match
        else None
    )

    when_match = re.search(r"When:\s*(.+?)\s*\((\d+)\s*hour", body)



    when_match = re.search(r"When:\s*(.+?)\s*\((\d+)\s*hour", body)
    if when_match:
        when_text = when_match.group(1).strip()
        duration_hours = int(when_match.group(2))

        dt = datetime.strptime(when_text, "%A, %B %d, %Y at %I:%M%p")

        booking["appointment_date"] = dt.date()
        booking["appointment_time"] = dt.time()
        booking["duration_minutes"] = duration_hours * 60
    else:
        booking["appointment_date"] = None
        booking["appointment_time"] = None
        booking["duration_minutes"] = None


    print(
        "[GODADDY PARSER] Provider:",
        booking.get("provider_name_at_booking")
    )

    payment_match = re.search(r"Payment status:\s*(.+)", body)
    booking["payment_status"] = payment_match.group(1).strip() if payment_match else None

    booking["external_source"] = "godaddy"

    return booking



######################################
#
#   TEST TEST. TODO
#
#   TEST GODADDY PARSER
#
#########################################################



@app.route("/booking-email-import", methods=["GET", "POST"])
@login_required
@spa_required
def booking_email_import():
    body = ""

    if request.method == "POST":
        body = request.form.get("email_body", "").strip()
        booking = parse_godaddy_email_body(body)

        return render_template(
            "booking_email_import.html",
            booking=booking,
            email_body=body
        )

    return render_template(
        "booking_email_import.html",
        booking=None,
        email_body=body
    )



















#   ------------------------------------------------
#
#    >>>>>>>>>>>>  GO DADDY  <<<<<<<<<<<<<<<<<
#
#    GoDaddy TEST CREATE APPOINTMENT
#           
#      GODADDY IMPORT
#               PARSER VERSION GODADDY_V2
#   -----------------------------------------------




def import_godaddy_booking(body, spa_id, subject=""):
    booking = parse_godaddy_booking_email(body)

    

    price_at_booking = booking.get("subtotal") or booking.get("order_total")
    from datetime import datetime

    conn = get_db_connection()
    cur = conn.cursor()

    # 1. Prevent duplicate appointment
    cur.execute("""
        SELECT appointment_id
        FROM appointments
        WHERE spa_id = %s
          AND external_source = %s
          AND external_order_id = %s
    """, (spa_id, "godaddy", booking["order_number"]))

    existing = cur.fetchone()

    if existing:
        cur.close()
        conn.close()
        return {
            "status": "duplicate",
            "message": "This GoDaddy booking was already imported.",
            "order_number": booking["order_number"]
        }

    # 2. Find existing client by email first
    client = None

    if booking.get("email"):
        cur.execute("""
            SELECT client_id
            FROM clients
            WHERE spa_id = %s
              AND LOWER(email) = LOWER(%s)
            LIMIT 1
        """, (spa_id, booking["email"]))

        client = cur.fetchone()

    # If no email match, try phone
    if not client and booking.get("phone"):
        cur.execute("""
            SELECT client_id
            FROM clients
            WHERE spa_id = %s
              AND phone = %s
            LIMIT 1
        """, (spa_id, booking["phone"]))

        client = cur.fetchone()

    # 3. Create client if not found
    if client:
        client_id = client[0]
    else:
        name_parts = booking["customer_name"].split(" ", 1)
        first_name = name_parts[0]
        last_name = name_parts[1] if len(name_parts) > 1 else ""

        cur.execute("""
            INSERT INTO clients (
                spa_id,
                first_name,
                last_name,
                phone,
                email
            )
            VALUES (%s, %s, %s, %s, %s)
            RETURNING client_id
        """, (
            spa_id,
            first_name,
            last_name,
            booking["phone"],
            booking["email"]
        ))

        client_id = cur.fetchone()[0]


    # Parser returns True/False
    paid_at_checkout = bool(booking.get("paid_at_checkout"))

    

    # 4. Insert appointment
    cur.execute("""
        INSERT INTO appointments (
            spa_id,
            client_id,
            appointment_date,
            appointment_time,
            duration_minutes,
            external_service_name,
            provider_name_at_booking,
            status,
            notes,
            external_source,
            external_order_id,
            external_email_subject,
            external_email_body,
            price_at_booking,
            subtotal,
            order_total,
            paid_at_checkout,
            imported_at,
            import_reviewed,
            parser_version,
            import_status
        )
        VALUES (%s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, 
                %s, %s, %s, %s, %s, 
                %s, %s, %s, %s, %s,
                %s
        )
        RETURNING appointment_id
    """, (
        spa_id,
        client_id,
        booking["appointment_datetime"].date(),
        booking["appointment_datetime"].time(),
        booking["duration_minutes"],
        booking["service"],
        booking.get("provider_name_at_booking"),
        "booked",
        (
            f"Imported from GoDaddy. "
            f"Service: {booking['service']}. "
            f"Payment status: {booking['payment_status']}"
        ),
        "godaddy",
        booking["order_number"],
        subject,
        body,
        price_at_booking,
        booking.get("subtotal"),
        booking.get("order_total"),
        paid_at_checkout,
        datetime.now(),
        False,
        "godaddy_v3",
        "Imported"
    ))

    appointment_id = cur.fetchone()[0]


    conn.commit()
    cur.close()
    conn.close()

    return {
        "status": "imported",
        "message": "GoDaddy booking imported successfully.",
        "order_number": booking["order_number"],
        "client_id": client_id,
        "appointment_id":appointment_id
    }








#   ------------------------------------------------
#     
#     
#     VIEW RAW GODADDY EMAILS
#     
#   
#   
#   
#   -----------------------------------------------


@app.route("/godaddy-imports/<int:appointment_id>/raw")
@login_required
@spa_required
def godaddy_import_raw(appointment_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            a.appointment_id,
            a.external_source,
            a.external_order_id,
            a.external_service_name,
            a.appointment_date,
            a.appointment_time,
            a.duration_minutes,
            a.status,
            a.notes,
            a.price_at_booking,
            a.external_email_subject,
            a.external_email_body,
            a.import_reviewed,
            a.import_reviewed_at,
            a.import_reviewed_by,
            a.parser_version,
            a.import_status,
            a.subtotal,
            a.order_total,
            a.paid_at_checkout,
            c.first_name,
            c.last_name,
            c.phone,
            c.email
        FROM appointments a
        LEFT JOIN clients c
            ON a.client_id = c.client_id
        WHERE a.appointment_id = %s
          AND a.spa_id = %s
          AND a.external_source = 'godaddy'
    """, (appointment_id, spa_id))

    row = cur.fetchone()

    cur.close()
    conn.close()

    if not row:
        return "Import not found.", 404

    appointment = {
        "appointment_id": row[0],
        "external_source": row[1],
        "external_order_id": row[2],
        "external_service_name": row[3],
        "appointment_date": row[4],
        "appointment_time": row[5],
        "duration_minutes": row[6],
        "status": row[7],
        "notes": row[8],
        "price_at_booking": row[9],
        "external_email_subject": row[10],
        "external_email_body": row[11],

        # Import audit fields
        "import_reviewed": row[12],
        "import_reviewed_at": row[13],
        "import_reviewed_by": row[14],
        "parser_version": row[15],
        "import_status": row[16],

        # Revenue import fields
        "subtotal": row[17],
        "order_total": row[18],
        "paid_at_checkout": row[19],

        # Client fields
        "client_name": f"{row[20] or ''} {row[21] or ''}".strip(),
        "phone": row[22],
        "email": row[23],
    }

    if appointment["phone"] and len(appointment["phone"]) == 10:
        appointment["phone"] = (
            f"({appointment['phone'][0:3]}) "
            f"{appointment['phone'][3:6]}-"
            f"{appointment['phone'][6:]}"
        )


    return render_template(
        "import_review.html",
        appointment=appointment
    )












##########################################
#
#   BOOKING EMIAL IMPORT --CONFIRM
#
#
###############################################



@app.route("/booking-email-import/confirm", methods=["POST"])
@login_required
@spa_required
def booking_email_import_confirm():
    spa_id = current_spa_id()
    body = request.form.get("email_body", "").strip()

    if not body:
        flash("No email body found to import.", "warning")
        return redirect(url_for("booking_email_import"))

    result = import_godaddy_booking(
        body,
        spa_id,
        "Forwarded GoDaddy booking email"
    )

    if result["status"] == "duplicate":
        flash(result["message"], "warning")
    else:
        flash(result["message"], "success")

    return redirect(url_for("calendar_view"))


















############################################################
#
#   TEST. TEST. TEST.   TODO
#
#   DEBUG
#
#
#########################################################

@app.route("/debug-db")
@login_required
def debug_db():
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT current_database(), current_user, inet_server_addr(), inet_server_port()")
    db_info = cur.fetchone()

    cur.execute("""
        SELECT column_name
        FROM information_schema.columns
        WHERE table_name = 'appointments'
          AND column_name IN (
            'subtotal',
            'order_total',
            'paid_at_checkout',
            'import_reviewed_at',
            'import_reviewed_by',
            'parser_version',
            'import_status'
          )
        ORDER BY column_name
    """)
    columns = cur.fetchall()

    cur.execute("""
        SELECT appointment_id, external_source, external_order_id, imported_at
        FROM appointments
        ORDER BY appointment_id DESC
        LIMIT 5
    """)
    latest = cur.fetchall()

    cur.close()
    conn.close()

    return {
        "db_info": db_info,
        "columns": columns,
        "latest_appointments": latest
    }




#   --------------------------
#
#   
#     GODADDY CREATE APPOINTMENT
#
#   
#
#   ----------------------
        

@app.route("/test-godaddy-create-appointment")
@login_required
@spa_required
def test_godaddy_create_appointment():
    spa_id = current_spa_id()

    with open("test_booking.txt", "r") as f:
        body = f.read()

    result = import_godaddy_booking(body, spa_id)

    return f"<pre>{result}</pre>"






#   --------------------------
#
#   
#     GODADDY IMPORTS REVIEWED
#
#   
#
#   ----------------------
        
@app.route("/godaddy-imports/reviewed/<int:appointment_id>", methods=["POST"])
@login_required
@spa_required
def mark_godaddy_import_reviewed(appointment_id):
    spa_id = current_spa_id()
    user_id = session.get("user_id")

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE appointments
            SET
                import_reviewed = TRUE,
                import_reviewed_at = COALESCE(import_reviewed_at, NOW()),
                import_reviewed_by = COALESCE(import_reviewed_by, %s),
                owner_reviewed = TRUE,
                owner_reviewed_at = COALESCE(owner_reviewed_at, NOW()),
                import_status = 'Reviewed'
            WHERE appointment_id = %s
            AND spa_id = %s
            AND LOWER(COALESCE(external_source, '')) = 'godaddy'
        """, (
            session.get("user_id"),
            appointment_id,
            spa_id
        ))

        if cur.rowcount == 0:
            conn.rollback()
            flash("The GoDaddy appointment could not be found.", "warning")
        else:
            conn.commit()
            flash("GoDaddy import marked as reviewed.", "success")

    except Exception as e:
        conn.rollback()
        print(f"MARK GODADDY IMPORT REVIEWED ERROR: {e}")
        flash("The GoDaddy import could not be marked as reviewed.", "danger")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("godaddy_imports"))
















#   --------------------------
#
#       RENDER GET MAIL
#    GET MAIL FROM GMAIL
#
#     GODADAY MAIL        
#
#   ----------------------



def poll_gmail_for_godaddy_bookings(spa_id):
    log_godaddy("Polling Gmail for GoDaddy booking emails...")

    gmail_user = os.getenv("GMAIL_BOOKING_EMAIL")
    gmail_pass = os.getenv("GMAIL_BOOKING_APP_PASSWORD")

    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(gmail_user, gmail_pass)
    mail.select("inbox")

    status, messages = mail.search(None, "ALL")
    email_ids = messages[0].split()

    log_godaddy(f"Booking emails found in inbox: {len(email_ids)}")

    if not email_ids:
        mail.logout()
        log_godaddy("No emails found.")
        return {"status": "no_emails_found"}

    results = []

    # Newest emails first
    for email_id in reversed(email_ids):
        started_at = time.perf_counter()

        log_godaddy(f"Checking email ID: {email_id.decode()}")

        status, msg_data = mail.fetch(email_id, "(RFC822)")
        raw_email = msg_data[0][1]

        email_message = email.message_from_bytes(raw_email)

        subject, encoding = decode_header(email_message["Subject"])[0]

        if isinstance(subject, bytes):
            subject = subject.decode(encoding or "utf-8")

        log_godaddy(f"Subject: {subject}")

        body = ""

        if email_message.is_multipart():
            for part in email_message.walk():
                content_type = part.get_content_type()
                content_disposition = str(part.get("Content-Disposition"))

                if (
                    content_type == "text/plain"
                    and "attachment" not in content_disposition
                ):
                    body = part.get_payload(decode=True).decode(errors="ignore")
                    break
        else:
            body = email_message.get_payload(decode=True).decode(errors="ignore")

        if (
            "Order #" not in subject
            and "Order #" not in body
        ):
            log_godaddy("Skipped: no Order # found.")
            results.append({
                "email_id": email_id.decode(),
                "subject": subject,
                "status": "skipped_no_order_number"
            })
            continue

        if (
            "Order #" not in body
            or "What:" not in body
            or "When:" not in body
            or "Payment status:" not in body
        ):
            log_godaddy("Skipped: email did not match booking format.")
            results.append({
                "email_id": email_id.decode(),
                "subject": subject,
                "status": "skipped_not_booking_email"
            })
            continue

        try:
            log_godaddy("Booking email matched. Starting import.")

            result = import_godaddy_booking(
                body,
                spa_id,
                subject
            )

            log_godaddy(f"Import result: {result}")

            processing_time_ms = int(
                (time.perf_counter() - started_at) * 1000
            )

            log_booking_import(
                spa_id=spa_id,
                source="godaddy",
                email_id=email_id.decode(),
                external_order_id=result.get("order_number"),
                email_subject=subject,
                status=result.get("status"),
                appointment_id=result.get("appointment_id"),
                client_id=result.get("client_id"),
                parser_version="godaddy_v1",
                error_message=None,
                processing_time_ms=processing_time_ms
            )

            results.append({
                "email_id": email_id.decode(),
                "subject": subject,
                "result": result
            })

        except KeyError as e:
            log_godaddy(f"Parse failed. Missing field: {e}")
            log_godaddy(
                "Email will be checked again during the next poll."
            )

            processing_time_ms = int(
                (time.perf_counter() - started_at) * 1000
            )

            log_booking_import(
                spa_id=spa_id,
                source="godaddy",
                email_id=email_id.decode(),
                external_order_id=None,
                email_subject=subject,
                status="parse_failed",
                parser_version="godaddy_v1",
                error_message=f"Missing field: {e}",
                processing_time_ms=processing_time_ms
            )

            results.append({
                "email_id": email_id.decode(),
                "subject": subject,
                "status": "parse_failed",
                "missing_field": str(e)
            })

        except Exception as e:
            log_godaddy(f"Import failed: {e}")
            log_godaddy(
                "Email will be checked again during the next poll."
            )

            processing_time_ms = int(
                (time.perf_counter() - started_at) * 1000
            )

            log_booking_import(
                spa_id=spa_id,
                source="godaddy",
                email_id=email_id.decode(),
                external_order_id=None,
                email_subject=subject,
                status="import_failed",
                appointment_id=None,
                client_id=None,
                parser_version="godaddy_v1",
                error_message=str(e),
                processing_time_ms=processing_time_ms
            )

            results.append({
                "email_id": email_id.decode(),
                "subject": subject,
                "status": "import_failed",
                "error": str(e)
            })

    mail.logout()

    log_godaddy(f"Polling completed. Processed count: {len(results)}")

    return {
        "status": "completed",
        "processed_count": len(results),
        "results": results
    }









#   ------------------------
#
#   TEST GMAIL
#
#
#
#
#   ------------------------





@app.route("/test-gmail-import")
@login_required
@spa_required
def test_gmail_import():
    spa_id = current_spa_id()
    return poll_gmail_for_godaddy_bookings(spa_id)















#   ------------------------------------------------
#
#
#           
#     SMS     CAMPAIGNS
#       
#      
#   
#   -----------------------------------------------




#   --------------------------
#
#
#     SMS WEBHOOK
#
#.  LEGACY TWILIO WEBHOOK
#       TODO
#   ----------------------

# PSP_CLEANUP:
# Legacy Twilio webhook.
# Safe to remove after confirming all messaging uses Telnyx.

@app.route("/sms/webhook", methods=["POST"])
def sms_webhook():
    from_number = request.form.get("From")
    body = request.form.get("Body", "").strip().lower()

    if body in ["stop", "unsubscribe", "cancel"]:
        conn = get_db_connection()
        cur = conn.cursor()

        # Get spa_id + client_id
        cur.execute("""
            SELECT spa_id, client_id
            FROM clients
            WHERE phone = %s
        """, (from_number,))
        
        row = cur.fetchone()

        if row:
            spa_id, client_id = row

            # Update client
            cur.execute("""
                UPDATE clients
                SET
                    ok_to_text = FALSE,
                    sms_opt_in = FALSE,
                    sms_opt_out = TRUE
                WHERE phone = %s
            """, (from_number,))

            # Log it
            cur.execute("""
                INSERT INTO sms_consent_log (
                    spa_id,
                    client_id,
                    phone_number,
                    consent_given,
                    consent_method,
                    consent_text
                )
                VALUES (%s, %s, %s, FALSE, %s, %s)
            """, (
                spa_id,
                client_id,
                from_number,
                "opt_out",
                "User replied STOP"
            ))

            conn.commit()

        cur.close()
        conn.close()

    return ("", 204)








            
#   ---------------------
#
#      SMS messaging settings
#
#
#   ----------------------

@app.route("/client/<int:client_id>/messaging-settings", methods=["GET", "POST"])
@login_required
@spa_required
def client_messaging_settings(client_id):
    spa_id = current_spa_id()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            client_id,
            first_name,
            last_name,
            phone,
            email,
            ok_to_call,
            sms_opt_in,
            sms_opt_out,
            email_opt_in,
            email_opt_out
        FROM clients
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, spa_id))

    client = cur.fetchone()

    if not client:
        cur.close()
        conn.close()
        flash("Client not found.", "error")
        return redirect(url_for("clients"))

    if request.method == "POST":
        ok_to_call = request.form.get("ok_to_call") == "on"

        sms_opt_in = request.form.get("sms_opt_in") == "on"
        sms_opt_out = request.form.get("sms_opt_out") == "on"

        email_opt_in = request.form.get("email_opt_in") == "on"
        email_opt_out = request.form.get("email_opt_out") == "on"

        phone = client[3]
        email = client[4]

        cur.execute("""
            UPDATE clients
            SET
                ok_to_call = %s,
                sms_opt_in = %s,
                sms_opt_out = %s,
                email_opt_in = %s,
                email_opt_out = %s
            WHERE client_id = %s
              AND spa_id = %s
        """, (
            ok_to_call,
            sms_opt_in,
            sms_opt_out,
            email_opt_in,
            email_opt_out,
            client_id,
            spa_id
        ))

        if phone:
            cur.execute("""
                INSERT INTO sms_consent_log (
                    spa_id,
                    client_id,
                    phone_number,
                    consent_given,
                    consent_method,
                    consent_text
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                spa_id,
                client_id,
                phone,
                sms_opt_in,
                "manual_admin_update",
                "Messaging preference updated by staff."
            ))

        if email:
            cur.execute("""
                INSERT INTO email_consent_log (
                    spa_id,
                    client_id,
                    email,
                    consent_given,
                    consent_method
                )
                VALUES (%s, %s, %s, %s, %s)
            """, (
                spa_id,
                client_id,
                email,
                email_opt_in,
                "manual_admin_update"
            ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Messaging settings updated.", "success")
        return redirect(url_for("client_messaging_settings", client_id=client_id))

    cur.execute("""
        SELECT
            phone_number,
            consent_given,
            consent_method,
            consent_text,
            created_at
        FROM sms_consent_log
        WHERE spa_id = %s
          AND client_id = %s
        ORDER BY created_at DESC
        LIMIT 20
    """, (spa_id, client_id))

    sms_logs = cur.fetchall()

    cur.execute("""
        SELECT
            email,
            consent_given,
            consent_method,
            created_at
        FROM email_consent_log
        WHERE spa_id = %s
          AND client_id = %s
        ORDER BY created_at DESC
        LIMIT 20
    """, (spa_id, client_id))

    email_logs = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "client_messaging_settings.html",
        client=client,
        sms_logs=sms_logs,
        email_logs=email_logs
    )










#   ---------------------
#
#         BIRTHDAY SMS
#
#
#   ----------------------



@app.route("/birthday-sms/send-month", methods=["POST"])
@login_required
@spa_required
def send_birthday_sms_month():
    spa_id = current_spa_id()
    spa_now = get_spa_now()
    today = spa_now.date()
    campaign_year = today.year
    campaign_month = today.month


    spa_id = current_spa_id()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    spa_name = get_spa_name(spa_id)

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            c.client_id,
            c.first_name,
            c.phone,
            c.birth_date
        FROM clients c
        WHERE c.spa_id = %s
          AND c.active_client = TRUE
          AND c.phone IS NOT NULL
          AND c.phone <> ''
          AND c.birth_date IS NOT NULL
          AND EXTRACT(MONTH FROM c.birth_date) = %s
          AND c.ok_to_text = TRUE
          AND c.sms_opt_in = TRUE
          AND c.sms_opt_out = FALSE

        ORDER BY EXTRACT(DAY FROM c.birth_date), c.first_name
    """, (spa_id, campaign_month))

    birthday_clients = cur.fetchall()


    sent_count = 0
    failed_count = 0

    for client in birthday_clients:
        client_id = client[0]
        first_name = client[1]
        phone = client[2]
        birth_date = client[3]

        merge_data = build_birthday_sms_merge_data(
            client_id=client_id,
            first_name=first_name,
            birth_date=birth_date,
            spa_name=spa_name
        )

        result = send_communication(
            spa_id=spa_id,
            channel="sms",
            recipient=phone,
            template_type="birthday_message",
            merge_data=merge_data,
            client_id=client_id,
            message_type="birthday_message"
        )

        if result.get("success"):
            sent_count += 1
        else:
            failed_count += 1
            print("Birthday SMS failed:", result)


    cur.close()
    conn.close()

    flash(
    f"Birthday SMS complete. Sent: {sent_count}. Failed: {failed_count}.",
    "success" if failed_count == 0 else "warning"
)


#   -------------------------
#
#   SMS PREVIEW
#
#   LEGACY. TODO
#   -------------------------


@app.route("/sms/preview/<int:client_id>", methods=["GET", "POST"])
@login_required
@spa_required
def sms_preview(client_id):
    spa_id = current_spa_id()


    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))


    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT 
            client_id,  
            first_name, 
            last_name, 
            phone,
            ok_to_text,
            sms_opt_in,
            sms_opt_out
        FROM clients
        WHERE client_id = %s
          AND spa_id = %s
          AND active_client = TRUE
    """, (client_id, spa_id))

    client = cur.fetchone()

    if not client:
        cur.close()
        conn.close()
        flash("Client not found.", "error")
        return redirect(url_for("clients_home"))

    message_body = request.form.get("message_body", "").strip()

    message_body = apply_sms_placeholders(message_body, {
        "first_name": first_name,
        "last_name": last_name,
        "phone": phone_number,
        "email": email,
        "spa_name": spa_name,
        "appointment_date": appointment_date,
        "appointment_time": appointment_time,
        "service_name": service_name
    })

    action = request.form.get("action")

    if request.method == "POST" and action == "send" and message_body:

        if not client or not client[3]:
            cur.close()
            conn.close()
            flash("Invalid client or missing phone.", "error")
            return redirect(request.referrer or url_for("clients_home"))



        if not client[4] or not client[5] or client[6]:
            cur.close()
            conn.close()
            flash("SMS not sent. This client is opted out of SMS messaging.", "error")
            return redirect(url_for("client_messaging_settings", client_id=client_id))

            # Optional extra safety if you later pass more fields
            # if sms_opt_out:
            #     skip

        result = send_communication(
            spa_id=spa_id,
            channel="sms",
            client_id=client[0],
            recipient=client[3],
            message_body=message_body,
            message_type="manual"
        )


        status = result.get("status") or "failed"

        cur.execute("""
            INSERT INTO sms_log (
                spa_id,
                client_id,
                phone_number,
                message_body,
                sms_type,
                status,
                provider_message_id,
                provider_error_code,
                provider_error_message,                
                sent_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,
                CASE WHEN %s = 'sent' THEN CURRENT_TIMESTAMP ELSE NULL END
            )
        """, (
            spa_id,
            client[0],
            client[3],
            result.get("final_message_body"),
            "manual",
            status,
            result.get("provider_message_id"),
            result.get("provider_error_code"),
            result.get("provider_error_message"),
            status
        ))

        conn.commit()

        cur.close()
        conn.close()

        if result.get("status") == "sent":
            flash("SMS sent successfully.", "success")
        elif result.get("status") == "logged":
            flash("SMS logged. Sending is currently disabled.", "success")
        else:
            flash("SMS failed to send.", "error")

        return redirect(url_for("sms_history_all"))

    cur.close()
    conn.close()

    return render_template(
        "sms_preview.html",
        client=client,
        message_body=message_body
    )


#   ------------------------
#
#    SMS CONVERSATION VIEW
#
#
#  
#   ---------------------
          

@app.route("/sms/conversation/<int:client_id>", methods=["GET", "POST"])
@login_required
@spa_required
def sms_conversation(client_id):
    spa_id = current_spa_id()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT client_id, first_name, last_name, phone
        FROM clients
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, spa_id))

    client = cur.fetchone()







###############################################
###############################################
#
#   HELPERS. HELPERS
#
#       GET ACTIVE MESSAGING TEMPLATES
#
###############################################


def get_active_messaging_templates(spa_id, channel):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            template_id,
            template_type,
            template_name,
            message_text,
            subject_text,
            language_code
        FROM messaging_templates
        WHERE spa_id = %s
          AND channel = %s
          AND is_active = TRUE
          AND approved_for_use = TRUE
          AND COALESCE(is_archived, FALSE) = FALSE
        ORDER BY template_type, template_name
    """, (spa_id, channel))

    templates = cur.fetchall()

    cur.close()
    conn.close()

    return templates


def get_active_messaging_template_types(spa_id, channel):
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT DISTINCT
            template_type,
            template_type
        FROM messaging_templates
        WHERE spa_id = %s
          AND channel = %s
          AND is_active = TRUE
          AND approved_for_use = TRUE
          AND COALESCE(is_archived, FALSE) = FALSE
        ORDER BY template_type
    """, (spa_id, channel))

    template_types = cur.fetchall()

    cur.close()
    conn.close()

    return template_types





####################################
#
#       SMS HOME
#
###################################



@app.route("/sms")
@login_required
@spa_required
def sms_home():
    spa_id = current_spa_id()
    language_code = get_request_language()

    template_id = request.args.get("template_id")
    template_type = request.args.get("template_type")


    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    selected_template_type = request.args.get("template_type", "").strip()
    selected_template_id = request.args.get("template_id", "").strip()
    search = request.args.get("search", "").strip()
    show_all = request.args.get("show_all") == "1"
    client_status = request.args.get("client_status", "").strip()

    sms_templates = get_active_messaging_templates(spa_id, "sms")
    sms_template_types = get_active_messaging_template_types(spa_id, "sms")

    conn = get_db_connection()
    cur = conn.cursor()

    client_statuses = get_client_statuses(spa_id)


    clients = get_sms_eligible_clients(
        spa_id=spa_id,
        search=search,
        client_status=client_status,
        show_all=show_all
    )   

    cur.close()
    conn.close()

    return render_template(
        "sms_home.html",
        sms_templates=sms_templates,
        sms_template_types=sms_template_types,
        selected_template_type=selected_template_type,
        selected_template_id=selected_template_id,
        client_statuses=client_statuses,
        clients=clients,
        search=search,
        show_all=show_all,
        client_status=client_status
    )




















                    
#   ------------------------
#
#    SMS TEMPLATES 
#
#
#   
#   ---------------------


@app.route("/sms/templates")
@login_required
@spa_required
def sms_templates_admin():
    spa_id = current_spa_id()
    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            sms_template_id,
            template_name,
            message_body,
            active,
            updated_at
        FROM sms_templates
        WHERE spa_id = %s
        ORDER BY template_name
    """, (spa_id,))

    templates = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("sms_templates_admin.html", templates=templates)




                


#   ------------------------
#
#   SMS TEMPLATES  ADD                
#
#
#
#   ---------------------

@app.route("/sms/templates/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_sms_template():
    spa_id = current_spa_id()

    if request.method == "POST":
        template_name = request.form.get("template_name", "").strip()
        message_body = request.form.get("message_body", "").strip()

        if not template_name or not message_body:
            flash("Template name and message are required.", "error")
            return redirect(url_for("add_sms_template"))

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO sms_templates (
                spa_id,
                template_name,
                message_body,
                active,
                created_at,
                updated_at
            )
            VALUES (%s, %s, %s, TRUE, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """, (spa_id, template_name, message_body))

        conn.commit()
        cur.close()
        conn.close()

        flash("SMS template added successfully.", "success")
        return redirect(url_for("sms_templates_admin"))

    return render_template("sms_template_form.html", template=None)
        
                    
                
            
            
#   ------------------------
#
#  SMS TEMPLATES EDIT
#
#
#       
#   ---------------------
        
#. LEGACY.  LEGACY.  LEGACY ROUTE.  LEGACY ROUTE
@app.route("/sms/templates/edit/<int:template_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_sms_template(template_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            sms_template_id,
            template_name,
            message_body,
            active
        FROM sms_templates
        WHERE sms_template_id = %s
          AND spa_id = %s
    """, (template_id, spa_id))

    template = cur.fetchone()

    if not template:
        cur.close()
        conn.close()
        flash("SMS template not found.", "error")
        return redirect(url_for("sms_templates_admin"))

    if request.method == "POST":
        template_name = request.form.get("template_name", "").strip()
        message_body = request.form.get("message_body", "").strip()
        active = True if request.form.get("active") == "on" else False

        if not template_name or not message_body:
            flash("Template name and message are required.", "error")
            return redirect(url_for("edit_sms_template", template_id=template_id))

        cur.execute("""
            UPDATE sms_templates
            SET template_name = %s,
                message_body = %s,
                active = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE sms_template_id = %s
              AND spa_id = %s
        """, (template_name, message_body, active, template_id, spa_id))

        conn.commit()
        cur.close()
        conn.close()

        flash("SMS template updated successfully.", "success")
        return redirect(url_for("sms_templates_admin"))

    cur.close()
    conn.close()

    return render_template("sms_template_form.html", template=template)






 


        
#   ------------------------
#
#    SMS GROUP PREVIEW
#               
#
#   
#   ---------------------
    
@app.route("/sms/group-preview", methods=["POST"])
@login_required
@spa_required
def sms_group_preview():
    spa_id = current_spa_id()

    selected_template_type = request.form.get("template_type", "").strip()
    selected_template_id = request.form.get("template_id", "").strip()
    search = request.form.get("search", "").strip()
    show_all = request.form.get("show_all", "").strip()
    client_status = request.form.get("client_status", "").strip()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    template_type = request.form.get("template_type")
    client_ids = [int(x) for x in request.form.getlist("client_ids")]

    print("GROUP PREVIEW TEMPLATE_TYPE:", template_type, flush=True)
    print("GROUP PREVIEW CLIENT_IDS:", client_ids, flush=True)

 
    if not template_type:
        flash("Please select an SMS template.", "error")
        return redirect(url_for("sms_home"))

    if not client_ids:
        flash("Please select at least one client.", "error")
        return redirect(url_for("sms_home"))

    if len(client_ids) > 5:
        flash("You can send SMS to a maximum of 5 clients at a time.", "error")
        return redirect(url_for("sms_home"))


    # Verify active approved SMS template exists in unified Communications Engine
    template = get_template_by_id(
        spa_id=spa_id,
        template_id=selected_template_id,
        channel="sms"
    )


    print("GROUP PREVIEW TEMPLATE FOUND:", template, flush=True)

    if not template or not template[6] or not template[7]:
        flash("SMS template not found, inactive, or not approved.", "error")
        return redirect(url_for("sms_home"))


    clients = get_sms_clients_by_ids(
        spa_id=spa_id,
        client_ids=client_ids
    )


    if not clients:
        flash("No eligible SMS clients found.", "error")
        return redirect(url_for("sms_home"))

    preview_messages = []

    for client in clients:
        (
            client_id,
            first_name,
            last_name,
            phone,
            sms_opt_in,
            sms_opt_out
        ) = client

        merge_data = {
            "client_first_name": first_name,
            "client_full_name": f"{first_name} {last_name}".strip(),
            "spa_name": get_spa_name(spa_id),
        }

        built = build_communication(
            spa_id=spa_id,
            channel="sms",
            template_type=template_type,
            merge_data=merge_data
        )

        print("BUILD COMMUNICATION RESULT:", built, flush=True)

        if built.get("body"):
            preview_messages.append({
                "client_id": client_id,
                "client_name": f"{first_name} {last_name}".strip(),
                "phone": phone,
                "message_body": built["body"],
                "template_type": template_type
            })

    if not preview_messages:
        flash("Unable to build preview messages.", "error")
        return redirect(url_for("sms_home"))

    return render_template(
        "sms_group_preview.html",
        selected_template_type=selected_template_type,
        selected_template_id=selected_template_id,
        search=search,
        show_all=show_all,
        client_status=client_status,
        clients=clients,
        template=template,
        preview_messages=preview_messages
    )



    
#   ------------------------
#
#    SMS GROUP SEND
#
#
#   6-30-26
#   ---------------------


@app.route("/sms/group-send", methods=["POST"])
@login_required
@spa_required
def sms_group_send():
    print("SMS GROUP SEND ROUTE HIT", flush=True)

    spa_id = current_spa_id()
    language_code = get_request_language()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    template_type = request.form.get("template_type")
    template_id_raw = request.form.get("template_id")
    client_ids = request.form.getlist("client_ids")

    template_id = int(template_id_raw) if template_id_raw else None

    if not template_type:
        flash("SMS template is required.", "error")
        return redirect(url_for(
            "sms_home",
            language_code=language_code,
            template_id=template_id
        ))

    if not client_ids:
        flash("Please select at least one client.", "error")
        return redirect(url_for(
            "sms_home",
            template_type=template_type,
            template_id=template_id,
            language_code=language_code
        ))

    if len(client_ids) > 5:
        flash("You can send SMS to a maximum of 5 clients at a time.", "error")
        return redirect(url_for(
            "sms_home",
            template_type=template_type,
            template_id=template_id,
            language_code=language_code
        ))

    client_ids = [int(x) for x in client_ids]

    sent_count = 0
    failed_count = 0

    try:
        clients = get_sms_clients_by_ids(
            spa_id=spa_id,
            client_ids=client_ids
        )

        if not clients:
            flash("No eligible SMS clients found.", "error")
            return redirect(url_for(
                "sms_home",
                template_type=template_type,
                template_id=template_id,
                language_code=language_code
            ))

        for client in clients:
            (
                client_id,
                first_name,
                last_name,
                phone,
                sms_opt_in,
                sms_opt_out
            ) = client

            merge_data = {
                "client_id": client_id,
                "language_code": language_code,
                "client_first_name": first_name,
                "client_full_name": f"{first_name} {last_name}".strip(),
                "spa_name": get_spa_name(spa_id),
            }

            print(
                "SMS SEND LANGUAGE:",
                language_code,
                "TEMPLATE ID:",
                template_id,
                "TYPE:",
                template_type,
                flush=True
            )

            result = send_communication(
                spa_id=spa_id,
                channel="sms",
                recipient=phone,
                template_type=template_type,
                merge_data=merge_data,
                client_id=client_id,
                message_type="group_send",
                language_code=language_code,
                template_id=template_id
            )

            log_sms_message(
                spa_id=spa_id,
                client_id=client_id,
                recipient_phone=phone,
                message_body=result.get("message_body",""),
                message_type="group_send",
                direction="outbound",
                status="sent" if result.get("success") else "failed",
                provider_message_id=result.get("provider_message_id"),
                provider_status=result.get("provider_status"),
                provider_error_code=result.get("provider_error_code"),
                provider_error_message=result.get("error")
            )


            print("SMS SEND RESULT:", result, flush=True)

            if result.get("success"):
                sent_count += 1
            else:
                failed_count += 1

        flash(f"SMS sent: {sent_count}. Failed: {failed_count}.", "success")
        return redirect(url_for("sms_history_all"))

    except Exception as error:
        print("SMS GROUP SEND ERROR:", error, flush=True)
        flash("Something went wrong while sending SMS messages.", "danger")
        return redirect(url_for(
            "sms_home",
            template_type=template_type,
            template_id=template_id,
            language_code=language_code
        ))








    
    
#   ------------------------
#       
#    SMS                  
#         
#         
#                            
#   ---------------------
          
        
    


    
    
    
#   ------------------------
#       
#    SMS                  
#         
#         
#                            
#   ---------------------
          
        
    






#   ------------------------
#
#    SMS HISTORY CLIENT ID
#
#
#   not a dupe of route below
#   ---------------------


@app.route("/sms/history/<int:client_id>")
@login_required
@spa_required
def sms_history(client_id):
    spa_id = current_spa_id()

    spa_timezone = get_spa_timezone(spa_id)

    conn = get_db_connection()
    cur = conn.cursor()

    # Get client
    cur.execute("""
        SELECT first_name, last_name
        FROM clients
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, spa_id))
    
    client = cur.fetchone()

    if not client:
        cur.close()
        conn.close()
        flash("Client not found.", "error")
        return redirect(url_for("clients_home"))

    # Get SMS logs
    cur.execute("""
        SELECT
            sl.created_at AT TIME ZONE %s AS created_at,      -- 0
            sl.sms_type,                 -- 1
            c.first_name,                -- 2
            c.last_name,                 -- 3
            sl.phone_number,             -- 4
            sl.message_body,             -- 5
            sl.status,                   -- 6
            sl.status,                   -- 7
            sl.provider_error_code,        -- 8
            sl.provider_error_message,     -- 9
            sl.sms_log_id                -- 10
        FROM sms_log sl
        JOIN clients c
            ON sl.client_id = c.client_id
           AND sl.spa_id = c.spa_id
        WHERE sl.spa_id = %s
          AND sl.client_id = %s
        ORDER BY sl.created_at DESC
    """, (spa_timezone, spa_id, client_id))

    sms_log = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "sms_history.html",
        client=client,
        sms_logs=sms_log
    )







#   ----------------------------
#
#     SMS HISTORY ALL
#
#   not a dupe of route above
#   ----------------------------
            




@app.route("/sms/history")
@login_required
@spa_required
def sms_history_all():
    spa_id = current_spa_id()

    spa_timezone = get_spa_timezone(spa_id)

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            sm.created_at AT TIME ZONE %s AS created_at,     -- 0
            sm.message_type,             -- 1
            c.first_name,                -- 2
            c.last_name,                 -- 3
            sm.recipient_phone,          -- 4
            sm.message_body,             -- 5
            sm.status,                   -- 6
            sm.provider_status,            -- 7  
            sm.provider_error_code,        -- 8
            sm.provider_error_message,     -- 9
            sm.sms_message_id            -- 10
        FROM sms_messages sm
        LEFT JOIN clients c
            ON sm.client_id = c.client_id
           AND sm.spa_id = c.spa_id
        WHERE sm.spa_id = %s
        ORDER BY sm.created_at DESC
    """, (spa_timezone, spa_id,))

    sms_log = cur.fetchall()

    print("SMS HISTORY USING sms_messages", flush=True)
    print("sms_messages count =", len(sms_log), flush=True)
    print("sms_messages =", sms_log, flush=True)

    print("sms messages count =", len(sms_log))
    print("sms_log =", sms_log)

    cur.close()
    conn.close()

    return render_template(
        "sms_history.html",
        sms_logs=sms_log,
        client=None
    )


 

#   ----------------------------
#
#     SMS LOGS
#
#
#   ----------------------------

@app.route("/sms_logs/<int:sms_log_id>/refresh", methods=["POST"])
@login_required
@spa_required
def refresh_sms_status(sms_log_id):
    flash(
        "SMS status refresh is not available after migrating to Telnyx. Status updates will be handled by Telnyx webhooks.",
        "info"
    )
    return redirect(url_for("sms_history_all"))



#   ---------------------------
#
#  SMS REFRESH HISTORY ALL
#
#
#
#   --------------------------


@app.route("/sms/refresh-all", methods=["POST"])
@login_required
@spa_required
def refresh_all_sms_statuses():
    flash(
        "SMS status refresh is not available after migrating to Telnyx. Status updates will be handled by Telnyx webhooks.",
        "info"
    )
    return redirect(url_for("sms_history_all"))



#   ------------------------------------------------
#
#
#           
#     SMS RESEND
#
#   
#       
#   -----------------------------------------------


@app.route("/sms/resend/<int:sms_log_id>", methods=["POST"])
@login_required
@spa_required
def resend_sms(sms_log_id):
    spa_id = current_spa_id()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            sms_log_id,
            client_id,
            phone_number,
            message_body,
            sms_type
        FROM sms_log
        WHERE sms_log_id = %s
          AND spa_id = %s
    """, (sms_log_id, spa_id))

    old_sms = cur.fetchone()

    if not old_sms:
        cur.close()
        conn.close()
        flash("SMS log not found.", "error")
        return redirect(url_for("sms_history_all"))

    old_sms_log_id, client_id, phone_number, message_body, sms_type = old_sms

    # Verify client still allows SMS
    cur.execute("""
        SELECT sms_opt_in, sms_opt_out
        FROM clients
        WHERE client_id = %s
          AND spa_id = %s
          AND active_client = TRUE
    """, (client_id, spa_id))

    client = cur.fetchone()

    if not client:
        cur.close()
        conn.close()
        flash("Client not found or inactive.", "error")
        return redirect(url_for("sms_history_all"))

    ok_to_text, sms_opt_in, sms_opt_out = client

    if not ok_to_text or not sms_opt_in or sms_opt_out:
        cur.close()
        conn.close()
        flash("SMS not resent. Client is not opted in for SMS.", "error")
        return redirect(url_for("sms_history", client_id=client_id))

    try:
        result = send_communication(
            spa_id=spa_id,
            channel="sms",
            client_id=client_id,
            recipient=phone_number,
            message_body=message_body,
            message_type=f"{sms_type or 'manual'}_resend"
        )
        cur.execute("""
            INSERT INTO sms_log (
                spa_id,
                client_id,
                phone_number,
                message_body,
                sms_type,
                status,
                provider_message_id,
                created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
        """, (
            spa_id,
            client_id,
            phone_number,
            result.get("final_message_body"),
            f"{sms_type or 'manual'}_resend",
            result.get("status"),
            result.get("provider_message_id")
        ))

        conn.commit()
        flash("SMS resent successfully.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"SMS resend failed: {e}", "error")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("sms_history", client_id=client_id))




#   ------------------------------------------------
#
#
#
#     EMAIL CAMPAIGNS
#
#
#
#   -----------------------------------------------






#   -----------------------------------
#
#
#     BIRTHDAY TEMPLATE 
#
# 
#  --------------------------------


@app.route("/email-templates")
@login_required
@spa_required
def email_templates_admin():
    return redirect(url_for("template_review", channel="email"))




#   ---------------------------------------
#
#    EMAIL TEMPLATE PREVIEW
#
#
#  
#   ----------------------------------------

@app.route("/email-template-preview/<int:email_template_id>")
@login_required
@spa_required
def email_template_preview(email_template_id):
    return redirect(url_for(
        "template_review",
        channel="email"
    ))






#   ------------------------------------------------
#
#
#    GIFT CERTIFICATES      EMAIL CAMPAIGN
#
#    GIFT CERTIFICATE SEND
#
#          4/23/26
#   -----------------------------------------------


@app.route("/gift-certificates/email/<int:gift_cert_id>", methods=["POST"])
@login_required
@spa_required
def send_gift_certificate_email(gift_cert_id):
    spa_id = current_spa_id()
    spa_name = get_spa_name(spa_id)


    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        print("GIFT CERT EMAIL ROUTE HIT", flush=True)
        print("SPA ID:", spa_id, flush=True)
        print("GIFT CERT ID:", gift_cert_id, flush=True)

        cur.execute("""
            SELECT
                gc.gift_cert_id,
                gc.certificate_number,
                gc.original_value,
                gc.remaining_balance,
                gc.expires_date,
                gc.recipient_name,
                gc.purchaser_email
            FROM gift_certificates gc
            WHERE gc.gift_cert_id = %s
              AND gc.spa_id = %s
        """, (gift_cert_id, spa_id))
        gift_cert = cur.fetchone()

        print("GIFT CERT ROW:", gift_cert, flush=True)

        if not gift_cert:
            flash("Gift certificate not found.", "error")
            return redirect(url_for("gift_certificates_home"))

        (
            gc_id,
            certificate_number,
            original_value,
            remaining_balance,
            expires_date,
            recipient_name,
            purchaser_email
        ) = gift_cert

        print("PURCHASER EMAIL:", purchaser_email, flush=True)

        if not purchaser_email or not purchaser_email.strip():
            flash("No purchaser email found for this gift certificate.", "error")
            return redirect(url_for("gift_certificates_home"))

        communication = build_communication(
            spa_id=spa_id,
            channel="email",
            template_type="gift_certificate",
            merge_data={
                "spa_name": spa_name or "",
                "certificate_number": certificate_number or "",
                "original_value": f"{float(original_value or 0):.2f}",
                "remaining_balance": f"{float(remaining_balance or 0):.2f}",
                "expires_date": expires_date.strftime("%Y-%m-%d") if expires_date else "",
                "recipient_name": recipient_name or ""
            }
        )

        if not communication.get("success"):
            flash(communication.get("error") or "Gift certificate email template is not available.", "error")
            return redirect(url_for("gift_certificates_home"))

        subject = communication.get("subject") or f"Your Gift Certificate from {spa_name}"
        body = communication.get("message_body")


        print("SUBJECT:", subject, flush=True)
        print("ABOUT TO SEND EMAIL", flush=True)

        response = send_email(
            to=purchaser_email,
            subject=subject,
            body=body
        )

        print("GIFT CERT EMAIL STATUS:", response.status_code, flush=True)
        print("GIFT CERT EMAIL BODY:", response.text, flush=True)

        if response.status_code == 200:
            flash("Gift certificate email sent.", "success")
        else:
            flash("Gift certificate email failed to send.", "error")

        return redirect(url_for("gift_certificates_home"))

    except Exception as e:
        print("GIFT CERT EMAIL ERROR:", repr(e), flush=True)
        flash("There was a problem sending the gift certificate email.", "error")
        return redirect(url_for("gift_certificates_home"))

    finally:
        cur.close()
        conn.close()









#   ---------------------------
#
#
#     GENERAL EMAILS
# THANK YOU - REMINDERS - ETC
#
#
#
#   ---------------------------


@app.route("/general-email", methods=["GET", "POST"])
@login_required
@spa_required
def general_email():
    spa_id = current_spa_id()

    search = request.args.get(
        "search",
        ""
    ).strip()

    show_all = request.args.get("show_all")

    filter_status = request.args.get("filter_status")

    client_status_id = request.args.get(
        "client_status",
        ""
    ).strip()


    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    template_type = request.args.get(
        "template_type",
        ""
    )

    template_id = request.args.get(
        "template_id",
        ""
    )

    search = request.args.get(
        "search",
        ""
    ).strip()

    show_all = request.args.get("show_all")

    client_status_id = request.args.get(
        "client_status",
        ""
    ).strip()

    if client_status_id:
        try:
            client_status_id = int(client_status_id)

        except ValueError:
            client_status_id = ""
    else:
        client_status_id = ""

    conn = get_db_connection()
    cur = conn.cursor()

    template_types = get_active_messaging_template_types(
        spa_id,
        "email"
    )

    templates = get_active_messaging_templates(
        spa_id,
        "email"
    )

    # Get active client statuses for the dropdown.
    cur.execute("""
        SELECT
            client_status_id,
            status_name
        FROM client_statuses
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY status_name
    """, (spa_id,))

    client_statuses = cur.fetchall()

    clients = get_email_eligible_clients(
        spa_id=spa_id,
        search=search if not show_all else "",
        client_status=(
            client_status_id
            if filter_status
            else ""
        ),
        show_all=bool(show_all)
    )

    cur.close()
    conn.close()

    return render_template(
        "general_email.html",
        template_types=template_types,
        templates=templates,
        clients=clients,
        template_type=template_type,
        template_id=template_id,
        search=search,
        show_all=show_all,
        client_statuses=client_statuses,
        client_status_id=client_status_id
    )




#   ----------------------------------
#
#
#  GENERAL EMAIL  PREVIEW
#
#
#    SPA_ID AND ROUTE GOOD   4/23/26
#   ------------------------------------

@app.route("/general-email/preview", methods=["GET"])
@login_required
@spa_required
def general_email_preview():
    spa_id = current_spa_id()
    spa_name = get_spa_name(spa_id)

    template_id = request.args.get("template_id")
    client_id = request.args.get("client_id")

    if not template_id:
        flash("Please select a template to preview.", "error")
        return redirect(url_for("general_email"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT template_type
            FROM messaging_templates
            WHERE template_id = %s
              AND spa_id = %s
              AND channel = 'email'
              AND is_active = TRUE
        """, (template_id, spa_id))

        template = cur.fetchone()

        if not template:
            flash("Template not found or inactive.", "error")
            return redirect(url_for("general_email"))

        template_type = template[0]

        if client_id:
            cur.execute("""
                SELECT first_name, last_name, email
                FROM clients
                WHERE client_id = %s
                  AND spa_id = %s
            """, (client_id, spa_id))
            client = cur.fetchone()
        else:
            client = None

        if client:
            first_name, last_name, email = client
        else:
            first_name = "Sample"
            last_name = "Client"
            email = "sample@example.com"

        spa_phone = "" #get_spa_phone(spa_id)

        communication = build_communication(
            spa_id=spa_id,
            channel="email",
            template_type=template_type,
            merge_data={
                "client_first_name": first_name or "",
                "client_last_name": last_name or "",
                "first_name": first_name or "",
                "last_name": last_name or "",
                "email": email or "",
                "spa_name": spa_name or "",
                "spa_phone": spa_phone or "",
                "appointment_date": "",
                "appointment_time": "",
                "service_name": "",
                "unsubscribe_link": ""
            }
        )

        subject = communication.get("subject") or f"Message from {spa_name}"
        body = communication.get("body") or communication.get("message_body")

        if not body:
            flash("Unable to build email preview.", "error")
            return redirect(url_for("general_email"))

        return render_template(
            "general_email_preview.html",
            subject=subject,
            body=body,
            first_name=first_name,
            last_name=last_name,
            email=email
        )

    finally:
        cur.close()
        conn.close()














#   --------------------------------
#
#
#    SEND GENERAL EMAILS
#
#
#           TODO
#  SPA_ID AND ROUTE GOOD   4/23/26
#   ---------------------------------

# PSP_REFACTOR:
# Refactor birthday month emails to use centralized Email Communications Pipeline.





@app.route("/general-email/send", methods=["POST"])
@login_required
@spa_required
def general_email_send():
    spa_id = current_spa_id()
    language_code = get_request_language()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    spa_name = get_spa_name(spa_id)

    template_id_raw = request.form.get("template_id")
    client_ids = request.form.getlist("client_ids")

    template_id = int(template_id_raw) if template_id_raw else None

    #----------------DEBUG TODO -----------
    print("GENERAL EMAIL SEND HIT", flush=True)
    print("TEMPLATE ID:", template_id, flush=True)
    print("CLIENT IDS:", client_ids, flush=True)
    print("LANGUAGE:", language_code, flush=True)
    print("MAILGUN DOMAIN:", MAILGUN_DOMAIN, flush=True)
    print("MAILGUN FROM:", MAILGUN_FROM, flush=True)
    print(
        "MAILGUN KEY STARTS:",
        MAILGUN_API_KEY[:4] if MAILGUN_API_KEY else None,
        flush=True
    )
    #--------------------------------------



    if not template_id:
        flash("Please select an email template.", "error")
        return redirect(url_for(
            "general_email",
            language_code=language_code
        ))

    if not client_ids:
        flash("Please select at least one client.", "error")
        return redirect(url_for(
            "general_email",
            template_id=template_id,
            language_code=language_code
        ))

    conn = get_db_connection()
    cur = conn.cursor()

    sent_count = 0
    failed_count = 0

    try:
        cur.execute("""
            SELECT template_type
            FROM messaging_templates
            WHERE template_id = %s
              AND spa_id = %s
              AND channel = 'email'
              AND is_active = TRUE
              AND approved_for_use = TRUE
              AND COALESCE(is_archived, FALSE) = FALSE
            LIMIT 1
        """, (template_id, spa_id))

        template = cur.fetchone()

        if not template:
            flash("Template not found or inactive.", "error")
            return redirect(url_for(
                "general_email",
                language_code=language_code
            ))

        template_type = template[0]

        print(
            "GENERAL EMAIL SEND:",
            "LANGUAGE:", language_code,
            "TEMPLATE ID:", template_id,
            "TYPE:", template_type,
            flush=True
        )

        for raw_client_id in client_ids:
            client_id = int(raw_client_id)

            cur.execute("""
                SELECT client_id, first_name, last_name, email
                FROM clients
                WHERE client_id = %s
                  AND spa_id = %s
                  AND email IS NOT NULL
                  AND email <> ''
            """, (client_id, spa_id))

            client = cur.fetchone()

            if not client:
                failed_count += 1
                continue

            client_id, first_name, last_name, email = client

            merge_data = {
                "client_id": client_id,
                "language_code": language_code,
                "client_first_name": first_name or "",
                "client_last_name": last_name or "",
                "client_full_name": f"{first_name or ''} {last_name or ''}".strip(),
                "first_name": first_name or "",
                "last_name": last_name or "",
                "email": email or "",
                "spa_name": spa_name or "",
                "business_phone": "",
                "spa_phone": "",
                "appointment_date": "",
                "appointment_time": "",
                "service_name": "",
                "unsubscribe_link": ""
            }

            result = send_communication(
                spa_id=spa_id,
                channel="email",
                recipient=email,
                template_type=template_type,
                merge_data=merge_data,
                client_id=client_id,
                message_type=template_type,
                language_code=language_code,
                template_id=template_id
            )


            print("GENERAL EMAIL SEND RESULT:", result, flush=True)

            subject_line = result.get("subject") if result else None

            if result and result.get("success"):
                sent_status = "Sent"
                error_message = None
                sent_count += 1
            else:
                sent_status = "Failed"
                error_message = result.get("error") if result else "Email send failed."
                failed_count += 1


            cur.execute("""
                INSERT INTO email_send_log (
                    spa_id,
                    client_id,
                    template_id,
                    email_type,
                    recipient_email,
                    subject_line,
                    sent_status,
                    error_message
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                spa_id,
                client_id,
                template_id,
                template_type,
                email,
                subject_line,
                sent_status,
                error_message
            ))

        conn.commit()

        flash(f"Emails sent: {sent_count}. Failed: {failed_count}.", "success")

        return redirect(url_for(
            "general_email",
            template_id=template_id,
            language_code=language_code
        ))

    except Exception as error:
        print("GENERAL EMAIL SEND ERROR:", error, flush=True)
        flash("Something went wrong while sending emails.", "danger")
        return redirect(url_for(
            "general_email",
            template_id=template_id,
            language_code=language_code
        ))

    finally:
        cur.close()
        conn.close()










#   ------------------------------------
#
#
#       EMAIL HISTORY
#
#
#
#   ----------------------------------



@app.route("/email-history")
@login_required
@spa_required
def email_history():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            e.sent_at,
            e.email_type,
            e.recipient_email,
            e.subject_line,
            e.sent_status,
            e.error_message,
            c.first_name,
            c.last_name
        FROM email_send_log e
        LEFT JOIN clients c
            ON e.client_id = c.client_id
           AND e.spa_id = c.spa_id
        WHERE e.spa_id = %s
        ORDER BY e.sent_at DESC
        LIMIT 100
    """, (spa_id,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("email_history.html", rows=rows)




#   ------------------------------------
#
#
#      CLEAR -  EMAIL HISTORY
#
#
#
#   ----------------------------------


@app.route("/email-history/clear", methods=["POST"])
@login_required
@spa_required
def clear_email_history():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            DELETE FROM email_send_log
            WHERE spa_id = %s
        """, (spa_id,))

        conn.commit()
        flash("Email history cleared.", "success")

    except Exception as e:
        conn.rollback()
        print("CLEAR EMAIL HISTORY ERROR:", str(e))
        flash("Error clearing email history.", "error")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("email_history"))










#   ----------------------------------------
#
#       SEND ONE BIRTHDAY EMAIL
#
#
#    4/25/26        TODO
#   -------------------------------------


# PSP_REFACTOR:
# Refactor birthday month emails to use centralized Email Communications Pipeline.





@app.route("/birthday-offers/send-one/<int:client_id>", methods=["POST"])
@login_required
@spa_required
def send_one_birthday_offer_email(client_id):
    spa_id = current_spa_id()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))


    spa_name = get_spa_name(spa_id)
    today = get_spa_today()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT client_id, first_name, email, birth_date
            FROM clients
            WHERE spa_id = %s
              AND client_id = %s
              AND active_client = TRUE
              AND email IS NOT NULL
              AND TRIM(email) <> ''
              AND birth_date IS NOT NULL
        """, (spa_id, client_id))
        client = cur.fetchone()

        if not client:
            flash("Client not found or missing email/birthday.", "error")
            return redirect(url_for("birthday_offers_home"))

        client_id, first_name, email, birth_date = client
        campaign_year = get_birthday_campaign_year(birth_date, today)


        communication = build_communication(
            spa_id=spa_id,
            channel="email",
            template_type="birthday_message",
            merge_data={
                "client_first_name": first_name or "",
                "first_name": first_name or "",
                "spa_name": spa_name or "",
                "birth_month": birth_date.strftime("%B") if birth_date else ""
            }
        )

        if not communication.get("success"):
            flash(communication.get("error") or "Birthday email template is not available.", "error")
            return redirect(url_for("birthday_offers_home"))

        subject = communication.get("subject") or f"{spa_name}, wishing you a Very Happy Birthday!"
        body = communication.get("message_body")


        response = send_email(to=email, subject=subject, body=body)

        if response.status_code == 200:
            cur.execute("""
                UPDATE client_birthday_offers
                SET offer_sent = TRUE,
                    offer_sent_date = CURRENT_DATE,
                    sent_status = 'Sent'
                WHERE spa_id = %s
                  AND client_id = %s
                  AND birthday_year = %s
            """, (spa_id, client_id, campaign_year))

            if cur.rowcount == 0:
                cur.execute("""
                    INSERT INTO client_birthday_offers (
                        spa_id,
                        client_id,
                        birthday_year,
                        offer_sent,
                        offer_sent_date,
                        sent_status
                    )
                    VALUES (%s, %s, %s, TRUE, CURRENT_DATE, 'Sent')
                """, (spa_id, client_id, campaign_year))

            conn.commit()
            flash("Birthday email sent.", "success")
        else:
            print("BIRTHDAY EMAIL FAILED:", response.status_code, response.text)
            flash("Birthday email failed to send.", "error")

        return redirect(url_for("birthday_offers_home"))

    finally:
        cur.close()
        conn.close()






#   --------------------------------------
#   
#   
#    SEND ALL  SEND ALL BIRTHDAY EMAILS
#   
#      4/25/26      TODO
#   --------------------------------------



# PSP_REFACTOR:
# Refactor birthday month emails to use centralized Email Communications Pipeline.

@app.route("/birthday-offers/send-all", methods=["POST"])
@login_required
@spa_required
def send_all_birthday_offer_emails():
    spa_id = current_spa_id()
    spa_name = get_spa_name(spa_id)
    today = get_spa_today()
    end_date = today + timedelta(days=45)

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    conn = get_db_connection()
    cur = conn.cursor()

    sent_count = 0
    failed_count = 0

    try:
        cur.execute("""
            SELECT
                c.client_id,
                c.first_name,
                c.email,
                c.birth_date
            FROM clients c
            WHERE c.spa_id = %s
              AND c.active_client = TRUE
              AND c.birth_date IS NOT NULL
              AND c.email IS NOT NULL
              AND c.email_opt_in = TRUE
              AND c.email_opt_out = FALSE      
              AND TRIM(c.email) <> ''
            ORDER BY c.last_name, c.first_name
        """, (spa_id,))
        clients = cur.fetchall()


        for client_id, first_name, email, birth_date in clients:
            this_year_birthday = birth_date.replace(year=today.year)

            if this_year_birthday < today:
                next_birth_date = birth_date.replace(year=today.year + 1)
            else:
                next_birth_date = this_year_birthday

            if not (today <= next_birth_date <= end_date):
                continue

            campaign_year = next_birth_date.year

            communication = build_communication(
                spa_id=spa_id,
                channel="email",
                template_type="birthday_message",
                merge_data={
                    "client_first_name": first_name or "",
                    "first_name": first_name or "",
                    "spa_name": spa_name or "",
                    "birth_month": birth_date.strftime("%B") if birth_date else ""
                }
            )

            if not communication.get("success"):
                failed_count += 1
                continue

            subject = communication.get("subject") or f"{spa_name}, wishing you a Very Happy Birthday!"
            body = communication.get("body") or communication.get("message_body")

            if not body:
                failed_count += 1
                continue

        

            cur.execute("""
                SELECT offer_sent
                FROM client_birthday_offers
                WHERE spa_id = %s
                  AND client_id = %s
                  AND birthday_year = %s
            """, (spa_id, client_id, campaign_year))
            offer_row = cur.fetchone()

            if offer_row and offer_row[0]:
                continue


            response = send_email(to=email, subject=subject, body=body)

            if response.status_code == 200:
                cur.execute("""
                    UPDATE client_birthday_offers
                    SET offer_sent = TRUE,
                        offer_sent_date = CURRENT_DATE,
                        sent_status = 'Sent'
                    WHERE spa_id = %s
                      AND client_id = %s
                      AND birthday_year = %s
                """, (spa_id, client_id, campaign_year))

                if cur.rowcount == 0:
                    cur.execute("""
                        INSERT INTO client_birthday_offers (
                            spa_id,
                            client_id,
                            birthday_year,
                            offer_sent,
                            offer_sent_date,
                            sent_status
                        )
                        VALUES (%s, %s, %s, TRUE, CURRENT_DATE, 'Sent')
                    """, (spa_id, client_id, campaign_year))

                sent_count += 1
            else:
                print("BIRTHDAY EMAIL FAILED:", response.status_code, response.text)
                failed_count += 1

        conn.commit()
        flash(f"Birthday emails sent: {sent_count}. Failed: {failed_count}.", "success")
        return redirect(url_for("birthday_offers_home"))

    finally:
        cur.close()
        conn.close()




#   --------------------------------------
#
#
#   SMS EMAIL REMINDER QUEUE
#
#
#   --------------------------------------

@app.route("/reminder_queue")
@login_required
@spa_required
def reminder_queue():
    spa_id = current_spa_id()
    status_filter = request.args.get("status", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            rq.reminder_id,
            rq.reminder_type,
            rq.send_method,
            rq.recipient_phone,
            rq.recipient_email,
            rq.message_body,
            rq.scheduled_for,
            rq.status,
            rq.sent_at,
            rq.error_message,
            c.first_name,
            c.last_name
        FROM reminder_queue rq
        LEFT JOIN clients c
            ON rq.client_id = c.client_id
           AND rq.spa_id = c.spa_id
        WHERE rq.spa_id = %s
    """

    params = [spa_id]

    if status_filter:
        query += " AND rq.status = %s"
        params.append(status_filter)

    query += """
        ORDER BY
            rq.status,
            rq.scheduled_for ASC,
            rq.created_at DESC
    """

    cur.execute(query, params)
    reminders = cur.fetchall()

    cur.execute("""
        SELECT status, COUNT(*)
        FROM reminder_queue
        WHERE spa_id = %s
        GROUP BY status
    """, (spa_id,))

    status_counts = dict(cur.fetchall())

    cur.close()
    conn.close()

    return render_template(
        "reminder_queue.html",
        reminders=reminders,
        status_filter=status_filter,
        status_counts=status_counts
    )






#   --------------------------------------
#
#
#   REMINDER QUEUE GENERATE APPOINTMENTS
#               
#
#   --------------------------------------


@app.route("/reminder_queue/generate-appointments", methods=["POST"])
@login_required
@spa_required
def generate_appointment_reminders():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                a.appointment_id,
                a.client_id,
                a.appointment_date,
                a.appointment_time,
                c.first_name,
                c.last_name,
                c.phone,
                c.ok_to_text
            FROM appointments a
            JOIN clients c
                ON a.client_id = c.client_id
               AND a.spa_id = c.spa_id
            WHERE a.spa_id = %s
              AND a.appointment_date >= CURRENT_DATE
              AND a.appointment_date <= CURRENT_DATE + INTERVAL '7 days'
            ORDER BY a.appointment_date, a.appointment_time
        """, (spa_id,))

        appointments = cur.fetchall()

        created_count = 0
        skipped_count = 0

        for appt in appointments:
            (
                appointment_id,
                client_id,
                appointment_date,
                appointment_time,
                first_name,
                last_name,
                phone,
                ok_to_text
            ) = appt

            if not phone or not ok_to_text:
                skipped_count += 1
                continue

            cur.execute("""
                SELECT 1
                FROM reminder_queue
                WHERE spa_id = %s
                  AND appointment_id = %s
                  AND reminder_type = 'appointment_reminder'
                  AND send_method = 'sms'
                  AND status IN ('pending', 'sent')
            """, (spa_id, appointment_id))

            existing = cur.fetchone()

            if existing:
                skipped_count += 1
                continue

            message_body = apply_sms_placeholders(
                "Hi {first_name}, this is Clear Skin Esthetics. This is a reminder of your appointment on {appointment_date} at {appointment_time}. Reply STOP to opt out.",
                {
                    "first_name": first_name,
                    "last_name": last_name,
                    "appointment_date": appointment_date,
                    "appointment_time": appointment_time
                }
            )

            cur.execute("""
                INSERT INTO reminder_queue (
                    spa_id,
                    client_id,
                    appointment_id,
                    reminder_type,
                    send_method,
                    recipient_phone,
                    message_body,
                    scheduled_for,
                    status
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s,
                        (%s::date + %s::time - INTERVAL '24 hours'),
                        'pending')
            """, (
                spa_id,
                client_id,
                appointment_id,
                "appointment_reminder",
                "sms",
                phone,
                message_body,
                appointment_date,
                appointment_time
            ))

            created_count += 1

        conn.commit()

        flash(
            f"Appointment reminders generated: {created_count}. Skipped: {skipped_count}.",
            "success"
        )

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("reminder_queue"))






#   --------------------------------------
#
#                   
#   REMINDER QUEUE SEND PENDING
#                   
#
#   --------------------------------------

@app.route("/reminder_queue/send-pending", methods=["POST"])
@login_required
@spa_required
def send_pending_reminders():
    spa_id = current_spa_id()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    conn = get_db_connection()
    cur = conn.cursor()

    sent_count = 0
    failed_count = 0
    skipped_count = 0

    try:
        cur.execute("""
            SELECT
                reminder_id,
                client_id,
                reminder_type,
                send_method,
                recipient_phone,
                recipient_email,
                message_body
            FROM reminder_queue
            WHERE spa_id = %s
              AND status = 'pending'
              AND scheduled_for <= NOW()
            ORDER BY scheduled_for ASC
        """, (spa_id,))

        reminders = cur.fetchall()

        for reminder in reminders:
            reminder_id, client_id, reminder_type, send_method, recipient_phone, recipient_email, message_body = reminder

            if send_method == "sms":
                if not recipient_phone:
                    cur.execute("""
                        UPDATE reminder_queue
                        SET status = 'skipped',
                            error_message = %s
                        WHERE reminder_id = %s
                          AND spa_id = %s
                    """, ("Missing phone number", reminder_id, spa_id))
                    skipped_count += 1
                    continue

                result = send_communication(
                    spa_id=spa_id,
                    channel="sms",
                    client_id=client_id,
                    recipient=recipient_phone,
                    message_body=message_body,
                    message_type=reminder_type
                )

                if result.get("success"):
                    cur.execute("""
                        UPDATE reminder_queue
                        SET status = 'sent',
                            sent_at = NOW(),
                            error_message = NULL
                        WHERE reminder_id = %s
                          AND spa_id = %s
                    """, (reminder_id, spa_id))


                    cur.execute("""
                        INSERT INTO sms_log (
                            spa_id,
                            client_id,
                            phone_number,
                            message_body,
                            sms_type,
                            status,
                            provider_message_id,
                            provider_error_code,
                            provider_error_message,
                            created_at,
                            sent_at
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """, (
                    spa_id,
                    client_id,
                    recipient_phone,
                    result.get("final_message_body"),
                    reminder_type,
                    "sent",
                    result.get("provider_message_id"),
                    result.get("provider_error_code"),
                    result.get("provider_error_message")
                ))

                    sent_count += 1
                else:
                    cur.execute("""
                        UPDATE reminder_queue
                        SET status = 'failed',
                            error_message = %s
                        WHERE reminder_id = %s
                          AND spa_id = %s
                    """, (
                        result.get("provider_error_message") or "SMS send failed",
                        reminder_id,
                        spa_id
                    ))
                    failed_count += 1

            else:
                cur.execute("""
                    UPDATE reminder_queue
                    SET status = 'skipped',
                        error_message = %s
                    WHERE reminder_id = %s
                      AND spa_id = %s
                """, ("Unsupported send method", reminder_id, spa_id))
                skipped_count += 1

        conn.commit()

        flash(
            f"Pending reminders processed. Sent: {sent_count}. Failed: {failed_count}. Skipped: {skipped_count}.",
            "success"
        )

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("reminder_queue"))



                
                    
                        
                        
                            
#   --------------------------------------
#
#
#   REMINDER QUEUE RETRY-FAILED
#
#               
#   --------------------------------------
                  
@app.route("/reminder_queue/retry-failed", methods=["POST"])
@login_required
@spa_required
def retry_failed_reminders():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            UPDATE reminder_queue
            SET status = 'pending',
                error_message = NULL
            WHERE spa_id = %s
              AND status = 'failed'
        """, (spa_id,))

        updated_count = cur.rowcount

        conn.commit()

        flash(
            f"Failed reminders reset to pending: {updated_count}.",
            "success"
        )

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("reminder_queue"))



#   --------------------------------------
#
#   
#   REMINDER QUEUE SEND ONE
#
#   
#   --------------------------------------


@app.route("/reminder_queue/send-one/<int:reminder_id>", methods=["POST"])
@login_required
@spa_required
def send_one_reminder(reminder_id):
    spa_id = current_spa_id()

    if not sms_email_terms_accepted(spa_id):
        flash(
            "You must accept the SMS and Email Terms and Conditions before using messaging features.",
            "warning"
        )
        return redirect(url_for("sms_email_terms"))

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                reminder_id,
                client_id,
                reminder_type,
                send_method,
                recipient_phone,
                recipient_email,
                message_body,
                status
            FROM reminder_queue
            WHERE reminder_id = %s
              AND spa_id = %s
        """, (reminder_id, spa_id))

        reminder = cur.fetchone()

        if not reminder:
            flash("Reminder not found.", "warning")
            return redirect(url_for("reminder_queue"))

        (
            reminder_id,
            client_id,
            reminder_type,
            send_method,
            recipient_phone,
            recipient_email,
            message_body,
            status
        ) = reminder

        if status == "sent":
            flash("This reminder has already been sent.", "warning")
            return redirect(url_for("reminder_queue"))

        if send_method == "sms":
            if reminder_type == "appointment_reminder":
                success, result = send_appointment_reminder_sms(
                    reminder_id=reminder_id,
                    spa_id=spa_id
                )

                if success:
                    flash("Appointment reminder SMS sent.", "success")
                else:
                    flash(f"Appointment reminder SMS failed: {result}", "danger")

                return redirect(url_for("reminder_queue"))

            if reminder_type == "birthday":
                success, result = send_birthday_reminder_sms(
                    reminder_id=reminder_id,
                    spa_id=spa_id
                )

                if success:
                    flash("Birthday SMS sent.", "success")
                else:
                    flash(f"Birthday SMS failed: {result}", "danger")

                return redirect(url_for("reminder_queue"))
     

            flash("This SMS reminder type is not connected to the new messaging pipeline yet.", "warning")
            return redirect(url_for("reminder_queue"))


        else:
            flash("Only SMS reminders are supported right now.", "warning")

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("reminder_queue"))











#   --------------------------------------
#
#
#   REMINDER QUEUE > AFTER APPOINTMENT
#
#
#   --------------------------------------
        

@app.route("/reminder_queue/generate-after-appointments", methods=["POST"])
@login_required
@spa_required
def generate_after_appointment_followups():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                a.appointment_id,
                a.client_id,
                a.appointment_date,
                a.appointment_time,
                c.first_name,
                c.last_name,
                c.phone,
                c.ok_to_text
            FROM appointments a
            JOIN clients c
                ON a.client_id = c.client_id
               AND a.spa_id = c.spa_id
            WHERE a.spa_id = %s
              AND a.appointment_date <= CURRENT_DATE
              AND a.appointment_date >= CURRENT_DATE - INTERVAL '7 days'
            ORDER BY a.appointment_date DESC, a.appointment_time DESC
        """, (spa_id,))

        appointments = cur.fetchall()

        created_count = 0
        skipped_count = 0

        for appt in appointments:
            (
                appointment_id,
                client_id,
                appointment_date,
                appointment_time,
                first_name,
                last_name,
                phone,
                ok_to_text
            ) = appt

            if not phone or not ok_to_text:
                skipped_count += 1
                continue

            cur.execute("""
                SELECT 1
                FROM reminder_queue
                WHERE spa_id = %s
                  AND appointment_id = %s
                  AND reminder_type = 'after_appointment_followup'
                  AND send_method = 'sms'
                  AND status IN ('pending', 'sent')
            """, (spa_id, appointment_id))

            existing = cur.fetchone()

            if existing:
                skipped_count += 1
                continue

            message_body = apply_sms_placeholders(
                "Hi {first_name}, thank you for visiting Clear Skin Esthetics. We hope you enjoyed your appointment. Please contact us if you have any questions about your aftercare. Reply STOP to opt out.",
                {
                    "first_name": first_name,
                    "last_name": last_name,
                    "appointment_date": appointment_date,
                    "appointment_time": appointment_time
                }
            )

            cur.execute("""
                INSERT INTO reminder_queue (
                    spa_id,
                    client_id,
                    appointment_id,
                    reminder_type,
                    send_method,
                    recipient_phone,
                    message_body,
                    scheduled_for,
                    status
                )
                VALUES (
                    %s, %s, %s, %s, %s, %s, %s,
                    (%s::date + %s::time + INTERVAL '2 hours'),
                    'pending'
                )
            """, (
                spa_id,
                client_id,
                appointment_id,
                "after_appointment_followup",
                "sms",
                phone,
                message_body,
                appointment_date,
                appointment_time
            ))

            created_count += 1

        conn.commit()

        flash(
            f"After appointment follow-ups generated: {created_count}. Skipped: {skipped_count}.",
            "success"
        )

    finally:
        cur.close()
        conn.close()

    return redirect(url_for("reminder_queue"))


            
            
        
            
                
                
#   --------------------------------------
#
#
#   REMINDER QUEUE DELETE/CANCEL
#
#
#   --------------------------------------
                    

@app.route("/reminders/cancel/<int:reminder_id>", methods=["POST"])
@login_required
@spa_required
def cancel_reminder(reminder_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE reminder_queue
        SET status = 'cancelled'
        WHERE reminder_id = %s
          AND spa_id = %s
          AND status = 'pending'
    """, (reminder_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Reminder cancelled.", "success")
    return redirect(url_for("reminder_queue"))





#   --------------------------------------
#
#
#   REMINDER QUEUE > BIRTHDAY REMINDER
#
#     AND   DEF         
#   --------------------------------------

def generate_birthday_reminders(spa_id):
    conn = get_db_connection()
    cur = conn.cursor()

    created_count = 0

    try:
        cur.execute("""
            SELECT client_id, first_name, last_name, phone, birth_date
            FROM clients
            WHERE spa_id = %s
              AND active_client = TRUE
              AND ok_to_text = TRUE
              AND birth_date IS NOT NULL
              AND EXTRACT(MONTH FROM birth_date) = EXTRACT(MONTH FROM CURRENT_DATE)
              AND EXTRACT(DAY FROM birth_date) = EXTRACT(DAY FROM CURRENT_DATE)
        """, (spa_id,))

        clients = cur.fetchall()

        for client in clients:
            client_id = client[0]
            first_name = client[1]
            phone = client[3]

            cur.execute("""
                SELECT reminder_id
                FROM reminder_queue
                WHERE spa_id = %s
                  AND client_id = %s
                  AND reminder_type = 'birthday'
                  AND DATE(scheduled_for) = CURRENT_DATE
                  AND status IN ('pending', 'sent')
            """, (spa_id, client_id))

            existing = cur.fetchone()

            if existing:
                continue

            message_body = f"Happy Birthday {first_name}! Clear Skin Esthetics hopes you have a wonderful day!"

            cur.execute("""
                INSERT INTO reminder_queue
                    (spa_id, client_id, reminder_type, send_method, recipient_phone, message_body, scheduled_for, status)
                VALUES
                    (%s, %s, %s, %s, %s, %s, NOW(), %s)
            """, (
                spa_id,
                client_id,
                "birthday",
                "sms",
                phone,
                message_body,
                "pending"
            ))

            created_count += 1

        conn.commit()
        return created_count

    finally:
        cur.close()
        conn.close()












@app.route("/reminders/create-birthday-reminders", methods=["POST"])
@login_required
@spa_required
def create_birthday_reminders():
    spa_id = current_spa_id()

    created_count = generate_birthday_reminders(spa_id)

    flash(f"{created_count} birthday reminder(s) added to the queue.", "success")
    return redirect(url_for("reminder_queue"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT client_id, first_name, last_name, phone, birth_date
        FROM clients
        WHERE spa_id = %s
          AND active_client = TRUE
          AND ok_to_text = TRUE
          AND birth_date IS NOT NULL
          AND EXTRACT(MONTH FROM birth_date) = EXTRACT(MONTH FROM CURRENT_DATE)
          AND EXTRACT(DAY FROM birth_date) = EXTRACT(DAY FROM CURRENT_DATE)
    """, (spa_id,))

    clients = cur.fetchall()

    created_count = 0

    for client in clients:

        client_id = client[0]
        first_name = client[1]
        phone = client[3]


        cur.execute("""
            SELECT reminder_id
            FROM reminder_queue
            WHERE spa_id = %s
              AND client_id = %s
              AND reminder_type = 'birthday'
              AND DATE(scheduled_for) = CURRENT_DATE
              AND status IN ('pending', 'sent')
        """, (spa_id, client_id))

        existing = cur.fetchone()

        if existing:
            continue

        message_body = f"Happy Birthday {first_name}! Clear Skin Esthetics hopes you have a wonderful day!"

        cur.execute("""
            INSERT INTO reminder_queue
                (spa_id, client_id, reminder_type, send_method, recipient_phone, message_body, scheduled_for, status)
            VALUES
                (%s, %s, %s, %s, %s, %s, NOW(), %s)
        """, (
            spa_id,
            client_id,
            "birthday",
            "sms",
            phone,
            message_body,
            "pending"
        ))

        created_count += 1

    conn.commit()
    cur.close()
    conn.close()

    flash(f"{created_count} birthday reminder(s) added to the queue.", "success")
    return redirect(url_for("reminder_queue"))





          
#   --------------------------------------
#
#
#   RUN DAILY BIRTHDAY > > AUTOMATION<<<
#
#   
#   --------------------------------------
    

@app.route("/reminders/run-daily-birthday-job", methods=["POST"])
@login_required
@spa_required
def run_daily_birthday_job():
    spa_id = current_spa_id()

    created_count = generate_birthday_reminders(spa_id)

    flash(f"Daily birthday job complete. {created_count} reminder(s) added.", "success")
    return redirect(url_for("reminder_queue"))




















#   --------------------------------------
#
#
#   ADD EMAIL TEMPLATE
#
#
#   --------------------------------------
@app.route("/email-templates/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_email_template():
    return redirect(url_for("template_review", channel="email"))






############################################################
#    EMAIL TEMPLATE ARCHIVE
#
#
#
#       6-30-26
#
############################################################


@app.route("/admin/messaging-compliance/templates/archive/<int:template_id>", methods=["POST"])
@login_required
@spa_required
def archive_messaging_template(template_id):
    spa_id = current_spa_id()
    language_code = get_request_language()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT channel, template_name
        FROM messaging_templates
        WHERE template_id = %s
          AND spa_id = %s
        LIMIT 1
    """, (template_id, spa_id))

    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        flash("Template not found.", "warning")
        return redirect(url_for(
            "template_review",
            channel="sms",
            language_code=language_code
        ))

    channel = row[0]
    template_name = row[1] or "Default"

    if template_name.strip().lower() == "default":
        cur.close()
        conn.close()
        flash("Default templates cannot be archived.", "warning")
        return redirect(url_for(
            "template_review",
            channel=channel,
            language_code=language_code
        ))

    cur.execute("""
        UPDATE messaging_templates
        SET
            is_archived = TRUE,
            is_active = FALSE,
            updated_at = NOW()
        WHERE template_id = %s
          AND spa_id = %s
    """, (template_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Template archived.", "success")

    return redirect(url_for(
        "template_review",
        channel=channel,
        language_code=language_code
    ))










############################################################
#    EMAIL TEMPLATE RESTORE
############################################################


@app.route(
    "/admin/messaging-compliance/templates/<channel>/<template_type>/restore",
    methods=["POST"]
)
@login_required
@spa_required
def restore_messaging_template(channel, template_type):

    spa_id = session.get("spa_id")

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE messaging_templates
           SET is_archived = FALSE,
               is_active = FALSE,
               updated_at = NOW()
         WHERE spa_id = %s
           AND channel = %s
           AND template_type = %s
           AND is_archived = TRUE
    """, (
        spa_id,
        channel,
        template_type
    ))

    conn.commit()

    cur.close()
    conn.close()

    flash(
        "Template restored and placed in Disabled status.",
        "success"
    )

    return redirect(
        url_for(
            "template_review",
            channel=channel,
            show_archived=1
        )
    )




############################################################
#    EMAIL TEMPLATE ENABLE
############################################################



@app.route(
    "/admin/messaging-compliance/templates/<int:template_id>/enable",
    methods=["POST"]
)
@login_required
@spa_required
def enable_messaging_template(template_id):
    spa_id = current_spa_id()
    language_code = get_request_language()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT channel, approved_for_use, COALESCE(is_archived, FALSE)
        FROM messaging_templates
        WHERE template_id = %s
          AND spa_id = %s
        LIMIT 1
    """, (template_id, spa_id))

    template = cur.fetchone()

    if not template:
        cur.close()
        conn.close()
        flash("Template not found.", "warning")
        return redirect(url_for(
            "template_review",
            channel="sms",
            language_code=language_code
        ))

    channel = template[0]
    approved_for_use = template[1]
    is_archived = template[2]

    if is_archived:
        cur.close()
        conn.close()
        flash("Archived templates must be restored before they can be enabled.", "warning")
        return redirect(url_for(
            "template_review",
            channel=channel,
            language_code=language_code
        ))

    if not approved_for_use:
        cur.close()
        conn.close()
        flash("Template must be approved for use before it can be enabled.", "warning")
        return redirect(url_for(
            "template_review",
            channel=channel,
            language_code=language_code
        ))

    cur.execute("""
        UPDATE messaging_templates
        SET
            is_active = TRUE,
            updated_at = NOW()
        WHERE template_id = %s
          AND spa_id = %s
    """, (template_id, spa_id))

    conn.commit()

    cur.close()
    conn.close()

    flash("Template enabled.", "success")

    return redirect(url_for(
        "template_review",
        channel=channel,
        language_code=language_code
    ))







############################################################
#    EMAIL TEMPLATE DISABLE
############################################################


@app.route(
    "/admin/messaging-compliance/templates/<int:template_id>/disable",
    methods=["POST"]
)
@login_required
@spa_required
def disable_messaging_template(template_id):

    spa_id = session.get("spa_id")

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE messaging_templates
           SET is_active = FALSE,
               updated_at = NOW()
         WHERE template_id = %s
           AND spa_id = %s
    """, (
        template_id,
        spa_id
    ))

    conn.commit()

    flash(
        "Template disabled.",
        "success"
    )

    cur.close()
    conn.close()

    return redirect(request.referrer or url_for(
        "template_review",
        channel="sms"
    ))








    
#   --------------------------------------------------
#     
#   >>>>>  BUSINESS GOALS   <<<<<<<<<<<<<<<
#       
#   
#       
#           
#
#
#
#   --------------------------------------------------




@app.route("/spa-management/business-goals", methods=["GET", "POST"])
@login_required
@spa_required
def business_goals():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        daily_revenue_goal = request.form.get("daily_revenue_goal") or 0
        weekly_revenue_goal = request.form.get("weekly_revenue_goal") or 0
        monthly_revenue_goal = request.form.get("monthly_revenue_goal") or 0
        average_ticket_goal = request.form.get("average_ticket_goal") or 0

        new_clients_goal = request.form.get("new_clients_goal") or 0
        completion_rate_goal = request.form.get("completion_rate_goal") or 95
        cancellation_rate_goal = request.form.get("cancellation_rate_goal") or 5
        no_show_goal = request.form.get("no_show_goal") or 2

        inactive_client_days = request.form.get("inactive_client_days") or 90
        low_inventory_threshold = request.form.get("low_inventory_threshold") or 5

        cur.execute("""
            INSERT INTO spa_business_goals (
                spa_id,
                daily_revenue_goal,
                weekly_revenue_goal,
                monthly_revenue_goal,
                average_ticket_goal,
                new_clients_goal,
                completion_rate_goal,
                cancellation_rate_goal,
                no_show_goal,
                inactive_client_days,
                low_inventory_threshold,
                updated_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (spa_id)
            DO UPDATE SET
                daily_revenue_goal = EXCLUDED.daily_revenue_goal,
                weekly_revenue_goal = EXCLUDED.weekly_revenue_goal,
                monthly_revenue_goal = EXCLUDED.monthly_revenue_goal,
                average_ticket_goal = EXCLUDED.average_ticket_goal,
                new_clients_goal = EXCLUDED.new_clients_goal,
                completion_rate_goal = EXCLUDED.completion_rate_goal,
                cancellation_rate_goal = EXCLUDED.cancellation_rate_goal,
                no_show_goal = EXCLUDED.no_show_goal,
                inactive_client_days = EXCLUDED.inactive_client_days,
                low_inventory_threshold = EXCLUDED.low_inventory_threshold,
                updated_at = CURRENT_TIMESTAMP
        """, (
            spa_id,
            daily_revenue_goal,
            weekly_revenue_goal,
            monthly_revenue_goal,
            average_ticket_goal,
            new_clients_goal,
            completion_rate_goal,
            cancellation_rate_goal,
            no_show_goal,
            inactive_client_days,
            low_inventory_threshold
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Business goals updated successfully.", "success")
        return redirect(url_for("business_goals"))

    cur.execute("""
        SELECT
            daily_revenue_goal,
            weekly_revenue_goal,
            monthly_revenue_goal,
            average_ticket_goal,
            new_clients_goal,
            completion_rate_goal,
            cancellation_rate_goal,
            no_show_goal,
            inactive_client_days,
            low_inventory_threshold
        FROM spa_business_goals
        WHERE spa_id = %s
    """, (spa_id,))

    goals = cur.fetchone()

    cur.close()
    conn.close()

    return render_template("business_goals.html", goals=goals)






















    






            
#   --------------------------------------------------
#
#   >>>>>  CLIENT CONTACT PREFERENCES   <<<<<<<<<<<<<<<
#
#
#
#       
#
#
#           
#   --------------------------------------------------

@app.route("/client-contact-preferences")
@login_required
@spa_required
def client_contact_preferences():
    spa_id = current_spa_id()

    search = request.args.get("search", "").strip()
    filter_status = request.args.get("filter_status", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            client_id,
            first_name,
            last_name,
            phone,
            email,
            sms_opt_in,
            email_opt_in,
            ok_to_call
        FROM clients
        WHERE spa_id = %s
          AND active_client = TRUE
    """

    params = [spa_id]

    if search:
        query += """
            AND (
                first_name ILIKE %s OR
                last_name ILIKE %s OR
                phone ILIKE %s OR
                email ILIKE %s
            )
        """
        like_search = f"%{search}%"
        params.extend([like_search, like_search, like_search, like_search])

    if filter_status == "sms_yes":
        query += """
            AND sms_opt_in = TRUE
            AND COALESCE(sms_opt_out, FALSE) = FALSE
        """
    elif filter_status == "sms_no":
        query += """
            AND (
                sms_opt_out = TRUE
                OR COALESCE(sms_opt_in, FALSE) = FALSE
            )
        """
    elif filter_status == "email_yes":
        query += """
            AND email_opt_in = TRUE
            AND COALESCE(email_opt_out, FALSE) = FALSE
        """
    elif filter_status == "email_no":
        query += """
            AND (
                email_opt_out = TRUE
                OR COALESCE(email_opt_in, FALSE) = FALSE
            )
        """
    elif filter_status == "missing_phone":
        query += " AND (phone IS NULL OR TRIM(phone) = '')"
    elif filter_status == "missing_email":
        query += " AND (email IS NULL OR TRIM(email) = '')"

    query += " ORDER BY last_name, first_name"

    cur.execute(query, params)
    clients = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "client_contact_preferences.html",
        clients=clients,
        search=search,
        filter_status=filter_status
    )















#   -------------------------------
#
#   EDIT PREFERENCES
#   FOR SMS - EMAIL -CALL
#
#   ------------------------------

@app.route("/client-contact-preferences/edit/<int:client_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_client_contact_preferences(client_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        sms_opt_in = request.form.get("sms_opt_in") == "on"
        sms_opt_out = request.form.get("sms_opt_out") == "on"
        email_opt_in = request.form.get("email_opt_in") == "on"
        email_opt_out = request.form.get("email_opt_out") == "on"
        ok_to_call = request.form.get("ok_to_call") == "on"

        print("FORM DATA:", request.form, flush=True)
        print("EMAIL OPT IN:", email_opt_in, flush=True)
        print("EMAIL OPT OUT:", email_opt_out, flush=True)




        cur.execute("""
            SELECT
                sms_opt_in,
                sms_opt_out,
                email_opt_in,
                email_opt_out,
                ok_to_call
            FROM clients
            WHERE client_id = %s
              AND spa_id = %s
        """, (client_id, spa_id))

        old_prefs = cur.fetchone()

        if not old_prefs:
            cur.close()
            conn.close()
            flash("Client not found.", "error")
            return redirect(url_for("client_contact_preferences"))

        old_sms_opt_in = old_prefs[0]
        old_sms_opt_out = old_prefs[1]
        old_email_opt_in = old_prefs[2]
        old_email_opt_out = old_prefs[3]
        old_call = old_prefs[4]

        cur.execute("""
            UPDATE clients
            SET
                sms_opt_in = %s,
                sms_opt_out = %s,
                email_opt_in = %s,
                email_opt_out = %s,
                ok_to_call = %s
            WHERE client_id = %s
              AND spa_id = %s
        """, (
            sms_opt_in,
            sms_opt_out,
            email_opt_in,
            email_opt_out,
            ok_to_call,
            client_id,
            spa_id
        ))

        updated_by = session.get("user_id")

        if old_sms_opt_in != sms_opt_in:
            add_consent_record(
                cur, spa_id, client_id,
                "SMS", sms_opt_in,
                "Admin Updated",
                "SMS opt-in updated from contact preferences page.",
                updated_by
            )

        if old_sms_opt_out != sms_opt_out:
            add_consent_record(
                cur, spa_id, client_id,
                "SMS", not sms_opt_out,
                "Admin Updated",
                "SMS opt-out updated from contact preferences page.",
                updated_by
            )

        if old_email_opt_in != email_opt_in:
            add_consent_record(
                cur, spa_id, client_id,
                "Email", email_opt_in,
                "Admin Updated",
                "Email opt-in updated from contact preferences page.",
                updated_by
            )

        if old_email_opt_out != email_opt_out:
            add_consent_record(
                cur, spa_id, client_id,
                "Email", not email_opt_out,
                "Admin Updated",
                "Email opt-out updated from contact preferences page.",
                updated_by
            )

        if old_call != ok_to_call:
            add_consent_record(
                cur, spa_id, client_id,
                "Phone", ok_to_call,
                "Admin Updated",
                "Phone contact consent updated from contact preferences page.",
                updated_by
            )

        conn.commit()
        cur.close()
        conn.close()

        flash("Contact preferences updated.", "success")
        return redirect(url_for("client_contact_preferences"))

    cur.execute("""
        SELECT
            client_id,
            first_name,
            last_name,
            phone,
            email,
            sms_opt_in,
            sms_opt_out,
            email_opt_in,
            email_opt_out,
            ok_to_call
        FROM clients
        WHERE client_id = %s
          AND spa_id = %s
          AND active_client = TRUE
    """, (client_id, spa_id))






    client = cur.fetchone()

    cur.close()
    conn.close()

    if not client:
        flash("Client not found.", "error")
        return redirect(url_for("client_contact_preferences"))

    return render_template(
        "edit_client_contact_preferences.html",
        client=client
    )






#   -----------------------------------
#
#   CLIENT CONSENT HISTORY
#
#
#   ------------------------------------



@app.route("/client-consent-history/<int:client_id>")
@login_required
@spa_required
def client_consent_history(client_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT first_name, last_name
        FROM clients
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, spa_id))

    client = cur.fetchone()

    if not client:
        cur.close()
        conn.close()
        flash("Client not found.", "error")
        return redirect(url_for("client_contact_preferences"))

    cur.execute("""
        SELECT
            consent_type,
            consent_status,
            consent_source,
            consent_note,
            created_at
        FROM consent_records
        WHERE client_id = %s
          AND spa_id = %s
        ORDER BY created_at DESC
    """, (client_id, spa_id))

    records = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "client_consent_history.html",
        client=client,
        records=records
    )










#   --------------------------------------------------
#
#             >>>>>  INVENTORY   <<<<<<<<<<<<<<<
#
#
#
#
#
#
#
#   --------------------------------------------------


@app.route("/inventory")
@login_required
@spa_required
def inventory_home():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            p.product_id,
            p.sku,
            p.product_name,
            p.vendor_company,
            p.product_category,
            p.product_sub_category,
            p.wholesale_cost,
            p.suggested_retail,
            p.expire_date,

            COALESCE(SUM(CASE WHEN m.movement_type = 'added' THEN m.quantity ELSE 0 END), 0) AS total_added,
            COALESCE(SUM(CASE WHEN m.movement_type = 'sold' THEN m.quantity ELSE 0 END), 0) AS total_sold,

            COALESCE(SUM(
                CASE
                    WHEN m.movement_type = 'added' THEN m.quantity
                    WHEN m.movement_type IN ('sold', 'expired', 'damaged') THEN -m.quantity
                    WHEN m.movement_type = 'returned' THEN m.quantity
                    WHEN m.movement_type = 'adjustment' THEN m.quantity
                    ELSE 0
                END
            ), 0) AS total_in_stock,

            COALESCE(SUM(
                CASE
                    WHEN m.movement_type = 'added' THEN m.quantity
                    WHEN m.movement_type IN ('sold', 'expired', 'damaged') THEN -m.quantity
                    WHEN m.movement_type = 'returned' THEN m.quantity
                    WHEN m.movement_type = 'adjustment' THEN m.quantity
                    ELSE 0
                END
            ), 0) * p.wholesale_cost AS inventory_value,

           
           COALESCE(SUM(
               CASE
                   WHEN m.movement_type = 'added' THEN m.quantity
                   WHEN m.movement_type IN ('sold', 'expired', 'damaged') THEN -m.quantity
                   WHEN m.movement_type = 'returned' THEN m.quantity
                   WHEN m.movement_type = 'adjustment' THEN m.quantity
                   ELSE 0
               END
            ), 0) * p.suggested_retail AS inventory_retail_value,


           COALESCE(SUM(
               CASE
                   WHEN m.movement_type = 'added' THEN m.quantity
                   WHEN m.movement_type IN ('sold', 'expired', 'damaged') THEN -m.quantity
                   WHEN m.movement_type = 'returned' THEN m.quantity
                   WHEN m.movement_type = 'adjustment' THEN m.quantity
                   ELSE 0
               END
           ), 0) * p.suggested_retail AS retail_value    


        FROM inventory_products p
        LEFT JOIN inventory_movements m
            ON p.product_id = m.product_id
           AND m.spa_id = %s
        WHERE p.spa_id = %s
          AND p.active = TRUE
        GROUP BY p.product_id
        ORDER BY p.product_name
    """, (spa_id, spa_id))

    inventory_rows = cur.fetchall()


    total_stock = sum(row[11] for row in inventory_rows)

    total_wholesale_value = sum(row[12] for row in inventory_rows)

    total_retail_value = sum(row[13] for row in inventory_rows)

    cur.close()
    conn.close()

    return render_template(
        "inventory_home.html",
        inventory_rows=inventory_rows,
        total_stock=total_stock,
        total_wholesale_value=total_wholesale_value,
        total_retail_value=total_retail_value
    )








            

#   --------------------------------------
#     >>  INVENTORY ADD
#
#   --------------------------------------


@app.route("/inventory/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_inventory_product():
    spa_id = current_spa_id()
    prefill_sku = request.args.get("sku", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":

        sku = request.form.get("sku", "").strip()
        product_name = request.form.get("product_name", "").strip()
        vendor_company = request.form.get("vendor_company", "").strip()
        product_category = request.form.get("product_category", "").strip()
        product_sub_category = request.form.get("product_sub_category", "").strip()

        wholesale_cost = request.form.get("wholesale_cost") or 0
        suggested_retail = request.form.get("suggested_retail") or 0

        expire_date = request.form.get("expire_date") or None
        note = request.form.get("note", "").strip()

        initial_quantity = request.form.get("initial_quantity") or 0

        if not sku:
            flash("SKU is required.", "error")
            return redirect(url_for("add_inventory_product"))

        if not product_name:
            flash("Product name is required.", "error")
            return redirect(url_for("add_inventory_product"))

        try:

            cur.execute("""
                SELECT product_id
                FROM inventory_products
                WHERE spa_id = %s
                  AND sku = %s
            """, (spa_id, sku))

            existing = cur.fetchone()

            if existing:
                flash("SKU already exists. Opening existing product", "warning")
                cur.close()
                conn.close()
                return redirect(url_for("add_inventory_product"))

            cur.execute("""
                INSERT INTO inventory_products (
                    spa_id,
                    sku,
                    expire_date,
                    vendor_company,
                    product_name,
                    product_category,
                    product_sub_category,
                    wholesale_cost,
                    suggested_retail,
                    note
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING product_id
            """, (
                spa_id,
                sku,
                expire_date,
                vendor_company if vendor_company else None,
                product_name,
                product_category if product_category else None,
                product_sub_category if product_sub_category else None,
                wholesale_cost,
                suggested_retail,
                note if note else None
            ))

            product_id = cur.fetchone()[0]

            if int(initial_quantity) > 0:

                cur.execute("""
                    INSERT INTO inventory_movements (
                        spa_id,
                        product_id,
                        movement_type,
                        quantity,
                        note
                    )
                    VALUES (%s, %s, %s, %s, %s)
                """, (
                    spa_id,
                    product_id,
                    "added",
                    initial_quantity,
                    "Initial inventory entry"
                ))

            conn.commit()

            flash("Inventory product added successfully.", "success")

            return redirect(url_for("inventory_home"))

        except Exception as e:
            conn.rollback()
            flash(f"Error adding inventory product: {e}", "error")

        finally:
            cur.close()
            conn.close()

    cur.close()
    conn.close()

    return render_template(
        "add_inventory_product.html",
        prefill_sku=prefill_sku
    )






#   --------------------------------------
#       INVENTORY ADD STOCK
#
#   --------------------------------------   


@app.route("/inventory/add_stock/<int:product_id>", methods=["GET", "POST"])
@login_required
@spa_required
def add_inventory_stock(product_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT product_id, sku, product_name, vendor_company
        FROM inventory_products
        WHERE product_id = %s
          AND spa_id = %s
          AND active = TRUE
    """, (product_id, spa_id))

    product = cur.fetchone()

    if not product:
        cur.close()
        conn.close()
        flash("Inventory product not found.", "error")
        return redirect(url_for("inventory_home"))

    if request.method == "POST":
        quantity = request.form.get("quantity") or 0
        note = request.form.get("note", "").strip()

        try:
            quantity = int(quantity)

            if quantity <= 0:
                flash("Quantity must be greater than zero.", "error")
                return redirect(url_for("add_inventory_stock", product_id=product_id))

            cur.execute("""
                INSERT INTO inventory_movements (
                    spa_id,
                    product_id,
                    movement_type,
                    quantity,
                    note
                )
                VALUES (%s, %s, %s, %s, %s)
            """, (
                spa_id,
                product_id,
                "added",
                quantity,
                note if note else "Stock added"
            ))

            conn.commit()
            flash("Stock added successfully.", "success")

            return redirect(url_for("inventory_home"))

        except Exception as e:
            conn.rollback()
            flash(f"Error adding stock: {e}", "error")

        finally:
            cur.close()
            conn.close()

    cur.close()
    conn.close()

    return render_template(
        "add_inventory_stock.html",
        product=product
    )









#   --------------------------------------
#         INVENTORY SOLD    
#
#   --------------------------------------   


@app.route("/inventory/sold/<int:product_id>", methods=["GET", "POST"])
@login_required
@spa_required
def record_inventory_sold(product_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT product_id, sku, product_name, vendor_company
        FROM inventory_products
        WHERE product_id = %s
          AND spa_id = %s
          AND active = TRUE
    """, (product_id, spa_id))

    product = cur.fetchone()

    if not product:
        cur.close()
        conn.close()
        flash("Inventory product not found.", "error")
        return redirect(url_for("inventory_home"))

    if request.method == "POST":
        quantity = request.form.get("quantity") or 0
        note = request.form.get("note", "").strip()

        try:
            quantity = int(quantity)

            if quantity <= 0:
                flash("Quantity must be greater than zero.", "error")
                return redirect(url_for("record_inventory_sold", product_id=product_id))

            cur.execute("""
                INSERT INTO inventory_movements (
                    spa_id,
                    product_id,
                    movement_type,
                    quantity,
                    note
                )
                VALUES (%s, %s, %s, %s, %s)
            """, (
                spa_id,
                product_id,
                "sold",
                quantity,
                note if note else "Inventory sold"
            ))

            conn.commit()
            flash("Inventory sale recorded.", "success")
            return redirect(url_for("inventory_home"))

        except Exception as e:
            conn.rollback()
            flash(f"Error recording sale: {e}", "error")

        finally:
            cur.close()
            conn.close()

    cur.close()
    conn.close()

    return render_template(
        "record_inventory_sold.html",
        product=product
    )










#   --------------------------------------
#      INVENTORY ADJUSTMENTS
#
#   --------------------------------------   


@app.route("/inventory/adjust/<int:product_id>", methods=["GET", "POST"])
@login_required
@spa_required
def adjust_inventory_stock(product_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT product_id, sku, product_name, vendor_company
        FROM inventory_products
        WHERE product_id = %s
          AND spa_id = %s
          AND active = TRUE
    """, (product_id, spa_id))

    product = cur.fetchone()

    if not product:
        cur.close()
        conn.close()
        flash("Inventory product not found.", "error")
        return redirect(url_for("inventory_home"))

    if request.method == "POST":
        movement_type = request.form.get("movement_type", "").strip()
        quantity = request.form.get("quantity") or 0
        note = request.form.get("note", "").strip()

        allowed_types = ["adjustment", "expired", "damaged", "returned"]

        if movement_type not in allowed_types:
            flash("Invalid adjustment type.", "error")
            return redirect(url_for("adjust_inventory_stock", product_id=product_id))

        try:
            quantity = int(quantity)

            if quantity <= 0:
                flash("Quantity must be greater than zero.", "error")
                return redirect(url_for("adjust_inventory_stock", product_id=product_id))

            cur.execute("""
                INSERT INTO inventory_movements (
                    spa_id,
                    product_id,
                    movement_type,
                    quantity,
                    note
                )
                VALUES (%s, %s, %s, %s, %s)
            """, (
                spa_id,
                product_id,
                movement_type,
                quantity,
                note if note else f"Inventory {movement_type}"
            ))

            conn.commit()
            flash("Inventory adjustment recorded.", "success")
            return redirect(url_for("inventory_home"))

        except Exception as e:
            conn.rollback()
            flash(f"Error recording adjustment: {e}", "error")

        finally:
            cur.close()
            conn.close()

    cur.close()
    conn.close()

    return render_template(
        "adjust_inventory_stock.html",
        product=product
    )








#   --------------------------------------
#      INVENTORY DETAILS
#   
#   --------------------------------------


@app.route("/inventory/product/<int:product_id>")
@login_required
@spa_required
def inventory_product_detail(product_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            product_id,
            sku,
            product_name,
            vendor_company,
            product_category,
            product_sub_category,
            wholesale_cost,
            suggested_retail,
            expire_date,
            note,
            date_added
        FROM inventory_products
        WHERE product_id = %s
          AND spa_id = %s
          AND active = TRUE
    """, (product_id, spa_id))

    product = cur.fetchone()

    if not product:
        cur.close()
        conn.close()
        flash("Inventory product not found.", "error")
        return redirect(url_for("inventory_home"))

    cur.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN movement_type = 'added' THEN quantity ELSE 0 END), 0) AS total_added,
            COALESCE(SUM(CASE WHEN movement_type = 'sold' THEN quantity ELSE 0 END), 0) AS total_sold,
            COALESCE(SUM(CASE WHEN movement_type = 'expired' THEN quantity ELSE 0 END), 0) AS total_expired,
            COALESCE(SUM(CASE WHEN movement_type = 'damaged' THEN quantity ELSE 0 END), 0) AS total_damaged,
            COALESCE(SUM(CASE WHEN movement_type = 'returned' THEN quantity ELSE 0 END), 0) AS total_returned,
            COALESCE(SUM(CASE WHEN movement_type = 'adjustment' THEN quantity ELSE 0 END), 0) AS total_adjusted,

            COALESCE(SUM(
                CASE
                    WHEN movement_type = 'added' THEN quantity
                    WHEN movement_type IN ('sold', 'expired', 'damaged') THEN -quantity
                    WHEN movement_type IN ('returned', 'adjustment') THEN quantity
                    ELSE 0
                END
            ), 0) AS total_in_stock
        FROM inventory_movements
        WHERE spa_id = %s
          AND product_id = %s
    """, (spa_id, product_id))

    totals = cur.fetchone()

    cur.execute("""
        SELECT
            movement_date,
            movement_type,
            quantity,
            note
        FROM inventory_movements
        WHERE spa_id = %s
          AND product_id = %s
        ORDER BY movement_date DESC, movement_id DESC
    """, (spa_id, product_id))

    movement_rows = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "inventory_product_detail.html",
        product=product,
        totals=totals,
        movement_rows=movement_rows
    )








#   --------------------------------------
#         EDIT INVENTORY  PRODUCT   
#   
#   --------------------------------------


@app.route("/inventory/edit/<int:product_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_inventory_product(product_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            product_id,
            sku,
            expire_date,
            vendor_company,
            product_name,
            product_category,
            product_sub_category,
            wholesale_cost,
            suggested_retail,
            note,
            active
        FROM inventory_products
        WHERE product_id = %s
          AND spa_id = %s
    """, (product_id, spa_id))

    product = cur.fetchone()

    if not product:
        cur.close()
        conn.close()
        flash("Inventory product not found.", "error")
        return redirect(url_for("inventory_home"))

    if request.method == "POST":
        sku = request.form.get("sku", "").strip()
        product_name = request.form.get("product_name", "").strip()
        vendor_company = request.form.get("vendor_company", "").strip()
        product_category = request.form.get("product_category", "").strip()
        product_sub_category = request.form.get("product_sub_category", "").strip()
        wholesale_cost = request.form.get("wholesale_cost") or 0
        suggested_retail = request.form.get("suggested_retail") or 0
        expire_date = request.form.get("expire_date") or None
        note = request.form.get("note", "").strip()
        active = True if request.form.get("active") == "on" else False

        if not sku:
            flash("SKU is required.", "error")
            return redirect(url_for("edit_inventory_product", product_id=product_id))

        if not product_name:
            flash("Product name is required.", "error")
            return redirect(url_for("edit_inventory_product", product_id=product_id))

        try:
            cur.execute("""
                UPDATE inventory_products
                SET
                    sku = %s,
                    expire_date = %s,
                    vendor_company = %s,
                    product_name = %s,
                    product_category = %s,
                    product_sub_category = %s,
                    wholesale_cost = %s,
                    suggested_retail = %s,
                    note = %s,
                    active = %s
                WHERE product_id = %s
                  AND spa_id = %s
            """, (
                sku,
                expire_date,
                vendor_company if vendor_company else None,
                product_name,
                product_category if product_category else None,
                product_sub_category if product_sub_category else None,
                wholesale_cost,
                suggested_retail,
                note if note else None,
                active,
                product_id,
                spa_id
            ))

            conn.commit()
            flash("Inventory product updated successfully.", "success")
            return redirect(url_for("inventory_product_detail", product_id=product_id))

        except Exception as e:
            conn.rollback()
            flash(f"Error updating inventory product: {e}", "error")

        finally:
            cur.close()
            conn.close()

    cur.close()
    conn.close()

    return render_template(
        "edit_inventory_product.html",
        product=product
    )











#   --------------------------------------
#   INVENTORY SCAN
#   
#   --------------------------------------


@app.route("/inventory/scan")
@login_required
@spa_required
def inventory_scan():
    return render_template("inventory_scan.html")




#   --------------------------------------
#   INVENTORY SCAN  RESULT
#
#   --------------------------------------

@app.route("/inventory/scan-result")
@login_required
@spa_required
def inventory_scan_result():
    spa_id = current_spa_id()
    scanned_sku = request.args.get("sku", "").strip()

    if not scanned_sku:
        flash("No SKU scanned.", "error")
        return redirect(url_for("inventory_scan"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT product_id
        FROM inventory_products
        WHERE spa_id = %s
          AND sku = %s
          AND active = TRUE
    """, (spa_id, scanned_sku))

    product = cur.fetchone()

    cur.close()
    conn.close()

    if product:
        flash(f"Product found: {scanned_sku}", "success")
        return redirect(url_for(
            "inventory_product_detail",
            product_id=product[0]
        ))

    flash(f"SKU not found: {scanned_sku}", "warning")
    return redirect(url_for("add_inventory_product") + f"?sku={scanned_sku}")










#   --------------------------------------
#     DEACTIVATE PRODUCT
#   
#   --------------------------------------


@app.route("/inventory/deactivate/<int:product_id>", methods=["POST"])
@login_required
@spa_required
def deactivate_inventory_product(product_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE inventory_products
        SET active = FALSE
        WHERE product_id = %s
          AND spa_id = %s
    """, (product_id, spa_id))

    conn.commit()

    cur.close()
    conn.close()

    flash("Inventory product deactivated.", "success")
    return redirect(url_for("inventory_home"))















#   --------------------------------------
#   
#   
#   --------------------------------------








#   --------------------------------------
#   
#   
#   --------------------------------------






#   --------------------------------------
#   
#   
#   --------------------------------------








#   --------------------------------------
#    
#
#   --------------------------------------   



#   --------------------------------------
#     LOAN CONTRIBUTIONS EXPORT   CSV
#
#   --------------------------------------

@app.route("/loan_contributions/export/csv")
@login_required
@spa_required
def export_loan_contributions_csv():
    spa_id = current_spa_id()

    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    rows = get_loan_contribution_rows(
        spa_id=spa_id,
        start_date=start_date or None,
        end_date=end_date or None
    )

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Date",
        "Type",
        "Description",
        "Amount",
        "Payment Method",
        "Loan",
        "Notes"
    ])

    for row in rows:
        writer.writerow([
            row[0].strftime("%Y-%m-%d") if row[0] else "",
            row[1] or "",
            row[2] or "",
            f"{float(row[3]):.2f}" if row[3] is not None else "0.00",
            row[4] or "",
            row[5] or "",
            row[6] or ""
        ])

    csv_data = output.getvalue()
    output.close()

    return Response(
        csv_data,
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=loan_contributions.csv"
        }
    )


#   --------------------------------------
#     LOAN CONTRIBUTIONS EXPORT   EXCEL
#
#   --------------------------------------

@app.route("/loan_contributions/export/excel")
@login_required
@spa_required
def export_loan_contributions_excel():
    spa_id = current_spa_id()

    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    rows = get_loan_contribution_rows(
        spa_id=spa_id,
        start_date=start_date or None,
        end_date=end_date or None
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Loans & Contributions"

    headers = [
        "Date",
        "Type",
        "Description",
        "Amount",
        "Payment Method",
        "Loan",
        "Notes"
    ]

    ws.append(headers)

    for row in rows:
        ws.append([
            row[0].strftime("%Y-%m-%d") if row[0] else "",
            row[1] or "",
            row[2] or "",
            float(row[3]) if row[3] is not None else 0.00,
            row[4] or "",
            row[5] or "",
            row[6] or ""
        ])

    # make columns fit nicely
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    # format amount column
    for cell in ws["D"][1:]:
        cell.number_format = "$#,##0.00"

    file_data = io.BytesIO()
    wb.save(file_data)
    file_data.seek(0)

    return send_file(
        file_data,
        as_attachment=True,
        download_name="loan_contributions.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )















#   -----------------------  
#
#      FUNDING     
#
#    spa_id good
#  ----------------------
    


@app.route("/funding")
@login_required
@spa_required
def funding_home():
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM owner_contributions
        WHERE spa_id = %s
    """, (spa_id,))
    total_contributions = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM owner_reimbursements
        WHERE spa_id = %s
    """, (spa_id,))
    total_reimbursements = cur.fetchone()[0]

    net_owner_funding = total_contributions - total_reimbursements

    cur.execute("""
        SELECT
            owner_contribution_id,
            contribution_date,
            amount,
            funding_source,
            notes
        FROM owner_contributions
        WHERE spa_id = %s
        ORDER BY contribution_date DESC, owner_contribution_id DESC
    """, (spa_id,))
    contributions = cur.fetchall()

    cur.execute("""
        SELECT
            owner_reimbursement_id,
            reimbursement_date,
            amount,
            payment_method,
            notes
        FROM owner_reimbursements
        WHERE spa_id = %s
        ORDER BY reimbursement_date DESC, owner_reimbursement_id DESC
    """, (spa_id,))
    reimbursements = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "funding_home.html",
        total_contributions=total_contributions,
        total_reimbursements=total_reimbursements,
        net_owner_funding=net_owner_funding,
        contributions=contributions,
        reimbursements=reimbursements
    )
                
                
                
#   -----------------------
#
#    OWNER CONTRIBUTIONS                     
#
#     spa_id good
#  ----------------------
                    
                    
                      
@app.route("/owner_contributions/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_owner_contribution():
    spa_id = current_spa_id()

    if request.method == "POST":
        contribution_date = request.form.get("contribution_date")
        amount = request.form.get("amount")
        funding_source = request.form.get("funding_source", "").strip()
        notes = request.form.get("notes", "").strip()

        conn = get_db_connection()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO owner_contributions (
                spa_id,
                contribution_date,
                amount,
                funding_source,
                notes
            )
            VALUES (%s, %s, %s, %s, %s)
        """, (
            spa_id,
            contribution_date,
            amount,
            funding_source,
            notes
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Owner contribution added successfully.", "success")
        return redirect(url_for("funding_home"))

    return render_template("add_owner_contribution.html")
                
                

                
#   -----------------------
#
#   OWNER REIMBURSEMENTS      
#
#    spa_id good
#  ----------------------
                    
                    
@app.route("/owner_reimbursements/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_owner_reimbursement():
    spa_id = current_spa_id()
        
    if request.method == "POST":
        reimbursement_date = request.form.get("reimbursement_date")
        amount = float(request.form.get("amount") or 0)
        payment_method = request.form.get("payment_method", "").strip()
        notes = request.form.get("notes", "").strip()

        if not reimbursement_date or amount <= 0:
            flash("Valid date and amount are required.", "error")
            return redirect(url_for("add_owner_reimbursement"))

        conn = get_db_connection()
        cur = conn.cursor()
             
        try:
            cur.execute("""
                INSERT INTO owner_reimbursements (
                    spa_id,
                    reimbursement_date,
                    amount,
                    payment_method,
                    notes
                )
                VALUES (%s, %s, %s, %s, %s)
            """, (
                spa_id, 
                reimbursement_date,
                amount,
                payment_method,
                notes
            ))
    
            conn.commit()
            flash("Owner reimbursement added successfully.", "success")

        except Exception as e:
            conn.rollback()
            flash(f"Error saving reimbursement: {e}", "error")

        finally:
            cur.close()
            conn.close()

        return redirect(url_for("funding_home"))

    return render_template("add_owner_reimbursement.html")

                
                
                
#   -----------------------
#
#   LOANS   HOME                   
#
#    spa_id good   
#  ----------------------
                    
                    
@app.route("/loans")
@login_required
@spa_required
def loans_home():
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            bl.loan_id,
            bl.loan_name,
            bl.lender_name,
            bl.loan_start_date,
            bl.original_amount,
            bl.interest_rate,
            bl.term_months,
            bl.notes,
            bl.is_active,
            COALESCE(SUM(lp.principal_paid), 0) AS principal_paid_total,
            COALESCE(SUM(lp.interest_paid), 0) AS interest_paid_total,
            bl.original_amount - COALESCE(SUM(lp.principal_paid), 0) AS remaining_balance
        FROM business_loans bl
        LEFT JOIN loan_payments lp
            ON bl.loan_id = lp.loan_id
           AND lp.spa_id = bl.spa_id
        WHERE bl.spa_id = %s
        GROUP BY
            bl.loan_id,
            bl.loan_name,
            bl.lender_name,
            bl.loan_start_date,
            bl.original_amount,
            bl.interest_rate,
            bl.term_months,
            bl.notes,
            bl.is_active
        ORDER BY bl.loan_start_date DESC NULLS LAST, bl.loan_id DESC
    """, (spa_id,))
    loans = cur.fetchall()

    cur.execute("""
        SELECT COALESCE(SUM(original_amount), 0)
        FROM business_loans
        WHERE spa_id = %s
    """, (spa_id,))
    total_original_loans = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(SUM(principal_paid), 0)
        FROM loan_payments
        WHERE spa_id = %s
    """, (spa_id,))
    total_principal_paid = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(SUM(interest_paid), 0)
        FROM loan_payments
        WHERE spa_id = %s
    """, (spa_id,))
    total_interest_paid = cur.fetchone()[0]

    total_remaining_balance = total_original_loans - total_principal_paid

    cur.close()
    conn.close()

    return render_template(
        "loans_home.html",
        loans=loans,
        total_original_loans=total_original_loans,
        total_principal_paid=total_principal_paid,
        total_interest_paid=total_interest_paid,
        total_remaining_balance=total_remaining_balance
    )                      

                
                

                
#   -----------------------
#
#    ADD BUSINESS LOANS                     
#
#    spa_id good
#  ----------------------
                    
                    
        
@app.route("/business_loans/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_business_loan(): 
    spa_id = current_spa_id()
            
    if request.method == "POST":
        loan_name = request.form.get("loan_name", "").strip()
        lender_name = request.form.get("lender_name", "").strip()
        loan_start_date = request.form.get("loan_start_date") or None
        original_amount = request.form.get("original_amount")
        interest_rate = request.form.get("interest_rate") or None
        term_months = request.form.get("term_months") or None
        notes = request.form.get("notes", "").strip()
        is_active = True if request.form.get("is_active") == "yes" else False
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            INSERT INTO business_loans (
                spa_id,  
                loan_name, 
                lender_name,   
                loan_start_date,
                original_amount,
                interest_rate,
                term_months,
                notes,  
                is_active
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            spa_id,
            loan_name,
            lender_name,   
            loan_start_date,
            original_amount,
            interest_rate,
            term_months,
            notes, 
            is_active
        ))
        
        conn.commit()
        cur.close()
        conn.close()
    
        flash("Business loan added successfully.", "success")
        return redirect(url_for("loans_home"))
        
    return render_template("add_business_loan.html")
                      


                
                
                
#   -----------------------
#
#    ADD LOAN PAYMENT                     
#
#   spa_id good
#  ----------------------
                    
                    
@app.route("/loan_payments/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_loan_payment():
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        loan_id = request.form.get("loan_id")
        payment_date = request.form.get("payment_date")
        principal_paid = float(request.form.get("principal_paid") or 0)
        interest_paid = float(request.form.get("interest_paid") or 0)
        total_payment = principal_paid + interest_paid
        notes = request.form.get("notes", "").strip()

        if not loan_id or not payment_date:
            cur.close()
            conn.close()
            flash("Loan and payment date are required.", "error")
            return redirect(url_for("add_loan_payment"))

        cur.execute("""
            SELECT loan_id
            FROM business_loans
            WHERE loan_id = %s
              AND spa_id = %s
        """, (loan_id, spa_id))
        valid_loan = cur.fetchone()

        if not valid_loan:
            cur.close()
            conn.close()
            flash("Invalid loan selected.", "error")
            return redirect(url_for("add_loan_payment"))

        cur.execute("""
            INSERT INTO loan_payments (
                spa_id,
                loan_id,
                payment_date,
                principal_paid,
                interest_paid,
                total_payment,
                notes
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            spa_id,
            loan_id,
            payment_date,
            principal_paid,
            interest_paid,
            total_payment,
            notes
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Loan payment added successfully.", "success")
        return redirect(url_for("loans_home"))

    cur.execute("""
        SELECT loan_id, loan_name
        FROM business_loans
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY loan_name
    """, (spa_id,))
    loans = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "add_loan_payment.html",
        loans=loans
    )                      






#   ------------------------------------
#   
#         EDIT OWNER CONTRIBUTIONS   
#  
#  spa id and route good  
#   --------------------------------


@app.route("/owner_contributions/edit/<int:owner_contribution_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_owner_contribution(owner_contribution_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        contribution_date = request.form.get("contribution_date")
        amount = request.form.get("amount")
        funding_source = request.form.get("funding_source", "").strip()
        notes = request.form.get("notes", "").strip()

        cur.execute("""
            UPDATE owner_contributions
            SET contribution_date = %s,
                amount = %s,
                funding_source = %s,
                notes = %s
            WHERE owner_contribution_id = %s
              AND spa_id = %s
        """, (
            contribution_date,
            amount,
            funding_source,
            notes,
            owner_contribution_id,
            spa_id
        ))

        if cur.rowcount == 0:
            conn.rollback()
            cur.close()
            conn.close()
            flash("Owner contribution not found.", "error")
            return redirect(url_for("funding_home"))

        conn.commit()
        cur.close()
        conn.close()

        flash("Owner contribution updated successfully.", "success")
        return redirect(url_for("funding_home"))

    cur.execute("""
        SELECT
            owner_contribution_id,
            contribution_date,
            amount,
            funding_source,
            notes
        FROM owner_contributions
        WHERE owner_contribution_id = %s
          AND spa_id = %s
    """, (owner_contribution_id, spa_id))

    contribution = cur.fetchone()

    cur.close()
    conn.close()

    if not contribution:
        flash("Owner contribution not found.", "error")
        return redirect(url_for("funding_home"))

    return render_template(
        "edit_owner_contribution.html",
        contribution=contribution
    )








#   ------------------------------------
#
#      DELETE OWNER CONTRIBUTIONS
#                    
#        spa id and route good
#   --------------------------------



@app.route("/owner_contributions/delete/<int:owner_contribution_id>", methods=["POST"])
@login_required
@spa_required
def delete_owner_contribution(owner_contribution_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM owner_contributions
        WHERE owner_contribution_id = %s
          AND spa_id = %s
    """, (owner_contribution_id, spa_id))

    if cur.rowcount == 0:
        conn.rollback()
        cur.close()
        conn.close()
        flash("Owner contribution not found.", "error")
        return redirect(url_for("funding_home"))

    conn.commit()
    cur.close()
    conn.close()

    flash("Owner contribution deleted successfully.", "success")
    return redirect(url_for("funding_home"))







        
#   ------------------------------------
#
#    EDIT OWNER REIMBURSEMENTS
#
#   spa id and route ok
#   --------------------------------
                     

@app.route("/owner_reimbursements/edit/<int:owner_reimbursement_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_owner_reimbursement(owner_reimbursement_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        reimbursement_date = request.form.get("reimbursement_date")
        amount = request.form.get("amount")
        payment_method = request.form.get("payment_method", "").strip()
        notes = request.form.get("notes", "").strip()

        cur.execute("""
            UPDATE owner_reimbursements
            SET reimbursement_date = %s,
                amount = %s,
                payment_method = %s,
                notes = %s
            WHERE owner_reimbursement_id = %s
              AND spa_id = %s
        """, (
            reimbursement_date,
            amount,
            payment_method,
            notes,
            owner_reimbursement_id,
            spa_id
        ))

        if cur.rowcount == 0:
            conn.rollback()
            cur.close()
            conn.close()
            flash("Owner reimbursement not found.", "error")
            return redirect(url_for("funding_home"))

        conn.commit()
        cur.close()
        conn.close()

        flash("Owner reimbursement updated successfully.", "success")
        return redirect(url_for("funding_home"))

    cur.execute("""
        SELECT
            owner_reimbursement_id,
            reimbursement_date,
            amount,
            payment_method,
            notes
        FROM owner_reimbursements
        WHERE owner_reimbursement_id = %s
          AND spa_id = %s
    """, (owner_reimbursement_id, spa_id))

    reimbursement = cur.fetchone()

    cur.close()
    conn.close()

    if not reimbursement:
        flash("Owner reimbursement not found.", "error")
        return redirect(url_for("funding_home"))

    return render_template(
        "edit_owner_reimbursement.html",
        reimbursement=reimbursement
    )







#   ------------------------------------
#   
#       DELETE OWNER REIMBURSEMENT        
#  
#   spa id and route good
#   --------------------------------


@app.route("/owner_reimbursements/delete/<int:owner_reimbursement_id>", methods=["POST"])
@login_required
@spa_required
def delete_owner_reimbursement(owner_reimbursement_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM owner_reimbursements
        WHERE owner_reimbursement_id = %s
          AND spa_id = %s
    """, (owner_reimbursement_id, spa_id))

    if cur.rowcount == 0:
        conn.rollback()
        cur.close()
        conn.close()
        flash("Owner reimbursement not found.", "error")
        return redirect(url_for("funding_home"))

    conn.commit()
    cur.close()
    conn.close()

    flash("Owner reimbursement deleted successfully.", "success")
    return redirect(url_for("funding_home"))






#   ------------------------------------
#   
#      EDIT LOAN    
#   
#    spa id and route ok 
#   --------------------------------

@app.route("/business_loans/edit/<int:loan_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_business_loan(loan_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        loan_name = request.form.get("loan_name", "").strip()
        lender_name = request.form.get("lender_name", "").strip()
        loan_start_date = request.form.get("loan_start_date") or None
        original_amount = request.form.get("original_amount")
        interest_rate = request.form.get("interest_rate") or None
        term_months = request.form.get("term_months") or None
        notes = request.form.get("notes", "").strip()
        is_active = True if request.form.get("is_active") == "yes" else False

        cur.execute("""
            UPDATE business_loans
            SET loan_name = %s,
                lender_name = %s,
                loan_start_date = %s,
                original_amount = %s,
                interest_rate = %s,
                term_months = %s,
                notes = %s,
                is_active = %s
            WHERE loan_id = %s
              AND spa_id = %s
        """, (
            loan_name,
            lender_name,
            loan_start_date,
            original_amount,
            interest_rate,
            term_months,
            notes,
            is_active,
            loan_id,
            spa_id
        ))

        if cur.rowcount == 0:
            conn.rollback()
            cur.close()
            conn.close()
            flash("Business loan not found.", "error")
            return redirect(url_for("loans_home"))

        conn.commit()
        cur.close()
        conn.close()

        flash("Business loan updated successfully.", "success")
        return redirect(url_for("loans_home"))

    cur.execute("""
        SELECT
            loan_id,
            loan_name,
            lender_name,
            loan_start_date,
            original_amount,
            interest_rate,
            term_months,
            notes,
            is_active
        FROM business_loans
        WHERE loan_id = %s
          AND spa_id = %s
    """, (loan_id, spa_id))

    loan = cur.fetchone()

    cur.close()
    conn.close()

    if not loan:
        flash("Business loan not found.", "error")
        return redirect(url_for("loans_home"))

    return render_template(
        "edit_business_loan.html",
        loan=loan
    )






#   ------------------------------------
#   
#   DELETE BUSINESS LOAN
#   
#   spa id and route  ok
#   ------------------------------------


@app.route("/business_loans/delete/<int:loan_id>", methods=["POST"])
@login_required
@spa_required
def delete_business_loan(loan_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM business_loans
        WHERE loan_id = %s
          AND spa_id = %s
    """, (loan_id, spa_id))

    if cur.rowcount == 0:
        conn.rollback()
        cur.close()
        conn.close()
        flash("Business loan not found.", "error")
        return redirect(url_for("loans_home"))

    conn.commit()
    cur.close()
    conn.close()

    flash("Business loan and all related payments deleted successfully.", "success")
    return redirect(url_for("loans_home"))












#   ------------------------------------
#
#
#   CLIENT MANAGEMENT
#  spa id and route good  4/22/26
#   --------------------------------


@app.route("/client_management")
@login_required
@spa_required
def client_management():
    spa_id = current_spa_id()
    search = request.args.get("search", "").strip()
    show_all = request.args.get("show_all")
    today = get_spa_today()

    conn = get_db_connection()
    cur = conn.cursor()

    rows = []
    record_count = 0

    if search:
        cur.execute("""
            SELECT
                c.client_id,
                c.first_name,
                c.last_name,
                c.phone,
                c.email,
                c.birth_date,

                (
                    SELECT MAX(a1.appointment_date)
                    FROM appointments a1
                    WHERE a1.client_id = c.client_id
                      AND a1.spa_id = c.spa_id
                      AND a1.appointment_date <= %s
                ) AS last_visit_date,

                (
                    SELECT MIN(a2.appointment_date)
                    FROM appointments a2
                    WHERE a2.client_id = c.client_id
                      AND a2.spa_id = c.spa_id
                      AND a2.appointment_date >= %s
                ) AS next_visit_date,

                c.ok_to_text,
                c.sms_opt_in,
                c.sms_opt_out,
                c.ok_to_email,
                c.email_opt_in,
                c.email_opt_out

            FROM clients c
            WHERE c.spa_id = %s
              AND c.active_client = TRUE
              AND (
                   LOWER(c.first_name) LIKE %s
                   OR LOWER(c.last_name) LIKE %s
                   OR c.phone LIKE %s
              )
            ORDER BY c.last_name, c.first_name
        """, (
            today,
            today,
            spa_id,
            f"%{search.lower()}%",
            f"%{search.lower()}%",
            f"%{search}%"
        ))

        rows = cur.fetchall()
        record_count = len(rows)

    elif show_all:
        cur.execute("""
            SELECT
                c.client_id,
                c.first_name,
                c.last_name,
                c.phone,
                c.email,
                c.birth_date,

                (
                    SELECT MAX(a1.appointment_date)
                    FROM appointments a1
                    WHERE a1.client_id = c.client_id
                      AND a1.spa_id = c.spa_id
                      AND a1.appointment_date <= %s
                ) AS last_visit_date,

                (
                    SELECT MIN(a2.appointment_date)
                    FROM appointments a2
                    WHERE a2.client_id = c.client_id
                      AND a2.spa_id = c.spa_id
                      AND a2.appointment_date >= %s
                ) AS next_visit_date,

                c.ok_to_text,
                c.sms_opt_in,
                c.sms_opt_out,
                c.ok_to_email,
                c.email_opt_in,
                c.email_opt_out

            FROM clients c
            WHERE c.spa_id = %s
            AND c.active_client = TRUE
            ORDER BY c.last_name, c.first_name
        """, (
            today,
            today,
            spa_id
        ))

        rows = cur.fetchall()
        record_count = len(rows)

    cur.close()
    conn.close()

    return render_template(
        "client_management.html",
        clients=rows,
        record_count=record_count,
        search=search,
        show_all=show_all
    )












#  --------------------------------------
#
#  SCHEDULE APPOINTMENT START
#
#  spa id and route good
#  -----------------------------------


#  --------------------------------------
#
#  SCHEDULE APPOINTMENT START
#
#  spa id and route good
#  -----------------------------------



@app.route("/schedule_appointment_start", methods=["GET", "POST"])
@login_required
@spa_required
def schedule_appointment_start():
    spa_id = current_spa_id()

    clients = []
    searched = False

    selected_date = request.args.get("selected_date") \
        or request.form.get("selected_date") \
        or ""

    last_name = ""
    birth_date = ""

    if request.method == "POST":
        searched = True
        last_name = request.form.get("last_name", "").strip()
        birth_date = request.form.get("birth_date", "").strip()

        conn = get_db_connection()
        cur = conn.cursor()

        query = """
            SELECT
                client_id,
                first_name,
                last_name,
                birth_date,
                phone
            FROM clients
            WHERE spa_id = %s
              AND active_client = TRUE
        """
        params = [spa_id]

        if last_name:
            query += " AND last_name ILIKE %s"
            params.append(f"%{last_name}%")

        if birth_date:
            query += " AND birth_date = %s"
            params.append(birth_date)

        query += " ORDER BY last_name, first_name"

        if last_name or birth_date:
            cur.execute(query, tuple(params))
            clients = cur.fetchall()

        cur.close()
        conn.close()

    return render_template(
        "schedule_appointment_start.html",
        clients=clients,
        searched=searched,
        selected_date=selected_date,
        last_name=last_name,
        birth_date=birth_date
    )







#  ------------------------------------------
#      INCOME HOME PAGE
#
# ROUTE: INCOME
# URL: /income
# SECTION: INCOME HOME PAGE
#
#    spa_id safe
#
#  ------------------------------------------

@app.route("/income")
@login_required
@spa_required
def income_home():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    # summary totals
    cur.execute("""
        SELECT COALESCE(SUM(total_amount), 0)
        FROM income
        WHERE spa_id = %s
    """, (spa_id,))
    total_income = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM expenses
        WHERE spa_id = %s
    """, (spa_id,))
    total_expenses = cur.fetchone()[0] or 0

    net_total = total_income - total_expenses

    # income list
    cur.execute("""
        SELECT
            i.income_id,
            i.income_date,
            i.client_id,
            c.first_name,
            c.last_name,
            i.total_amount,
            i.payment_method
        FROM income i
        LEFT JOIN clients c
            ON i.client_id = c.client_id
           AND i.spa_id = c.spa_id
        WHERE i.spa_id = %s
        ORDER BY i.income_date DESC, i.income_id DESC
    """, (spa_id,))
    income_records = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "income_home.html",
        income_records=income_records,
        total_income=total_income,
        total_expenses=total_expenses,
        net_total=net_total
    )







#  ------------------------------------------
#           
#           CLIENT FORMS 
#
#    spa_id good
#  ------------------------------------------


@app.route("/client_forms/<int:client_id>", methods=["GET", "POST"])
@login_required
@spa_required
def client_forms(client_id):
    spa_id = current_spa_id()
    appointment_id = request.args.get("appointment_id") or request.form.get("appointment_id")
    selected_date = request.args.get("date") or request.form.get("date")

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        form_type_id = request.form.get("form_type_id")
        date_given = request.form.get("date_given") or None
        date_signed = request.form.get("date_signed") or None
        form_given = "form_given" in request.form
        form_signed = "form_signed" in request.form
        notes = request.form.get("notes", "").strip()

        # Validate client belongs to this spa
        cur.execute("""
            SELECT 1
            FROM clients
            WHERE client_id = %s
              AND spa_id = %s
        """, (client_id, spa_id))
        valid_client = cur.fetchone()

        if not valid_client:
            cur.close()
            conn.close()
            flash("Client not found.", "error")
            return redirect(url_for("client_history"))

        # Validate form type belongs to this spa
        cur.execute("""
            SELECT 1
            FROM form_types
            WHERE form_type_id = %s
              AND spa_id = %s
              AND is_active = TRUE
        """, (form_type_id, spa_id))
        valid_form_type = cur.fetchone()

        if not valid_form_type:
            cur.close()
            conn.close()
            flash("Invalid form type selected.", "error")
            return redirect(url_for(
                "client_forms",
                client_id=client_id,
                appointment_id=appointment_id,
                date=selected_date
            ))

        # Optional: validate appointment belongs to this spa/client if present
        if appointment_id:
            cur.execute("""
                SELECT 1
                FROM appointments
                WHERE appointment_id = %s
                  AND client_id = %s
                  AND spa_id = %s
            """, (appointment_id, client_id, spa_id))
            valid_appointment = cur.fetchone()

            if not valid_appointment:
                cur.close()
                conn.close()
                flash("Invalid appointment selected.", "error")
                return redirect(url_for(
                    "client_forms",
                    client_id=client_id,
                    appointment_id=appointment_id,
                    date=selected_date
                ))

        cur.execute("""
            INSERT INTO client_forms_log (
                spa_id,
                client_id,
                appointment_id,
                form_type_id,
                date_given,
                date_signed,
                form_given,
                form_signed,
                notes
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            spa_id,
            client_id,
            appointment_id,
            form_type_id,
            date_given,
            date_signed,
            form_given,
            form_signed,
            notes
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Client form log saved successfully.", "success")

        return redirect(
            url_for(
                "client_forms",
                client_id=client_id,
                appointment_id=appointment_id,
                date=selected_date
            )
        )

    form_types = get_dropdown_options(
        "client_form_names",
        spa_id
    )


    cur.execute("""
        SELECT client_id, first_name, last_name
        FROM clients
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, spa_id))
    client = cur.fetchone()

    if not client:
        cur.close()
        conn.close()
        flash("Client not found.", "error")
        return redirect(url_for("client_history"))

    cur.execute("""
        SELECT
            cfl.client_form_log_id,
            ft.form_name,
            cfl.date_given,
            cfl.date_signed,
            cfl.form_given,
            cfl.form_signed,
            cfl.notes
        FROM client_forms_log cfl
        JOIN form_types ft
            ON cfl.form_type_id = ft.form_type_id
           AND cfl.spa_id = ft.spa_id
        WHERE cfl.client_id = %s
          AND cfl.spa_id = %s
        ORDER BY cfl.created_at DESC
    """, (client_id, spa_id))
    form_history = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "client_forms.html",
        client=client,
        form_types=form_types,
        form_history=form_history,
        appointment_id=appointment_id,
        selected_date=selected_date
    )







#  ------------------------------------------
#      GIFT CERTIFICATES
#
#  spa id and route good
#  ------------------------------------------


@app.route("/gift_certificates")
@login_required
@spa_required
def gift_certificates_home():
    spa_id = current_spa_id()
    spa_now = get_spa_now()
    today = spa_now.date()

    certificate_search = request.args.get("certificate_search", "").strip()
    sort_by = request.args.get("sort_by", "date_desc")
    filter_by = request.args.get("filter", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    where_clauses = ["gc.spa_id = %s"]
    params = [spa_id]

    if certificate_search:
        where_clauses.append("gc.certificate_number ILIKE %s")
        params.append(f"%{certificate_search}%")

    if filter_by == "expiring_soon":
        where_clauses.append("gc.expires_date IS NOT NULL")
        where_clauses.append("gc.expires_date >= CURRENT_DATE")
        where_clauses.append("gc.expires_date <= CURRENT_DATE + INTERVAL '60 days'")
        where_clauses.append("gcs.status_name IN ('Active', 'Printed')")
        where_clauses.append("gc.remaining_balance > 0")

    where_sql = "WHERE " + " AND ".join(where_clauses)

    if sort_by == "date_asc":
        order_sql = "ORDER BY gc.date_issued ASC, gc.gift_cert_id ASC"
    elif sort_by == "cert_asc":
        order_sql = "ORDER BY gc.certificate_number ASC"
    elif sort_by == "cert_desc":
        order_sql = "ORDER BY gc.certificate_number DESC"
    else:
        order_sql = "ORDER BY gc.date_issued DESC, gc.gift_cert_id DESC"

    query = f"""
        SELECT
            gc.gift_cert_id,
            gc.certificate_number,
            gc.date_issued,
            gc.expires_date,
            gc.original_value,
            gc.amount_paid,
            gc.remaining_balance,
            gc.purchased_by_first_name,
            gc.purchased_by_last_name,
            gc.purchaser_phone,
            gc.purchaser_email,
            gc.recipient_name,
            gcs.status_name,
            gc.notes,
            COALESCE(r.sent_status, 'Not Sent') AS email_sent_status,
            r.sent_date
        FROM gift_certificates gc
        LEFT JOIN gift_certificate_statuses gcs
            ON gc.gift_certificate_status_id = gcs.gift_certificate_status_id
           AND gc.spa_id = gcs.spa_id 
        LEFT JOIN (
            SELECT DISTINCT ON (gift_cert_id, spa_id)
                gift_cert_id,
                spa_id,
                sent_status,
                sent_date
            FROM gift_certificate_email_reminders
            ORDER BY gift_cert_id, spa_id, sent_date DESC
        ) r
            ON gc.gift_cert_id = r.gift_cert_id
           AND gc.spa_id = r.spa_id
        {where_sql}
        {order_sql}
    """

    spa_now = get_spa_now()
    today = spa_now.date()

    expiring_gc_count = 0

    # Count for alert/banner/card
    cur.execute("""
        SELECT COUNT(*)
        FROM gift_certificates gc
        JOIN gift_certificate_statuses gcs
          ON gc.gift_certificate_status_id = gcs.gift_certificate_status_id
         AND gc.spa_id = gcs.spa_id          
        WHERE gc.spa_id = %s
          AND gcs.status_name = 'Active'
          AND gc.amount_paid > 0
          AND gc.is_redeemed = FALSE
          AND gc.remaining_balance > 0
          AND gc.expires_date BETWEEN %s AND (%s + INTERVAL '60 days')
    """, (spa_id, today, today))
    expiring_gc_count = cur.fetchone()[0] or 0

    cur.execute("""
        SELECT
            COUNT(*) FILTER (
                WHERE gcs.status_name = 'Active'
                  AND gc.amount_paid > 0
            ) AS issued_count,

            COUNT(*) FILTER (
                WHERE gcs.status_name = 'Printed'
            ) AS on_hand_count,

            COALESCE(SUM(gc.remaining_balance) FILTER (
                WHERE gcs.status_name = 'Active'
                  AND gc.amount_paid > 0
                  AND gc.is_redeemed = FALSE
                  AND gc.remaining_balance > 0
            ), 0) AS issued_not_redeemed_value

        FROM gift_certificates gc
        JOIN gift_certificate_statuses gcs
          ON gc.gift_certificate_status_id = gcs.gift_certificate_status_id
         AND gc.spa_id = gcs.spa_id
        WHERE gc.spa_id = %s
    """, (spa_id,))

    gift_certificate_summary = cur.fetchone()

    cur.execute(query, params)
    gift_certificates = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "gift_certificates_home.html",
        gift_certificates=gift_certificates,
        certificate_search=certificate_search,
        sort_by=sort_by,
        filter_by=filter_by,
        expiring_gc_count=expiring_gc_count,
        gift_certificate_summary=gift_certificate_summary
    )







#  ------------------------------------------
#      ADD GIFT CERTIFICATE
#
#
#    spa_id good
#  ------------------------------------------


@app.route("/add_gift_certificate", methods=["GET", "POST"])
@login_required
@spa_required
def add_gift_certificate():
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        certificate_number = (request.form.get("certificate_number") or "").strip()
        date_issued = request.form.get("date_issued") or None
        expires_date = request.form.get("expires_date") or None
        original_value = request.form.get("original_value") or None
        amount_paid = request.form.get("amount_paid") or None
        remaining_balance = request.form.get("remaining_balance") or None
        purchased_by_first_name = request.form.get("purchased_by_first_name") or None
        purchased_by_last_name = request.form.get("purchased_by_last_name") or None
        purchaser_phone = request.form.get("purchaser_phone") or None
        purchaser_email = request.form.get("purchaser_email") or None
        recipient_name = request.form.get("recipient_name") or None
        notes = request.form.get("notes") or None

        if not certificate_number:
            cur.close()
            conn.close()
            flash("Certificate number is required.", "danger")
            return redirect(url_for("add_gift_certificate"))

        active_status_id = get_status_id("Active")
        if not active_status_id:
            cur.close()
            conn.close()
            flash("Active gift certificate status not found.", "danger")
            return redirect(url_for("gift_certificates_home"))

        try:
            cur.execute("""
                INSERT INTO gift_certificates (
                    spa_id,
                    certificate_number,
                    date_issued,
                    expires_date,
                    original_value,
                    amount_paid,
                    remaining_balance,
                    purchased_by_first_name,
                    purchased_by_last_name,
                    purchaser_phone,
                    purchaser_email,
                    recipient_name,
                    notes,
                    gift_certificate_status_id
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                spa_id,
                certificate_number,
                date_issued,
                expires_date,
                original_value,
                amount_paid,
                remaining_balance,
                purchased_by_first_name,
                purchased_by_last_name,
                purchaser_phone,
                purchaser_email,
                recipient_name,
                notes,
                active_status_id
            ))

            conn.commit()
            flash("Gift certificate added to inventory.", "success")

        except Exception as e:
            conn.rollback()
            flash(f"Error adding gift certificate: {e}", "danger")

        finally:
            cur.close()
            conn.close()

        return redirect(url_for("gift_certificates_home"))

    cur.close()
    conn.close()
    return render_template("add_gift_certificate.html")








#  ------------------------------------------
#         EDIT GIFT CERTIFICATE
#
#
#
#      spa_id good
#  ------------------------------------------



@app.route("/edit_gift_certificate/<int:certificate_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_gift_certificate(certificate_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        certificate_number = request.form.get("certificate_number") or None
        date_issued = request.form.get("date_issued") or None
        expires_date = request.form.get("expires_date") or None
        original_value = request.form.get("original_value") or None
        amount_paid = request.form.get("amount_paid") or None
        remaining_balance = request.form.get("remaining_balance") or None
        purchased_by_first_name = request.form.get("purchased_by_first_name") or None
        purchased_by_last_name = request.form.get("purchased_by_last_name") or None
        purchaser_phone = request.form.get("purchaser_phone") or None
        purchaser_email = request.form.get("purchaser_email") or None
        recipient_name = request.form.get("recipient_name") or None
        gift_certificate_status_id = request.form.get("gift_certificate_status_id") or None
        notes = request.form.get("notes") or None

        if gift_certificate_status_id:
            cur.execute("""
                SELECT 1
                FROM gift_certificate_statuses
                WHERE gift_certificate_status_id = %s
                  AND spa_id = %s
            """, (gift_certificate_status_id, spa_id))
            valid_status = cur.fetchone()

            if not valid_status:
                cur.close()
                conn.close()
                flash("Invalid gift certificate status selected.", "danger")
                return redirect(url_for("edit_gift_certificate", certificate_id=certificate_id))

        cur.execute("""
            UPDATE gift_certificates
            SET certificate_number = %s,
                date_issued = %s,
                expires_date = %s,
                original_value = %s,
                amount_paid = %s,
                remaining_balance = %s,
                purchased_by_first_name = %s,
                purchased_by_last_name = %s,
                purchaser_phone = %s,
                purchaser_email = %s,
                recipient_name = %s,
                gift_certificate_status_id = %s,
                notes = %s
            WHERE gift_cert_id = %s
              AND spa_id = %s
        """, (
            certificate_number,
            date_issued,
            expires_date,
            original_value,
            amount_paid,
            remaining_balance,
            purchased_by_first_name,
            purchased_by_last_name,
            purchaser_phone,
            purchaser_email,
            recipient_name,
            gift_certificate_status_id,
            notes,
            certificate_id,
            spa_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Gift certificate updated successfully.", "success")
        return redirect(url_for("gift_certificates_home"))

    cur.execute("""
        SELECT
            gift_cert_id,
            certificate_number,
            date_issued,
            expires_date,
            original_value,
            amount_paid,
            remaining_balance,
            purchased_by_first_name,
            purchased_by_last_name,
            purchaser_phone,
            purchaser_email,
            recipient_name,
            gift_certificate_status_id,
            notes
        FROM gift_certificates
        WHERE gift_cert_id = %s
          AND spa_id = %s
    """, (certificate_id, spa_id))
    gift_certificate = cur.fetchone()

    if not gift_certificate:
        cur.close()
        conn.close()
        flash("Gift certificate not found.", "danger")
        return redirect(url_for("gift_certificates_home"))

    cur.execute("""
        SELECT gift_certificate_status_id, status_name
        FROM gift_certificate_statuses
        WHERE spa_id = %s
        ORDER BY gift_certificate_status_id
    """, (spa_id,))
    statuses = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "edit_gift_certificate.html",
        gift_certificate=gift_certificate,
        statuses=statuses
    )









#  ------------------------------------------
#       REDEEM GIFT CERTIFICATE
#
#
#    spa_id good
#  ------------------------------------------



@app.route("/redeem_gift_certificate/<int:certificate_id>", methods=["GET", "POST"])
@login_required
@spa_required
def redeem_gift_certificate(certificate_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        redeemed_on = request.form.get("redeemed_on") or None
        redeemed_by = request.form.get("redeemed_by") or None

        redeemed_status_id = get_status_id("Redeemed")
        if not redeemed_status_id:
            cur.close()
            conn.close()
            flash("Redeemed status not found.", "danger")
            return redirect(url_for("gift_certificates_home"))

        cur.execute("""
            UPDATE gift_certificates
            SET redeemed_on = %s,
                redeemed_by = %s,
                remaining_balance = 0,
                gift_certificate_status_id = %s
            WHERE gift_cert_id = %s
              AND spa_id = %s
        """, (
            redeemed_on,
            redeemed_by,
            redeemed_status_id,
            certificate_id,
            spa_id
        ))

        if cur.rowcount == 0:
            conn.rollback()
            cur.close()
            conn.close()
            flash("Gift certificate not found.", "danger")
            return redirect(url_for("gift_certificates_home"))

        conn.commit()
        flash("Gift certificate redeemed successfully.", "success")
        cur.close()
        conn.close()

        return redirect(url_for("gift_certificates_home"))

    cur.execute("""
        SELECT
            gc.gift_cert_id,
            gc.certificate_number,
            gc.original_value,
            gc.remaining_balance,
            gc.purchased_by_first_name,
            gc.purchased_by_last_name,
            gc.recipient_name,
            gcs.status_name
        FROM gift_certificates gc
        LEFT JOIN gift_certificate_statuses gcs
            ON gc.gift_certificate_status_id = gcs.gift_certificate_status_id
           AND gc.spa_id = gcs.spa_id
        WHERE gc.gift_cert_id = %s
          AND gc.spa_id = %s
    """, (certificate_id, spa_id))
    gift_certificate = cur.fetchone()

    cur.close()
    conn.close()

    if not gift_certificate:
        flash("Gift certificate not found.", "danger")
        return redirect(url_for("gift_certificates_home"))

    return render_template(
        "redeem_gift_certificate.html",
        gift_certificate=gift_certificate
    )









#   ------------------------------------------
#
#   GIFT CERTIFICATES REMINDER
#
#     spa_id good
#   ------------------------------------------


@app.route("/gift_certificate_reminders")
@login_required
@spa_required
def gift_certificate_reminders():
    spa_id = current_spa_id()
    active_status_id = get_status_id("Active")

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            gc.gift_cert_id,
            gc.purchased_by_first_name,
            gc.purchased_by_last_name,
            gc.purchaser_email,
            gc.recipient_name,
            gc.date_issued,
            gc.expires_date,
            gc.remaining_balance,
            CASE
                WHEN gc.expires_date = CURRENT_DATE THEN 'same_day'
                WHEN gc.expires_date <= CURRENT_DATE + INTERVAL '7 days' THEN '7_day'
                ELSE '30_day'
            END AS reminder_type
        FROM gift_certificates gc
        WHERE gc.spa_id = %s
          AND gc.gift_certificate_status_id = %s
          AND gc.remaining_balance > 0
          AND gc.purchaser_email IS NOT NULL
          AND gc.purchaser_email <> ''
          AND gc.expires_date BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '30 days'
        ORDER BY gc.expires_date ASC, gc.gift_cert_id ASC
    """, (spa_id, active_status_id))

    rows = cur.fetchall()

    eligible_rows = []

    for row in rows:
        gift_cert_id = row[0]
        reminder_type = row[8]

        cur.execute("""
            SELECT 1
            FROM gift_certificate_email_reminders
            WHERE spa_id = %s
              AND gift_cert_id = %s
              AND reminder_type = %s
            LIMIT 1
        """, (spa_id, gift_cert_id, reminder_type))

        already_sent = cur.fetchone()

        if not already_sent:
            expires_date = row[6]
            days_left = (expires_date - date.today()).days if expires_date else None

            eligible_rows.append({
                "gift_cert_id": row[0],
                "purchased_by_first_name": row[1],
                "purchased_by_last_name": row[2],
                "purchaser_email": row[3],
                "recipient_name": row[4],
                "date_issued": row[5],
                "expires_date": row[6],
                "remaining_balance": row[7],
                "reminder_type": row[8],
                "days_left": days_left
            })

    cur.close()
    conn.close()

    return render_template(
        "gift_certificate_reminders.html",
        reminders=eligible_rows
    )




#   ----------------------------
#
#  
#   ---------------------------








#   --------------------------------------------
#
#     GIFT CERTIFICATE REMINDER HISTORY
#
#     spa id and route good
#
#   ---------------------------------------------



@app.route("/gift_certificate_reminder_history")
@login_required
@spa_required
def gift_certificate_reminder_history():
    spa_id = current_spa_id()
    print("DEBUG current spa_id:", spa_id)

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            r.gc_email_reminder_id,
            r.gift_cert_id,
            r.spa_id,
            gc.certificate_number,
            gc.purchased_by_first_name,
            gc.purchased_by_last_name,
            r.recipient_email,
            gc.recipient_name,
            gc.expires_date,
            r.reminder_type,
            r.sent_status,
            r.sent_date,
            r.notes
        FROM gift_certificate_email_reminders r
        LEFT JOIN gift_certificates gc
            ON r.gift_cert_id = gc.gift_cert_id
           AND r.spa_id = gc.spa_id
        WHERE r.spa_id = %s
        ORDER BY r.sent_date DESC NULLS LAST, r.gc_email_reminder_id DESC
    """, (spa_id,))

    reminders = cur.fetchall()
    print("DEBUG reminders found:", reminders)

    cur.close()
    conn.close()

    return render_template(
        "sent_reminder_history.html",
        reminders=reminders
    )








                
#  ------------------------------------------
#      CLIENTS   HOME PAGE
#
#
#
#  spa id and route good
#
#  ------------------------------------------




from datetime import date

@app.route("/clients")
@login_required
@spa_required
def clients():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    today = date.today()
    current_month = today.month
    current_day = today.day
    current_year = today.year

    months_to_show = [current_month]

    if current_day >= 15:
        next_month = 1 if current_month == 12 else current_month + 1
        months_to_show.append(next_month)

    cur.execute("""
        SELECT
            c.client_id,
            c.first_name,
            c.last_name,
            c.birth_date,
            EXTRACT(MONTH FROM c.birth_date) AS birth_month,
            cbo.birthday_offer_id,
            cbo.birthday_year,
            COALESCE(cbo.offer_sent, FALSE) AS offer_sent,
            cbo.offer_sent_date,
            cbo.acknowledged_by,
            cbo.notes
        FROM clients c
        LEFT JOIN client_birthday_offers cbo
            ON c.client_id = cbo.client_id
           AND c.spa_id = cbo.spa_id
           AND cbo.birthday_year = %s
        WHERE c.spa_id = %s
          AND c.birth_date IS NOT NULL
          AND EXTRACT(MONTH FROM c.birth_date) = ANY(%s)
          AND COALESCE(cbo.offer_sent, FALSE) = FALSE
        ORDER BY
            EXTRACT(MONTH FROM c.birth_date),
            EXTRACT(DAY FROM c.birth_date),
            c.last_name,
            c.first_name
    """, (current_year, spa_id, months_to_show))

    birthday_clients = cur.fetchall()

    print("birthday_clients =", birthday_clients)

    cur.execute("""
        SELECT 
            client_id, 
            first_name, 
            last_name, 
            phone, 
            email, 
            birth_date,
            ok_to_text,
            sms_opt_in,
            sms_opt_out,
            ok_to_email,
            email_opt_in,
            email_opt_out
        FROM clients
        WHERE spa_id = %s
        ORDER BY last_name, first_name
    """, (spa_id,))
    clients = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "clients_home.html",
        clients=clients,
        birthday_clients=birthday_clients
    )











#  ------------------------------------------
#          
#        FULL EDIT CLIENT   
#  
#     spa id and full route good
#  ------------------------------------------



@app.route("/edit-client-full/<int:client_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_client_full(client_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # Always verify the client belongs to the current spa first
        cur.execute("""
            SELECT
                client_id,
                first_name,
                last_name,
                phone,
                email,
                birth_date,
                address,
                city,
                state,
                zip,
                emergency_contact_name,
                emergency_contact_phone,
                referred_by,
                notes_one,
                notes_two,
                notes_three,
                active_client,
                created_at,
                updated_at
            FROM clients
            WHERE client_id = %s
              AND spa_id = %s
        """, (client_id, spa_id))
        client = cur.fetchone()

        if not client:
            flash("Client not found.", "error")
            return redirect(url_for("clients_home"))

        if request.method == "POST":
            # ----- clients table fields -----
            first_name = request.form.get("first_name", "").strip()
            last_name = request.form.get("last_name", "").strip()
            phone = request.form.get("phone", "").strip()
            email = request.form.get("email", "").strip()
            birth_date = request.form.get("birth_date") or None
            address = request.form.get("address", "").strip()
            city = request.form.get("city", "").strip()
            state = request.form.get("state", "").strip()
            zip_code = request.form.get("zip", "").strip()
            emergency_contact_name = request.form.get("emergency_contact_name", "").strip()
            emergency_contact_phone = request.form.get("emergency_contact_phone", "").strip()
            referred_by = request.form.get("referred_by", "").strip()
            notes_one = request.form.get("notes_one", "").strip()
            notes_two = request.form.get("notes_two", "").strip()
            notes_three = request.form.get("notes_three", "").strip()
            active_client = True if request.form.get("active_client") == "on" else False

            # ----- client_health_profile fields -----
            sex = request.form.get("sex", "").strip()
            skin_type_id = request.form.get("skin_type_id") or None
            fitzpatrick_id = request.form.get("fitzpatrick_id") or None
            skin_concerns = request.form.get("skin_concerns", "").strip()
            skin_conditions = request.form.get("skin_conditions", "").strip()
            allergies = request.form.get("allergies", "").strip()
            medications = request.form.get("medications", "").strip()
            current_medical_conditions = request.form.get("current_medical_conditions", "").strip()
            past_medical_treatments = request.form.get("past_medical_treatments", "").strip()

            recent_injections = parse_bool(request.form.get("recent_injections"))
            recent_laser = parse_bool(request.form.get("recent_laser"))
            pregnant = parse_bool(request.form.get("pregnant"))
            nursing = parse_bool(request.form.get("nursing"))
            using_retinol = parse_bool(request.form.get("using_retinol"))
            using_accutane = parse_bool(request.form.get("using_accutane"))

            sun_exposure_level = request.form.get("sun_exposure_level", "").strip()
            last_facial_date = request.form.get("last_facial_date") or None
            health_notes1 = request.form.get("health_notes1", "").strip()
            health_notes2 = request.form.get("health_notes2", "").strip()
            health_notes3 = request.form.get("health_notes3", "").strip()

            # ----- update clients -----
            cur.execute("""
                UPDATE clients
                SET
                    first_name = %s,
                    last_name = %s,
                    phone = %s,
                    email = %s,
                    birth_date = %s,
                    address = %s,
                    city = %s,
                    state = %s,
                    zip = %s,
                    emergency_contact_name = %s,
                    emergency_contact_phone = %s,
                    referred_by = %s,
                    notes_one = %s,
                    notes_two = %s,
                    notes_three = %s,
                    active_client = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE client_id = %s
                  AND spa_id = %s
            """, (
                first_name,
                last_name,
                phone,
                email,
                birth_date,
                address,
                city,
                state,
                zip_code,
                emergency_contact_name,
                emergency_contact_phone,
                referred_by,
                notes_one,
                notes_two,
                notes_three,
                active_client,
                client_id,
                spa_id
            ))

            if cur.rowcount == 0:
                conn.rollback()
                flash("Client not found.", "error")
                return redirect(url_for("clients_home"))

            # ----- check whether health profile exists -----
            cur.execute("""
                SELECT health_profile_id
                FROM client_health_profile
                WHERE client_id = %s
                  AND spa_id = %s
            """, (client_id, spa_id))
            existing_health = cur.fetchone()

            if existing_health:
                # update existing health profile
                cur.execute("""
                    UPDATE client_health_profile
                    SET
                        sex = %s,
                        skin_type_id = %s,
                        fitzpatrick_id = %s,
                        skin_concerns = %s,
                        skin_conditions = %s,
                        allergies = %s,
                        medications = %s,
                        current_medical_conditions = %s,
                        past_medical_treatments = %s,
                        recent_injections = %s,
                        recent_laser = %s,
                        pregnant = %s,
                        nursing = %s,
                        using_retinol = %s,
                        using_accutane = %s,
                        sun_exposure_level = %s,
                        last_facial_date = %s,
                        notes1 = %s,
                        notes2 = %s,
                        notes3 = %s,
                        last_updated = CURRENT_DATE
                    WHERE client_id = %s
                      AND spa_id = %s
                """, (
                    sex,
                    skin_type_id,
                    fitzpatrick_id,
                    skin_concerns,
                    skin_conditions,
                    allergies,
                    medications,
                    current_medical_conditions,
                    past_medical_treatments,
                    recent_injections,
                    recent_laser,
                    pregnant,
                    nursing,
                    using_retinol,
                    using_accutane,
                    sun_exposure_level,
                    last_facial_date,
                    health_notes1,
                    health_notes2,
                    health_notes3,
                    client_id,
                    spa_id
                ))
            else:
                # insert new health profile
                cur.execute("""
                    INSERT INTO client_health_profile (
                        spa_id,
                        client_id,
                        sex,
                        skin_type_id,
                        fitzpatrick_id,
                        skin_concerns,
                        skin_conditions,
                        allergies,
                        medications,
                        current_medical_conditions,
                        past_medical_treatments,
                        recent_injections,
                        recent_laser,
                        pregnant,
                        nursing,
                        using_retinol,
                        using_accutane,
                        sun_exposure_level,
                        last_facial_date,
                        notes1,
                        notes2,
                        notes3,
                        last_updated,
                        created_at
                    )
                    VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        CURRENT_DATE, CURRENT_TIMESTAMP
                    )
                """, (
                    spa_id,
                    client_id,
                    sex,
                    skin_type_id,
                    fitzpatrick_id,
                    skin_concerns,
                    skin_conditions,
                    allergies,
                    medications,
                    current_medical_conditions,
                    past_medical_treatments,
                    recent_injections,
                    recent_laser,
                    pregnant,
                    nursing,
                    using_retinol,
                    using_accutane,
                    sun_exposure_level,
                    last_facial_date,
                    health_notes1,
                    health_notes2,
                    health_notes3
                ))

            conn.commit()
            flash("Client full record updated successfully.", "success")
            return redirect(url_for("clients_home"))

        # ---------------- GET request ----------------

        cur.execute("""
            SELECT
                health_profile_id,
                client_id,
                sex,
                skin_type_id,
                fitzpatrick_id,
                skin_concerns,
                skin_conditions,
                allergies,
                medications,
                current_medical_conditions,
                past_medical_treatments,
                recent_injections,
                recent_laser,
                pregnant,
                nursing,
                using_retinol,
                using_accutane,
                sun_exposure_level,
                last_facial_date,
                notes1,
                notes2,
                notes3,
                last_updated,
                created_at
            FROM client_health_profile
            WHERE client_id = %s
              AND spa_id = %s
        """, (client_id, spa_id))
        health = cur.fetchone()

        # global lookup tables
        cur.execute("""
            SELECT sex_type_id, sex_type
            FROM sex
            ORDER BY sex_type
        """, (spa_id,))
        sex_options = cur.fetchall()

        cur.execute("""
            SELECT skin_type_id, skin_type_name
            FROM skin_types
            ORDER BY skin_type_name
        """, (spa_id,))
        skin_types = cur.fetchall()

        cur.execute("""
            SELECT fitzpatrick_id, fitzpatrick_level
            FROM fitzpatrick_types
            ORDER BY fitzpatrick_id
        """, (spa_id,))
        fitzpatrick_types = cur.fetchall()

        # spa-owned lookup table
        cur.execute("""
            SELECT referral_source_id, referral_source_name
            FROM referral_sources
            WHERE spa_id = %s
            ORDER BY referral_source_name
        """, (spa_id,))
        referral_sources = cur.fetchall()

        return render_template(
            "edit_client_full.html",
            client=client,
            health=health,
            sex_options=sex_options,
            skin_types=skin_types,
            fitzpatrick_types=fitzpatrick_types,
            referral_sources=referral_sources
        )

    except Exception as e:
        conn.rollback()
        flash(f"Error updating client record: {e}", "error")
        return redirect(url_for("clients_home"))

    finally:
        cur.close()
        conn.close()







#  ------------------------------------------
#           BIRTHDAYS            
#
#
#  
#  ------------------------------------------






#  ------------------------------------------
#           BIRTHDAY OFFERS
#
#
#     spa id and route good
#  ------------------------------------------


@app.route("/birthday_offers_home")
@login_required
@spa_required
def birthday_offers_home():
    spa_id = current_spa_id()
    spa_now = get_spa_now()
    today = spa_now.date()
    end_date = today + timedelta(days=45)

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            c.client_id,
            c.first_name,
            c.last_name,
            c.phone,
            c.email,
            c.birth_date,
            cbo.offer_sent,
            cbo.offer_sent_date,
            cbo.sent_status,
            cbo.notes
        FROM clients c
        LEFT JOIN client_birthday_offers cbo
            ON c.client_id = cbo.client_id
           AND c.spa_id = cbo.spa_id
           AND cbo.birthday_year = %s
        WHERE c.spa_id = %s
          AND c.birth_date IS NOT NULL
          AND c.active_client = TRUE
        ORDER BY c.last_name, c.first_name
    """, (today.year, spa_id))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    upcoming_birthdays = []

    for row in rows:
        (
            client_id,
            first_name,
            last_name,
            phone,
            email,
            birth_date,
            offer_sent,
            offer_sent_date,
            sent_status,
            notes
        ) = row

        this_year_birthday = birth_date.replace(year=today.year)

        if this_year_birthday < today:
            next_birth_date = birth_date.replace(year=today.year + 1)
        else:
            next_birth_date = this_year_birthday

        if today <= next_birth_date <= end_date:
            days_until = (next_birth_date - today).days

            upcoming_birthdays.append({
                "client_id": client_id,
                "first_name": first_name,
                "last_name": last_name,
                "phone": phone,
                "email": email,
                "birth_date": birth_date,
                "next_birth_date": next_birth_date,
                "days_until": days_until,
                "offer_sent": bool(offer_sent),
                "offer_sent_date": offer_sent_date,
                "sent_status": sent_status,
                "notes": notes,
                "acknowledged_by": notes
            })

    return render_template(
        "birthday_offers_home.html",
        upcoming_birthdays=upcoming_birthdays,
        today=today
    )











        
#  ------------------------------------------
#        BIRTHDAY OFFERS MARK SENT
#
#
#  >>>>>>>>> when all working
#  ------------------------------------------




#@app.route("/birthday-offers/mark-sent", methods=["POST"])
@login_required
@spa_required
def mark_birthday_offer_sent_disabled():
    spa_id = current_spa_id()
    client_id = request.form.get("client_id")

    if not client_id:
        flash("Missing client ID.", "error")
        return redirect(url_for("birthday_offers_home"))

    spa_now = get_spa_now()
    today = spa_now.date()
    current_year = today.year
    current_month = today.month

    if current_month == 12:
        next_month = 1
        next_month_year = current_year + 1
    else:
        next_month = current_month + 1
        next_month_year = current_year

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        # Verify client belongs to current spa
        cur.execute("""
            SELECT birth_date
            FROM clients
            WHERE client_id = %s
              AND spa_id = %s
        """, (client_id, spa_id))

        client = cur.fetchone()

        if not client or not client[0]:
            flash("Client birthday not found.", "error")
            return redirect(url_for("birthday_offers_home"))

        birth_date = client[0]
        campaign_year = get_birthday_campaign_year(birth_date, today)

        if campaign_year is None:
            flash("Client is not in the active birthday campaign window.", "error")
            return redirect(url_for("birthday_offers_home"))

        acknowledged_by = "System Admin"

        cur.execute("""
            INSERT INTO client_birthday_offers_home (
                spa_id,
                client_id,
                birthday_year,
                offer_sent,
                offer_sent_date,
                acknowledged_by,
                notes,
                created_at,
                email_template_id,
                sent_status
            )
            VALUES (%s, %s, %s, TRUE, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (spa_id, client_id, birthday_year)
            DO UPDATE SET
                offer_sent = TRUE,
                offer_sent_date = EXCLUDED.offer_sent_date,
                acknowledged_by = EXCLUDED.acknowledged_by,
                notes = EXCLUDED.notes,
                sent_status = EXCLUDED.sent_status
        """, (
            spa_id,
            client_id,
            campaign_year,
            spa_now,
            acknowledged_by,
            "Marked sent manually",
            spa_now,
            None,
            "manual"
        ))

        conn.commit()
        flash("Birthday offer marked as sent.", "success")
        return redirect(url_for("birthday_offers_home"))

    except Exception as e:
        conn.rollback()
        flash(f"Error marking birthday offer sent: {e}", "error")
        return redirect(url_for("birthday_offers_home"))

    finally:
        cur.close()
        conn.close()










#    --------------------------------------
#
#      BIRTHDAY OFFERS SEND
#
#   
#   ---------------------------------------


@app.route("/birthday-offers/send/<int:client_id>", methods=["POST"])
@login_required
@spa_required
def send_birthday_offer(client_id):

    spa_id = current_spa_id()
    spa_now = get_spa_now()
    today = spa_now.date()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT first_name, last_name, email, birth_date
        FROM clients
        WHERE spa_id = %s
          AND client_id = %s
          AND active_client = TRUE
    """, (spa_id, client_id))

    client = cur.fetchone()

    if not client:
        cur.close()
        conn.close()
        flash("Client not found.", "error")
        return redirect(url_for("birthday_offers"))

    first_name, last_name, email, birth_date = client

    if not birth_date:
        cur.close()
        conn.close()
        flash("Client birthday not found.", "error")
        return redirect(url_for("birthday_offers"))

    campaign_year = get_birthday_campaign_year(birth_date, today)
    if campaign_year is None:
        cur.close()
        conn.close()
        flash("Client is not in the active birthday campaign window.", "error")
        return redirect(url_for("birthday_offers"))

    if not email or not email.strip():
        cur.close()
        conn.close()
        flash("Client does not have an email address.", "error")
        return redirect(url_for("birthday_offers"))

    subject, body = build_birthday_email(first_name)

    email_sent_successfully = False
    error_note = None
    email_template_id = None

    try:
        # Replace later with your actual Mailgun function
        # email_sent_successfully = send_email_via_mailgun(email, subject, body)
        pass
    except Exception as e:
        error_note = str(e)

    if email_sent_successfully:
        cur.execute("""
            INSERT INTO client_birthday_offers (
                spa_id,
                client_id,
                birthday_year,
                offer_sent,
                offer_sent_date,
                acknowledged_by,
                notes,
                created_at,
                email_template_id,
                sent_status
            )
            VALUES (%s, %s, %s, TRUE, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (spa_id, client_id, birthday_year)
            DO UPDATE SET
                offer_sent = TRUE,
                offer_sent_date = EXCLUDED.offer_sent_date,
                acknowledged_by = EXCLUDED.acknowledged_by,
                sent_status = EXCLUDED.sent_status,
                email_template_id = EXCLUDED.email_template_id,
                notes = EXCLUDED.notes
        """, (
            spa_id,
            client_id,
            campaign_year,
            spa_now,
            "system",
            "Birthday email sent successfully",
            spa_now,
            email_template_id,
            "sent"
        ))

        conn.commit()
        flash(f"Birthday email sent to {first_name} {last_name}.", "success")
    else:
        cur.execute("""
            INSERT INTO client_birthday_offers (
                spa_id,
                client_id,
                birthday_year,
                offer_sent,
                offer_sent_date,
                acknowledged_by,
                notes,
                created_at,
                email_template_id,
                sent_status
            )
            VALUES (%s, %s, %s, FALSE, NULL, %s, %s, %s, %s, %s)
            ON CONFLICT (spa_id, client_id, birthday_year)
            DO UPDATE SET
                offer_sent = FALSE,
                acknowledged_by = EXCLUDED.acknowledged_by,
                sent_status = EXCLUDED.sent_status,
                email_template_id = EXCLUDED.email_template_id,
                notes = EXCLUDED.notes
        """, (
            spa_id,
            client_id,
            campaign_year,
            "system",
            error_note or "Email send failed",
            spa_now,
            email_template_id,
            "failed"
        ))

        conn.commit()
        flash("Birthday email could not be sent.", "error")

    cur.close()
    conn.close()

    return redirect(url_for("birthday_offers_home"))









#  ------------------------------------------
#      EMPLOYEES  
#
#
#
#  ------------------------------------------




#  ------------------------------------------
#      EMPLOYEES   HOME PAGE
#
#
#   ROUTE GOOD
#  ------------------------------------------



@app.route("/employees")
@login_required
@spa_required
def employees_home():
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            employee_id,
            first_name,
            last_name,
            phone,
            email,
            job_title,
            hire_date,
            termination_date,
            status,
            birthday,
            esthetician_license_number,
            license_expiration_date,
            pay_type,
            pay_rate,
            created_at
        FROM employees
        WHERE spa_id = %s
        ORDER BY last_name ASC, first_name ASC
    """, (spa_id,))
    employees = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*)
        FROM employees
        WHERE spa_id = %s
    """, (spa_id,))
    total_employees = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*)
        FROM employees
        WHERE spa_id = %s
          AND status = 'Active'
    """, (spa_id,))
    active_employees = cur.fetchone()[0]

    cur.close()
    conn.close()

    return render_template(
        "employees_home.html",
        employees=employees,
        total_employees=total_employees,
        active_employees=active_employees
    )





#   ---------------------------------------
#
#   EMPLOYEE PAY SUMMARY
#
#   route good
#   ---------------------------------------



from datetime import date

@app.route("/employee_pay_summary")
@login_required
@spa_required
def employee_pay_summary():
    spa_id = current_spa_id()

    today = date.today()
    first_day = today.replace(day=1)

    start_date = request.args.get("start_date") or first_day.strftime("%Y-%m-%d")
    end_date = request.args.get("end_date") or today.strftime("%Y-%m-%d")

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            e.employee_id,
            e.first_name || ' ' || e.last_name AS employee_name,
            COUNT(i.income_id) AS sessions_worked,
            COALESCE(SUM(i.service_amount), 0.00) AS service_sales,
            COALESCE(SUM(i.retail_amount), 0.00) AS retail_sales,
            COALESCE(SUM(i.tip_amount), 0.00) AS tips_earned,
            COALESCE(SUM(i.total_amount), 0.00) AS gross_collected
        FROM employees e
        LEFT JOIN income i
            ON e.employee_id = i.employee_id
           AND i.spa_id = e.spa_id
           AND i.income_date BETWEEN %s AND %s
        WHERE e.spa_id = %s
        GROUP BY e.employee_id, e.first_name, e.last_name
        ORDER BY e.last_name, e.first_name
    """, (start_date, end_date, spa_id))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "employee_pay_summary.html",
        rows=rows,
        start_date=start_date,
        end_date=end_date
    )






#   ----------------------------------------
#
#   ADD  EMPLOYEE COMPENSATION
#
#
#   ---------------------------------------



@app.route("/add_employee_compensation", methods=["GET", "POST"])
@login_required
@spa_required
def add_employee_compensation():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    # Load employees
    cur.execute("""
        SELECT employee_id, first_name || ' ' || last_name AS employee_name
        FROM employees
        WHERE spa_id = %s
        ORDER BY first_name, last_name
    """, (spa_id,))
    employees = cur.fetchall()

    # Load compensation types for all 4 dropdowns
    cur.execute("""
        SELECT compensation_type_id, compensation_type_name
        FROM compensation_types
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY compensation_type_name
    """, (spa_id,))
    compensation_types = cur.fetchall()

    if request.method == "POST":
        payment_date = request.form.get("payment_date")
        employee_id = request.form.get("employee_id") or None
        notes = (request.form.get("notes") or "").strip()

        # Gather up to 4 compensation lines
        detail_lines = []

        for i in range(1, 5):
            comp_type_id = request.form.get(f"comp_type_{i}") or None
            amount_raw = request.form.get(f"amount_{i}") or "0"

            try:
                amount = float(amount_raw)
            except ValueError:
                amount = 0

            if comp_type_id and amount > 0:
                detail_lines.append((comp_type_id, amount))

        # Validation
        if not payment_date:
            flash("Payment date is required.", "error")
        elif not employee_id:
            flash("Employee is required.", "error")
        elif not detail_lines:
            flash("Enter at least one compensation line with type and amount.", "error")
        else:
            # Insert header
            cur.execute("""
                INSERT INTO employee_compensation (
                    spa_id,
                    employee_id,
                    compensation_date,
                    notes
                )
                VALUES (%s, %s, %s, %s)
                RETURNING compensation_id
            """, (
                spa_id,
                employee_id,
                payment_date,
                notes
            ))
            compensation_id = cur.fetchone()[0]

            # Insert detail lines
            for comp_type_id, amount in detail_lines:
                cur.execute("""
                    INSERT INTO employee_compensation_lines (
                        compensation_id,
                        compensation_type_id,
                        amount
                    )
                    VALUES (%s, %s, %s)
                """, (
                    compensation_id,
                    comp_type_id,
                    amount
                ))

            conn.commit()
            cur.close()
            conn.close()

            flash("Compensation saved successfully.", "success")
            return redirect(url_for("employee_compensation_report"))

    cur.close()
    conn.close()

    return render_template(
        "add_employee_compensation.html",
        employees=employees,
        compensation_types=compensation_types
    )





#   -------------------------------
#  
#     EMPLOYEE  ADMIN
#   
#   -------------------------------

@app.route("/employee_admin")
@login_required
@spa_required
def employee_admin():
    spa_id = current_spa_id()

    return render_template("employee_admin.html")









#   ------------------------------------------
#
#   EMPLOYEE COMPENSATION HELPER
# 
#     HELPER         HELPER        HELPER
#   -----------------------------------------

def get_employee_compensation_history_data(spa_id, employee_id="", start_date="", end_date=""):
    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            ec.compensation_id,
            ec.compensation_date,
            e.first_name,
            e.last_name,
            ct.compensation_type_name,
            ecl.amount,
            ec.notes,
            ec.created_at
        FROM employee_compensation ec
        JOIN employees e
            ON ec.employee_id = e.employee_id
        JOIN employee_compensation_lines ecl
            ON ec.compensation_id = ecl.compensation_id
        LEFT JOIN compensation_types ct
            ON ecl.compensation_type_id = ct.compensation_type_id
        WHERE ec.spa_id = %s
          AND ec.compensation_date BETWEEN %s AND %s
    """
    params = [spa_id, start_date, end_date]

    if employee_id:
        query += " AND ec.employee_id = %s"
        params.append(employee_id)

    query += """
        ORDER BY ec.compensation_date DESC,
                 ec.compensation_id DESC,
                 ct.compensation_type_name ASC
    """

    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    cur.close()
    conn.close()

    return rows











#   -------------------------------
#  
#  COMPENSATION TYPES
#               route good 4/27
#   -------------------------------


@app.route("/compensation_types")
@login_required
@spa_required
def compensation_types_report():
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            compensation_type_id,
            compensation_type_name,
            is_active
        FROM compensation_types
        WHERE spa_id = %s
        ORDER BY compensation_type_name
    """, (spa_id,))
    rows = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "compensation_types_report.html",
        rows=rows
    )




#   -------------------------------
#  
#   ADD   COMPENSATION TYPE
#              route good 4/27
#   -------------------------------



@app.route("/add_compensation_type", methods=["GET", "POST"])
@login_required
@spa_required
def add_compensation_type():
    spa_id = current_spa_id()

    if request.method == "POST":
        compensation_type_name = (request.form.get("compensation_type_name") or "").strip()

        if not compensation_type_name:
            flash("Compensation type name is required.", "error")
        else:
            conn = get_db_connection()
            cur = conn.cursor()

            cur.execute("""
                SELECT 1
                FROM compensation_types
                WHERE spa_id = %s
                  AND LOWER(compensation_type_name) = LOWER(%s)
            """, (spa_id, compensation_type_name))
            existing = cur.fetchone()

            if existing:
                cur.close()
                conn.close()
                flash("That compensation type already exists.", "warning")
            else:
                cur.execute("""
                    INSERT INTO compensation_types (
                        spa_id,
                        compensation_type_name,
                        is_active
                    )
                    VALUES (%s, %s, TRUE)
                """, (spa_id, compensation_type_name))

                conn.commit()
                cur.close()
                conn.close()

                flash("Compensation type added successfully.", "success")
                return redirect(url_for("compensation_types_report"))

    return render_template("add_compensation_type.html")




#   -------------------------------
#  
#    EDIT COMPENSATION  TYPE
#   
#   -------------------------------


@app.route("/edit_compensation_type/<int:compensation_type_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_compensation_type(compensation_type_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        compensation_type_name = (request.form.get("compensation_type_name") or "").strip()

        if not compensation_type_name:
            flash("Compensation type name is required.", "error")
        else:
            cur.execute("""
                SELECT 1
                FROM compensation_types
                WHERE spa_id = %s
                  AND LOWER(compensation_type_name) = LOWER(%s)
                  AND compensation_type_id <> %s
            """, (spa_id, compensation_type_name, compensation_type_id))
            existing = cur.fetchone()

            if existing:
                flash("That compensation type already exists.", "warning")
            else:
                cur.execute("""
                    UPDATE compensation_types
                    SET compensation_type_name = %s
                    WHERE compensation_type_id = %s
                      AND spa_id = %s
                """, (compensation_type_name, compensation_type_id, spa_id))

                conn.commit()
                cur.close()
                conn.close()

                flash("Compensation type updated successfully.", "success")
                return redirect(url_for("compensation_types_report"))

    cur.execute("""
        SELECT
            compensation_type_id,
            compensation_type_name,
            is_active
        FROM compensation_types
        WHERE compensation_type_id = %s
          AND spa_id = %s
    """, (compensation_type_id, spa_id))
    row = cur.fetchone()

    cur.close()
    conn.close()

    if not row:
        flash("Compensation type not found.", "error")
        return redirect(url_for("compensation_types_report"))

    return render_template(
        "edit_compensation_type.html",
        row=row
    )







#   -------------------------------
#  
#     TOGGLE  COMPENSATION TYPE
#   route good 4/27   
#   -------------------------------


@app.route("/toggle_compensation_type/<int:compensation_type_id>", methods=["POST"])
@login_required
@spa_required
def toggle_compensation_type(compensation_type_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE compensation_types
        SET is_active = NOT is_active
        WHERE compensation_type_id = %s
          AND spa_id = %s
    """, (compensation_type_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Compensation type status updated.", "success")
    return redirect(url_for("compensation_types_report"))









#   -------------------------------
#
#  EMPLOYEE COMPENSATION REPORT
#   route good 4/27
#   -------------------------------



@app.route("/employee_compensation_report")
@login_required
@spa_required
def employee_compensation_report():
    spa_id = current_spa_id()
    today = date.today()
    first_day = today.replace(day=1)

    start_date = request.args.get("start_date") or first_day.strftime("%Y-%m-%d")
    end_date = request.args.get("end_date") or today.strftime("%Y-%m-%d")

    conn = get_db_connection()
    cur = conn.cursor()

    # Summary totals from income
    cur.execute("""
        SELECT
            COALESCE(SUM(service_amount), 0.00) AS services_billed,
            COALESCE(SUM(tip_amount), 0.00) AS tips_earned
        FROM income
        WHERE spa_id = %s
          AND income_date BETWEEN %s AND %s
    """, (spa_id, start_date, end_date))

    income_totals = cur.fetchone()
    services_billed_total = income_totals[0]
    tips_earned_total = income_totals[1]

    # Summary totals from compensation payments
    cur.execute("""
        SELECT
            COALESCE(SUM(CASE WHEN ct.compensation_type_name = 'Tip Payout' THEN ecl.amount ELSE 0 END), 0.00) AS tip_payouts_paid,
            COALESCE(SUM(CASE WHEN ct.compensation_type_name IN ('Draw', 'Owner Draw') THEN ecl.amount ELSE 0 END), 0.00) AS owner_draws_paid,
            COALESCE(SUM(CASE WHEN ct.compensation_type_name = 'Bonus' THEN ecl.amount ELSE 0 END), 0.00) AS bonus_paid,
            COALESCE(SUM(CASE WHEN ct.compensation_type_name = 'Extra Pay' THEN ecl.amount ELSE 0 END), 0.00) AS extra_pay_paid,
            COALESCE(SUM(ecl.amount), 0.00) AS total_comp_paid
        FROM employee_compensation ec
        JOIN employee_compensation_lines ecl
            ON ec.compensation_id = ecl.compensation_id
        JOIN compensation_types ct
            ON ecl.compensation_type_id = ct.compensation_type_id
        WHERE ec.spa_id = %s
          AND ec.compensation_date BETWEEN %s AND %s
    """, (spa_id, start_date, end_date))

    comp_totals = cur.fetchone()
    tip_payouts_paid_total = comp_totals[0]
    owner_draws_paid_total = comp_totals[1]
    bonus_paid_total = comp_totals[2]
    extra_pay_paid_total = comp_totals[3]
    total_comp_paid = comp_totals[4]

    outstanding_tips = tips_earned_total - tip_payouts_paid_total

    # Employee summary
    cur.execute("""
        SELECT
            e.employee_id,
            e.first_name || ' ' || e.last_name AS employee_name,

            COALESCE(inc.services_billed, 0.00) AS services_billed,
            COALESCE(inc.tips_earned, 0.00) AS tips_earned,

            0.00 AS commission_earned,

            COALESCE(comp.tip_payouts_paid, 0.00) AS tip_payouts_paid,
            COALESCE(comp.owner_draws_paid, 0.00) AS owner_draws_paid,
            COALESCE(comp.bonus_paid, 0.00) AS bonus_paid,
            COALESCE(comp.extra_pay_paid, 0.00) AS extra_pay_paid,
            COALESCE(comp.total_comp_paid, 0.00) AS total_paid,

            COALESCE(inc.tips_earned, 0.00) - COALESCE(comp.tip_payouts_paid, 0.00) AS outstanding_tips,

            (
                0.00
                + COALESCE(inc.tips_earned, 0.00)
                - COALESCE(comp.tip_payouts_paid, 0.00)
            ) AS net_balance

        FROM employees e

        LEFT JOIN (
            SELECT
                employee_id,
                COALESCE(SUM(service_amount), 0.00) AS services_billed,
                COALESCE(SUM(tip_amount), 0.00) AS tips_earned
            FROM income
            WHERE spa_id = %s
              AND income_date BETWEEN %s AND %s
            GROUP BY employee_id
        ) inc ON e.employee_id = inc.employee_id

        LEFT JOIN (
            SELECT
                ec.employee_id,
                COALESCE(SUM(CASE WHEN ct.compensation_type_name = 'Tip Payout' THEN ecl.amount ELSE 0 END), 0.00) AS tip_payouts_paid,
                COALESCE(SUM(CASE WHEN ct.compensation_type_name IN ('Draw', 'Owner Draw') THEN ecl.amount ELSE 0 END), 0.00) AS owner_draws_paid,
                COALESCE(SUM(CASE WHEN ct.compensation_type_name = 'Bonus' THEN ecl.amount ELSE 0 END), 0.00) AS bonus_paid,
                COALESCE(SUM(CASE WHEN ct.compensation_type_name = 'Extra Pay' THEN ecl.amount ELSE 0 END), 0.00) AS extra_pay_paid,
                COALESCE(SUM(ecl.amount), 0.00) AS total_comp_paid
            FROM employee_compensation ec
            JOIN employee_compensation_lines ecl
                ON ec.compensation_id = ecl.compensation_id
            JOIN compensation_types ct
                ON ecl.compensation_type_id = ct.compensation_type_id
            WHERE ec.spa_id = %s
              AND ec.compensation_date BETWEEN %s AND %s
            GROUP BY ec.employee_id
        ) comp ON e.employee_id = comp.employee_id

        WHERE e.spa_id = %s
        ORDER BY e.last_name, e.first_name
    """, (
        spa_id, start_date, end_date,
        spa_id, start_date, end_date,
        spa_id
    ))

    summary_rows = cur.fetchall()

    # Detailed ledger rows
    cur.execute("""
        SELECT
            ec.compensation_id,
            ec.compensation_date,
            e.first_name || ' ' || e.last_name AS employee_name,
            ct.compensation_type_name,
            ecl.amount,
            COALESCE(ec.notes, '') AS notes
        FROM employee_compensation ec
        JOIN employee_compensation_lines ecl
            ON ec.compensation_id = ecl.compensation_id
        JOIN compensation_types ct
            ON ecl.compensation_type_id = ct.compensation_type_id
        LEFT JOIN employees e
            ON ec.employee_id = e.employee_id
        WHERE ec.spa_id = %s
          AND ec.compensation_date BETWEEN %s AND %s
        ORDER BY ec.compensation_date DESC, ec.compensation_id DESC, ct.compensation_type_name
    """, (spa_id, start_date, end_date))

    ledger_rows = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "employee_compensation_report.html",
        summary_rows=summary_rows,
        ledger_rows=ledger_rows,
        start_date=start_date,
        end_date=end_date,
        services_billed_total=services_billed_total,
        tips_earned_total=tips_earned_total,
        tip_payouts_paid_total=tip_payouts_paid_total,
        owner_draws_paid_total=owner_draws_paid_total,
        bonus_paid_total=bonus_paid_total,
        extra_pay_paid_total=extra_pay_paid_total,
        total_comp_paid=total_comp_paid,
        outstanding_tips=outstanding_tips
    )





#   ----------------------------------
#
#     EMPLOYEE COMPENSATION HISTORY
#
#  route good 4/27
#   --------------------------------




@app.route("/employee_compensation_history")
@login_required
@spa_required
def employee_compensation_history():
    spa_id = current_spa_id()

    today = date.today()
    first_day = today.replace(day=1)

    employee_id = request.args.get("employee_id", "").strip()
    start_date = request.args.get("start_date", first_day.strftime("%Y-%m-%d")).strip()
    end_date = request.args.get("end_date", today.strftime("%Y-%m-%d")).strip()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT employee_id, first_name, last_name
        FROM employees
        WHERE spa_id = %s
        ORDER BY last_name, first_name
    """, (spa_id,))
    employees = cur.fetchall()

    cur.close()
    conn.close()

    rows = get_employee_compensation_history_data(
        spa_id=spa_id,
        employee_id=employee_id,
        start_date=start_date,
        end_date=end_date
    )

    total_amount = sum((row[5] or 0) for row in rows)

    return render_template(
        "employee_compensation_history.html",
        employees=employees,
        rows=rows,
        employee_id=employee_id,
        start_date=start_date,
        end_date=end_date,
        total_amount=total_amount
    )









#   -----------------------------------------
#       DELETE  EMPLOYEE COMPENSATION HISTORY
#
#    
#
#  route good 4/27
#   -----------------------------------------


@app.route("/delete_employee_compensation/<int:compensation_id>", methods=["POST"])
@login_required
@spa_required
def delete_employee_compensation(compensation_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT compensation_id
        FROM employee_compensation
        WHERE compensation_id = %s
          AND spa_id = %s
    """, (compensation_id, spa_id))
    record = cur.fetchone()

    if not record:
        cur.close()
        conn.close()
        flash("Compensation record not found.", "error")
        return redirect(url_for("employee_compensation_history"))

    cur.execute("""
        DELETE FROM employee_compensation_lines
        WHERE compensation_id = %s
    """, (compensation_id,))

    cur.execute("""
        DELETE FROM employee_compensation
        WHERE compensation_id = %s
          AND spa_id = %s
    """, (compensation_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Full compensation entry deleted successfully.", "success")
    return redirect(url_for("employee_compensation_history"))







#   -----------------------------------------
#       EDIT   EMPLOYEE COMPENSATION HISTORY
#           
#
#
#  route good 4/27
#   -----------------------------------------



@app.route("/edit_employee_compensation/<int:compensation_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_employee_compensation(compensation_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    # Load employees
    cur.execute("""
        SELECT employee_id, first_name || ' ' || last_name AS employee_name
        FROM employees
        WHERE spa_id = %s
        ORDER BY first_name, last_name
    """, (spa_id,))
    employees = cur.fetchall()

    # Load active compensation types
    cur.execute("""
        SELECT compensation_type_id, compensation_type_name
        FROM compensation_types
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY compensation_type_name
    """, (spa_id,))
    compensation_types = cur.fetchall()

    if request.method == "POST":
        payment_date = request.form.get("payment_date")
        employee_id = request.form.get("employee_id") or None
        notes = (request.form.get("notes") or "").strip()

        detail_lines = []

        for i in range(1, 5):
            comp_type_id = request.form.get(f"comp_type_{i}") or None
            amount_raw = request.form.get(f"amount_{i}") or "0"

            try:
                amount = float(amount_raw)
            except ValueError:
                amount = 0

            if comp_type_id and amount > 0:
                detail_lines.append((comp_type_id, amount))

        # Validation
        if not payment_date:
            flash("Payment date is required.", "error")
        elif not employee_id:
            flash("Employee is required.", "error")
        elif not detail_lines:
            flash("Enter at least one compensation line with type and amount.", "error")
        else:
            # Update header
            cur.execute("""
                UPDATE employee_compensation
                SET employee_id = %s,
                    compensation_date = %s,
                    notes = %s
                WHERE compensation_id = %s
                  AND spa_id = %s
            """, (
                employee_id,
                payment_date,
                notes,
                compensation_id,
                spa_id
            ))

            # Remove old detail lines
            cur.execute("""
                DELETE FROM employee_compensation_lines
                WHERE compensation_id = %s
            """, (compensation_id,))

            # Insert updated detail lines
            for comp_type_id, amount in detail_lines:
                cur.execute("""
                    INSERT INTO employee_compensation_lines (
                        compensation_id,
                        compensation_type_id,
                        amount
                    )
                    VALUES (%s, %s, %s)
                """, (
                    compensation_id,
                    comp_type_id,
                    amount
                ))

            conn.commit()
            cur.close()
            conn.close()

            flash("Compensation updated successfully.", "success")
            return redirect(url_for("employee_compensation_history"))

    # Load header record
    cur.execute("""
        SELECT compensation_id, employee_id, compensation_date, notes
        FROM employee_compensation
        WHERE compensation_id = %s
          AND spa_id = %s
    """, (compensation_id, spa_id))
    compensation = cur.fetchone()

    # Load detail lines
    cur.execute("""
        SELECT compensation_type_id, amount
        FROM employee_compensation_lines
        WHERE compensation_id = %s
        ORDER BY compensation_type_id
    """, (compensation_id,))
    existing_lines = cur.fetchall()

    cur.close()
    conn.close()

    # Pad to 4 rows for form display
    detail_rows = list(existing_lines)
    while len(detail_rows) < 4:
        detail_rows.append((None, ""))

    return render_template(
        "edit_employee_compensation.html",
        compensation=compensation,
        employees=employees,
        compensation_types=compensation_types,
        detail_rows=detail_rows
    )














#   -----------------------------------------
#       EXPORT EMPLOYEE COMPENSATION HISTORY
#
#     EXPORT TO CSV   EPXORT
#
#
#   -----------------------------------------



@app.route("/export_employee_compensation_history_csv")
@login_required
@spa_required
def export_employee_compensation_history_csv():
    spa_id = current_spa_id()

    today = date.today()
    first_day = today.replace(day=1)

    employee_id = request.args.get("employee_id", "").strip()
    start_date = request.args.get("start_date", first_day.strftime("%Y-%m-%d")).strip()
    end_date = request.args.get("end_date", today.strftime("%Y-%m-%d")).strip()

    rows = get_employee_compensation_history_data(
        spa_id=spa_id,
        employee_id=employee_id,
        start_date=start_date,
        end_date=end_date
    )

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Compensation ID",
        "Compensation Date",
        "Employee First Name",
        "Employee Last Name",
        "Compensation Type",
        "Amount",
        "Notes",
        "Created At"
    ])

    for row in rows:
        writer.writerow([
            row[0],
            row[1].strftime("%Y-%m-%d") if row[1] else "",
            row[2] or "",
            row[3] or "",
            row[4] or "",
            float(row[5] or 0),
            row[6] or "",
            row[7].strftime("%Y-%m-%d %I:%M %p") if row[7] else ""
        ])

    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=employee_compensation_history.csv"
        }
    )









#   ----------------------------------------
#   EXPORT   EXPORT  EXPORT TO EXCEL
#
#   EXPORT EMPLOYEE COMPENSATION HISTORY
#
#            EXPORT TO EXCEL
#   route good 4/27
#   ---------------------------------------



@app.route("/export_employee_compensation_history_excel")
@login_required
@spa_required
def export_employee_compensation_history_excel():
    spa_id = current_spa_id()

    today = date.today()
    first_day = today.replace(day=1)

    employee_id = request.args.get("employee_id", "").strip()
    start_date = request.args.get("start_date", first_day.strftime("%Y-%m-%d")).strip()
    end_date = request.args.get("end_date", today.strftime("%Y-%m-%d")).strip()

    rows = get_employee_compensation_history_data(
        spa_id=spa_id,
        employee_id=employee_id,
        start_date=start_date,
        end_date=end_date
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Compensation History"

    ws.append([
        "Compensation ID",
        "Compensation Date",
        "Employee First Name",
        "Employee Last Name",
        "Compensation Type",
        "Amount",
        "Notes",
        "Created At"
    ])

    for row in rows:
        ws.append([
            row[0],
            row[1].strftime("%Y-%m-%d") if row[1] else "",
            row[2] or "",
            row[3] or "",
            row[4] or "",
            float(row[5] or 0),
            row[6] or "",
            row[7].strftime("%Y-%m-%d %I:%M %p") if row[7] else ""
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=employee_compensation_history.xlsx"
        }
    )










#  ------------------------------------------
#          ADD EMPLOYEE
#
#
#
#   spa_id good   route good 4/27
#  ------------------------------------------

@app.route("/employees/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_employee():
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        first_name = request.form.get("first_name")
        last_name = request.form.get("last_name")
        address_line1 = request.form.get("address_line1")
        address_line2 = request.form.get("address_line2")
        city = request.form.get("city")
        state = request.form.get("state")
        zip_code = request.form.get("zip_code")
        phone = request.form.get("phone")
        email = request.form.get("email")
        job_title = request.form.get("job_title")
        hire_date = request.form.get("hire_date")
        termination_date = request.form.get("termination_date")
        status = request.form.get("status")
        birthday = request.form.get("birthday")
        ssn_on_file = True if request.form.get("ssn_on_file") == "on" else False
        esthetician_license_number = request.form.get("esthetician_license_number")
        license_expiration_date = request.form.get("license_expiration_date")
        year_graduated = request.form.get("year_graduated")
        certifications = request.form.get("certifications")
        pay_type = request.form.get("pay_type")
        pay_rate = request.form.get("pay_rate")
        notes = request.form.get("notes")

        if not first_name or not last_name:
            flash("First name and last name are required.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_employee"))

        cur.execute("""
            INSERT INTO employees (
                spa_id,
                first_name,
                last_name,
                address_line1,
                address_line2,
                city,
                state,
                zip_code,
                phone,
                email,
                job_title,
                hire_date,
                termination_date,
                status,
                birthday,
                ssn_on_file,
                esthetician_license_number,
                license_expiration_date,
                year_graduated,
                certifications,
                pay_type,
                pay_rate,
                notes
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            spa_id,
            first_name,
            last_name,
            address_line1,
            address_line2,
            city,
            state,
            zip_code,
            phone,
            email,
            job_title,
            hire_date or None,
            termination_date or None,
            status,
            birthday or None,
            ssn_on_file,
            esthetician_license_number,
            license_expiration_date or None,
            year_graduated or None,
            certifications,
            pay_type,
            pay_rate or None,
            notes
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Employee added successfully.", "success")
        return redirect(url_for("employees_home"))

    cur.execute("""
        SELECT status_name
        FROM employee_status
        WHERE spa_id = %s
        ORDER BY status_name ASC
    """, (spa_id,))
    statuses = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("add_employee.html", statuses=statuses)












#  ------------------------------------------
#          EDIT EMPLOYEE
# route good 4/27
#  ------------------------------------------

@app.route("/employees/edit/<int:employee_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_employee(employee_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        first_name = request.form.get("first_name")
        last_name = request.form.get("last_name")
        address_line1 = request.form.get("address_line1")
        address_line2 = request.form.get("address_line2")
        city = request.form.get("city")
        state = request.form.get("state")
        zip_code = request.form.get("zip_code")
        phone = request.form.get("phone")
        email = request.form.get("email")
        job_title = request.form.get("job_title")
        hire_date = request.form.get("hire_date")
        termination_date = request.form.get("termination_date")
        status = request.form.get("status")
        birthday = request.form.get("birthday")
        ssn_on_file = True if request.form.get("ssn_on_file") == "on" else False
        esthetician_license_number = request.form.get("esthetician_license_number")
        license_expiration_date = request.form.get("license_expiration_date")
        year_graduated = request.form.get("year_graduated")
        certifications = request.form.get("certifications")
        pay_type = request.form.get("pay_type")
        pay_rate = request.form.get("pay_rate")
        notes = request.form.get("notes")

        cur.execute("""
            UPDATE employees
            SET first_name = %s,
                last_name = %s,
                address_line1 = %s,
                address_line2 = %s,
                city = %s,
                state = %s,
                zip_code = %s,
                phone = %s,
                email = %s,
                job_title = %s,
                hire_date = %s,
                termination_date = %s,
                status = %s,
                birthday = %s,
                ssn_on_file = %s,
                esthetician_license_number = %s,
                license_expiration_date = %s,
                year_graduated = %s,
                certifications = %s,
                pay_type = %s,
                pay_rate = %s,
                notes = %s
            WHERE employee_id = %s
              AND spa_id = %
        """, (
            first_name,
            last_name,
            address_line1,
            address_line2,
            city,
            state,
            zip_code,
            phone,
            email,
            job_title,
            hire_date or None,
            termination_date or None,
            status,
            birthday or None,
            ssn_on_file,
            esthetician_license_number,
            license_expiration_date or None,
            year_graduated or None,
            certifications,
            pay_type,
            pay_rate or None,
            notes,
            employee_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Employee updated successfully.", "success")
        return redirect(url_for("employees_home"))

    cur.execute("""
        SELECT
            employee_id,
            first_name,
            last_name,
            address_line1,
            address_line2,
            city,
            state,
            zip_code,
            phone,
            email,
            job_title,
            hire_date,
            termination_date,
            status,
            birthday,
            ssn_on_file,
            esthetician_license_number,
            license_expiration_date,
            year_graduated,
            certifications,
            pay_type,
            pay_rate,
            notes,
            created_at
        FROM employees
        WHERE employee_id = %s
          AND spa_id = %s
    """, (employee_id,))
    employee = cur.fetchone()

    if not employee:
        cur.close()
        conn.close()
        flash("Employee not found.", "error")
        return redirect(url_for("employees_home"))

    cur.execute("SELECT status_name FROM employee_status ORDER BY status_name ASC")
    statuses = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("edit_employee.html", employee=employee, statuses=statuses)



            
#  ------------------------------------------
#            DELETE EMPLOYEE
#  good 4/27
#  ------------------------------------------
            
@app.route("/employees/delete/<int:employee_id>", methods=["POST"])
@login_required
@spa_required
def delete_employee(employee_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM employees WHERE employee_id = %s", (employee_id,))
    conn.commit()

    cur.close()
    conn.close()

    flash("Employee deleted successfully.", "success")
    return redirect(url_for("employees_home"))




#  ------------------------------------------
#
#      EXPENSES  SECTION
#
#
#  ------------------------------------------



    
#  ------------------------------------------
#      EXPENSES  HOME
# good 4/27
#  ------------------------------------------




@app.route("/expenses")
@login_required
@spa_required
def expenses_home():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            expense_id,
            expense_date,
            vendor_name,
            category,
            description,
            amount,
            payment_method,
            receipt_file,
            notes,
            created_at
        FROM expenses
        WHERE spa_id = %s
        ORDER BY expense_date DESC, expense_id DESC
        LIMIT 25
    """, (spa_id,))
    expenses = cur.fetchall()

    cur.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM expenses
        WHERE spa_id = %s
          AND expense_date = CURRENT_DATE
    """, (spa_id,))
    today_total = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM expenses
        WHERE spa_id = %s
          AND DATE_TRUNC('month', expense_date) = DATE_TRUNC('month', CURRENT_DATE)
    """, (spa_id,))
    month_total = cur.fetchone()[0]

    cur.close()
    conn.close()

    return render_template(
        "expenses_home.html",
        expenses=expenses,
        today_total=today_total,
        month_total=month_total
    )



















        
#  ------------------------------------------
#      ADD  EXPENSES
#
#    good 4/27
#  ------------------------------------------

@app.route("/expenses/add", methods=["GET", "POST"])
def add_expense(): 
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()  
            
    if request.method == "POST":
        expense_date = request.form.get("expense_date")
        vendor_name = request.form.get("vendor_name")
        category = request.form.get("category")
        description = request.form.get("description")
        amount = request.form.get("amount")
        payment_method = request.form.get("payment_method")
        receipt_file = request.form.get("receipt_file")
        notes = request.form.get("notes")
        
        if not expense_date or not vendor_name or not amount:
            flash("Expense date, vendor name, and amount are required.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_expense"))
        
        try:
            amount = Decimal(amount)
        except:
            flash("Amount must be a valid number.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_expense"))
            
        cur.execute("""
            INSERT INTO expenses (
                spa_id,
                expense_date,
                vendor_name,
                category,
                description,
                amount,
                payment_method,
                receipt_file,
                notes
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            spa_id,
            expense_date,
            vendor_name,
            category,  
            description,
            amount,
            payment_method,
            receipt_file,
            notes
        ))
        
        conn.commit()
        cur.close()
        conn.close()
    
        flash("Expense added successfully.", "success")
        return redirect(url_for("expenses_home"))
            
    cur.execute("""
        SELECT vendors_name
        FROM vendor_name
        ORDER BY vendors_name ASC
    """)
    vendors = cur.fetchall()

    cur.execute("""
        SELECT expense_cat_name
        FROM expense_categories
        WHERE spa_id = %s
        ORDER BY expense_cat_name ASC
    """, (spa_id,))
    categories = cur.fetchall()
        
    cur.execute("""
        SELECT payment_method
        FROM payment_methods
        WHERE spa_id =%s
        ORDER BY payment_method ASC
    """, (spa_id,))
    payment_methods = cur.fetchall()
            
    cur.close()
    conn.close()
        
    return render_template(
        "add_expense.html",  
        today=date.today().isoformat(),
        vendors=vendors, 
        categories=categories,
        payment_methods=payment_methods
    )




        
#  ------------------------------------------
#      EXPENSE REPORT
#  good 4/27
#  ------------------------------------------


@app.route("/expenses/report", methods=["GET"])
@login_required
@spa_required
def expense_report():
    spa_id = current_spa_id()

    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    category = request.args.get("category", "").strip()
    vendor_name = request.args.get("vendor_name", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT expense_cat_id, expense_cat_name
        FROM expense_categories
        ORDER BY expense_cat_name
    """)
    category_options = cur.fetchall()

    query = """
        SELECT
            expense_id,
            expense_date,
            vendor_name,
            category,
            description,
            amount,
            payment_method,
            receipt_file,
            notes,
            created_at
        FROM expenses
        WHERE spa_id = %s
    """
    params = [spa_id]

    if start_date:
        query += " AND expense_date >= %s"
        params.append(start_date)

    if end_date:
        query += " AND expense_date <= %s"
        params.append(end_date)

    if category:
        query += " AND category = %s"
        params.append(category)

    if vendor_name:
        query += " AND vendor_name ILIKE %s"
        params.append(f"%{vendor_name}%")

    query += " ORDER BY expense_date DESC, created_at DESC"

    cur.execute(query, tuple(params))
    expenses = cur.fetchall()

    total_query = """
        SELECT COALESCE(SUM(amount), 0)
        FROM expenses
        WHERE spa_id = %s
    """
    total_params = [spa_id]

    if start_date:
        total_query += " AND expense_date >= %s"
        total_params.append(start_date)

    if end_date:
        total_query += " AND expense_date <= %s"
        total_params.append(end_date)

    if category:
        total_query += " AND category = %s"
        total_params.append(category)

    if vendor_name:
        total_query += " AND vendor_name ILIKE %s"
        total_params.append(f"%{vendor_name}%")

    cur.execute(total_query, tuple(total_params))
    report_total = cur.fetchone()[0]

    category_totals_query = """
        SELECT
            category,
            COALESCE(SUM(amount), 0)
        FROM expenses
        WHERE spa_id = %s
    """
    category_totals_params = [spa_id]

    if start_date:
        category_totals_query += " AND expense_date >= %s"
        category_totals_params.append(start_date)

    if end_date:
        category_totals_query += " AND expense_date <= %s"
        category_totals_params.append(end_date)

    if category:
        category_totals_query += " AND category = %s"
        category_totals_params.append(category)

    if vendor_name:
        category_totals_query += " AND vendor_name ILIKE %s"
        category_totals_params.append(f"%{vendor_name}%")

    category_totals_query += """
        GROUP BY category
        ORDER BY category
    """

    cur.execute(category_totals_query, tuple(category_totals_params))
    category_totals_rows = cur.fetchall()

    category_totals = {}
    for row in category_totals_rows:
        category_name = row[0] if row[0] else "Uncategorized"
        category_totals[category_name] = row[1]

    cur.close()
    conn.close()

    return render_template(
        "expense_report.html",
        start_date=start_date,
        end_date=end_date,
        category=category,
        vendor_name=vendor_name,
        category_options=category_options,
        expenses=expenses,
        report_total=report_total,
        category_totals=category_totals
    )




















#  ------------------------------------------
#         EDIT EXPENSES
#  good 4/27/26
#  ------------------------------------------



@app.route("/expenses/edit/<int:expense_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_expense(expense_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        expense_date = request.form.get("expense_date")
        vendor_name = request.form.get("vendor_name")
        category = request.form.get("category")
        description = request.form.get("description")
        amount = request.form.get("amount")
        payment_method = request.form.get("payment_method")
        receipt_file = request.form.get("receipt_file")
        notes = request.form.get("notes")

        cur.execute("""
            UPDATE expenses
            SET expense_date = %s,
                vendor_name = %s,
                category = %s,
                description = %s,
                amount = %s,
                payment_method = %s,
                receipt_file = %s,
                notes = %s
            WHERE spa_id = %s
              AND expense_id = %s
        """, (
            expense_date,
            vendor_name,
            category,
            description,
            amount,
            payment_method,
            receipt_file,
            notes,
            spa_id,
            expense_id
        ))

        if cur.rowcount == 0:
            conn.rollback()
            cur.close()
            conn.close()

            flash(
                "Expense not found or not authorized.",
                "error"
            )
            return redirect(url_for("expenses_home"))

        conn.commit()
        cur.close()
        conn.close()

        flash("Expense updated successfully.", "success")
        return redirect(url_for("expenses_home"))

    cur.execute("""
        SELECT
            expense_id,
            expense_date,
            vendor_name,
            category,
            description,
            amount,
            payment_method,
            receipt_file,
            notes,
            created_at
        FROM expenses
        WHERE spa_id = %s
          AND expense_id = %s
    """, (
        spa_id,
        expense_id
    ))

    expense = cur.fetchone()

    if not expense:
        cur.close()
        conn.close()

        flash("Expense not found.", "error")
        return redirect(url_for("expenses_home"))

    cur.execute("""
        SELECT vendors_name
        FROM vendor_name
        ORDER BY vendors_name ASC
    """)
    vendors = cur.fetchall()

    cur.execute("""
        SELECT expense_cat_name
        FROM expense_categories
        ORDER BY expense_cat_name ASC
    """)
    categories = cur.fetchall()

    cur.execute("""
        SELECT payment_method
        FROM payment_methods
        ORDER BY payment_method ASC
    """)
    payment_methods = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "edit_expense.html",
        expense=expense,
        vendors=vendors,
        categories=categories,
        payment_methods=payment_methods
    )





#  ------------------------------------------
#         DELETE EXPENSES
#
#
#   good 4/27/26
#  ------------------------------------------
            
@app.route("/expenses/delete/<int:expense_id>", methods=["POST"])
@login_required
@spa_required
def delete_expense(expense_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM expenses WHERE expense_id = %s", (expense_id,))
    conn.commit()

    cur.close()
    conn.close()

    flash("Expense deleted successfully.", "success")
    return redirect(url_for("expenses_home"))


#  ------------------------------------------
#         EXPORT  EXPENSES TO CSV
#  ------------------------------------------

from flask import Response, request
import csv
import io

@app.route("/export_expense_report_csv")
@login_required
@spa_required
def export_expense_report_csv():
    spa_id = current_spa_id()
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            expense_date,
            vendor_name,
            category,
            description,
            amount,
            payment_method,
            notes
        FROM expenses
        WHERE spa_id =%s
           AND 1=1
    """
    params = []

    if start_date:
        query += " AND expense_date >= %s"
        params.append(start_date)

    if end_date:
        query += " AND expense_date <= %s"
        params.append(end_date)

    query += " ORDER BY expense_date DESC, expense_id DESC"

    cur.execute(query, params)
    rows = cur.fetchall()

    cur.close()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        "Expense Date",
        "Vendor Name",
        "Category",
        "Description",
        "Amount",
        "Payment Method",
        "Notes"
    ])

    # Data rows
    for row in rows:
        writer.writerow(row)

    csv_data = output.getvalue()
    output.close()

    return Response(
        csv_data,
        mimetype="text/csv",
        headers={
            "Content-Disposition": "attachment; filename=expense_report.csv"
        }
    )



#  ------------------------------------------  
#        EXPORT EXPENSES TO XLSX FORMAT
#  ------------------------------------------

from flask import send_file, request
from openpyxl import Workbook
from openpyxl.styles import Font
from io import BytesIO
from collections import defaultdict

@app.route("/export_expense_report_xlsx")
@login_required
@spa_required
def export_expense_report_xlsx():
    spa_id = current_spa_id()
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT
            expense_date,
            vendor_name,
            category,
            description,
            amount,
            payment_method,
            notes
        FROM expenses
        WHERE spa_id =%s
           AND 1=1
    """
    params = []

    if start_date:
        query += " AND expense_date >= %s"
        params.append(start_date)

    if end_date:
        query += " AND expense_date <= %s"
        params.append(end_date)

    query += " ORDER BY expense_date DESC, expense_id DESC"

    cur.execute(query, params)
    rows = cur.fetchall()

    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    # Title
    ws["A1"] = "Expense Report"
    ws["A1"].font = Font(bold=True, size=14)

    # Date range
    if start_date and end_date:
        ws["A2"] = f"Date Range: {start_date} to {end_date}"
    elif start_date:
        ws["A2"] = f"Date Range: From {start_date}"
    elif end_date:
        ws["A2"] = f"Date Range: Through {end_date}"
    else:
        ws["A2"] = "Date Range: All Expenses"

    # Headers
    headers = [
        "Expense Date",
        "Vendor Name",
        "Category",
        "Description",
        "Amount",
        "Payment Method",
        "Notes"
    ]

    header_row = 4
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Data rows
    data_start_row = 5
    grand_total = 0
    category_totals = defaultdict(float)

    for row_index, row in enumerate(rows, start=data_start_row):
        expense_date, vendor_name, category, description, amount, payment_method, notes = row

        ws.cell(row=row_index, column=1, value=expense_date)
        ws.cell(row=row_index, column=2, value=vendor_name)
        ws.cell(row=row_index, column=3, value=category)
        ws.cell(row=row_index, column=4, value=description)
        ws.cell(row=row_index, column=5, value=float(amount) if amount is not None else 0)
        ws.cell(row=row_index, column=6, value=payment_method)
        ws.cell(row=row_index, column=7, value=notes)

        ws.cell(row=row_index, column=5).number_format = '$#,##0.00'

        amt = float(amount) if amount is not None else 0
        grand_total += amt
        category_totals[category or "Uncategorized"] += amt

    # Grand total
    total_row = data_start_row + len(rows) + 1
    ws.cell(row=total_row, column=4, value="GRAND TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=grand_total).font = Font(bold=True)
    ws.cell(row=total_row, column=5).number_format = '$#,##0.00'

    # Category totals section
    category_start_row = total_row + 3
    ws.cell(row=category_start_row, column=1, value="Category Totals").font = Font(bold=True, size=12)

    ws.cell(row=category_start_row + 1, column=1, value="Category").font = Font(bold=True)
    ws.cell(row=category_start_row + 1, column=2, value="Total").font = Font(bold=True)

    current_row = category_start_row + 2
    for category, total in sorted(category_totals.items()):
        ws.cell(row=current_row, column=1, value=category)
        ws.cell(row=current_row, column=2, value=total)
        ws.cell(row=current_row, column=2).number_format = '$#,##0.00'
        current_row += 1

    # Column widths
    widths = {
        "A": 15,
        "B": 22,
        "C": 20,
        "D": 30,
        "E": 14,
        "F": 18,
        "G": 30
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Filename
    filename = "expense_report.xlsx"
    if start_date and end_date:
        filename = f"expense_report_{start_date}_to_{end_date}.xlsx"
    elif start_date:
        filename = f"expense_report_from_{start_date}.xlsx"
    elif end_date:
        filename = f"expense_report_to_{end_date}.xlsx"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )





################################
#
#   AUTOMATIC EXPENSES
#
#
################################



@app.route("/automatic-expenses")
@login_required
@spa_required
def automatic_expenses():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            ae.automatic_expense_id,          -- 0
            ae.expense_name,                  -- 1
            ae.vendor_name,                   -- 2
            ec.expense_cat_name,              -- 3
            ae.amount,                        -- 4
            ae.frequency,                     -- 5
            ae.next_post_date,                -- 6
            ae.last_processed_date,           -- 7
            ae.processing_type,               -- 8
            ae.is_active,                     -- 9
            ae.skip_next_occurrence,           -- 10
            ae.skipped_occurrence_date,        -- 11
            ae.last_error_message,             -- 12
            ae.last_error_at,                  -- 13
            ae.last_success_at,                -- 14
            pm.payment_method                  -- 15
        FROM automatic_expenses ae
        LEFT JOIN expense_categories ec
            ON ae.expense_cat_id = ec.expense_cat_id
           AND ec.spa_id = ae.spa_id
        LEFT JOIN payment_methods pm
            ON ae.payment_method_id = pm.payment_method_id
           AND pm.spa_id = ae.spa_id
        WHERE ae.spa_id = %s
        ORDER BY
            ae.is_active DESC,
            ae.next_post_date ASC,
            ae.expense_name ASC
    """, (spa_id,))

    automatic_expense_rows = cur.fetchall()

    cur.close()
    conn.close()

    automatic_expense_list = []

    spa_now = get_spa_now(spa_id)
    today = spa_now.date()
    now_time = spa_now.time()


    annual_multipliers = {
        "weekly": 52,
        "monthly": 12,
        "quarterly": 4,
        "annual": 1,
    }

    for row in automatic_expense_rows:
        amount = row[4] or 0
        frequency = row[5]
        next_post_date = row[6]
        is_active = row[9]
        skip_next = row[10]
        last_error_message = row[12]

        annual_total = amount * annual_multipliers.get(frequency, 0)

        if not is_active:
            payment_health = "Not Scheduled"
            payment_health_class = "inactive"
        elif last_error_message:
            payment_health = "Processing Failed"
            payment_health_class = "failed"
        elif next_post_date and next_post_date < today:
            payment_health = "Past Due"
            payment_health_class = "past-due"
        elif next_post_date == today:
            payment_health = "Due Today"
            payment_health_class = "due-today"
        elif skip_next:
            payment_health = "Next Occurrence Skipped"
            payment_health_class = "skipped"
        else:
            payment_health = "On Schedule"
            payment_health_class = "on-schedule"

        operating_status = "Active" if is_active else "Paused"

        processing_labels = {
            "auto_alert": "Auto-post + Coach Alert",
            "auto_silent": "Auto-post Silently",
            "reminder_only": "Reminder Only",
        }

        automatic_expense_list.append({
            "automatic_expense_id": row[0],
            "expense_name": row[1],
            "vendor_name": row[2],
            "expense_category": row[3],
            "amount": amount,
            "frequency": frequency,
            "next_post_date": next_post_date,
            "last_processed_date": row[7],
            "processing_type": row[8],
            "processing_label": processing_labels.get(
                row[8],
                row[8]
            ),
            "is_active": is_active,
            "skip_next_occurrence": skip_next,
            "skipped_occurrence_date": row[11],
            "last_error_message": last_error_message,
            "last_error_at": row[13],
            "last_success_at": row[14],
            "payment_method": row[15],
            "estimated_annual_total": annual_total,
            "payment_health": payment_health,
            "payment_health_class": payment_health_class,
            "operating_status": operating_status,
        })

    from datetime import timedelta

    week_end = today + timedelta(days=6)

    active_expenses = [
        expense
        for expense in automatic_expense_list
        if expense["is_active"]
    ]

    monthly_total = 0

    for expense in active_expenses:
        amount = expense["amount"] or 0
        frequency = expense["frequency"]

        if frequency == "weekly":
            monthly_total += amount * 52 / 12
        elif frequency == "monthly":
            monthly_total += amount
        elif frequency == "quarterly":
            monthly_total += amount * 4 / 12
        elif frequency == "annual":
            monthly_total += amount / 12

    due_this_week = sum(
        1
        for expense in active_expenses
        if expense["next_post_date"]
        and today <= expense["next_post_date"] <= week_end
    )

    reminder_only = sum(
        1
        for expense in active_expenses
        if expense["processing_type"] == "reminder_only"
    )

    summary = {
        "active_expenses": len(active_expenses),
        "monthly_total": monthly_total,
        "due_this_week": due_this_week,
        "reminder_only": reminder_only,
    }

    #coach message 

    past_due_expenses = [
        expense
        for expense in active_expenses
        if expense["next_post_date"]
        and expense["next_post_date"] < today
    ]

    due_today_expenses = [
        expense
        for expense in active_expenses
        if expense["next_post_date"] == today
    ]

    failed_expenses = [
        expense
        for expense in active_expenses
        if expense["last_error_message"]
    ]

    if failed_expenses:
        count = len(failed_expenses)

        if count == 1:
            coach_message = (
                f"{failed_expenses[0]['expense_name']} encountered a processing problem. "
                "Please review the expense before its next scheduled posting."
            )
        else:
            coach_message = (
                f"{count} recurring expenses encountered processing problems. "
                "Please review them before their next scheduled postings."
            )

    elif past_due_expenses:
        count = len(past_due_expenses)

        if count == 1:
            coach_message = (
                f"{past_due_expenses[0]['expense_name']} is past due. "
                "Please review its schedule and processing status."
            )
        else:
            coach_message = (
                f"{count} recurring expenses are past due. "
                "Please review their schedules and processing status."
            )

    elif due_today_expenses:
        count = len(due_today_expenses)

        if count == 1:
            expense = due_today_expenses[0]

            if expense["processing_type"] == "reminder_only":
                coach_message = (
                    f"{expense['expense_name']} is due today. "
                    "This expense requires you to record it manually."
                )
            else:
                coach_message = (
                    f"{expense['expense_name']} is scheduled to post today. "
                    "I’ll continue monitoring its processing status."
                )
        else:
            coach_message = (
                f"{count} recurring expenses are scheduled today. "
                "I’ll continue monitoring their processing status."
            )

    elif due_this_week == 1:
        upcoming_expense = next(
            expense
            for expense in active_expenses
            if expense["next_post_date"]
            and today <= expense["next_post_date"] <= week_end
        )

        formatted_date = upcoming_expense["next_post_date"].strftime(
            "%A, %B %d"
        )

        if upcoming_expense["processing_type"] == "reminder_only":
            coach_message = (
                f"{upcoming_expense['expense_name']} is due {formatted_date}. "
                "I’ll remind you to record it."
            )
        elif upcoming_expense["processing_type"] == "auto_silent":
            coach_message = (
                f"{upcoming_expense['expense_name']} is scheduled to post "
                f"{formatted_date}. No action is needed."
            )
        else:
            coach_message = (
                f"{upcoming_expense['expense_name']} is scheduled to post "
                f"{formatted_date}. I’ll let you know after it has been recorded."
            )

    elif due_this_week > 1:
        coach_message = (
            f"You have {due_this_week} recurring expenses scheduled during "
            "the next seven days. Everything is currently on schedule."
        )

    elif active_expenses:
        coach_message = (
            "Everything looks good. None of your recurring expenses are "
            "scheduled during the next seven days."
        )

    else:
        coach_message = (
            "You do not currently have any active recurring expenses."
        )

    return render_template(
        "automatic_expenses.html",
        automatic_expenses=automatic_expense_list,
        summary=summary,
        coach_message=coach_message
    )







################################
#
#   ADD AUTOMATIC EXPENSES
#
################################



@app.route("/automatic-expenses/add", methods=["GET", "POST"])
@login_required
@spa_required
def add_automatic_expense():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    # Load tenant-safe dropdown values.
    cur.execute("""
        SELECT
            expense_cat_id,
            expense_cat_name
        FROM expense_categories
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY expense_cat_name
    """, (spa_id,))
    expense_categories = cur.fetchall()

    cur.execute("""
        SELECT
            payment_method_id,
            payment_method
        FROM payment_methods
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY payment_method
    """, (spa_id,))
    payment_methods = cur.fetchall()

    form_data = {
        "expense_name": "",
        "vendor_name": "",
        "expense_cat_id": "",
        "payment_method_id": "",
        "amount": "",
        "frequency": "",
        "start_date": "",
        "end_date": "",
        "processing_type": "auto_alert",
        "description": "",
        "notes": "",
    }

    if request.method == "POST":
        form_data = {
            "expense_name": request.form.get(
                "expense_name",
                ""
            ).strip(),

            "vendor_name": request.form.get(
                "vendor_name",
                ""
            ).strip(),

            "expense_cat_id": request.form.get(
                "expense_cat_id",
                ""
            ).strip(),

            "payment_method_id": request.form.get(
                "payment_method_id",
                ""
            ).strip(),

            "amount": request.form.get(
                "amount",
                ""
            ).strip(),

            "frequency": request.form.get(
                "frequency",
                ""
            ).strip().lower(),

            "start_date": request.form.get(
                "start_date",
                ""
            ).strip(),

            "end_date": request.form.get(
                "end_date",
                ""
            ).strip(),

            "processing_type": request.form.get(
                "processing_type",
                ""
            ).strip().lower(),

            "description": request.form.get(
                "description",
                ""
            ).strip(),

            "notes": request.form.get(
                "notes",
                ""
            ).strip(),
        }

        errors = []

        allowed_frequencies = {
            "weekly",
            "monthly",
            "quarterly",
            "annual",
        }

        allowed_processing_types = {
            "auto_alert",
            "auto_silent",
            "reminder_only",
        }

        if not form_data["expense_name"]:
            errors.append("Expense Name is required.")

        if not form_data["expense_cat_id"]:
            errors.append("Expense Category is required.")

        if form_data["frequency"] not in allowed_frequencies:
            errors.append("Select a valid frequency.")

        if form_data["processing_type"] not in allowed_processing_types:
            errors.append("Select a valid processing option.")

        amount = None

        try:
            amount = Decimal(form_data["amount"])

            if amount <= 0:
                errors.append("Amount must be greater than zero.")

        except (InvalidOperation, TypeError):
            errors.append("Enter a valid amount.")

        start_date = None
        end_date = None

        try:
            start_date = date.fromisoformat(
                form_data["start_date"]
            )
        except (TypeError, ValueError):
            errors.append("Start Date is required.")

        if form_data["end_date"]:
            try:
                end_date = date.fromisoformat(
                    form_data["end_date"]
                )
            except (TypeError, ValueError):
                errors.append("Enter a valid End Date.")

        if (
            start_date
            and end_date
            and end_date < start_date
        ):
            errors.append(
                "End Date cannot be before Start Date."
            )

        expense_cat_id = None

        if form_data["expense_cat_id"]:
            try:
                expense_cat_id = int(
                    form_data["expense_cat_id"]
                )
            except ValueError:
                errors.append(
                    "Select a valid expense category."
                )

        payment_method_id = None

        if form_data["payment_method_id"]:
            try:
                payment_method_id = int(
                    form_data["payment_method_id"]
                )
            except ValueError:
                errors.append(
                    "Select a valid payment method."
                )

        # Confirm the category belongs to this business.
        if expense_cat_id is not None:
            cur.execute("""
                SELECT 1
                FROM expense_categories
                WHERE expense_cat_id = %s
                  AND spa_id = %s
                  AND is_active = TRUE
            """, (
                expense_cat_id,
                spa_id,
            ))

            if cur.fetchone() is None:
                errors.append(
                    "The selected expense category is not available."
                )

        # Confirm the payment method belongs to this business.
        if payment_method_id is not None:
            cur.execute("""
                SELECT 1
                FROM payment_methods
                WHERE payment_method_id = %s
                  AND spa_id = %s
                  AND is_active = TRUE
            """, (
                payment_method_id,
                spa_id,
            ))

            if cur.fetchone() is None:
                errors.append(
                    "The selected payment method is not available."
                )

        if errors:
            for error in errors:
                flash(error, "error")

            cur.close()
            conn.close()

            return render_template(
                "add_automatic_expense.html",
                expense_categories=expense_categories,
                payment_methods=payment_methods,
                form_data=form_data,
                page_title="Add Recurring Expense",
                form_action=url_for("add_automatic_expense"),
                submit_label="Save Recurring Expense",
                is_edit=False
            )

        try:
            cur.execute("""
                INSERT INTO automatic_expenses (
                    spa_id,
                    expense_name,
                    vendor_name,
                    expense_cat_id,
                    payment_method_id,
                    amount,
                    description,
                    notes,
                    frequency,
                    start_date,
                    end_date,
                    next_post_date,
                    processing_type,
                    is_active
                )
                VALUES (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, TRUE
                )
                RETURNING automatic_expense_id
            """, (
                spa_id,
                form_data["expense_name"],
                form_data["vendor_name"] or None,
                expense_cat_id,
                payment_method_id,
                amount,
                form_data["description"] or None,
                form_data["notes"] or None,
                form_data["frequency"],
                start_date,
                end_date,
                start_date,
                form_data["processing_type"],
            ))

            automatic_expense_id = cur.fetchone()[0]

            conn.commit()

        except Exception:
            conn.rollback()
            cur.close()
            conn.close()

            app.logger.exception(
                "Failed to add recurring expense."
            )

            flash(
                "The recurring expense could not be saved.",
                "error"
            )

            return render_template(
                "add_automatic_expense.html",
                expense_categories=expense_categories,
                payment_methods=payment_methods,
                form_data=form_data,
                page_title="Add Recurring Expense",
                form_action=url_for("add_automatic_expense"),
                submit_label="Save Recurring Expense",
                is_edit=False
            )

        cur.close()
        conn.close()

        flash(
            "Recurring expense added successfully.",
            "success"
        )

        return redirect(
            url_for("automatic_expenses")
        )

    cur.close()
    conn.close()

    return render_template(
        "add_automatic_expense.html",
        expense_categories=expense_categories,
        payment_methods=payment_methods,
        form_data=form_data,
        page_title="Add Recurring Expense",
        form_action=url_for("add_automatic_expense"),
        submit_label="Save Recurring Expense",
        is_edit=False
    )











################################
#
#   EDIT AUTOMATIN EXPENSE
#
################################



@app.route(
    "/automatic-expenses/<int:automatic_expense_id>/edit",
    methods=["GET", "POST"]
)
@login_required
@spa_required
def edit_automatic_expense(automatic_expense_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    # Load the recurring expense tenant-safely.
    cur.execute("""
        SELECT
            automatic_expense_id,       -- 0
            expense_name,               -- 1
            vendor_name,                -- 2
            expense_cat_id,             -- 3
            payment_method_id,          -- 4
            amount,                     -- 5
            frequency,                  -- 6
            start_date,                 -- 7
            end_date,                   -- 8
            next_post_date,             -- 9
            processing_type,            -- 10
            description,                -- 11
            notes,                      -- 12
            is_active                   -- 13
        FROM automatic_expenses
        WHERE automatic_expense_id = %s
          AND spa_id = %s
    """, (
        automatic_expense_id,
        spa_id,
    ))

    automatic_expense = cur.fetchone()

    if automatic_expense is None:
        cur.close()
        conn.close()

        flash(
            "Recurring expense not found or not authorized.",
            "error"
        )

        return redirect(
            url_for("automatic_expenses")
        )

    original_frequency = automatic_expense[6]
    original_start_date = automatic_expense[7]
    original_next_post_date = automatic_expense[9]

    # Load tenant-safe dropdown values.
    cur.execute("""
        SELECT
            expense_cat_id,
            expense_cat_name
        FROM expense_categories
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY expense_cat_name
    """, (spa_id,))

    expense_categories = cur.fetchall()

    cur.execute("""
        SELECT
            payment_method_id,
            payment_method
        FROM payment_methods
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY payment_method
    """, (spa_id,))

    payment_methods = cur.fetchall()

    form_data = {
        "expense_name": automatic_expense[1] or "",
        "vendor_name": automatic_expense[2] or "",
        "expense_cat_id": automatic_expense[3] or "",
        "payment_method_id": automatic_expense[4] or "",
        "amount": automatic_expense[5] or "",
        "frequency": automatic_expense[6] or "",
        "start_date": (
            automatic_expense[7].isoformat()
            if automatic_expense[7]
            else ""
        ),
        "end_date": (
            automatic_expense[8].isoformat()
            if automatic_expense[8]
            else ""
        ),
        "processing_type": automatic_expense[10] or "auto_alert",
        "description": automatic_expense[11] or "",
        "notes": automatic_expense[12] or "",
    }

    if request.method == "POST":
        form_data = {
            "expense_name": request.form.get(
                "expense_name",
                ""
            ).strip(),

            "vendor_name": request.form.get(
                "vendor_name",
                ""
            ).strip(),

            "expense_cat_id": request.form.get(
                "expense_cat_id",
                ""
            ).strip(),

            "payment_method_id": request.form.get(
                "payment_method_id",
                ""
            ).strip(),

            "amount": request.form.get(
                "amount",
                ""
            ).strip(),

            "frequency": request.form.get(
                "frequency",
                ""
            ).strip().lower(),

            "start_date": request.form.get(
                "start_date",
                ""
            ).strip(),

            "end_date": request.form.get(
                "end_date",
                ""
            ).strip(),

            "processing_type": request.form.get(
                "processing_type",
                ""
            ).strip().lower(),

            "description": request.form.get(
                "description",
                ""
            ).strip(),

            "notes": request.form.get(
                "notes",
                ""
            ).strip(),
        }

        errors = []

        allowed_frequencies = {
            "weekly",
            "monthly",
            "quarterly",
            "annual",
        }

        allowed_processing_types = {
            "auto_alert",
            "auto_silent",
            "reminder_only",
        }

        if not form_data["expense_name"]:
            errors.append("Expense Name is required.")

        if not form_data["expense_cat_id"]:
            errors.append("Expense Category is required.")

        if form_data["frequency"] not in allowed_frequencies:
            errors.append("Select a valid frequency.")

        if (
            form_data["processing_type"]
            not in allowed_processing_types
        ):
            errors.append("Select a valid processing option.")

        amount = None

        try:
            amount = Decimal(form_data["amount"])

            if amount <= 0:
                errors.append(
                    "Amount must be greater than zero."
                )

        except (InvalidOperation, TypeError):
            errors.append("Enter a valid amount.")

        start_date = None
        end_date = None

        try:
            start_date = date.fromisoformat(
                form_data["start_date"]
            )

        except (TypeError, ValueError):
            errors.append("Start Date is required.")

        if form_data["end_date"]:
            try:
                end_date = date.fromisoformat(
                    form_data["end_date"]
                )

            except (TypeError, ValueError):
                errors.append("Enter a valid End Date.")

        if (
            start_date
            and end_date
            and end_date < start_date
        ):
            errors.append(
                "End Date cannot be before Start Date."
            )

        expense_cat_id = None

        if form_data["expense_cat_id"]:
            try:
                expense_cat_id = int(
                    form_data["expense_cat_id"]
                )

            except ValueError:
                errors.append(
                    "Select a valid expense category."
                )

        payment_method_id = None

        if form_data["payment_method_id"]:
            try:
                payment_method_id = int(
                    form_data["payment_method_id"]
                )

            except ValueError:
                errors.append(
                    "Select a valid payment method."
                )

        # Confirm category belongs to this business.
        if expense_cat_id is not None:
            cur.execute("""
                SELECT 1
                FROM expense_categories
                WHERE expense_cat_id = %s
                  AND spa_id = %s
                  AND is_active = TRUE
            """, (
                expense_cat_id,
                spa_id,
            ))

            if cur.fetchone() is None:
                errors.append(
                    "The selected expense category is not available."
                )

        # Confirm payment method belongs to this business.
        if payment_method_id is not None:
            cur.execute("""
                SELECT 1
                FROM payment_methods
                WHERE payment_method_id = %s
                  AND spa_id = %s
                  AND is_active = TRUE
            """, (
                payment_method_id,
                spa_id,
            ))

            if cur.fetchone() is None:
                errors.append(
                    "The selected payment method is not available."
                )

        if errors:
            for error in errors:
                flash(error, "error")

            cur.close()
            conn.close()

            return render_template(
                "add_automatic_expense.html",
                expense_categories=expense_categories,
                payment_methods=payment_methods,
                form_data=form_data,
                page_title="Edit Recurring Expense",
                form_action=url_for(
                    "edit_automatic_expense",
                    automatic_expense_id=automatic_expense_id
                ),
                submit_label="Save Changes",
                is_edit=True
            )

        schedule_changed = (
            form_data["frequency"] != original_frequency
            or start_date != original_start_date
        )

        if schedule_changed:
            next_post_date = start_date
        else:
            next_post_date = original_next_post_date

        try:
            cur.execute("""
                UPDATE automatic_expenses
                SET
                    expense_name = %s,
                    vendor_name = %s,
                    expense_cat_id = %s,
                    payment_method_id = %s,
                    amount = %s,
                    description = %s,
                    notes = %s,
                    frequency = %s,
                    start_date = %s,
                    end_date = %s,
                    next_post_date = %s,
                    processing_type = %s,
                    updated_at = CURRENT_TIMESTAMP
                WHERE automatic_expense_id = %s
                  AND spa_id = %s
            """, (
                form_data["expense_name"],
                form_data["vendor_name"] or None,
                expense_cat_id,
                payment_method_id,
                amount,
                form_data["description"] or None,
                form_data["notes"] or None,
                form_data["frequency"],
                start_date,
                end_date,
                next_post_date,
                form_data["processing_type"],
                automatic_expense_id,
                spa_id,
            ))

            if cur.rowcount == 0:
                raise ValueError(
                    "Recurring expense was not updated."
                )

            conn.commit()

        except Exception:
            conn.rollback()

            app.logger.exception(
                "Failed to edit recurring expense."
            )

            flash(
                "The recurring expense could not be updated.",
                "error"
            )

            cur.close()
            conn.close()

            return render_template(
                "add_automatic_expense.html",
                expense_categories=expense_categories,
                payment_methods=payment_methods,
                form_data=form_data,
                page_title="Edit Recurring Expense",
                form_action=url_for(
                    "edit_automatic_expense",
                    automatic_expense_id=automatic_expense_id
                ),
                submit_label="Save Changes",
                is_edit=True
            )

        cur.close()
        conn.close()

        flash(
            "Recurring expense updated successfully.",
            "success"
        )

        return redirect(
            url_for("automatic_expenses")
        )

    cur.close()
    conn.close()

    return render_template(
        "add_automatic_expense.html",
        expense_categories=expense_categories,
        payment_methods=payment_methods,
        form_data=form_data,
        page_title="Edit Recurring Expense",
        form_action=url_for(
            "edit_automatic_expense",
            automatic_expense_id=automatic_expense_id
        ),
        submit_label="Save Changes",
        is_edit=True
    )









################################
#
#   
#
################################



################################
#
#   
#
################################
















#  ------------------------------------------
#
#      INCOME SECTION
#
#
#  ------------------------------------------



#  ------------------------------------------
#        
#        ADD INCOME
#
#
#    spa_id good.... route good  4/23/26
#  ------------------------------------------



@app.route("/add_income/<int:appointment_id>", methods=["GET", "POST"])
@login_required
@spa_required
def add_income(appointment_id):
    spa_id = current_spa_id()
    selected_date = request.args.get("date") or request.form.get("date") or ""

    def money_value(field_name):
        return float(request.form.get(field_name) or 0)

    conn = get_db_connection()
    cur = conn.cursor()

    # Get appointment and client info
    cur.execute("""
        SELECT
            a.appointment_id,
            a.client_id,
            a.appointment_date,
            a.appointment_time,
            c.first_name,
            c.last_name
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
           AND c.spa_id = a.spa_id
        WHERE a.appointment_id = %s
          AND a.spa_id = %s
    """, (appointment_id, spa_id))
    appt = cur.fetchone()

    if not appt:
        cur.close()
        conn.close()
        flash("Appointment not found.", "error")
        if selected_date:
            return redirect(url_for("daily_schedule", date=selected_date))
        return redirect(url_for("appointments"))

    cur.execute("""
        SELECT employee_id,
               first_name || ' ' || last_name AS employee_name
        FROM employees
        WHERE spa_id = %s
        ORDER BY employee_name
    """, (spa_id,))
    employees = cur.fetchall()

    cur.execute("""
        SELECT credit_processor_id, credit_processor_name
        FROM credit_processors
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY credit_processor_name
    """, (spa_id,))
    credit_processors = cur.fetchall()

    cur.execute("""
        SELECT COALESCE(SUM(amount), 0.00)
        FROM client_credit_transactions
        WHERE spa_id = %s
          AND client_id = %s
    """, (spa_id, appt[1]))
    credit_balance = float(cur.fetchone()[0] or 0.00)

    if request.method == "POST":
        income_date = request.form.get("income_date")
        income_type = request.form.get("income_type")
        description = request.form.get("description")

        service_amount = money_value("service_amount")
        retail_amount = money_value("retail_amount")
        tax_amount = money_value("tax_amount")
        tip_amount = money_value("tip_amount")
        credit_applied = money_value("credit_applied")

        total_amount = round(service_amount + retail_amount + tax_amount + tip_amount, 2)
        discountable_total = round(service_amount + retail_amount, 2)

        payment_method = request.form.get("payment_method", "").strip()
        credit_processor_id = request.form.get("credit_processor_id") or None
        processor_payment_id = request.form.get("processor_payment_id") or None
        employee_id = request.form.get("employee_id") or None
        notes = request.form.get("notes") or ""
        visit_id = None

        if credit_applied < 0:
            flash("Credit applied cannot be negative.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_income", appointment_id=appointment_id, date=selected_date))

        if credit_applied > credit_balance:
            flash("Credit applied cannot exceed available credit.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_income", appointment_id=appointment_id, date=selected_date))

        if credit_applied > discountable_total:
            flash("Credit applied cannot exceed service and retail total.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_income", appointment_id=appointment_id, date=selected_date))

        if employee_id:
            cur.execute("""
                SELECT employee_id
                FROM employees
                WHERE employee_id = %s
                  AND spa_id = %s
            """, (employee_id, spa_id))

            if not cur.fetchone():
                flash("Invalid employee selected.", "error")
                cur.close()
                conn.close()
                return redirect(url_for("add_income", appointment_id=appointment_id, date=selected_date))

        processing_fee_amount = 0.00
        net_received = total_amount
        processor_percentage_fee = 0.00
        processor_flat_fee = 0.00
        processor_additional_fee = 0.00

        card_based_methods = ["card", "credit card", "apple pay", "google pay", "square"]

        if payment_method.lower() in card_based_methods and credit_processor_id:
            cur.execute("""
                SELECT percentage_fee, flat_fee, additional_fee
                FROM credit_processors
                WHERE credit_processor_id = %s
                  AND spa_id = %s
                  AND is_active = TRUE
            """, (credit_processor_id, spa_id))

            processor_row = cur.fetchone()

            if not processor_row:
                flash("Invalid credit processor selected.", "error")
                cur.close()
                conn.close()
                return redirect(url_for("add_income", appointment_id=appointment_id, date=selected_date))

            processor_percentage_fee = float(processor_row[0] or 0)
            processor_flat_fee = float(processor_row[1] or 0)
            processor_additional_fee = float(processor_row[2] or 0)

            processing_fee_amount = round(
                (total_amount * (processor_percentage_fee / 100))
                + processor_flat_fee
                + processor_additional_fee,
                2
            )

            net_received = round(total_amount - processing_fee_amount, 2)

        elif payment_method.lower() not in card_based_methods:
            credit_processor_id = None
            processor_payment_id = None

        if credit_applied > 0:
            cur.execute("""
                INSERT INTO client_credit_transactions (
                    spa_id,
                    client_id,
                    source_type,
                    source_id,
                    transaction_date,
                    transaction_type,
                    amount,
                    notes
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                spa_id,
                appt[1],
                "Income",
                None,
                income_date,
                "Applied",
                -abs(credit_applied),
                f"Client credit applied to income for appointment {appt[0]}."
            ))

        cur.execute("""
            INSERT INTO income (
                income_date,
                client_id,
                appointment_id,
                visit_id,
                income_type,
                description,
                service_amount,
                retail_amount,
                tax_amount,
                tip_amount,
                total_amount,
                payment_method,
                processor_payment_id,
                notes,
                spa_id,
                employee_id,
                credit_processor_id,
                processing_fee_amount,
                net_received,
                processor_percentage_fee,
                processor_flat_fee,
                processor_additional_fee,
                created_at
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, CURRENT_TIMESTAMP
            )
        """, (
            income_date,
            appt[1],
            appt[0],
            visit_id,
            income_type,
            description,
            service_amount,
            retail_amount,
            tax_amount,
            tip_amount,
            total_amount,
            payment_method,
            processor_payment_id,
            notes,
            spa_id,
            employee_id,
            credit_processor_id,
            processing_fee_amount,
            net_received,
            processor_percentage_fee,
            processor_flat_fee,
            processor_additional_fee
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Income added successfully.", "success")
        return redirect(url_for(
            "post_appointment_wrap_up",
            appointment_id=appt[0],
            date=selected_date
        ))

    cur.close()
    conn.close()

    return render_template(
        "add_income.html",
        appt=appt,
        selected_date=selected_date,
        credit_processors=credit_processors,
        employees=employees,
        credit_balance=credit_balance
    )








#  --------------------------
#
#     EDIT  INCOME
#good 4/27
#  ------------------------



@app.route("/edit_income/<int:income_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_income(income_id):
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        income_date = request.form.get("income_date") or None
        income_type = request.form.get("income_type", "").strip()  or "General"
        employee_id = request.form.get("employee_id") or None

        service_amount = float(request.form.get("service_amount") or 0)
        retail_amount = float(request.form.get("retail_amount") or 0)
        tax_amount = float(request.form.get("tax_amount") or 0)
        tip_amount = float(request.form.get("tip_amount") or 0)

        total_amount = round(service_amount + retail_amount + tax_amount + tip_amount, 2)

        payment_method = request.form.get("payment_method", "").strip()
        credit_processor_id = request.form.get("credit_processor_id") or None
        processor_payment_id = request.form.get("processor_payment_id") or None

        processing_fee_amount = 0.00
        net_received = total_amount
        processor_percentage_fee = 0.00
        processor_flat_fee = 0.00
        processor_additional_fee = 0.00

        card_based_methods = ["card", "credit card", "apple pay", "google pay", "square"]

        if payment_method.lower() in card_based_methods:
            if credit_processor_id:
                cur.execute("""
                    SELECT percentage_fee, flat_fee, additional_fee
                    FROM credit_processors
                    WHERE credit_processor_id = %s
                      AND spa_id = %s
                      AND is_active = TRUE
                """, (credit_processor_id, spa_id))
                processor_row = cur.fetchone()

                if processor_row:
                    processor_percentage_fee = float(processor_row[0] or 0)
                    processor_flat_fee = float(processor_row[1] or 0)
                    processor_additional_fee = float(processor_row[2] or 0)

                    processing_fee_amount = round(
                        (total_amount * (processor_percentage_fee / 100))
                        + processor_flat_fee
                        + processor_additional_fee,
                        2
                    )
                    net_received = round(total_amount - processing_fee_amount, 2)
        else:
            credit_processor_id = None
            processor_payment_id = None

        description = request.form.get("description") or None
        notes = request.form.get("notes") or None
        client_id = request.form.get("client_id") or None

        cur.execute("""
            UPDATE income
            SET
                income_date = %s,
                income_type = %s,
                description = %s,
                service_amount = %s,
                retail_amount = %s,
                tax_amount = %s,
                tip_amount = %s,
                total_amount = %s,
                payment_method = %s,
                processor_payment_id = %s,
                notes = %s,
                employee_id = %s,
                credit_processor_id = %s,
                processing_fee_amount = %s,
                net_received = %s,
                processor_percentage_fee = %s,
                processor_flat_fee = %s,
                processor_additional_fee = %s,
                client_id = %s
            WHERE income_id = %s
              AND spa_id = %s
        """, (
            income_date,
            income_type,
            description,
            service_amount,
            retail_amount,
            tax_amount,
            tip_amount,
            total_amount,
            payment_method,
            processor_payment_id,
            notes,
            employee_id,
            credit_processor_id,
            processing_fee_amount,
            net_received,
            processor_percentage_fee,
            processor_flat_fee,
            processor_additional_fee,
            client_id,
            income_id,
            spa_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Income record updated successfully.", "success")
        return redirect(url_for("income_report"))

    cur.execute("""
        SELECT
            i.income_id,
            i.income_date,
            i.income_type,
            i.description,
            i.service_amount,
            i.retail_amount,
            i.tax_amount,
            i.tip_amount,
            i.total_amount,
            i.payment_method,
            i.processor_payment_id,
            i.notes,
            i.client_id,
            i.employee_id,
            i.credit_processor_id,
            c.first_name,
            c.last_name
        FROM income i
        LEFT JOIN clients c ON i.client_id = c.client_id
        WHERE i.income_id = %s
          AND i.spa_id = %s
    """, (income_id, spa_id))
    income_record = cur.fetchone()

    cur.execute("""
        SELECT client_id, first_name, last_name
        FROM clients
        WHERE spa_id = %s
        ORDER BY last_name, first_name
    """, (spa_id,))
    clients = cur.fetchall()

    cur.execute("""
        SELECT employee_id, first_name, last_name
        FROM employees
        WHERE spa_id = %s
        ORDER BY last_name, first_name
    """, (spa_id,))
    employees = cur.fetchall()

    cur.execute("""
        SELECT credit_processor_id, credit_processor_name
        FROM credit_processors
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY credit_processor_name
    """, (spa_id,))
    credit_processors = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "edit_income.html",
        income_record=income_record,
        clients=clients,
        employees=employees,
        credit_processors=credit_processors
    )





    
#  --------------------------
#     INCOME EXPORT TO CSV      
# ROUTE: income_report/csv    
#  good 4/27
#  ------------------------


@app.route("/income_report/csv")
@login_required
@spa_required
def income_report_csv(): 
    spa_id = current_spa_id()
    role = session.get("role")

    today = date.today()
    first_day = today.replace(day=1)
            
    start_date = request.args.get("start_date", first_day.strftime("%Y-%m-%d"))
    end_date = request.args.get("end_date", today.strftime("%Y-%m-%d"))
    income_type = request.args.get("income_type", "").strip()
            
    conn = get_db_connection()
    cur = conn.cursor()

    filter_sql = "WHERE i.income_date BETWEEN %s AND %s"
    params = [start_date, end_date]

    if role != "master_admin":
        filter_sql += " AND i.spa_id = %s"
        params.append(spa_id)

    if income_type:
        filter_sql += " AND i.income_type = %s"
        params.append(income_type)
        
    cur.execute(f"""
        SELECT
            i.income_id,
            i.income_date,
            COALESCE(c.first_name || ' ' || c.last_name, 'No Client') AS client_name,
            COALESCE(e.first_name || ' ' || e.last_name, 'Unassigned') AS employee_name,
            COALESCE(i.income_type, '') AS income_type,
            COALESCE(i.description, '') AS description,
            COALESCE(i.payment_method, '') AS payment_method,
            COALESCE(i.service_amount, 0.00) AS service_amount,
            COALESCE(i.tip_amount, 0.00) AS tip_amount,
            COALESCE(i.retail_amount, 0.00) AS retail_amount,
            COALESCE(i.tax_amount, 0.00) AS tax_amount,
            COALESCE(i.total_amount, 0.00) AS total_amount,
            COALESCE(i.notes, '') AS notes
        FROM income i
        LEFT JOIN clients c 
            ON i.client_id = c.client_id
           AND i.spa_id = c.spa_id
        LEFT JOIN employees e 
            ON i.employee_id = e.employee_id
           AND i.spa_id = e.spa_id
        {filter_sql}
        ORDER BY i.income_date DESC, i.income_id DESC
    """, params)
            
    rows = cur.fetchall()
    
    cur.close()
    conn.close()
        
    output = io.StringIO()
    writer = csv.writer(output)
        
    writer.writerow([
        "Income ID",
        "Income Date",
        "Client Name",
        "Employee",
        "Income Type",
        "Description",
        "Payment Method",
        "Service Amount",
        "Tip Amount",
        "Retail Amount",  
        "Tax Amount",
        "Total Amount",
        "Notes"
    ])      

    for row in rows:
        writer.writerow(row)

    output.seek(0)
    
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment; filename=income_report.csv"}
    )






# -----------------------
# INCOME EXPORT TO EXCEL
#  ROUTE: income_report/excel
#
#  good 4/27
#  ----------------------


@app.route("/income_report/excel")
@login_required
@spa_required
def income_report_excel():
    spa_id = current_spa_id()
    role = session.get("role")

    today = date.today()
    first_day = today.replace(day=1)
        
    start_date = request.args.get("start_date", first_day.strftime("%Y-%m-%d"))
    end_date = request.args.get("end_date", today.strftime("%Y-%m-%d"))
    income_type = request.args.get("income_type", "").strip()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    filter_sql = "WHERE i.income_date BETWEEN %s AND %s"
    params = [start_date, end_date]

    if role != "master_admin":
        filter_sql += " AND i.spa_id = %s"
        params.append(spa_id)
        
    if income_type:
        filter_sql += " AND i.income_type = %s"
        params.append(income_type)

    cur.execute(f"""
        SELECT
            i.income_id,
            i.income_date,
            COALESCE(c.first_name || ' ' || c.last_name, 'No Client') AS client_name,
            COALESCE(e.first_name || ' ' || e.last_name, 'Unassigned') AS employee_name,
            COALESCE(i.income_type, '') AS income_type,
            COALESCE(i.description, '') AS description,
            COALESCE(i.payment_method, '') AS payment_method,
            COALESCE(i.service_amount, 0.00) AS service_amount,
            COALESCE(i.tip_amount, 0.00) AS tip_amount,
            COALESCE(i.retail_amount, 0.00) AS retail_amount,
            COALESCE(i.tax_amount, 0.00) AS tax_amount,
            COALESCE(i.total_amount, 0.00) AS total_amount,
            COALESCE(i.notes, '') AS notes
        FROM income i
        LEFT JOIN clients c 
            ON i.client_id = c.client_id
           AND i.spa_id = c.spa_id
        LEFT JOIN employees e 
            ON i.employee_id = e.employee_id
           AND i.spa_id = e.spa_id
        {filter_sql}
        ORDER BY i.income_date DESC, i.income_id DESC
    """, params)

    rows = cur.fetchall()   
            
    cur.close()   
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Report"
     
    ws.append([
        "Income ID",
        "Income Date",
        "Client Name",
        "Employee",
        "Income Type",
        "Description",    
        "Payment Method",
        "Service Amount",
        "Tip Amount",
        "Retail Amount",
        "Tax Amount",
        "Total Amount",
        "Notes"
    ])

    for row in rows:
        ws.append(list(row))
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name="income_report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )






#  --------------------------
#
#     DELETE  INCOME
#   4/27
#  ------------------------

@app.route("/delete_income/<int:income_id>", methods=["POST"])
@login_required
@spa_required
def delete_income(income_id):
    spa_id = current_spa_id()
    start_date = request.form.get("start_date", "").strip()
    end_date = request.form.get("end_date", "").strip()
    income_type = request.form.get("income_type", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM income WHERE income_id = %s", (income_id,))
    conn.commit()

    cur.close()
    conn.close()

    flash("Income record deleted.", "success")

    return redirect(url_for(
        "income_report",
        start_date=start_date,
        end_date=end_date,
        income_type=income_type
    ))






#-----------------------------
#
#  INCOME REPORT
#
#   4/28
#  --------------------------




from datetime import date
from flask import render_template, request
from db import get_db_connection


@app.route("/income_report")
@login_required  
@spa_required
def income_report():
    spa_id = current_spa_id()
    role = session.get("role")

    today = date.today()
    first_day = today.replace(day=1)
        
    start_date = request.args.get("start_date") or first_day.strftime("%Y-%m-%d")
    end_date = request.args.get("end_date") or today.strftime("%Y-%m-%d")
    income_type = request.args.get("income_type", "").strip()
        
    conn = get_db_connection() 
    cur = conn.cursor()

    # Income type dropdown
    if role == "master_admin":
        cur.execute("""
            SELECT DISTINCT income_type
            FROM income
            WHERE income_type IS NOT NULL
              AND income_type <> ''
            ORDER BY income_type
        """)
    else:
        cur.execute("""
            SELECT DISTINCT income_type
            FROM income
            WHERE spa_id = %s
              AND income_type IS NOT NULL
              AND income_type <> ''
            ORDER BY income_type
        """, (spa_id,))

    income_type_options = [row[0] for row in cur.fetchall()]

    # Base filters
    filter_sql = "WHERE income_date BETWEEN %s AND %s"
    params = [start_date, end_date]

    if role != "master_admin":
        filter_sql += " AND spa_id = %s"
        params.append(spa_id)

    if income_type:
        filter_sql += " AND income_type = %s"
        params.append(income_type)

    # Alias version for joined income queries
    filter_sql_i = "WHERE i.income_date BETWEEN %s AND %s"
    params_i = [start_date, end_date]

    if role != "master_admin":
        filter_sql_i += " AND i.spa_id = %s"
        params_i.append(spa_id)

    if income_type:
        filter_sql_i += " AND i.income_type = %s"
        params_i.append(income_type)

    # Summary totals
    cur.execute(f"""
        SELECT    
            COUNT(*) AS total_entries,
            COALESCE(SUM(service_amount), 0.00) AS total_services,
            COALESCE(SUM(retail_amount), 0.00) AS total_retail,
            COALESCE(SUM(tip_amount), 0.00) AS total_tips,
            COALESCE(SUM(tax_amount), 0.00) AS total_tax,
            COALESCE(SUM(total_amount), 0.00) AS gross_collected,
            COALESCE(SUM(processing_fee_amount), 0.00) AS total_processing_fees,
            COALESCE(SUM(net_received), 0.00) AS total_net_received,
            COALESCE(SUM(service_amount + retail_amount), 0.00) AS spa_income
        FROM income
        {filter_sql}
    """, params)
    summary = cur.fetchone()

    # Income type breakdown
    cur.execute(f"""
        SELECT
            COALESCE(income_type, 'Unspecified') AS income_type,
            COUNT(*) AS entry_count,
            COALESCE(SUM(service_amount + retail_amount), 0.00) AS spa_income,
            COALESCE(SUM(tip_amount), 0.00) AS total_tips,
            COALESCE(SUM(total_amount), 0.00) AS gross_collected,
            COALESCE(SUM(processing_fee_amount), 0.00) AS total_processing_fees,
            COALESCE(SUM(net_received), 0.00) AS total_net_received
        FROM income
        {filter_sql}
        GROUP BY income_type
        ORDER BY gross_collected DESC
    """, params)
    income_type_breakdown = cur.fetchall()
        
    # Payment method breakdown
    cur.execute(f"""
        SELECT
            COALESCE(payment_method, 'Unspecified') AS payment_method,
            COUNT(*) AS entry_count,
            COALESCE(SUM(service_amount + retail_amount), 0.00) AS spa_income,
            COALESCE(SUM(tip_amount), 0.00) AS total_tips,
            COALESCE(SUM(total_amount), 0.00) AS gross_collected,
            COALESCE(SUM(processing_fee_amount), 0.00) AS total_processing_fees, 
            COALESCE(SUM(net_received), 0.00) AS total_net_received
        FROM income
        {filter_sql}
        GROUP BY payment_method
        ORDER BY gross_collected DESC
    """, params)
    payment_breakdown = cur.fetchall()
    
    # Credit processor breakdown   
    cur.execute(f"""
        SELECT
            COALESCE(cp.credit_processor_name, 'None') AS credit_processor_name,
            COUNT(*) AS entry_count,
            COALESCE(SUM(i.total_amount), 0.00) AS gross_collected,
            COALESCE(SUM(i.processing_fee_amount), 0.00) AS total_processing_fees,
            COALESCE(SUM(i.net_received), 0.00) AS total_net_received
        FROM income i
        LEFT JOIN credit_processors cp
            ON i.credit_processor_id = cp.credit_processor_id
           AND i.spa_id = cp.spa_id
        {filter_sql_i}
        GROUP BY cp.credit_processor_name
        ORDER BY total_processing_fees DESC, gross_collected DESC
    """, params_i)
    processor_breakdown = cur.fetchall()
    
    # Detailed report rows
    cur.execute(f"""
        SELECT    
            i.income_id,
            i.income_date,
            COALESCE(c.first_name || ' ' || c.last_name, 'No Client') AS client_name,
            COALESCE(e.first_name || ' ' || e.last_name, 'Unassigned') AS employee_name,
            COALESCE(i.income_type, '') AS income_type,  
            COALESCE(i.description, '') AS description,
            COALESCE(i.payment_method, '') AS payment_method,
            COALESCE(cp.credit_processor_name, '') AS credit_processor_name,
            COALESCE(i.processor_payment_id, '') AS processor_payment_id,
            COALESCE(i.service_amount, 0.00) AS service_amount,
            COALESCE(i.tip_amount, 0.00) AS tip_amount,
            COALESCE(i.retail_amount, 0.00) AS retail_amount,
            COALESCE(i.tax_amount, 0.00) AS tax_amount,
            COALESCE(i.total_amount, 0.00) AS total_amount,
            COALESCE(i.processing_fee_amount, 0.00) AS processing_fee_amount,
            COALESCE(i.net_received, 0.00) AS net_received,
            COALESCE(i.notes, '') AS notes
        FROM income i
        LEFT JOIN clients c 
            ON i.client_id = c.client_id
           AND i.spa_id = c.spa_id
        LEFT JOIN employees e 
            ON i.employee_id = e.employee_id
           AND i.spa_id = e.spa_id
        LEFT JOIN credit_processors cp 
            ON i.credit_processor_id = cp.credit_processor_id
           AND i.spa_id = cp.spa_id
        {filter_sql_i}
        ORDER BY i.income_date DESC, i.income_id DESC
    """, params_i)
    income_rows = cur.fetchall()
        
    cur.close()
    conn.close()
    
    return render_template(
        "income_report.html",
        start_date=start_date,
        end_date=end_date,
        income_type=income_type,
        income_type_options=income_type_options,
        summary=summary,
        income_type_breakdown=income_type_breakdown,
        payment_breakdown=payment_breakdown,
        processor_breakdown=processor_breakdown,
        income_rows=income_rows
    )








#  -----------------------------
#     ADD GENERAL INCOME
#    4/28 
#  -----------------------------

from flask import render_template, request, redirect, url_for, flash
from decimal import Decimal, InvalidOperation
from db import get_db_connection


@app.route("/add_general_income", methods=["GET", "POST"])
@login_required
@spa_required  
def add_general_income():
    spa_id = current_spa_id()
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()

    if role == "master_admin":
        cur.execute("""
            SELECT client_id, first_name, last_name 
            FROM clients
            ORDER BY last_name, first_name
        """)
    else:
        cur.execute("""
            SELECT client_id, first_name, last_name 
            FROM clients
            WHERE spa_id = %s
            ORDER BY last_name, first_name
        """, (spa_id,))
    clients = cur.fetchall()   

    cur.execute("""
        SELECT income_type_name
        FROM income_types
        WHERE spa_id = %s
        ORDER BY income_type_name
    """, (spa_id,))
    income_types = cur.fetchall()

    cur.execute("""
        SELECT payment_method
        FROM payment_methods
        WHERE spa_id = %s
        ORDER BY payment_method
    """, (spa_id,))
    payment_methods = cur.fetchall()

    if request.method == "POST":
        income_date = request.form.get("income_date")
        client_id = request.form.get("client_id")
        income_type = request.form.get("income_type")
        description = request.form.get("description", "").strip()
        service_amount = request.form.get("service_amount", "0").strip()
        retail_amount = request.form.get("retail_amount", "0").strip()
        tax_amount = request.form.get("tax_amount", "0").strip()
        total_amount = request.form.get("total_amount", "0").strip()
        payment_method = request.form.get("payment_method")
        square_payment_id = request.form.get("square_payment_id", "").strip()
        notes = request.form.get("notes", "").strip()

        try:
            service_amount = Decimal(service_amount) if service_amount else Decimal("0.00")
            retail_amount = Decimal(retail_amount) if retail_amount else Decimal("0.00")
            tax_amount = Decimal(tax_amount) if tax_amount else Decimal("0.00")
            total_amount = Decimal(total_amount) if total_amount else Decimal("0.00")
        except InvalidOperation:
            flash("Please enter valid numeric amounts.", "error")
            cur.close()
            conn.close()
            return render_template(
                "add_general_income.html",
                clients=clients,
                income_types=income_types,
                payment_methods=payment_methods
            )

        if not income_date:
            flash("Income date is required.", "error")
            cur.close()
            conn.close()
            return render_template("add_general_income.html", clients=clients, income_types=income_types, payment_methods=payment_methods)

        if not income_type:   
            flash("Income type is required.", "error")
            cur.close()
            conn.close()
            return render_template("add_general_income.html", clients=clients, income_types=income_types, payment_methods=payment_methods)

        if not payment_method:
            flash("Payment method is required.", "error")
            cur.close()
            conn.close() 
            return render_template("add_general_income.html", clients=clients, income_types=income_types, payment_methods=payment_methods)

        if total_amount < 0:
            flash("Total amount cannot be negative.", "error")
            cur.close()
            conn.close()
            return render_template("add_general_income.html", clients=clients, income_types=income_types, payment_methods=payment_methods)

        if client_id == "":
            client_id = None

        # Safety check: client must belong to this spa
        if client_id:
            cur.execute("""
                SELECT client_id
                FROM clients
                WHERE client_id = %s
                  AND spa_id = %s
            """, (client_id, spa_id))

            if not cur.fetchone():
                flash("Selected client does not belong to this spa.", "error")
                cur.close()
                conn.close()
                return render_template(
                    "add_general_income.html",
                    clients=clients,
                    income_types=income_types,
                    payment_methods=payment_methods
                )

        cur.execute("""
            INSERT INTO income (
                spa_id,
                income_date,
                client_id,
                appointment_id,
                visit_id,
                income_type,
                description,
                service_amount,
                retail_amount,  
                tax_amount,
                total_amount,
                payment_method,
                square_payment_id, 
                notes
            )
            VALUES (%s, %s, %s, NULL, NULL, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            spa_id,
            income_date,
            client_id,
            income_type,   
            description,
            service_amount,
            retail_amount,
            tax_amount,
            total_amount,
            payment_method,
            square_payment_id,
            notes
        ))   
    
        conn.commit()
        flash("General income entry added successfully.", "success")
            
        cur.close()
        conn.close()
        return redirect(url_for("add_general_income"))
                
    cur.close()
    conn.close()
             
    return render_template(
        "add_general_income.html",
        clients=clients,
        income_types=income_types,
        payment_methods=payment_methods
    )



#  -----------------------------
#
#
#     CALENDAR
#
#  -----------------------------



from datetime import timedelta, datetime
from flask import render_template, request, redirect


@app.route("/calendar")
@login_required  
@spa_required
def calendar_view():
    spa_id = current_spa_id()
    role = session.get("role")

    week_start_str = request.args.get("week_start")
    goto_date = request.args.get("goto_date")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
        
    spa_now = get_spa_now()
    today = spa_now.date()
    now_time = spa_now.time()
    current_timezone = get_current_spa_timezone()
    
    today_days_since_sunday = (today.weekday() + 1) % 7
    current_week_start = today - timedelta(days=today_days_since_sunday)
        
    if goto_date:
        selected_date = datetime.strptime(goto_date, "%Y-%m-%d").date()
        days_since_sunday = (selected_date.weekday() + 1) % 7
        start_of_week = selected_date - timedelta(days=days_since_sunday)
    elif week_start_str:
        start_of_week = datetime.strptime(week_start_str, "%Y-%m-%d").date()
    else:
        start_of_week = current_week_start

    week_days = [start_of_week + timedelta(days=i) for i in range(7)]
            
    prev_week_start = start_of_week - timedelta(days=7)
    next_week_start = start_of_week + timedelta(days=7)

    conn = get_db_connection()
    cur = conn.cursor()
                    
    filtered_appointments = []

    # Base spa filters
    filter_sql = "WHERE a.appointment_date BETWEEN %s AND %s"
    week_params = [week_days[0], week_days[-1]]


    filter_sql += " AND a.spa_id = %s"
    week_params.append(spa_id)

    if start_date and end_date:
        date_filter_sql = "WHERE a.appointment_date BETWEEN %s AND %s"
        date_params = [start_date, end_date]

        date_filter_sql += " AND a.spa_id = %s"
        date_params.append(spa_id)

        
        cur.execute(f"""
            SELECT
                a.appointment_date,                             -- 0
                a.appointment_time,                             -- 1
                c.first_name,                                   -- 2
                c.last_name,                                    -- 3
                COALESCE(
                    NULLIF(snt.service_name, ''),
                    NULLIF(a.service_type, ''),
                    NULLIF(a.external_service_name, ''),
                    s.service_name,
                    'Service not entered'
                ) AS service_name,                              -- 4
                a.status,                                       -- 5
                a.appointment_id,                               -- 6
                a.duration_minutes,                             -- 7
                a.price_at_booking,                             -- 8
                a.owner_reviewed,                               -- 9
                a.owner_reviewed_at                             -- 10
            FROM appointments a
            JOIN clients c
                ON a.client_id = c.client_id
            AND a.spa_id = c.spa_id

            LEFT JOIN service_name_types snt
                ON a.service_type_id = snt.service_type_id
            AND a.spa_id = snt.spa_id

            LEFT JOIN services s
                ON a.service_id = s.service_id
            AND a.spa_id = s.spa_id

            {date_filter_sql}

            ORDER BY a.appointment_date, a.appointment_time
        """, week_params)

        filtered_appointments = cur.fetchall()
            
    # Show booked appointments for the displayed week
    cur.execute(f"""
        SELECT
            a.appointment_date,                             -- 0
            a.appointment_time,                             -- 1
            c.first_name,                                   -- 2
            c.last_name,                                    -- 3
            COALESCE(
                NULLIF(snt.service_name, ''),
                NULLIF(a.service_type, ''),
                NULLIF(a.external_service_name, ''),
                s.service_name,
                'Service not entered'
            ) AS service_name,                              -- 4
            a.status,                                       -- 5
            a.appointment_id,                               -- 6
            a.duration_minutes,                             -- 7
            a.price_at_booking,                             -- 8
            a.owner_reviewed,                               -- 9
            a.owner_reviewed_at                             -- 10
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
        AND a.spa_id = c.spa_id

        LEFT JOIN service_name_types snt
            ON a.service_type_id = snt.service_type_id
        AND a.spa_id = snt.spa_id

        LEFT JOIN services s
            ON a.service_id = s.service_id
        AND a.spa_id = s.spa_id

        {filter_sql}

        ORDER BY a.appointment_date, a.appointment_time
    """, week_params)

    appointments = cur.fetchall()
    
    # Next booked appointment banner
    next_filter_sql = """
        WHERE a.status = 'booked'
          AND (
                a.appointment_date > %s
                OR (
                    a.appointment_date = %s
                    AND a.appointment_time >= %s
                )
              )
    """
    next_params = [today, today, now_time]

    next_filter_sql += " AND a.spa_id = %s"
    next_params.append(spa_id)

    
    cur.execute(f"""
        SELECT
            c.first_name,                                   -- 0
            c.last_name,                                    -- 1
            a.appointment_date,                             -- 2
            a.appointment_time,                             -- 3
            COALESCE(
                NULLIF(snt.service_name, ''),
                NULLIF(a.service_type, ''),
                NULLIF(a.external_service_name, ''),
                s.service_name,
                'Service not entered'
            ) AS service_name                               -- 4
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
        AND a.spa_id = c.spa_id

        LEFT JOIN service_name_types snt
            ON a.service_type_id = snt.service_type_id
        AND a.spa_id = snt.spa_id

        LEFT JOIN services s
            ON a.service_id = s.service_id
        AND a.spa_id = s.spa_id

        {next_filter_sql}

        ORDER BY a.appointment_date, a.appointment_time
        LIMIT 1
    """, next_params)

    next_appt = cur.fetchone()
            
    # Overdue booked appointment count
    overdue_filter_sql = """
        WHERE status = 'booked'
          AND (
                appointment_date < %s
                OR (
                    appointment_date = %s
                    AND appointment_time < %s
                )
              )
    """
    overdue_params = [today, today, now_time]

    overdue_filter_sql += " AND spa_id = %s"
    overdue_params.append(spa_id)

    cur.execute(f"""
        SELECT COUNT(*)
        FROM appointments
        {overdue_filter_sql}
    """, overdue_params)

    overdue_count = cur.fetchone()[0]  
                
    cur.close()
    conn.close()
              
    formatted_spa_time = spa_now.strftime("%A, %B %d, %Y %I:%M %p")
        
    return render_template(
        "calendar.html",
        week_days=week_days,
        appointments=appointments,
        today=today,
        now_time=now_time,
        spa_now=spa_now, 
        formatted_spa_time=formatted_spa_time,
        current_timezone=current_timezone,
        next_appt=next_appt,
        overdue_count=overdue_count,
        goto_date=goto_date,
        start_of_week=start_of_week,
        current_week_start=current_week_start,
        prev_week_start=prev_week_start,
        next_week_start=next_week_start
    )








#  -----------------------------
#
#
#   MODAL QUICK RESCHEDULE
#  6/2/26
#  -----------------------------

@app.route("/quick_reschedule_appointment/<int:appointment_id>", methods=["POST"])
@login_required
@spa_required
def quick_reschedule_appointment(appointment_id):
    spa_id = current_spa_id()
    user_id = session.get("user_id")
    role = session.get("role")
        
    appointment_date = request.form.get("appointment_date")
    appointment_time = request.form.get("appointment_time")
        
    if not appointment_date or not appointment_time:
        flash("Date and time are required to reschedule.", "error")
        return redirect(url_for("calendar_view"))   
        
    conn = get_db_connection()
    cur = conn.cursor()

    filter_sql = "WHERE appointment_id = %s"
    params = [appointment_id]

    if role != "master_admin":
        filter_sql += " AND spa_id = %s"
        params.append(spa_id)

    cur.execute(f"""
        SELECT spa_id, client_id, appointment_date, appointment_time, status
        FROM appointments
        {filter_sql}
    """, params)

    old_appt = cur.fetchone()

    if not old_appt:
        cur.close()
        conn.close()
        flash("Appointment not found or not authorized.", "error")
        return redirect(url_for("calendar_view"))

    appointment_spa_id = old_appt[0]
    client_id = old_appt[1]
    old_date = old_appt[2]
    old_time = old_appt[3]
    old_status = old_appt[4]
        
    cur.execute(f"""
        UPDATE appointments
        SET appointment_date = %s,
            appointment_time = %s,
            updated_at = CURRENT_TIMESTAMP
        {filter_sql}
    """, (
        appointment_date,
        appointment_time,
        *params
    ))

    if cur.rowcount == 0:
        conn.rollback()
        cur.close()
        conn.close()
        flash("Appointment not found or not authorized.", "error")
        return redirect(url_for("calendar_view"))

    action_type = "rescheduled"

    if str(old_date) == appointment_date and str(old_time)[:5] == appointment_time[:5]:
        action_type = "updated"

    log_audit(
        cur,
        spa_id=appointment_spa_id,
        user_id=user_id,
        action_type=f"appointment_{action_type}",
        table_name="appointments",
        record_id=appointment_id,
        old_value=f"{old_date} {old_time} {old_status}",
        new_value=f"{appointment_date} {appointment_time} {old_status}",
        notes="Appointment quick rescheduled" if action_type == "rescheduled" else "Appointment quick updated"
    )

    log_appointment_history(
        cur,
        spa_id=appointment_spa_id,
        appointment_id=appointment_id,
        client_id=client_id,
        user_id=user_id,
        action_type=action_type,
        old_date=old_date,
        old_time=old_time,
        new_date=appointment_date,
        new_time=appointment_time,
        old_status=old_status,
        new_status=old_status,
        notes="Appointment quick rescheduled" if action_type == "rescheduled" else "Appointment quick updated"
    )

    conn.commit()
    cur.close()
    conn.close()
    
    flash("Appointment rescheduled successfully.", "success")
    return redirect(url_for("calendar_view"))








#  -----------------------------
#
#   MODAL ROUTE
#. APPOINTMENT COMMAND MODAL
#   CALENDAR MODAL
#      APPOINTMENT DETAILS
#     
#  4/28 cleaned
#  -----------------------------


     
@app.route("/appointment-details/<int:appointment_id>")
@login_required
@spa_required
def appointment_details(appointment_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            a.appointment_id,
            a.appointment_date,
            a.appointment_time,
            a.status,
            c.first_name,
            c.last_name,
            c.phone,
            c.email,
            COALESCE(
                NULLIF(snt.service_name, ''),
                NULLIF(a.service_type, ''),
                NULLIF(a.external_service_name, ''),
                s.service_name,
                'Service not entered'
            ) AS service_name,
            a.notes,
            a.external_source,
            a.duration_minutes,
            a.price_at_booking
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
        AND a.spa_id = c.spa_id
        LEFT JOIN service_name_types snt
            ON a.service_type_id = snt.service_type_id
        AND a.spa_id = snt.spa_id
        LEFT JOIN services s
            ON a.service_id = s.service_id
        AND a.spa_id = s.spa_id
        WHERE a.appointment_id = %s
        AND a.spa_id = %s
    """, (appointment_id, spa_id))

    appt = cur.fetchone()

    if not appt:
        cur.close()
        conn.close()
        return {"error": "Appointment not found."}, 404

    # Opening the Appointment Command Center counts as owner review.
    cur.execute("""
        UPDATE appointments
        SET
            import_reviewed = TRUE,
            import_reviewed_at = COALESCE(import_reviewed_at, NOW()),
            import_reviewed_by = COALESCE(import_reviewed_by, %s),
            owner_reviewed = TRUE,
            owner_reviewed_at = COALESCE(owner_reviewed_at, NOW()),
            import_status = 'Reviewed'
        WHERE appointment_id = %s
        AND spa_id = %s
        AND LOWER(COALESCE(external_source, '')) = 'godaddy'
        AND (
                COALESCE(import_reviewed, FALSE) = FALSE
            OR COALESCE(owner_reviewed, FALSE) = FALSE
            OR COALESCE(import_status, '') <> 'Reviewed'
        )
    """, (
        session.get("user_id"),
        appointment_id,
        spa_id
    ))

    conn.commit()

    cur.close()
    conn.close()

    return {
        "appointment_id": appt[0],
        "appointment_date": appt[1].strftime("%Y-%m-%d") if appt[1] else "",
        "start_time": appt[2].strftime("%I:%M %p") if appt[2] else "",
        "raw_time": appt[2].strftime("%H:%M") if appt[2] else "",
        "status": appt[3] or "",
        "client_name": f"{appt[4]} {appt[5]}",
        "phone": appt[6] or "",
        "email": appt[7] or "",
        "service_name": appt[8] or "",
        "provider_name": "",
        "notes": appt[9] or "",
        "external_source": appt[10] or "",
        "duration_minutes": appt[11],
        "price_at_booking": appt[12]
    }






#  -----------------------------
#     
#     
#     DAILY SCHEDULE
#  4/27
#  -----------------------------



from datetime import datetime, date

@app.route("/daily_schedule")
@login_required
@spa_required
def daily_schedule():
    spa_id = current_spa_id()
    role = session.get("role")

    selected_date = request.args.get("date")
    
    if selected_date:
        display_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
    else:
        display_date = get_spa_today()
        
    conn = get_db_connection()
    cur = conn.cursor()

    filter_sql = "WHERE a.appointment_date = %s"
    params = [display_date]

    if role != "master_admin":
        filter_sql += " AND a.spa_id = %s"
        params.append(spa_id)
            
    cur.execute(f"""
        SELECT
            a.appointment_id,                              -- 0
            a.client_id,                                   -- 1
            c.first_name,                                  -- 2
            c.last_name,                                   -- 3
            COALESCE(a.service_type, s.service_name),       -- 4
            a.appointment_time,                            -- 5
            a.duration_minutes,                            -- 6
            a.room_number,                                 -- 7
            a.status,                                      -- 8
            a.notes,                                       -- 9
            a.price_at_booking                             -- 10
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
        AND a.spa_id = c.spa_id
        LEFT JOIN services s
            ON a.service_id = s.service_id
        AND a.spa_id = s.spa_id
        {filter_sql}
        ORDER BY a.appointment_time
    """, params)

    appointments = cur.fetchall()
        
    cur.close()
    conn.close()
            
    return render_template(   
        "daily_schedule.html",
        appointments=appointments,
        display_date=display_date
    ) 






#  -----------------------------
#    
#
#     DASHBOARD
#    4/28 cleaned
#  -----------------------------



from flask import render_template
from db import get_db_connection
from datetime import date, timedelta




@app.route("/dashboard")
@login_required
@spa_required
def dashboard():

    if session.get("role") == "master_admin":
       return redirect(url_for("feedback_admin"))

    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    spa_now = get_spa_now()
    today = spa_now.date()
    year_start = date(today.year, 1, 1)
    now_time = spa_now.time()

    # Safe defaults
    birthday_alert_count = 0
    expiring_gc_count = 0
    next_appt = None
    today_count = 0
    upcoming_count = 0
    completed_count = 0
    total_clients = 0
    revenue_today = 0.00
    todays_appointments = []
    upcoming_appointments = []
    top_services = []

    # Total clients
    cur.execute("""
        SELECT COUNT(*)
        FROM clients
        WHERE spa_id = %s
    """, (spa_id,))
    total_clients = cur.fetchone()[0] or 0

    # Revenue today
    cur.execute("""
        SELECT COALESCE(SUM(total_amount), 0)
        FROM income
        WHERE spa_id = %s
          AND income_date = %s
    """, (spa_id, today))
    revenue_today = cur.fetchone()[0] or 0.00

    # Birthday alert count
    # Adjust this query if your birthday table/logic differs
    cur.execute("""
        SELECT COUNT(*)
        FROM client_birthday_offers cbo
        JOIN clients c 
            ON c.client_id = cbo.client_id
           AND c.spa_id = cbo.spa_id
        WHERE c.spa_id = %s
          AND cbo.offer_sent = FALSE
          AND cbo.birthday_year = %s
    """, (spa_id, today.year))
    birthday_alert_count = cur.fetchone()[0] or 0

    # Expiring gift certificate count
    cur.execute("""
        SELECT COUNT(*)
        FROM gift_certificates gc 
        JOIN gift_certificate_statuses gcs
        ON gc.gift_certificate_status_id = gcs.gift_certificate_status_id
        WHERE gc.spa_id = %s
        AND gcs.status_name = 'Active'
        AND gc.amount_paid > 0
        AND gc.is_redeemed = FALSE
        AND gc.remaining_balance > 0
        AND gc.expires_date BETWEEN %s AND (%s + INTERVAL '60 days')
    """, (spa_id, today, today))
    expiring_gc_count = cur.fetchone()[0] or 0

    # Next appointment
    cur.execute("""
        SELECT
            a.appointment_id,
            c.first_name,
            c.last_name,
            a.appointment_date,
            a.appointment_time,
            s.service_name
        FROM appointments a
        JOIN clients c 
            ON a.client_id = c.client_id
           AND a.spa_id = c.spa_id
        LEFT JOIN services s 
            ON a.service_id = s.service_id
           AND a.spa_id = s.spa_id
        WHERE a.spa_id = %s
          AND a.status = 'booked'
          AND (
                a.appointment_date > %s
                OR (
                    a.appointment_date = %s
                    AND a.appointment_time >= %s
                )
              )
        ORDER BY a.appointment_date, a.appointment_time
        LIMIT 1
    """, (spa_id, today, today, now_time))
    row = cur.fetchone()
    if row:
        # Make it match template indexes:
        # next_appt[0] first_name
        # next_appt[1] last_name
        # next_appt[2] appointment_date
        # next_appt[3] appointment_time
        # next_appt[4] service_name
        next_appt = (row[1], row[2], row[3], row[4], row[5])

    # Today's appointments count
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE spa_id = %s
          AND appointment_date = %s
    """, (spa_id, today))
    today_count = cur.fetchone()[0] or 0

    # Upcoming appointments count
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE spa_id = %s
          AND status = 'booked'
          AND (
                appointment_date > %s
                OR (
                    appointment_date = %s
                    AND appointment_time > %s
                )
              )
    """, (spa_id, today, today, now_time))
    upcoming_count = cur.fetchone()[0] or 0

    # Completed appointments count
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE spa_id = %s
          AND status = 'completed'
    """, (spa_id,))
    completed_count = cur.fetchone()[0] or 0


    cur.execute("""
        SELECT COALESCE(SUM(total_amount), 0)
        FROM income
        WHERE spa_id = %s
          AND income_date BETWEEN %s AND %s
    """, (spa_id, year_start, today))

    ytd_income = cur.fetchone()[0]


    cur.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM expenses
        WHERE spa_id = %s
          AND expense_date BETWEEN %s AND %s
    """, (spa_id, year_start, today))

    ytd_expenses = cur.fetchone()[0]


    cur.execute("""
        SELECT COALESCE(SUM(amount), 0)
        FROM employee_compensation
        WHERE spa_id = %s
          AND compensation_date BETWEEN %s AND %s
    """, (spa_id, year_start, today))

    ytd_employee_compensation = cur.fetchone()[0]


    # Today's appointment table
    cur.execute("""
        SELECT
            a.appointment_id,      -- row[0]
            c.first_name,          -- row[1]
            c.last_name,           -- row[2]
            s.service_name,        -- row[3]
            a.appointment_date,    -- row[4]
            a.appointment_time,    -- row[5]
            a.status               -- row[6]
        FROM appointments a
        JOIN clients c ON a.client_id = c.client_id
        LEFT JOIN services s ON a.service_id = s.service_id
        WHERE a.spa_id = %s
          AND a.appointment_date = %s
        ORDER BY a.appointment_time
    """, (spa_id, today))
    todays_appointments = cur.fetchall()

    # Upcoming appointments table
    cur.execute("""
        SELECT
            a.appointment_id,      -- row[0]
            c.first_name,          -- row[1]
            c.last_name,           -- row[2]
            s.service_name,        -- row[3]
            a.appointment_date,    -- row[4]
            a.appointment_time,    -- row[5]
            a.status               -- row[6]
        FROM appointments a
        JOIN clients c ON a.client_id = c.client_id
        LEFT JOIN services s ON a.service_id = s.service_id
        WHERE a.spa_id = %s
          AND (
                a.appointment_date > %s
                OR (
                    a.appointment_date = %s
                    AND a.appointment_time > %s
                )
              )
        ORDER BY a.appointment_date, a.appointment_time
        LIMIT 10
    """, (spa_id, today, today, now_time))
    upcoming_appointments = cur.fetchall()

    # Top services
    cur.execute("""
        SELECT
            COALESCE(s.service_name, 'Unknown Service') AS service_name,
            COUNT(*) AS booked_count
        FROM appointments a
        LEFT JOIN services s ON a.service_id = s.service_id
        WHERE a.spa_id = %s
        GROUP BY COALESCE(s.service_name, 'Unknown Service')
        ORDER BY booked_count DESC, service_name ASC
        LIMIT 5
    """, (spa_id,))
    top_services = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "home.html",
        ytd_income=ytd_income,
        ytd_expenses=ytd_expenses,
        ytd_employee_compensation=ytd_employee_compensation,
        birthday_alert_count=birthday_alert_count,
        expiring_gc_count=expiring_gc_count,
        next_appt=next_appt,
        today_count=today_count,
        upcoming_count=upcoming_count,
        completed_count=completed_count,
        total_clients=total_clients,
        revenue_today=revenue_today,
        todays_appointments=todays_appointments,
        upcoming_appointments=upcoming_appointments,
        top_services=top_services
    )



        
#  ------------------------------
#       
#      MORNING BRIEFING
#  
#  -----------------------------
   



@app.route("/morning_briefing")
@login_required
@spa_required
def morning_briefing():

    spa_id = session["spa_id"]
    user_id = session["user_id"]

    spa_now = get_spa_now(spa_id)
    today = spa_now.date()
    now_time = spa_now.time()

    dashboard = get_dashboard_data(
    spa_id,
    spa_now=spa_now
)

    business_health = {
        "score": None,
        "label": "Not Calculated"
    }

    conn = get_db_connection()
    cur = conn.cursor()


    # ---------------------------------------------------------
    # Start or resume today's Coach session
    # ---------------------------------------------------------
    coach_session = get_or_create_coach_daily_session(
        cur=cur,
        spa_id=spa_id,
        user_id=user_id,
        session_date=today
    )


    # ---------------------------------------------------------
    # Business schedule items due now
    # ---------------------------------------------------------
    cur.execute("""
        SELECT
            schedule_id,
            category,
            title,
            description,
            due_date,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            is_required
        FROM business_schedule
        WHERE spa_id = %s
          AND is_active = TRUE
          AND COALESCE(is_completed, FALSE) = FALSE
          AND due_date IS NOT NULL
          AND due_date <= %s
        ORDER BY
            due_date ASC,
            category,
            title
        LIMIT 5
    """, (spa_id, today))

    business_schedule_due = cur.fetchall()

    # ---------------------------------------------------------
    # Business schedule items coming up within 14 days
    # ---------------------------------------------------------
    cur.execute("""
        SELECT
            schedule_id,
            category,
            title,
            description,
            due_date,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            is_required
        FROM business_schedule
        WHERE spa_id = %s
          AND is_active = TRUE
          AND COALESCE(is_completed, FALSE) = FALSE
          AND due_date IS NOT NULL
          AND due_date > %s
          AND due_date <= %s + INTERVAL '14 days'
        ORDER BY
            due_date ASC,
            category,
            title
        LIMIT 5
    """, (spa_id, today, today))

    business_schedule_upcoming = cur.fetchall()

    # ---------------------------------------------------------
    # Priority action review
    # ---------------------------------------------------------
    priority_actions = []

    if dashboard.get("birthdays_today", 0) > 0:
        priority_actions.append({
            "icon": "🎂",
            "label": "Birthday today",
            "value": dashboard.get("birthdays_today", 0),
            "url": None,
            "category": "Client Care",
            "priority": 80
        })

    if dashboard.get("appointments_tomorrow", 0) > 0:
        priority_actions.append({
            "icon": "📅",
            "label": "Appointments tomorrow",
            "value": dashboard.get("appointments_tomorrow", 0),
            "url": url_for("calendar_view"),
            "category": "Appointments",
            "priority": 60
        })

    # ---------------------------------------------------------
    # GoDaddy imports awaiting review
    # ---------------------------------------------------------
    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE spa_id = %s
          AND external_source = 'godaddy'
          AND COALESCE(import_reviewed, FALSE) = FALSE
    """, (spa_id,))

    godaddy_unreviewed_count = cur.fetchone()[0] or 0

    if godaddy_unreviewed_count > 0:
        priority_actions.append({
            "icon": "📨",
            "label": "GoDaddy imports need review",
            "value": godaddy_unreviewed_count,
            "url": url_for("godaddy_imports"),
            "category": "Appointments",
            "priority": 90
        })


    # ---------------------------------------------------------
    # Overdue appointments awaiting closeout
    # ---------------------------------------------------------
    cur.execute("""
        SELECT
            a.appointment_id,
            a.appointment_date,
            a.appointment_time,
            a.status,
            c.first_name,
            c.last_name
        FROM appointments a
        LEFT JOIN clients c
            ON c.client_id = a.client_id
        WHERE a.spa_id = %s
        AND a.appointment_date < %s
        AND LOWER(COALESCE(a.status, '')) = 'booked'
        ORDER BY
            a.appointment_date ASC,
            a.appointment_time ASC
    """, (spa_id, today))

    overdue_appointments = cur.fetchall()

    dashboard["overdue_appointments"] = overdue_appointments


   # ---------------------------------------------------------
    # Coach performs the review only after all data is collected
    # ---------------------------------------------------------
    coach = build_coach(
        dashboard=dashboard,
        business_schedule_due=business_schedule_due,
        business_schedule_upcoming=business_schedule_upcoming,
        priority_actions=priority_actions,
        spa_now=spa_now,
        coach_session=coach_session
    )


    record_coach_interaction(
        cur=cur,
        coach_session_id=coach_session["coach_session_id"],
        spa_id=spa_id,
        user_id=user_id,
        coach=coach,
        message_type="briefing_open"
    )


    action_cards = build_action_cards(
        dashboard=dashboard,
        priority_actions=priority_actions
    )

    conn.commit()



    cur.close()
    conn.close()

    return render_template(
        "morning_briefing.html",
        dashboard=dashboard,
        business_health=business_health,
        today=today,
        business_schedule_due=business_schedule_due,
        business_schedule_upcoming=business_schedule_upcoming,
        priority_actions=priority_actions,
        godaddy_unreviewed_count=godaddy_unreviewed_count,
        coach=coach,
        coach_session=coach_session,
        spa_now=spa_now,
        action_cards=action_cards
    )







######################################
#
#   DAILY BRIEFING - - TODAY
#
#
###########################################


@app.route("/daily-briefing/today")
@login_required
@spa_required
def daily_briefing_today():

    spa_id = session["spa_id"]
    spa_now = get_spa_now(spa_id)
    today = spa_now.date()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            a.appointment_id,
            a.appointment_date,
            a.appointment_time,
            a.service_type,
            a.duration_minutes,
            a.price_at_booking,
            a.status,
            c.client_id,
            c.first_name,
            c.last_name,
            c.phone,
            c.email
        FROM appointments a
        LEFT JOIN clients c
            ON a.client_id = c.client_id
        WHERE a.spa_id = %s
          AND a.appointment_date = %s
          AND COALESCE(a.status, '') NOT IN ('Cancelled', 'Canceled')
        ORDER BY a.appointment_time ASC
    """, (spa_id, today))

    appointments = cur.fetchall()

    projected_revenue = sum(
        appt[5] or 0 for appt in appointments
    )

    appointment_count = len(appointments)

    cur.close()
    conn.close()

    return render_template(
        "daily_briefing_today.html",
        today=today,
        appointments=appointments,
        appointment_count=appointment_count,
        projected_revenue=projected_revenue,
        action_url=url_for("priority_actions")
    )









####################################
#
#   BUSINESS COACH PROFILE
#
#
##########################################



@app.route("/business-coach/profile", methods=["GET", "POST"])
@login_required
@spa_required
def business_coach_profile():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        business_type = request.form.get("business_type")
        business_description = request.form.get("business_description")
        business_address = request.form.get("business_address")
        city = request.form.get("city")
        state = request.form.get("state")
        zip_code = request.form.get("zip_code")
        service_area = request.form.get("service_area")
        business_hours = request.form.get("business_hours")

        website_url = request.form.get("website_url")
        facebook_url = request.form.get("facebook_url")
        instagram_url = request.form.get("instagram_url")
        yelp_url = request.form.get("yelp_url")
        google_business_url = request.form.get("google_business_url")

        website_intro = request.form.get("website_intro")
        facebook_intro = request.form.get("facebook_intro")
        instagram_intro = request.form.get("instagram_intro")
        yelp_intro = request.form.get("yelp_intro")

        has_business_cards = bool(request.form.get("has_business_cards"))
        has_flyers = bool(request.form.get("has_flyers"))
        has_referral_cards = bool(request.form.get("has_referral_cards"))
        has_gift_certificates = bool(request.form.get("has_gift_certificates"))
        has_promotional_banner = bool(request.form.get("has_promotional_banner"))
        has_signage = bool(request.form.get("has_signage"))

        primary_goal = request.form.get("primary_goal")

        cur.execute("""
            INSERT INTO business_coach_profiles (
                spa_id,
                business_type,
                business_description,
                business_address,
                city,
                state,
                zip_code,
                service_area,
                business_hours,
                website_url,
                facebook_url,
                instagram_url,
                yelp_url,
                google_business_url,
                website_intro,
                facebook_intro,
                instagram_intro,
                yelp_intro,
                has_business_cards,
                has_flyers,
                has_referral_cards,
                has_gift_certificates,
                has_promotional_banner,
                has_signage,
                primary_goal,
                updated_at
            )
            VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, NOW()
            )
            ON CONFLICT (spa_id)
            DO UPDATE SET
                business_type = EXCLUDED.business_type,
                business_description = EXCLUDED.business_description,
                business_address = EXCLUDED.business_address,
                city = EXCLUDED.city,
                state = EXCLUDED.state,
                zip_code = EXCLUDED.zip_code,
                service_area = EXCLUDED.service_area,
                business_hours = EXCLUDED.business_hours,
                website_url = EXCLUDED.website_url,
                facebook_url = EXCLUDED.facebook_url,
                instagram_url = EXCLUDED.instagram_url,
                yelp_url = EXCLUDED.yelp_url,
                google_business_url = EXCLUDED.google_business_url,
                website_intro = EXCLUDED.website_intro,
                facebook_intro = EXCLUDED.facebook_intro,
                instagram_intro = EXCLUDED.instagram_intro,
                yelp_intro = EXCLUDED.yelp_intro,
                has_business_cards = EXCLUDED.has_business_cards,
                has_flyers = EXCLUDED.has_flyers,
                has_referral_cards = EXCLUDED.has_referral_cards,
                has_gift_certificates = EXCLUDED.has_gift_certificates,
                has_promotional_banner = EXCLUDED.has_promotional_banner,
                has_signage = EXCLUDED.has_signage,
                primary_goal = EXCLUDED.primary_goal,
                updated_at = NOW()
        """, (
            spa_id,
            business_type,
            business_description,
            business_address,
            city,
            state,
            zip_code,
            service_area,
            business_hours,
            website_url,
            facebook_url,
            instagram_url,
            yelp_url,
            google_business_url,
            website_intro,
            facebook_intro,
            instagram_intro,
            yelp_intro,
            has_business_cards,
            has_flyers,
            has_referral_cards,
            has_gift_certificates,
            has_promotional_banner,
            has_signage,
            primary_goal
        ))

        conn.commit()
        flash("Business Coach Profile saved.", "success")
        return redirect(url_for("business_coach_profile"))

    cur.execute("""
        SELECT *
        FROM business_coach_profiles
        WHERE spa_id = %s
    """, (spa_id,))

    profile = cur.fetchone()

    cur.close()
    conn.close()

    return render_template(
        "business_coach_profile.html",
        profile=profile
    )




###################################
#
#   BUSINESS SCHEDULE
#
#
#####################################


@app.route("/business-schedule", methods=["GET", "POST"])
@login_required
@spa_required
def business_schedule():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    show_archived = request.args.get("show_archived") == "1"

    if request.method == "POST":
        category = request.form.get("category")
        title = request.form.get("title")
        description = request.form.get("description")
        due_date = request.form.get("due_date") or None
        recurrence_type = request.form.get("recurrence_type")
        recurrence_interval = request.form.get("recurrence_interval") or 1
        reminder_days = request.form.get("reminder_days") or 14
        is_required = bool(request.form.get("is_required"))

        cur.execute("""
            INSERT INTO business_schedule (
                spa_id,
                category,
                title,
                description,
                due_date,
                recurrence_type,
                recurrence_interval,
                reminder_days,
                is_required
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            spa_id,
            category,
            title,
            description,
            due_date,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            is_required
        ))

        conn.commit()
        flash("Business Schedule item added.", "success")
        return redirect(url_for("business_schedule"))

        

    if show_archived:
        cur.execute("""
            SELECT
                schedule_id,
                category,
                title,
                description,
                due_date,
                recurrence_type,
                recurrence_interval,
                reminder_days,
                is_required,
                is_completed,
                is_active
            FROM business_schedule
            WHERE spa_id = %s
            ORDER BY
                is_active DESC,
                due_date ASC NULLS LAST,
                category,
                title
        """, (spa_id,))
    else:
        cur.execute("""
            SELECT
                schedule_id,
                category,
                title,
                description,
                due_date,
                recurrence_type,
                recurrence_interval,
                reminder_days,
                is_required,
                is_completed,
                is_active
            FROM business_schedule
            WHERE spa_id = %s
              AND is_active = TRUE
            ORDER BY
                due_date ASC NULLS LAST,
                category,
                title
        """, (spa_id,))

    schedule_items = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "business_schedule.html",
        schedule_items=schedule_items,
        show_archived=show_archived
    )





######################################
#
#   BUSINESS SCHEDULE BUILD
#
#
#########################################




@app.route("/business-schedule/build", methods=["POST"])
@login_required
@spa_required
def build_business_schedule():
    spa_id = current_spa_id()

    recommended_items = [
        ("Financial", "Texas Sales Tax Report", "Review and file Texas sales tax report.", "Monthly", 1, 14),
        ("Financial", "Quarterly Estimated Tax Review", "Review estimated tax payments with your accountant.", "Quarterly", 1, 21),
        ("Financial", "Monthly Financial Review", "Review income, expenses, and profit trends.", "Monthly", 1, 7),

        ("Licensing", "Professional License Renewal", "Review professional license renewal requirements.", "Yearly", 1, 60),
        ("Insurance", "Business Insurance Review", "Review liability and business insurance coverage.", "Yearly", 1, 30),

        ("Operations", "Monthly Inventory Count", "Review retail and backbar inventory levels.", "Monthly", 1, 7),
        ("Operations", "Equipment Maintenance Review", "Check treatment equipment, supplies, and maintenance needs.", "Quarterly", 1, 14),

        ("Marketing", "Google Review Reminder", "Ask recent happy clients for Google reviews.", "Weekly", 1, 2),
        ("Marketing", "Facebook / Instagram Post Review", "Review recent social activity and plan new posts.", "Weekly", 1, 2),
        ("Marketing", "Monthly Promotion Planning", "Plan next month's featured service or retail promotion.", "Monthly", 1, 10),

        ("Growth", "Quarterly Pricing Review", "Review service pricing and compare against business goals.", "Quarterly", 1, 21),
        ("Growth", "Website / Online Presence Review", "Review website, Google profile, and social media presence.", "Quarterly", 1, 21),
    ]

    conn = get_db_connection()
    cur = conn.cursor()

    for category, title, description, recurrence_type, recurrence_interval, reminder_days in recommended_items:
        cur.execute("""
            INSERT INTO business_schedule (
                spa_id,
                category,
                title,
                description,
                recurrence_type,
                recurrence_interval,
                reminder_days,
                is_required,
                is_active
            )
            SELECT %s, %s, %s, %s, %s, %s, %s, TRUE, TRUE
            WHERE NOT EXISTS (
                SELECT 1
                FROM business_schedule
                WHERE spa_id = %s
                  AND title = %s
                  AND is_active = TRUE
            )
        """, (
            spa_id,
            category,
            title,
            description,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            spa_id,
            title
        ))

    conn.commit()
    cur.close()
    conn.close()

    flash("Recommended Business Schedule created.", "success")
    return redirect(url_for("business_schedule"))




##################################
#
#   EDIT BUSINESS SCHEDULE ITEM
#
#
##################################


@app.route("/business-schedule/edit/<int:schedule_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_business_schedule(schedule_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        category = request.form.get("category")
        title = request.form.get("title")
        description = request.form.get("description")
        due_date = request.form.get("due_date") or None
        recurrence_type = request.form.get("recurrence_type")
        recurrence_interval = request.form.get("recurrence_interval") or 1
        reminder_days = request.form.get("reminder_days") or 14
        is_required = bool(request.form.get("is_required"))

        cur.execute("""
            UPDATE business_schedule
            SET
                category = %s,
                title = %s,
                description = %s,
                due_date = %s,
                recurrence_type = %s,
                recurrence_interval = %s,
                reminder_days = %s,
                is_required = %s,
                updated_at = NOW()
            WHERE schedule_id = %s
              AND spa_id = %s
        """, (
            category,
            title,
            description,
            due_date,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            is_required,
            schedule_id,
            spa_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Business Schedule item updated.", "success")
        return redirect(url_for("business_schedule"))

    cur.execute("""
        SELECT
            schedule_id,
            category,
            title,
            description,
            due_date,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            is_required
        FROM business_schedule
        WHERE schedule_id = %s
          AND spa_id = %s
    """, (schedule_id, spa_id))

    item = cur.fetchone()

    cur.close()
    conn.close()

    if not item:
        flash("Business Schedule item not found.", "warning")
        return redirect(url_for("business_schedule"))

    return render_template(
        "edit_business_schedule.html",
        item=item
    )






##################################
#
#   ARCHIVE BUSINESS SCHEDULE CONFIRMATION
#
#
##################################



@app.route("/business-schedule/<int:schedule_id>/archive", methods=["GET"])
@login_required
@spa_required
def confirm_archive_business_schedule(schedule_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            schedule_id,
            category,
            title,
            description,
            due_date,
            recurrence_type,
            is_required
        FROM business_schedule
        WHERE schedule_id = %s
          AND spa_id = %s
          AND is_active = TRUE
    """, (schedule_id, spa_id))

    item = cur.fetchone()

    cur.close()
    conn.close()

    if not item:
        flash("Business Schedule item not found.", "warning")
        return redirect(url_for("business_schedule"))

    return render_template(
        "archive_business_schedule.html",
        item=item
    )














##################################
#
#   ARCHIVE BUSINESS SCHEDULE ITEM
#
#
##################################


@app.route("/business-schedule/<int:schedule_id>/archive", methods=["POST"])
@login_required
@spa_required
def archive_business_schedule(schedule_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE business_schedule
        SET
            is_active = FALSE,
            updated_at = NOW()
        WHERE schedule_id = %s
          AND spa_id = %s
    """, (schedule_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Business Schedule item archived.", "success")
    return redirect(url_for("business_schedule"))









##################################
#
#   RESTORE BUSINESS SCHEDULE ITEM
#
#
##################################



@app.route("/business-schedule/<int:schedule_id>/restore", methods=["POST"])
@login_required
@spa_required
def restore_business_schedule(schedule_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE business_schedule
        SET
            is_active = TRUE,
            updated_at = NOW()
        WHERE schedule_id = %s
          AND spa_id = %s
    """, (schedule_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Business Schedule item restored.", "success")
    return redirect(url_for("business_schedule"))










##################################
#
#   DELETE BUSINESS SCHEDULE ITEM
#       CONFIRM DELETE
#
##################################



@app.route("/business-schedule/<int:schedule_id>/delete", methods=["GET"])
@login_required
@spa_required
def confirm_delete_business_schedule(schedule_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            schedule_id,
            category,
            title,
            description,
            due_date,
            recurrence_type,
            is_required
        FROM business_schedule
        WHERE schedule_id = %s
          AND spa_id = %s
    """, (schedule_id, spa_id))

    item = cur.fetchone()

    cur.close()
    conn.close()

    if not item:
        flash("Business Schedule item not found.", "warning")
        return redirect(url_for("business_schedule"))

    return render_template(
        "delete_business_schedule.html",
        item=item
    )










##################################
#
#   DELETE BUSINESS SCHEDULE ITEM
#
#
##################################


@app.route("/business-schedule/<int:schedule_id>/delete", methods=["POST"])
@login_required
@spa_required
def delete_business_schedule(schedule_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        DELETE FROM business_schedule
        WHERE schedule_id = %s
          AND spa_id = %s
    """, (schedule_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Business Schedule item deleted.", "success")
    return redirect(url_for("business_schedule"))





##################################
#
#   DUPLICATE BUSINESS SCHEDULE ITEM
#   CONFIRM DUPLICATE
#
##################################


@app.route("/business-schedule/<int:schedule_id>/duplicate", methods=["GET"])
@login_required
@spa_required
def confirm_duplicate_business_schedule(schedule_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            schedule_id,
            category,
            title,
            description,
            due_date,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            is_required
        FROM business_schedule
        WHERE schedule_id = %s
          AND spa_id = %s
    """, (schedule_id, spa_id))

    item = cur.fetchone()

    cur.close()
    conn.close()

    if not item:
        flash("Business Schedule item not found.", "warning")
        return redirect(url_for("business_schedule"))

    return render_template(
        "duplicate_business_schedule.html",
        item=item
    )









##################################
#
#   DUPLICATE BUSINESS SCHEDULE ITEM
#
#
##################################


@app.route("/business-schedule/<int:schedule_id>/duplicate", methods=["POST"])
@login_required
@spa_required
def duplicate_business_schedule(schedule_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO business_schedule (
            spa_id,
            category,
            title,
            description,
            due_date,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            is_required,
            is_completed,
            is_active
        )
        SELECT
            spa_id,
            category,
            title || ' Copy',
            description,
            due_date,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            is_required,
            FALSE,
            TRUE
        FROM business_schedule
        WHERE schedule_id = %s
          AND spa_id = %s
    """, (schedule_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Business Schedule item duplicated.", "success")
    return redirect(url_for("business_schedule"))






##################################
#
#   VIEW BUSINESS SCHEDULE ITEM
#
#
##################################


@app.route("/business-schedule/<int:schedule_id>")
@login_required
@spa_required
def view_business_schedule(schedule_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            schedule_id,
            category,
            title,
            description,
            due_date,
            recurrence_type,
            recurrence_interval,
            reminder_days,
            is_required,
            is_completed,
            is_active,
            completed_at,
            created_at,
            updated_at
        FROM business_schedule
        WHERE schedule_id = %s
          AND spa_id = %s
    """, (schedule_id, spa_id))

    item = cur.fetchone()

    cur.close()
    conn.close()

    if not item:
        flash("Business Schedule item not found.", "warning")
        return redirect(url_for("business_schedule"))

    return render_template(
        "view_business_schedule.html",
        item=item
    )








##################################
#
#
#
#
##################################









##################################
#
#
#
#
##################################










#  ------------------------------
#
#       REPORTS
#   4/28 clean
#  -----------------------------


from datetime import date, datetime, timedelta

@app.route("/reports")
@login_required
@spa_required
def reports():
    spa_id = session.get("spa_id")
    spa_now = get_spa_now(spa_id)
    dashboard = get_dashboard_data(
    spa_id,
    spa_now=spa_now
)
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()

        # Business goals
    cur.execute("""
        SELECT
            daily_revenue_goal,
            weekly_revenue_goal,
            monthly_revenue_goal,
            average_ticket_goal,
            new_clients_goal,
            completion_rate_goal,
            cancellation_rate_goal,
            no_show_goal,
            inactive_client_days,
            low_inventory_threshold
        FROM spa_business_goals
        WHERE spa_id = %s
    """, (spa_id,))

    goals = cur.fetchone()

    if goals:
        daily_revenue_goal = goals[0] or 0
        weekly_revenue_goal = goals[1] or 0
        monthly_revenue_goal = goals[2] or 0
        average_ticket_goal = goals[3] or 0
        new_clients_goal = goals[4] or 0
        completion_rate_goal = goals[5] or 95
        cancellation_rate_goal = goals[6] or 5
        no_show_goal = goals[7] or 2
        inactive_client_days = goals[8] or 90
        low_inventory_threshold = goals[9] or 5
    else:
        daily_revenue_goal = 0
        weekly_revenue_goal = 0
        monthly_revenue_goal = 0
        average_ticket_goal = 0
        new_clients_goal = 0
        completion_rate_goal = 95
        cancellation_rate_goal = 5
        no_show_goal = 2
        inactive_client_days = 90
        low_inventory_threshold = 5
        
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    month_start = today.replace(day=1)
            
    if today.month == 12:
        next_month_start = date(today.year + 1, 1, 1)
    else:
        next_month_start = date(today.year, today.month + 1, 1)

    spa_filter = ""
    spa_params = []

    if role != "master_admin":
        spa_filter = " AND a.spa_id = %s"
        spa_params = [spa_id]

    # Today's completed appointments
    cur.execute(f"""
        SELECT
            a.appointment_id,
            c.first_name,
            c.last_name,
            s.service_name,
            a.appointment_date,
            a.appointment_time,  
            a.room_number,
            a.price_at_booking
        FROM appointments a
        JOIN clients c 
            ON a.client_id = c.client_id
           AND a.spa_id = c.spa_id
        LEFT JOIN services s 
            ON a.service_id = s.service_id
           AND a.spa_id = s.spa_id
        WHERE a.appointment_date = %s
          AND a.status = 'completed'
          {spa_filter}
        ORDER BY a.appointment_time
    """, [today] + spa_params)
    daily_completed = cur.fetchall() or []

    # Weekly totals
    cur.execute(f"""
        SELECT
            COUNT(*) AS total_appointments,
            COUNT(*) FILTER (WHERE status = 'booked') AS booked_count,
            COUNT(*) FILTER (WHERE status = 'completed') AS completed_count,
            COUNT(*) FILTER (WHERE status = 'cancelled') AS cancelled_count
        FROM appointments a
        WHERE a.appointment_date BETWEEN %s AND %s
          {spa_filter}
    """, [week_start, week_end] + spa_params)
    weekly_totals = cur.fetchone() or (0, 0, 0, 0)
    

    # Most booked services with revenue and average ticket
    cur.execute(f"""
        SELECT
            COALESCE(s.service_name, 'Unknown Service') AS service_name,
            COUNT(*) AS total_booked,
            COALESCE(SUM(CASE 
                WHEN a.status = 'completed' THEN a.price_at_booking 
                ELSE 0 
            END), 0) AS total_revenue,
            COALESCE(AVG(CASE 
                WHEN a.status = 'completed' THEN a.price_at_booking 
                ELSE NULL 
            END), 0) AS average_ticket
        FROM appointments a
        LEFT JOIN services s
            ON a.service_id = s.service_id
           AND a.spa_id = s.spa_id
        WHERE a.status IN ('booked', 'completed')
          {spa_filter}
        GROUP BY COALESCE(s.service_name, 'Unknown Service')
        ORDER BY total_booked DESC, service_name ASC
        LIMIT 10
    """, spa_params)
    most_booked_services = cur.fetchall() or []


    # Cancelled appointments count
    cur.execute(f"""
        SELECT COUNT(*)
        FROM appointments a
        WHERE a.status = 'cancelled'
          {spa_filter}
    """, spa_params)
    cancelled_result = cur.fetchone()
    cancelled_count = cancelled_result[0] if cancelled_result else 0
    
    # Daily revenue
    cur.execute(f"""
        SELECT COALESCE(SUM(a.price_at_booking), 0)
        FROM appointments a
        WHERE a.appointment_date = %s
          AND a.status = 'completed'
          {spa_filter}
    """, [today] + spa_params)
    daily_revenue = cur.fetchone()[0] or 0
        
    # Weekly revenue
    cur.execute(f"""
        SELECT COALESCE(SUM(a.price_at_booking), 0)
        FROM appointments a
        WHERE a.appointment_date BETWEEN %s AND %s
          AND a.status = 'completed'
          {spa_filter}
    """, [week_start, week_end] + spa_params)
    weekly_revenue = cur.fetchone()[0] or 0
        
    # Monthly revenue
    cur.execute(f"""
        SELECT COALESCE(SUM(a.price_at_booking), 0)
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          AND a.status = 'completed'
          {spa_filter}
    """, [month_start, next_month_start] + spa_params)
    monthly_revenue = cur.fetchone()[0] or 0
        
    # Monthly completed count
    cur.execute(f"""
        SELECT COUNT(*)
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          AND a.status = 'completed'
          {spa_filter}
    """, [month_start, next_month_start] + spa_params)
    monthly_completed_count = cur.fetchone()[0] or 0
        
    # Average ticket this month
    cur.execute(f"""
        SELECT COALESCE(AVG(a.price_at_booking), 0)
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          AND a.status = 'completed'
          AND a.price_at_booking IS NOT NULL
          {spa_filter}
    """, [month_start, next_month_start] + spa_params)
    average_ticket = cur.fetchone()[0] or 0
        

    # Total clients
    client_spa_filter = ""
    client_spa_params = []

    if role != "master_admin":
        client_spa_filter = "WHERE spa_id = %s"
        client_spa_params = [spa_id]

    cur.execute(f"""
        SELECT COUNT(*)
        FROM clients
        {client_spa_filter}
    """, client_spa_params)
    total_clients = cur.fetchone()[0] or 0


    # New clients this month
    if role != "master_admin":
        cur.execute("""
            SELECT COUNT(*)
            FROM clients
            WHERE spa_id = %s
              AND created_at >= %s
              AND created_at < %s
        """, [spa_id, month_start, next_month_start])
    else:
        cur.execute("""
            SELECT COUNT(*)
            FROM clients
            WHERE created_at >= %s
              AND created_at < %s
        """, [month_start, next_month_start])

    new_clients_month = cur.fetchone()[0] or 0


    # Returning clients this month
    cur.execute(f"""
        SELECT COUNT(DISTINCT a.client_id)
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          AND a.status = 'completed'
          {spa_filter}
    """, [month_start, next_month_start] + spa_params)
    returning_clients = cur.fetchone()[0] or 0


    # YTD Revenue
    year_start = date(today.year, 1, 1)

    cur.execute(f"""
        SELECT COALESCE(SUM(a.price_at_booking), 0)
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date <= %s
          AND a.status = 'completed'
          {spa_filter}
    """, [year_start, today] + spa_params)
    ytd_revenue = cur.fetchone()[0] or 0


    # Upcoming appointments - next 7 days
    next_7_days = today + timedelta(days=7)

    cur.execute(f"""
        SELECT COUNT(*)
        FROM appointments a
        WHERE a.appointment_date > %s
          AND a.appointment_date <= %s
          AND a.status = 'booked'
          {spa_filter}
    """, [today, next_7_days] + spa_params)
    upcoming_appointments_7_days = cur.fetchone()[0] or 0


    # No shows this month
    cur.execute(f"""
        SELECT COUNT(*)
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          AND a.status = 'no show'
          {spa_filter}
    """, [month_start, next_month_start] + spa_params)
    no_shows_month = cur.fetchone()[0] or 0


    # Cancellation rate this month
    cur.execute(f"""
        SELECT COUNT(*)
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          {spa_filter}
    """, [month_start, next_month_start] + spa_params)
    total_month_appointments = cur.fetchone()[0] or 0

    cur.execute(f"""
        SELECT COUNT(*)
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          AND a.status = 'cancelled'
          {spa_filter}
    """, [month_start, next_month_start] + spa_params)
    cancelled_month = cur.fetchone()[0] or 0

    cancellation_rate = (
        cancelled_month / total_month_appointments * 100
        if total_month_appointments > 0 else 0
    )


    # Completion rate this month
    completion_rate = (
        monthly_completed_count / total_month_appointments * 100
        if total_month_appointments > 0 else 0
    )


    # Revenue by service for current month
    cur.execute(f"""
        SELECT
            COALESCE(s.service_name, 'Unknown Service') AS service_name,
            COUNT(*) AS completed_count,
            COALESCE(SUM(a.price_at_booking), 0) AS total_revenue
        FROM appointments a
        LEFT JOIN services s 
            ON a.service_id = s.service_id
           AND a.spa_id = s.spa_id
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          AND a.status = 'completed'
          {spa_filter}
        GROUP BY COALESCE(s.service_name, 'Unknown Service')
        ORDER BY total_revenue DESC, service_name ASC
    """, [month_start, next_month_start] + spa_params)
    revenue_by_service = cur.fetchall() or []



    # Top 10 Clients by Revenue (Current Month)
    cur.execute(f"""
        SELECT
            c.client_id,
            c.first_name,
            c.last_name,
            COUNT(a.appointment_id) AS visits,
            COALESCE(SUM(a.price_at_booking), 0) AS total_revenue
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
           AND a.spa_id = c.spa_id
        WHERE a.status = 'completed'
          AND a.appointment_date >= %s
          AND a.appointment_date < %s
          {spa_filter}
        GROUP BY
            c.client_id,
            c.first_name,
            c.last_name
        ORDER BY
            total_revenue DESC,
            visits DESC,
            c.last_name
        LIMIT 10
    """, [month_start, next_month_start] + spa_params)

    top_clients = cur.fetchall() or []

    # Revenue by Day of Week - Current Month
    cur.execute(f"""
        SELECT
            TO_CHAR(a.appointment_date, 'Day') AS day_name,
            EXTRACT(DOW FROM a.appointment_date) AS day_number,
            COUNT(*) AS completed_count,
            COALESCE(SUM(a.price_at_booking), 0) AS total_revenue
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          AND a.status = 'completed'
          {spa_filter}
        GROUP BY day_name, day_number
        ORDER BY day_number
    """, [month_start, next_month_start] + spa_params)
    revenue_by_day = cur.fetchall() or []


    # Peak Appointment Hours - Current Month
    cur.execute(f"""
        SELECT
            TO_CHAR(a.appointment_time, 'HH12:00 AM') AS appointment_hour,
            COUNT(*) AS total_appointments
        FROM appointments a
        WHERE a.appointment_date >= %s
          AND a.appointment_date < %s
          AND a.status IN ('booked', 'completed')
          AND a.appointment_time IS NOT NULL
          {spa_filter}
        GROUP BY TO_CHAR(a.appointment_time, 'HH12:00 AM'),
                 EXTRACT(HOUR FROM a.appointment_time)
        ORDER BY EXTRACT(HOUR FROM a.appointment_time)
    """, [month_start, next_month_start] + spa_params)
    peak_hours = cur.fetchall() or []



    # Goal progress percentages
    weekly_revenue_progress = progress_percent(
        weekly_revenue,
        weekly_revenue_goal
    )



    daily_revenue_progress = progress_percent(daily_revenue, daily_revenue_goal)
    weekly_revenue_progress = progress_percent(weekly_revenue, weekly_revenue_goal)
    monthly_revenue_progress = progress_percent(monthly_revenue, monthly_revenue_goal)
    average_ticket_progress = progress_percent(average_ticket, average_ticket_goal)
    new_clients_progress = progress_percent(new_clients_month, new_clients_goal)
    completion_rate_progress = progress_percent(completion_rate, completion_rate_goal)



    ############################################################
    # BUSINESS GOAL PROGRESS
    ############################################################

    daily_revenue_progress = progress_percent(
        daily_revenue,
        daily_revenue_goal
    )

    weekly_revenue_progress = progress_percent(
        weekly_revenue,
        weekly_revenue_goal
    )

    monthly_revenue_progress = progress_percent(
        monthly_revenue,
        monthly_revenue_goal
    )

    average_ticket_progress = progress_percent(
        average_ticket,
        average_ticket_goal
    )

    new_clients_progress = progress_percent(
        new_clients_month,
        new_clients_goal
    )

    completion_rate_progress = progress_percent(
        completion_rate,
        completion_rate_goal
    )



     ############################################################
     # BUSINESS HEALTH SCORE
     ############################################################

    def score_from_progress(progress, max_points):
        progress = float(progress or 0)

        if progress >= 100:
            return max_points
        elif progress >= 75:
            return round(max_points * 0.75, 1)
        elif progress >= 50:
            return round(max_points * 0.50, 1)
        else:
            return round(max_points * 0.25, 1)


    weekly_revenue_score = score_from_progress(weekly_revenue_progress, 20)
    monthly_revenue_score = score_from_progress(monthly_revenue_progress, 20)
    average_ticket_score = score_from_progress(average_ticket_progress, 20)
    new_clients_score = score_from_progress(new_clients_progress, 15)

    if completion_rate >= 95:
        completion_score = 15
    elif completion_rate >= 85:
        completion_score = 10
    else:
        completion_score = 5

    if cancellation_rate <= 5:
        cancellation_score = 5
    elif cancellation_rate <= 10:
        cancellation_score = 3
    else:
        cancellation_score = 1

    if no_shows_month <= 2:
        no_show_score = 5
    elif no_shows_month <= 4:
        no_show_score = 3
    else:
        no_show_score = 1

    business_health_score = round(
        weekly_revenue_score
        + monthly_revenue_score
        + average_ticket_score
        + new_clients_score
        + completion_score
        + cancellation_score
        + no_show_score
    )

    if business_health_score >= 90:
        business_health_label = "Excellent"
        business_health_class = "kpi-green"
    elif business_health_score >= 75:
        business_health_label = "Good"
        business_health_class = "kpi-yellow"
    elif business_health_score >= 60:
        business_health_label = "Needs Attention"
        business_health_class = "kpi-yellow"
    else:
        business_health_label = "Critical"
        business_health_class = "kpi-red"

    cur.execute("""
        SELECT COALESCE(SUM(price_at_booking), 0)
        FROM appointments
        WHERE appointment_date = %s
          AND status IN ('booked', 'completed')
          AND spa_id = %s
    """, (today, spa_id))

    expected_revenue = cur.fetchone()[0] or 0
        



    cur.close()
    conn.close()
          
    return render_template(
        "reports.html",
        today=today,
        week_start=week_start,
        week_end=week_end,
        month_start=month_start,
        daily_completed=daily_completed,
        weekly_totals=weekly_totals,
        most_booked_services=most_booked_services,
        cancelled_count=cancelled_count,
        daily_revenue=daily_revenue,
        weekly_revenue=weekly_revenue,
        monthly_revenue=monthly_revenue,   
        monthly_completed_count=monthly_completed_count,
        average_ticket=average_ticket,
        total_clients=total_clients,
        new_clients_month=new_clients_month,
        returning_clients=returning_clients,
        ytd_revenue=ytd_revenue,
        revenue_by_service=revenue_by_service,
        upcoming_appointments_7_days=upcoming_appointments_7_days,
        no_shows_month=no_shows_month,
        cancellation_rate=cancellation_rate,
        completion_rate=completion_rate,
        top_clients=top_clients,
        revenue_by_day=revenue_by_day,
        peak_hours=peak_hours,
        daily_revenue_goal=daily_revenue_goal,
        weekly_revenue_goal=weekly_revenue_goal,
        monthly_revenue_goal=monthly_revenue_goal,
        average_ticket_goal=average_ticket_goal,
        new_clients_goal=new_clients_goal,
        completion_rate_goal=completion_rate_goal,
        cancellation_rate_goal=cancellation_rate_goal,
        no_show_goal=no_show_goal,
        inactive_client_days=inactive_client_days,
        low_inventory_threshold=low_inventory_threshold,
        daily_revenue_progress=daily_revenue_progress,
        weekly_revenue_progress=weekly_revenue_progress,
        monthly_revenue_progress=monthly_revenue_progress,
        average_ticket_progress=average_ticket_progress,
        new_clients_progress=new_clients_progress,
        completion_rate_progress=completion_rate_progress,
        business_health_score=business_health_score,
        business_health_label=business_health_label,
        business_health_class=business_health_class,
        weekly_revenue_score=weekly_revenue_score,
        monthly_revenue_score=monthly_revenue_score,
        average_ticket_score=average_ticket_score,
        new_clients_score=new_clients_score,
        completion_score=completion_score,
        cancellation_score=cancellation_score,
        no_show_score=no_show_score,
        dashboard=dashboard,
        # appointments_today=appointments_today,
        # expected_revenue=expected_revenue,
        # birthdays_today=birthdays_today,
        # business_alerts=business_alerts,
        # opportunity=opportunity
    )









#  -------------------------
#
#    REPORTS BY DATE/RANGE
#   4/28 cleaned
#  ------------------------

from datetime import datetime

@app.route("/reports/range", methods=["GET", "POST"])
@login_required
@spa_required
def reports_range():
    spa_id = current_spa_id()
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()
        
    today = date.today()
        
    preset = request.form.get("preset") or request.args.get("preset")
    start_date = request.form.get("start_date") or request.args.get("start_date")
    end_date = request.form.get("end_date") or request.args.get("end_date")
        
    if preset == "today":
        start_date = today.strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")

    elif preset == "this_week":
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        start_date = week_start.strftime("%Y-%m-%d")
        end_date = week_end.strftime("%Y-%m-%d")

    elif preset == "this_month":
        month_start = today.replace(day=1)
        if today.month == 12:
            next_month_start = date(today.year + 1, 1, 1)
        else:
            next_month_start = date(today.year, today.month + 1, 1)

        month_end = next_month_start - timedelta(days=1)
        start_date = month_start.strftime("%Y-%m-%d")
        end_date = month_end.strftime("%Y-%m-%d")

    elif preset == "last_30":
        start_date = (today - timedelta(days=29)).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
        
    formatted_start = None
    formatted_end = None
        
    if start_date and end_date:
        try:
            formatted_start = datetime.strptime(start_date, "%Y-%m-%d").strftime("%B %d, %Y")
            formatted_end = datetime.strptime(end_date, "%Y-%m-%d").strftime("%B %d, %Y")
        except ValueError:
            formatted_start = start_date
            formatted_end = end_date
     
    report_data = None
            
    if start_date and end_date:
        spa_filter = ""
        spa_params = []

        if role != "master_admin":
            spa_filter = " AND a.spa_id = %s"
            spa_params = [spa_id]

        # Totals
        cur.execute(f"""
            SELECT
                COUNT(*) AS total_appointments,
                COUNT(*) FILTER (WHERE status = 'booked') AS booked_count,
                COUNT(*) FILTER (WHERE status = 'completed') AS completed_count,
                COUNT(*) FILTER (WHERE status = 'cancelled') AS cancelled_count
            FROM appointments a
            WHERE a.appointment_date BETWEEN %s AND %s
              {spa_filter}
        """, [start_date, end_date] + spa_params)
        totals = cur.fetchone() or (0, 0, 0, 0)

        # Revenue
        cur.execute(f"""
            SELECT COALESCE(SUM(a.price_at_booking), 0)
            FROM appointments a
            WHERE a.appointment_date BETWEEN %s AND %s
              AND a.status = 'completed'
              {spa_filter}
        """, [start_date, end_date] + spa_params)
        total_revenue = cur.fetchone()[0] or 0
    
        # Average ticket
        cur.execute(f"""
            SELECT COALESCE(AVG(a.price_at_booking), 0)
            FROM appointments a
            WHERE a.appointment_date BETWEEN %s AND %s
              AND a.status = 'completed'
              AND a.price_at_booking IS NOT NULL
              {spa_filter}
        """, [start_date, end_date] + spa_params)
        average_ticket = cur.fetchone()[0] or 0
    
        # Most booked services
        cur.execute(f"""
            SELECT
                COALESCE(s.service_name, 'Unknown Service') AS service_name,
                COUNT(*) AS total_booked
            FROM appointments a 
            LEFT JOIN services s 
                ON a.service_id = s.service_id
               AND a.spa_id = s.spa_id
            WHERE a.appointment_date BETWEEN %s AND %s   
              AND a.status IN ('booked', 'completed')
              {spa_filter}
            GROUP BY COALESCE(s.service_name, 'Unknown Service')   
            ORDER BY total_booked DESC, service_name ASC
            LIMIT 10
        """, [start_date, end_date] + spa_params)
        most_booked_services = cur.fetchall() or []

        # Revenue by service 
        cur.execute(f"""
            SELECT
                COALESCE(s.service_name, 'Unknown Service') AS service_name,
                COUNT(*) AS completed_count,
                COALESCE(SUM(a.price_at_booking), 0) AS total_revenue
            FROM appointments a
            LEFT JOIN services s 
                ON a.service_id = s.service_id
               AND a.spa_id = s.spa_id
            WHERE a.appointment_date BETWEEN %s AND %s
              AND a.status = 'completed'
              {spa_filter}
            GROUP BY COALESCE(s.service_name, 'Unknown Service')
            ORDER BY total_revenue DESC, service_name ASC
        """, [start_date, end_date] + spa_params)
        revenue_by_service = cur.fetchall() or []
    
        report_data = {
            "totals": totals,  
            "total_revenue": total_revenue,
            "average_ticket": average_ticket,
            "most_booked_services": most_booked_services,
            "revenue_by_service": revenue_by_service
        }
                
    cur.close()
    conn.close()
            
    return render_template(
        "reports_range.html",
        start_date=start_date,
        end_date=end_date,
        formatted_start=formatted_start,
        formatted_end=formatted_end,
        preset=preset,
        report_data=report_data
    )









#  -----------------------------
#     CLIENT SECTION
#
#     CLIENT HEALTH PROFILE
#
#  4/28 clean
#  -----------------------------


@app.route("/client_health_profile/<int:client_id>", methods=["GET", "POST"])
@login_required
@spa_required
def client_health_profile(client_id):
    spa_id = current_spa_id()
    role = session.get("role")

    appointment_id = request.args.get("appointment_id") or request.form.get("appointment_id")
    selected_date = request.args.get("date") or request.form.get("date")
    
    conn = get_db_connection()
    cur = conn.cursor()

    client_filter = "WHERE client_id = %s"
    client_params = [client_id]

    if role != "master_admin":
        client_filter += " AND spa_id = %s"
        client_params.append(spa_id)

    cur.execute(f"""
        SELECT client_id, first_name, last_name, spa_id
        FROM clients
        {client_filter}
    """, client_params)
    client = cur.fetchone()

    if not client:
        cur.close()
        conn.close()
        flash("Client not found or not authorized.", "error")
        return redirect(url_for("clients_home"))

    client_spa_id = client[3]

    if request.method == "POST":
        sex = request.form.get("sex") or None
        skin_type_id = request.form.get("skin_type_id") or None
        fitzpatrick_id = request.form.get("fitzpatrick_id") or None
        skin_concerns = request.form.get("skin_concerns")
        skin_conditions = request.form.get("skin_conditions")
        allergies = request.form.get("allergies")
        medications = request.form.get("medications")
        current_medical_conditions = request.form.get("current_medical_conditions")
        past_medical_treatments = request.form.get("past_medical_treatments")
        
        recent_injections = "recent_injections" in request.form
        recent_laser = "recent_laser" in request.form
        pregnant = "pregnant" in request.form
        nursing = "nursing" in request.form  
        using_retinol = "using_retinol" in request.form
        using_accutane = "using_accutane" in request.form
        
        sun_exposure_level = request.form.get("sun_exposure_level")
        last_facial_date = request.form.get("last_facial_date") or None
        
        notes1 = request.form.get("notes1")
        notes2 = request.form.get("notes2")
        notes3 = request.form.get("notes3")
        
        cur.execute("""
            SELECT health_profile_id
            FROM client_health_profile
            WHERE client_id = %s
              AND spa_id = %s
        """, (client_id, client_spa_id))

        existing_profile = cur.fetchone()
        
        if existing_profile:
            cur.execute(""" 
                UPDATE client_health_profile
                SET
                    sex = %s,
                    skin_type_id = %s,
                    fitzpatrick_id = %s,
                    skin_concerns = %s,
                    skin_conditions = %s,
                    allergies = %s,
                    medications = %s,
                    current_medical_conditions = %s,
                    past_medical_treatments = %s,
                    recent_injections = %s,
                    recent_laser = %s,
                    pregnant = %s,
                    nursing = %s,
                    using_retinol = %s,
                    using_accutane = %s,
                    sun_exposure_level = %s,
                    last_facial_date = %s,
                    notes1 = %s,
                    notes2 = %s,      
                    notes3 = %s,
                    last_updated = CURRENT_DATE
                WHERE client_id = %s
                  AND spa_id = %s
            """, (
                sex, skin_type_id, fitzpatrick_id, skin_concerns,
                skin_conditions, allergies, medications,
                current_medical_conditions, past_medical_treatments,
                recent_injections, recent_laser, pregnant, nursing,
                using_retinol, using_accutane, sun_exposure_level,
                last_facial_date, notes1, notes2, notes3,
                client_id, client_spa_id
            ))
        else:
            cur.execute("""
                INSERT INTO client_health_profile (
                    spa_id,
                    client_id,
                    sex,
                    skin_type_id,
                    fitzpatrick_id,
                    skin_concerns,
                    skin_conditions,
                    allergies,
                    medications,
                    current_medical_conditions,
                    past_medical_treatments,
                    recent_injections,
                    recent_laser,
                    pregnant,   
                    nursing,
                    using_retinol,
                    using_accutane,
                    sun_exposure_level,
                    last_facial_date,
                    notes1,
                    notes2,
                    notes3   
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                client_spa_id, client_id, sex, skin_type_id, fitzpatrick_id,
                skin_concerns, skin_conditions, allergies, medications,
                current_medical_conditions, past_medical_treatments,
                recent_injections, recent_laser, pregnant, nursing,
                using_retinol, using_accutane, sun_exposure_level,
                last_facial_date, notes1, notes2, notes3
            ))
                
        conn.commit()
        cur.close()
        conn.close()
                
        flash("Pre-session intake saved successfully.", "success")
                
        if appointment_id:   
            return redirect(url_for(
                "post_appointment_wrap_up",
                appointment_id=appointment_id,
                date=selected_date
            ))
                
        return redirect(url_for("clients_home"))
                    
    cur.execute("""
        SELECT sex_type_id, sex_type
        FROM sex
        ORDER BY sex_type
    """)
    sex_options = cur.fetchall()
                    
    cur.execute("""
        SELECT skin_type_id, skin_type_name
        FROM skin_types
        WHERE spa_id = %s
        ORDER BY skin_type_name
    """, (client_spa_id,))
    skin_types = cur.fetchall()
                    
    cur.execute("""
        SELECT fitzpatrick_id, fitzpatrick_level, description
        FROM fitzpatrick_types
        ORDER BY fitzpatrick_id
    """)
    fitzpatrick_types = cur.fetchall()
            
    cur.execute("""
        SELECT *
        FROM client_health_profile
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, client_spa_id))

    profile = cur.fetchone()
                
    cur.close()
    conn.close()
                
    return render_template(
        "client_health_profile.html",
        client=client,   
        appointment_id=appointment_id,
        selected_date=selected_date,  
        profile=profile,
        sex_options=sex_options,   
        skin_types=skin_types,      
        fitzpatrick_types=fitzpatrick_types
    )






    
#  ------------------------------------
#      APPOINTMEENTS
#
#    spa_id good
#   4/28 cleaned
#  -----------------------------------


from datetime import date



@app.route("/appointments")
@login_required
@spa_required
def appointments():
    spa_id = current_spa_id()
    role = session.get("role")
    
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    show_all = request.args.get("show_all", "").strip()

    from_coach = request.args.get("from_coach", "").strip() == "1"


    status_filter = request.args.get(
        "status",
        request.args.get("filter", "")
    ).strip().lower()

    # Use the same business-local time helper used by Daily Briefing.
    spa_now = get_spa_now(spa_id)
    today = spa_now.date()
    today_str = today.isoformat()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    query = """
        SELECT
            a.appointment_id,
            a.client_id, 
            c.first_name,
            c.last_name,   
            COALESCE(s.service_name, a.external_service_name) AS service_name,
            a.appointment_date,
            a.appointment_time,
            a.status,
            a.notes
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
           AND a.spa_id = c.spa_id
        LEFT JOIN service_name_types s
            ON a.service_type_id = s.service_type_id
           AND a.spa_id = s.spa_id
        WHERE 1=1
    """

    params = []

    if role != "master_admin":
        query += " AND a.spa_id = %s"
        params.append(spa_id)
                
    # Overdue is a calculated status and always searches past dates.
    if status_filter == "overdue":
        query += """
            AND a.appointment_date < %s
            AND LOWER(COALESCE(a.status, '')) = 'booked'
        """
        params.append(today)

    else:
        # Standard status filter.
        if status_filter:
            query += """
                AND LOWER(COALESCE(a.status, '')) = %s
            """
            params.append(status_filter)

        # Apply an entered date range.
        if start_date and end_date:
            query += """
                AND a.appointment_date BETWEEN %s AND %s
            """
            params.extend([start_date, end_date])

        # With no status, no dates, and no Show All request,
        # default the page to today's appointments.
        elif not status_filter and show_all != "1":
            query += " AND a.appointment_date = %s"
            params.append(today)
        
    query += " ORDER BY a.appointment_date, a.appointment_time"
        
    cur.execute(query, tuple(params))
    appointments = cur.fetchall()
     
    cur.close()
    conn.close()
        
    return render_template(
        "appointments.html",
        appointments=appointments,
        start_date=start_date,
        end_date=end_date,
        today_str=today_str,
        status_filter=status_filter,
        status_count=len(appointments),
        from_coach=from_coach
    )









#  --------------------
#   ADD APPOINTMENT
#
#
#
#    4/28
#  ---------------------


@app.route("/add_appointment", methods=["GET", "POST"])
@login_required
@spa_required
def add_appointment():
    user_id = session.get("user_id")
    spa_id = current_spa_id()

    client_id = request.args.get("client_id") or request.form.get("client_id") or ""
    selected_date = request.args.get("selected_date") or request.form.get("selected_date") or ""
    service_type = request.form.get("service_type", "").strip()

    conn = get_db_connection() 
    cur = conn.cursor()
            
    if request.method == "POST":
        client_id = (request.form.get("client_id") or "").strip()
        service_type_id = (request.form.get("service_type_id") or "").strip()

        duration_minutes_raw = (
            request.form.get("duration_minutes") or ""
        ).strip()

        price_at_booking_raw = (
            request.form.get("price_at_booking") or ""
        ).strip()

        appointment_date = (
            request.form.get("appointment_date") or ""
        ).strip()
        
        appointment_time = (
            request.form.get("appointment_time") or ""
        ).strip()

        status = (request.form.get("status") or "booked").strip()
        notes = (request.form.get("notes") or "").strip()
        incoming_booking_id = (
            request.form.get("incoming_booking_id") or ""
        ).strip()

        if not client_id:
            flash("Client is required.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_appointment", selected_date=selected_date))
     
        if not service_type_id:
            flash("Service is required.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_appointment", selected_date=selected_date))
            
        if not appointment_date or not appointment_time:
            flash("Appointment date and time are required.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_appointment", selected_date=selected_date))
        
        try:
            duration_minutes = int(duration_minutes_raw)
        except (TypeError, ValueError):
            flash("Session length must be entered in minutes.", "error")
            cur.close()
            conn.close()

            return redirect(url_for(
                "add_appointment",
                selected_date=selected_date
            ))

        if duration_minutes <= 0:
            flash("Session length must be greater than zero.", "error")
            cur.close()
            conn.close()

            return redirect(url_for(
                "add_appointment",
                selected_date=selected_date
            ))

        try:
            price_at_booking = float(price_at_booking_raw)
        except (TypeError, ValueError):
            flash("Service price must be a valid amount.", "error")
            cur.close()
            conn.close()

            return redirect(url_for(
                "add_appointment",
                selected_date=selected_date
            ))

        if price_at_booking < 0:
            flash("Service price cannot be negative.", "error")
            cur.close()
            conn.close()

            return redirect(url_for(
                "add_appointment",
                selected_date=selected_date
            ))

    
        cur.execute("""
            SELECT 1
            FROM clients
            WHERE client_id = %s
              AND spa_id = %s
        """, (client_id, spa_id))

        if not cur.fetchone():
            flash("Invalid client selected.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_appointment", selected_date=selected_date))

        cur.execute("""
            SELECT 1
            FROM service_name_types
            WHERE service_type_id = %s
              AND spa_id = %s
        """, (service_type_id, spa_id))

        if not cur.fetchone():
            flash("Invalid service selected.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("add_appointment", selected_date=selected_date))
    
        cur.execute("""
            INSERT INTO appointments (
                spa_id,
                client_id,
                service_type_id,
                service_type,
                duration_minutes,
                price_at_booking,
                appointment_date,
                appointment_time,
                status,
                notes,
                owner_reviewed,
                owner_reviewed_at
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                TRUE,
                CURRENT_TIMESTAMP
            )
            RETURNING appointment_id
        """, (
            spa_id,
            client_id,
            service_type_id,
            service_type,
            duration_minutes,
            price_at_booking,
            appointment_date,
            appointment_time,
            status,
            notes
        ))

        appointment_id = cur.fetchone()[0]

        log_audit(
            cur,
            spa_id=spa_id,
            user_id=user_id,
            action_type="appointment_created",
            table_name="appointments",
            record_id=appointment_id,
            new_value=f"{appointment_date} {appointment_time}",
            notes="Appointment created"
        )

        log_appointment_history(
            cur,
            spa_id=spa_id,
            appointment_id=appointment_id,
            client_id=client_id,
            user_id=user_id,
            action_type="created",
            new_date=appointment_date,
            new_time=appointment_time,
            new_status=status,
            notes="Appointment created"
        )


        if incoming_booking_id:
            cur.execute("""
                UPDATE incoming_square_bookings  
                SET status = 'imported',
                    reviewed_at = CURRENT_TIMESTAMP
                WHERE incoming_booking_id = %s
                  AND spa_id = %s
            """, (incoming_booking_id, spa_id))

            session.pop("incoming_booking_data", None)
            
        conn.commit()   
        cur.close()
        conn.close()
        
        return redirect(url_for("daily_schedule", date=appointment_date))
            
    incoming_booking_id = request.args.get("incoming_booking_id", "")
    prefill_date = request.args.get("appointment_date", "") or selected_date
    prefill_time = request.args.get("appointment_time", "")
    prefill_service_name = request.args.get("service_name", "")
            

    client_search = request.args.get("client_search", "").strip()

    if client_search:
        cur.execute("""
            SELECT client_id, first_name, last_name
            FROM clients
            WHERE spa_id = %s
              AND active_client = TRUE
              AND last_name ILIKE %s
            ORDER BY last_name, first_name
        """, (spa_id, f"%{client_search}%"))

        clients = cur.fetchall()

        if len(clients) == 1:
            client_id = clients[0][0]

        elif len(clients) == 0:
            cur.close()
            conn.close()

            flash("No client found. Please add the client first.", "warning")

            return redirect(url_for(
                "add_new_client", 
                selected_date=selected_date 
            ))

    else:

        cur.execute("""
            SELECT client_id, first_name, last_name
            FROM clients
            WHERE spa_id = %s
              AND active_client = TRUE
            ORDER BY last_name, first_name
            LIMIT 25
        """, (spa_id,))

        clients = cur.fetchall()              

    cur.execute("""
        SELECT
            service_type_id,
            service_name,
            default_duration_minutes,
            default_price
        FROM service_name_types
        WHERE spa_id = %s
        AND is_active = TRUE
        ORDER BY service_name
    """, (spa_id,))

    service_types = cur.fetchall()    
                
    cur.close()
    conn.close()
                
    return render_template(
        "add_appointment.html",
        clients=clients,
        service_types=service_types,
        selected_date=selected_date,
        client_id=client_id,
        incoming_booking_id=incoming_booking_id,
        prefill_date=prefill_date,
        prefill_time=prefill_time,
        prefill_service_name=prefill_service_name
    )










#  ---------------------
#
#   EDIT  APPOINTMENT
#
#
#    spa_id good
#  4/28 cleaned
#  -------------------- 


@app.route("/edit_appointment/<int:appointment_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_appointment(appointment_id):
    spa_id = current_spa_id()
    user_id = session.get("user_id")
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()

    filter_sql = "WHERE appointment_id = %s"
    params = [appointment_id]

    if role != "master_admin":
        filter_sql += " AND spa_id = %s"
        params.append(spa_id)


    if request.method == "POST":
        service_type = request.form["service_type"].strip()
        price_at_booking = request.form["price_at_booking"].strip()
        appointment_date = request.form["appointment_date"]
        appointment_time = request.form["appointment_time"]
        duration = request.form["duration"]
        room_number = request.form["room_number"]
        notes = request.form["notes"].strip()

        duration_value = int(duration) if duration else None
        price_value = float(price_at_booking) if price_at_booking else None

        # Get the appointment before changing it
        cur.execute(f"""
            SELECT
                client_id,
                appointment_date,
                appointment_time,
                status,
                service_type,
                price_at_booking,
                duration_minutes
            FROM appointments
            {filter_sql}
        """, params)

        old_appt = cur.fetchone()

        if not old_appt:
            cur.close()
            conn.close()
            flash("Appointment not found.", "warning")
            return redirect(url_for("calendar_view"))

        client_id = old_appt[0]
        old_date = old_appt[1]
        old_time = old_appt[2]
        old_status = old_appt[3]
        old_service_type = old_appt[4]
        old_price = old_appt[5]
        old_duration = old_appt[6]

        cur.execute(f"""
            UPDATE appointments
            SET
                service_type = %s,
                price_at_booking = %s,
                appointment_date = %s,
                appointment_time = %s,
                duration_minutes = %s,
                room_number = %s,
                notes = %s,
                updated_at = CURRENT_TIMESTAMP
            {filter_sql}
        """, (
            service_type,
            price_value,
            appointment_date,
            appointment_time,
            duration_value,
            room_number,
            notes,
            *params
        ))

        log_audit(
            cur,
            spa_id=spa_id,
            user_id=user_id,
            action_type="appointment_updated",
            table_name="appointments",
            record_id=appointment_id,
            old_value=(
                f"{old_date} {old_time} | "
                f"{old_service_type or 'No service'} | "
                f"${old_price or 0} | "
                f"{old_duration or 0} minutes"
            ),
            new_value=(
                f"{appointment_date} {appointment_time} | "
                f"{service_type or 'No service'} | "
                f"${price_value or 0} | "
                f"{duration_value or 0} minutes"
            ),
            notes="Appointment updated"
        )

        was_rescheduled = (
            str(old_date) != appointment_date
            or old_time.strftime("%H:%M") != appointment_time
        )

        log_appointment_history(
            cur,
            spa_id=spa_id,
            appointment_id=appointment_id,
            client_id=client_id,
            user_id=user_id,
            action_type="rescheduled" if was_rescheduled else "updated",
            old_date=old_date,
            old_time=old_time,
            new_date=appointment_date,
            new_time=appointment_time,
            old_status=old_status,
            new_status=old_status,
            notes="Appointment edited"
        )

        conn.commit()
        cur.close()
        conn.close()

        flash("Appointment updated successfully.", "success")
        return redirect(url_for("daily_schedule", date=appointment_date))


    cur.execute(f"""
        SELECT
            appointment_id,
            appointment_date,
            appointment_time,
            duration_minutes,
            room_number,
            notes,
            service_type,
            price_at_booking
        FROM appointments
        {filter_sql}
    """, params)

    appt = cur.fetchone()

    cur.close()
    conn.close()
 
    if not appt:
        flash("Appointment not found or not authorized.", "error")
        return redirect(url_for("appointments"))

    return render_template("edit_appointment.html", appt=appt)







#  ---------------------
#   
#   DELETE  APPOINTMENT
#
#
#   spa_id good  
#   4/28
#  --------------------


@app.route("/delete_appointment/<int:appointment_id>", methods=["POST"])
@login_required
@spa_required
def delete_appointment(appointment_id):
    spa_id = current_spa_id()
    user_id = session.get("user_id")
    role = session.get("role")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    filter_sql = "WHERE appointment_id = %s"
    params = [appointment_id]
    
    if role != "master_admin":
        filter_sql += " AND spa_id = %s"
        params.append(spa_id)
    
    cur.execute(f"""
        SELECT client_id, appointment_date, appointment_time, status
        FROM appointments
        {filter_sql}
    """, params)
            
    appt = cur.fetchone()

    if not appt:
        cur.close()
        conn.close()
        flash("Appointment not found or not authorized.", "error")
        return redirect(url_for("appointments"))

    client_id = appt[0]
    old_date = appt[1]
    old_time = appt[2]
    old_status = appt[3]

    log_audit(
        cur,
        spa_id=spa_id,
        user_id=user_id,
        action_type="appointment_deleted",
        table_name="appointments",
        record_id=appointment_id,
        old_value=f"{old_date} {old_time} {old_status}",
        new_value=None,
        notes="Appointment deleted"
    )

    log_appointment_history(
        cur,
        spa_id=spa_id,
        appointment_id=appointment_id,
        client_id=client_id,
        user_id=user_id,
        action_type="deleted",
        old_date=old_date,
        old_time=old_time,
        old_status=old_status,
        notes="Appointment deleted"
    )

    cur.execute(f"""   
        DELETE FROM appointments
        {filter_sql}
    """, params)

    conn.commit()
    cur.close()
    conn.close()
    
    flash("Appointment deleted successfully.", "success")
    
    return redirect(url_for("appointments", date=old_date))











#  ---------------------
#
#   CANCEL APPOINTMENT
#       spa_id good
#  4/28
#  --------------------


@app.route("/cancel_appointment/<int:appointment_id>", methods=["POST"])
@login_required 
@spa_required
def cancel_appointment(appointment_id):
    spa_id = current_spa_id()
    user_id = session.get("user_id")
    role = session.get("role")
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    filter_sql = """
        WHERE appointment_id = %s
          AND (appointment_date + appointment_time) > CURRENT_TIMESTAMP
    """
    params = [appointment_id]
        
    if role != "master_admin":
        filter_sql += " AND spa_id = %s"
        params.append(spa_id)

    cur.execute(f"""
        SELECT client_id, appointment_date, appointment_time, status
        FROM appointments
        {filter_sql}
    """, params)

    appt = cur.fetchone()

    if not appt:
        conn.rollback()
        cur.close()
        conn.close()   
        flash("Appointment not found or can no longer be cancelled.", "error")
        return redirect(url_for("calendar_view", offset=0))

    client_id = appt[0]
    old_date = appt[1]
    old_time = appt[2]
    old_status = appt[3]

    cur.execute(f"""
        UPDATE appointments
        SET status = 'cancelled',
            updated_at = CURRENT_TIMESTAMP
        {filter_sql}
    """, params)
    
    if cur.rowcount == 0:
        conn.rollback()
        cur.close()
        conn.close()   
        flash("Appointment not found or can no longer be cancelled.", "error")
        return redirect(url_for("calendar_view", offset=0))

    log_audit(
        cur,
        spa_id=spa_id,
        user_id=user_id,
        action_type="appointment_cancelled",
        table_name="appointments",
        record_id=appointment_id,
        old_value=old_status,
        new_value="cancelled",
        notes="Appointment cancelled"
    )

    log_appointment_history(
        cur,
        spa_id=spa_id,
        appointment_id=appointment_id,
        client_id=client_id,
        user_id=user_id,
        action_type="cancelled",
        old_date=old_date,
        old_time=old_time,
        new_date=old_date,
        new_time=old_time,
        old_status=old_status,
        new_status="cancelled",
        notes="Appointment cancelled"
    )
    
    conn.commit()
    cur.close()
    conn.close()
        
    flash("Appointment cancelled.", "warning")
    return redirect(url_for("calendar_view", offset=0))






#  -----------------
#     
#   RESCHEDULE   APPOINTMENT
#  
#  spa_id good
#   4/28
#  -----------------



@app.route("/reschedule_appointment/<int:appointment_id>", methods=["GET", "POST"])
@login_required
@spa_required
def reschedule_appointment(appointment_id):
    spa_id = current_spa_id()
    user_id = session.get("user_id")
    role = session.get("role")
        
    conn = get_db_connection()
    cur = conn.cursor()

    appt_filter = "WHERE appointment_id = %s"
    appt_params = [appointment_id]

    if role != "master_admin":
        appt_filter += " AND spa_id = %s"
        appt_params.append(spa_id)
    
    if request.method == "POST":
        service_type_id = (request.form.get("service_type_id") or "").strip()
        appointment_date = (request.form.get("appointment_date") or "").strip()
        appointment_time = (request.form.get("appointment_time") or "").strip()
        status = (request.form.get("status") or "").strip()
        notes = (request.form.get("notes") or "").strip()
        original_date = (request.form.get("original_date") or "").strip()
    
        if not service_type_id or not appointment_date or not appointment_time or not status:
            flash("Please complete all required fields.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("reschedule_appointment", appointment_id=appointment_id))

        if role == "master_admin":
            cur.execute("""
                SELECT spa_id
                FROM appointments
                WHERE appointment_id = %s
            """, (appointment_id,))
            appt_spa = cur.fetchone()

            if not appt_spa:
                flash("Appointment not found.", "error")
                cur.close()
                conn.close()
                return redirect(url_for("appointments"))

            service_spa_id = appt_spa[0]
        else:
            service_spa_id = spa_id
    

        cur.execute(f"""
            SELECT client_id, appointment_date, appointment_time, status
            FROM appointments
            {appt_filter}
        """, appt_params)

        old_appt = cur.fetchone()

        if not old_appt:
            cur.close()
            conn.close()
            flash("Appointment not found or not authorized.", "error")
            return redirect(url_for("appointments"))

        client_id = old_appt[0]
        old_date = old_appt[1]
        old_time = old_appt[2]
        old_status = old_appt[3]

        cur.execute("""
            SELECT 1
            FROM service_name_types
            WHERE service_type_id = %s
              AND spa_id = %s
        """, (service_type_id, service_spa_id))

        if not cur.fetchone():
            flash("Invalid service selected.", "error")
            cur.close()
            conn.close()
            return redirect(url_for("reschedule_appointment", appointment_id=appointment_id))

        cur.execute(f"""
            UPDATE appointments
            SET
                service_type_id = %s,
                appointment_date = %s,
                appointment_time = %s,
                status = %s,
                notes = %s,
                updated_at = CURRENT_TIMESTAMP
            {appt_filter}
        """, (
            service_type_id,
            appointment_date,
            appointment_time,
            status,
            notes,
            *appt_params
        ))

        if cur.rowcount == 0:
            conn.rollback()
            cur.close()
            conn.close()
            flash("Appointment not found or not authorized.", "error")
            return redirect(url_for("appointments"))


        action_type = "rescheduled"

        if str(old_date) == appointment_date and str(old_time)[:5] == appointment_time[:5]:
            action_type = "updated"
        else:
            action_type = "rescheduled"


        log_audit(
            cur,
            spa_id=service_spa_id,
            user_id=user_id,
            action_type=f"appointment_{action_type}",
            table_name="appointments",
            record_id=appointment_id,
            old_value=f"{old_date} {old_time} {old_status}",
            new_value=f"{appointment_date} {appointment_time} {status}",
            notes="Appointment rescheduled" if action_type == "rescheduled" else "Appointment updated"
        )

        log_appointment_history(
            cur,
            spa_id=service_spa_id,
            appointment_id=appointment_id,
            client_id=client_id,
            user_id=user_id,
            action_type=action_type,
            old_date=old_date,
            old_time=old_time,
            new_date=appointment_date,
            new_time=appointment_time,
            old_status=old_status,
            new_status=status,
            notes="Appointment rescheduled" if action_type == "rescheduled" else "Appointment updated"
        )


        conn.commit()
        cur.close()
        conn.close()
    
        flash("Appointment rescheduled successfully.", "success")
        return redirect(url_for("daily_schedule", date=appointment_date or original_date))
        

    select_filter = "WHERE a.appointment_id = %s"
    select_params = [appointment_id]

    if role != "master_admin":
        select_filter += " AND a.spa_id = %s"
        select_params.append(spa_id)



    cur.execute(f"""
        SELECT
            a.appointment_id,
            a.client_id,
            c.first_name,
            c.last_name,
            a.service_type_id,
            a.appointment_date,
            a.appointment_time,
            a.status,
            a.notes,
            a.spa_id
        FROM appointments a
        JOIN clients c 
            ON a.client_id = c.client_id
           AND a.spa_id = c.spa_id   
        {select_filter}
    """, select_params)
        
    appointment = cur.fetchone()
        
    if not appointment:
        cur.close()
        conn.close()
        flash("Appointment not found or not authorized.", "error")
        return redirect(url_for("appointments"))

    appointment_spa_id = appointment[9]
        
    cur.execute("""
        SELECT service_type_id, service_name
        FROM service_name_types
        WHERE spa_id = %s
        ORDER BY service_name
    """, (appointment_spa_id,))

    service_types = cur.fetchall()
                
    cur.close()
    conn.close()
        
    return render_template( 
        "reschedule_appointment.html",
        appointment=appointment,
        service_types=service_types
    )










#  -----------------
#
#     COMPLETE APPOINTMENT
#
#     6/2/26  good
#  -----------------

@app.route("/complete_appointment/<int:appointment_id>", methods=["POST"])
@login_required
@spa_required
def complete_appointment(appointment_id):
    spa_id = current_spa_id()
    user_id = session.get("user_id")
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()
            
    filter_sql = "WHERE appointment_id = %s"
    params = [appointment_id]

    if role != "master_admin":
        filter_sql += " AND spa_id = %s"
        params.append(spa_id) 
            
    cur.execute(f"""
        SELECT client_id, appointment_date, appointment_time, status
        FROM appointments
        {filter_sql}
    """, params)
          
    appt = cur.fetchone()
        
    if not appt:
        cur.close()
        conn.close()
        flash("Appointment not found or not authorized.", "error")
        return redirect(url_for("appointments"))

    client_id = appt[0]
    old_date = appt[1]
    old_time = appt[2]
    old_status = appt[3]
            
    complete_filter_sql = """
        WHERE appointment_id = %s
          AND (appointment_date + appointment_time) <= CURRENT_TIMESTAMP
    """
    complete_params = [appointment_id]
    
    if role != "master_admin":
        complete_filter_sql += " AND spa_id = %s"
        complete_params.append(spa_id)
    
    cur.execute(f"""
        UPDATE appointments
        SET status = 'completed',
            updated_at = CURRENT_TIMESTAMP
        {complete_filter_sql}
    """, complete_params)
        
    if cur.rowcount == 0:
        conn.rollback()
        cur.close()
        conn.close()
        flash("Appointment cannot be completed yet.", "error")
        return redirect(url_for("daily_schedule", date=old_date))

    log_audit(
        cur,
        spa_id=spa_id,
        user_id=user_id,
        action_type="appointment_completed",
        table_name="appointments",
        record_id=appointment_id,
        old_value=old_status,
        new_value="completed",
        notes="Appointment completed"
    )

    log_appointment_history(
        cur,
        spa_id=spa_id,
        appointment_id=appointment_id,
        client_id=client_id,
        user_id=user_id,
        action_type="completed",
        old_date=old_date,
        old_time=old_time,
        new_date=old_date,
        new_time=old_time,
        old_status=old_status,
        new_status="completed",
        notes="Appointment completed"
    )
        
    conn.commit()
    cur.close()
    conn.close()

    flash("Appointment marked completed.", "success")
    return redirect(url_for("daily_schedule", date=old_date))










#  ------------------
#     
#   COMPLETE OVERDUE APPOINTMENTS 
#   6/2/26  spa_id good
#  -----------------


@app.route("/complete_overdue_appointments", methods=["POST"])
@login_required
@spa_required
def complete_overdue_appointments():
    spa_id = current_spa_id()
    user_id = session.get("user_id")
    role = session.get("role")
        
    conn = get_db_connection() 
    cur = conn.cursor()
     
    filter_sql = """
        WHERE status = 'booked'
          AND (appointment_date + appointment_time) < CURRENT_TIMESTAMP
    """
        
    params = []
    
    if role != "master_admin":
        filter_sql += " AND spa_id = %s"
        params.append(spa_id)

    cur.execute(f"""
        SELECT appointment_id, spa_id, client_id, appointment_date, appointment_time, status
        FROM appointments
        {filter_sql}
    """, params)

    overdue_appointments = cur.fetchall()

    cur.execute(f"""
        UPDATE appointments
        SET status = 'completed',
            updated_at = CURRENT_TIMESTAMP
        {filter_sql}
    """, params)

    updated_count = cur.rowcount

    if updated_count:
        for appt in overdue_appointments:
            appt_id = appt[0]
            appt_spa_id = appt[1]
            client_id = appt[2]
            old_date = appt[3]
            old_time = appt[4]
            old_status = appt[5]

            log_audit(
                cur,
                spa_id=appt_spa_id,
                user_id=user_id,
                action_type="appointment_completed_overdue",
                table_name="appointments",
                record_id=appt_id,
                old_value=old_status,
                new_value="completed",
                notes="Overdue appointment marked completed"
            )

            log_appointment_history(
                cur,
                spa_id=appt_spa_id,
                appointment_id=appt_id,
                client_id=client_id,
                user_id=user_id,
                action_type="completed_overdue",
                old_date=old_date,
                old_time=old_time,
                new_date=old_date,
                new_time=old_time,
                old_status=old_status,
                new_status="completed",
                notes="Overdue appointment marked completed"
            )

    conn.commit()
    cur.close()
    conn.close()
        
    if updated_count:
        flash(f"{updated_count} overdue appointment(s) marked completed.", "success")
    else:
        flash("No overdue appointments to complete.", "info")

    return redirect(url_for("calendar_view", offset=0))






#   -----------------------------
#
#    POST APPOINTMENT WRAP UP
#
#
#  4/28     spa_id good
#   ---------------------------


@app.route("/post_appointment_wrap_up/<int:appointment_id>", methods=["GET", "POST"])
@login_required
@spa_required
def post_appointment_wrap_up(appointment_id):
    spa_id = current_spa_id()
    user_id = session.get("user_id")
    role = session.get("role")
    selected_date = request.args.get("date") or request.form.get("date") or ""
            
    conn = get_db_connection()
    cur = conn.cursor()

    appt_filter = "WHERE a.appointment_id = %s"
    appt_params = [appointment_id]

    if role != "master_admin":
        appt_filter += " AND a.spa_id = %s"
        appt_params.append(spa_id)
                
    if request.method == "POST":
        treatment_notes = request.form.get("treatment_notes", "")
        products_used = request.form.get("products_used", "")
        home_care_advice = request.form.get("home_care_advice", "")
        provider_notes = request.form.get("provider_notes", "")

        cur.execute(f"""   
            SELECT a.spa_id, a.client_id, a.appointment_date, appointment_time, a.status
            FROM appointments a
            {appt_filter}
        """, appt_params)

        valid_appointment = cur.fetchone()
            
        if not valid_appointment:
            cur.close()
            conn.close()
            flash("Appointment not found or not authorized.", "error")
            if selected_date:
                return redirect(url_for("daily_schedule", date=selected_date))
            return redirect(url_for("appointments"))

        appointment_spa_id, referred_client_id, completed_date, appointment_time, old_status  = valid_appointment
            
        cur.execute("""
            INSERT INTO appointment_wrap_up (
                spa_id,
                appointment_id,
                treatment_notes,
                products_used,
                home_care_advice,
                provider_notes
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (spa_id, appointment_id)
            DO UPDATE SET
                treatment_notes = EXCLUDED.treatment_notes,
                products_used = EXCLUDED.products_used,
                home_care_advice = EXCLUDED.home_care_advice,
                provider_notes = EXCLUDED.provider_notes
        """, (
            appointment_spa_id,
            appointment_id,
            treatment_notes,
            products_used,
            home_care_advice,
            provider_notes
        ))
        
        cur.execute("""
            UPDATE appointments
            SET status = 'completed',
                updated_at = CURRENT_TIMESTAMP
            WHERE appointment_id = %s
              AND spa_id = %s
        """, (appointment_id, appointment_spa_id))


        if cur.rowcount == 0:
            conn.rollback()
            cur.close()
            conn.close()
            flash("Appointment could not be marked completed.", "error")
            return redirect(url_for("appointments"))

        log_audit(
            cur,
            spa_id=appointment_spa_id,
            user_id=user_id,
            action_type="appointment_wrap_up_saved",
            table_name="appointments",
            record_id=appointment_id,
            old_value=old_status,
            new_value="completed",
            notes="Post appointment wrap-up saved and appointment marked completed"
        )

        log_appointment_history(
            cur,
            spa_id=appointment_spa_id,
            appointment_id=appointment_id,
            client_id=referred_client_id,
            user_id=user_id,
            action_type="wrap_up_saved",
            old_date=completed_date,
            old_time=appointment_time,
            new_date=completed_date,
            new_time=appointment_time,
            old_status=old_status,
            new_status="completed",
            notes="Post appointment wrap-up saved"
        )        

        cur.execute("""
            SELECT
                referral_id,
                referrer_type,
                referrer_client_id,
                reward_amount,
                credit_earned
            FROM referrals
            WHERE spa_id = %s
              AND referred_client_id = %s
            ORDER BY referral_id DESC
            LIMIT 1
        """, (appointment_spa_id, referred_client_id))

        referral_row = cur.fetchone()
            
        if referral_row:
            referral_id, referrer_type, referrer_client_id, reward_amount, credit_earned = referral_row
            
            if not credit_earned:
                cur.execute("""
                    UPDATE referrals
                    SET credit_earned = TRUE,
                        first_completed_appointment_date = %s
                    WHERE referral_id = %s
                      AND spa_id = %s
                """, (completed_date, referral_id, appointment_spa_id))
                
                if referrer_type == "Client" and referrer_client_id:
                    cur.execute("""
                        INSERT INTO client_credit_transactions (
                            spa_id,
                            client_id,
                            source_type,
                            source_id,
                            transaction_date,
                            transaction_type,
                            amount,
                            notes
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        appointment_spa_id,
                        referrer_client_id,
                        "Referral",
                        referral_id,
                        completed_date,
                        "Earned",
                        reward_amount,
                        f"Referral credit earned when referred client {referred_client_id} completed first appointment."
                    ))
              
        conn.commit()
        cur.close()
        conn.close()
        
        flash("Wrap-Up was saved successfully.", "success")
        return redirect(url_for("post_appointment_wrap_up_saved", appointment_id=appointment_id))
                
    cur.execute(f"""
        SELECT
            a.appointment_id,  
            a.appointment_date,
            a.appointment_time,
            c.client_id,
            c.first_name,   
            c.last_name,
            a.spa_id
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
           AND a.spa_id = c.spa_id
        {appt_filter}
    """, appt_params)

    appointment = cur.fetchone()
                            
    if not appointment:
        cur.close()
        conn.close()
        flash("Appointment not found or not authorized.", "error")
        if selected_date:
            return redirect(url_for("daily_schedule", date=selected_date))
        return redirect(url_for("appointments"))

    appointment_spa_id = appointment[6]
                
    cur.execute("""
        SELECT
            treatment_notes,
            products_used,
            home_care_advice,
            provider_notes
        FROM appointment_wrap_up
        WHERE appointment_id = %s
          AND spa_id = %s
    """, (appointment_id, appointment_spa_id))

    wrap_up = cur.fetchone()
                    
    cur.close()
    conn.close()
                    
    return render_template(
        "post_appointment_wrap_up.html",
        appointment=appointment,
        wrap_up=wrap_up,
        selected_date=selected_date
    )








#  ----------------------------
#      POST APPOINTMENT SAVED
#    
#   4/28
#   ---------------------------




@app.route("/post_appointment_wrap_up_saved/<int:appointment_id>")
def post_appointment_wrap_up_saved(appointment_id):


    return render_template(
        "post_appointment_wrap_up_saved.html",
        appointment_id=appointment_id
    )










                        
                        
                        
                        
#  ----------------------------
#      APPOINTMENT HISTORY
#
#   6/2/26
#   ---------------------------
            
        
@app.route("/appointment_history/<int:appointment_id>")
@login_required
@spa_required
def appointment_history(appointment_id):
    spa_id = current_spa_id()
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()

    filter_sql = "WHERE a.appointment_id = %s"
    params = [appointment_id]

    if role != "master_admin":
        filter_sql += " AND a.spa_id = %s"
        params.append(spa_id)

    cur.execute(f"""
        SELECT
            a.appointment_id,
            a.appointment_date,
            a.appointment_time,
            a.status,
            c.first_name,
            c.last_name
        FROM appointments a
        JOIN clients c
            ON a.client_id = c.client_id
           AND a.spa_id = c.spa_id
        {filter_sql}
    """, params)

    appointment = cur.fetchone()

    if not appointment:
        cur.close()
        conn.close()
        flash("Appointment not found or not authorized.", "error")
        return redirect(url_for("appointments"))

    cur.execute("""
        SELECT
            ah.action_type,
            ah.old_date,
            ah.old_time,
            ah.new_date,
            ah.new_time,
            ah.old_status,
            ah.new_status,
            ah.notes,
            ah.created_at,
            u.username
        FROM appointment_history ah
        LEFT JOIN users u
            ON ah.user_id = u.user_id
        WHERE ah.appointment_id = %s
        ORDER BY ah.created_at DESC
    """, (appointment_id,))

    history_rows = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "appointment_history.html",
        appointment=appointment,
        history_rows=history_rows
    )
        
            









#  ------------------
#      CLIENT SECTION
#
#    Client History
#
#  spa_id safe
# 4/28
#  -----------------


@app.route("/client_history")
@login_required
@spa_required
def client_history():
    spa_id = current_spa_id()
    role = session.get("role")

    search = request.args.get("search", "").strip()

    conn = get_db_connection()
    cur = conn.cursor()

    query = """
        SELECT client_id, first_name, last_name
        FROM clients
        WHERE 1=1
    """
    params = []

    if role != "master_admin":
        query += " AND spa_id = %s"
        params.append(spa_id)

    if search:
        query += """
            AND (
                LOWER(first_name) LIKE %s  
                OR LOWER(last_name) LIKE %s
                OR phone LIKE %s
            )
        """
        params.extend([
            f"%{search.lower()}%",
            f"%{search.lower()}%",
            f"%{search}%"
        ])

    query += " ORDER BY last_name, first_name"

    cur.execute(query, params)
    rows = cur.fetchall()
        
    cur.close()
    conn.close()
        
    return render_template(
        "client_history.html",
        rows=rows,      
        search=search
    )




    




#  ------------------
#   Client History Detail page 1
#
#
#   Spa_id good
#  4/28
#  -----------------


@app.route("/client_history/<int:client_id>")
@login_required
@spa_required
def client_history_detail(client_id):
    spa_id = current_spa_id()
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()

    client_filter = "WHERE client_id = %s"
    client_params = [client_id]

    if role != "master_admin":
        client_filter += " AND spa_id = %s"
        client_params.append(spa_id)
    
    cur.execute(f"""
        SELECT client_id, first_name, last_name, phone, email, birth_date, spa_id
        FROM clients
        {client_filter}
    """, client_params)

    client = cur.fetchone()
        
    if not client:
        cur.close()
        conn.close()
        return "Client not found", 404

    client_spa_id = client[6]
        
    cur.execute("""
        SELECT
            a.appointment_id,
            a.appointment_date,
            a.appointment_time,
            s.service_name,
            a.duration_minutes, 
            a.room_number,
            a.status,
            a.notes,   
            aw.treatment_notes,   
            aw.products_used,
            aw.home_care_advice,
            aw.provider_notes
        FROM appointments a
        LEFT JOIN services s
            ON a.service_id = s.service_id
           AND a.spa_id = s.spa_id
        LEFT JOIN appointment_wrap_up aw
            ON a.appointment_id = aw.appointment_id
           AND a.spa_id = aw.spa_id
        WHERE a.client_id = %s
          AND a.spa_id = %s
        ORDER BY a.appointment_date DESC NULLS LAST,
                 a.appointment_time DESC NULLS LAST
    """, (client_id, client_spa_id))

    rows = cur.fetchall()
     
    cur.close()
    conn.close()

    return render_template(
        "client_history_detail.html",
        rows=rows,
        client=client
    )



#  ------------------
#    Client History Detail page 2
#
#
#    spa_id good
#  4/28
#  -----------------


@app.route("/client_history_two/<int:client_id>")
@login_required
@spa_required
def client_history_detail_two(client_id):
    spa_id = current_spa_id()
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()

    client_filter = "WHERE client_id = %s"
    client_params = [client_id]

    if role != "master_admin":
        client_filter += " AND spa_id = %s"
        client_params.append(spa_id)
            
    cur.execute(f"""
        SELECT client_id, first_name, last_name, phone, email, birth_date, spa_id
        FROM clients
        {client_filter}
    """, client_params)

    client = cur.fetchone()
    
    if not client:
        cur.close()
        conn.close()
        return "Client not found", 404

    client_spa_id = client[6]
    
    cur.execute("""
        SELECT
            a.appointment_id,
            a.appointment_date,
            a.appointment_time,
            s.service_name,
            a.duration_minutes,
            a.room_number,
            a.status,
            a.notes,
            aw.treatment_notes,
            aw.products_used,
            aw.home_care_advice,
            aw.provider_notes
        FROM appointments a
        LEFT JOIN services s
            ON a.service_id = s.service_id
           AND a.spa_id = s.spa_id
        LEFT JOIN appointment_wrap_up aw
            ON a.appointment_id = aw.appointment_id
           AND a.spa_id = aw.spa_id
        WHERE a.client_id = %s
          AND a.spa_id = %s
        ORDER BY a.appointment_date DESC NULLS LAST,
                 a.appointment_time DESC NULLS LAST
    """, (client_id, client_spa_id))

    rows = cur.fetchall()

    cur.close()
    conn.close()
    
    return render_template(
        "client_history_detail_two.html",
        rows=rows,
        client=client,
        client_id=client_id
    )



#  ---------------------------------
#   ADD NEW CLIENT STEP 1  
#         PAGE 1
#
#
#
#    spa_id good
#
#   4/28
#  --------------------------------


@app.route("/add_new_client", methods=["GET", "POST"])
@login_required
@spa_required
def add_new_client():
    spa_id = current_spa_id()
    selected_date = request.args.get("selected_date") or request.form.get("selected_date") or ""
    
    conn = get_db_connection()
    cur = conn.cursor()
            
    cur.execute("""
        SELECT spa_location_id, location_name
        FROM spa_locations
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY location_name
    """, (spa_id,))

    locations = cur.fetchall()
        
            

    if request.method == "POST":
        action = request.form.get("action", "next")

        step1_data = {
            "first_name": request.form.get("first_name", "").strip(),
            "last_name": request.form.get("last_name", "").strip(),
            "phone": request.form.get("phone", "").strip(),
            "email": request.form.get("email", "").strip(),
            "birth_date": request.form.get("birth_date", ""),
            "address": request.form.get("address", "").strip(),
            "city": request.form.get("city", "").strip(),
            "state": request.form.get("state", "TX").strip(),
            "zip": request.form.get("zip", "").strip(),
            "spa_location_id": request.form.get("spa_location_id") or "",
            "preferred_location_id": request.form.get("preferred_location_id") or "",
            "client_status": request.form.get("client_status", "Current").strip(),
            "preferred_language": request.form.get("preferred_language", "").strip(),
            "ok_to_call": "ok_to_call" in request.form,
            "ok_to_text": "ok_to_text" in request.form,
            "ok_to_email": "ok_to_email" in request.form,
            "preferred_contact_method": request.form.get("preferred_contact_method", "").strip()
        }

        if action == "next":
            session["new_client_step1"] = step1_data
            return redirect(url_for("add_new_client_step2", selected_date=selected_date))    


 

        if action == "save":
            try:
                cur.execute("""
                    INSERT INTO clients (
                        spa_id,
                        first_name,
                        last_name,
                        phone,
                        email,
                        birth_date,
                        address,
                        city,
                        state,
                        zip,
                        spa_location_id,
                        preferred_location_id,
                        client_status,
                        preferred_language,
                    	ok_to_call,
                    	ok_to_text,
                    	ok_to_email,
                    	preferred_contact_method,
                    	emergency_contact_name,
                    	emergency_contact_phone,
                    	referred_by,
                    	notes_one,
                    	notes_two,
                    	notes_three,
                    	active_client
                    )   
                    VALUES (
                        %s, %s, %s, %s, %s,
                    	%s, %s, %s, %s, %s,
                    	%s, %s, %s, %s, %s,
                    	%s, %s, %s, %s, %s,
                    	%s, %s, %s, %s, %s
                    )
                    RETURNING client_id
                """, (
                    spa_id,
                    step1_data.get("first_name", ""),
                    step1_data.get("last_name", ""),
                    step1_data.get("phone", ""),
                    step1_data.get("email", ""),
                    step1_data.get("birth_date") or None,
                    step1_data.get("address", ""),
                    step1_data.get("city", ""),
                    step1_data.get("state", ""),
                    step1_data.get("zip", ""),
                    step1_data.get("spa_location_id") or None,
                    step1_data.get("preferred_location_id") or None,
                    step1_data.get("client_status", "Current"),
                    step1_data.get("preferred_language") or None,
                    step1_data.get("ok_to_call", True),
                    step1_data.get("ok_to_text", True),
                    step1_data.get("ok_to_email", True),
                    step1_data.get("preferred_contact_method") or None,
                    "",
                    "",
                    None,
                    "",
                    "",
                    "",
                    True
                ))

                new_client_id = cur.fetchone()[0]
                conn.commit()

            finally:
                cur.close()
                conn.close()

            flash("Client added successfully!", "success")

            session.pop("new_client_step1", None)
            session.pop("new_client_step2", None)

            if selected_date:
                return redirect(url_for(
                    "add_appointment",
                    client_id=new_client_id,
                    selected_date=selected_date
                ))

            return redirect(url_for("client_history"))    

    step1_data = session.get("new_client_step1", {})
    
    if not step1_data:
        incoming_booking_data = session.get("incoming_booking_data", {})
        if incoming_booking_data:
            step1_data = {
                "first_name": incoming_booking_data.get("first_name", ""),
                "last_name": incoming_booking_data.get("last_name", ""),
                "phone": incoming_booking_data.get("phone", ""),
                "email": incoming_booking_data.get("email", ""),
                "birth_date": "",
                "address": "",
                "city": "",
                "state": "TX",
                "zip": "",
                "spa_location_id": "",
                "preferred_location_id": "",
                "client_status": "Current",
                "preferred_language": "",
                "ok_to_call": True,
                "ok_to_text": True,
                "ok_to_email": True,
                "preferred_contact_method": ""
            }
    
    locations = get_dropdown_options("spa_locations", spa_id)
    client_statuses = get_dropdown_options("client_statuses", spa_id)
    preferred_languages = get_dropdown_options("preferred_languages", spa_id)
    preferred_contact_methods = get_dropdown_options("preferred_contact_methods", spa_id)


    cur.close()
    conn.close()

    return render_template(
        "add_new_client.html",
        selected_date=selected_date,
        step1_data=step1_data,
        locations=locations,
        preferred_languages=preferred_languages,
        preferred_contact_methods=preferred_contact_methods,
        client_statuses=client_statuses
    )






#  ---------------------------------
#   ADD NEW CLIENT STEP 2  
#         PAGE 2
#
#    spa_id good
#  4/28
#  --------------------------------


@app.route("/add_new_client_step2", methods=["GET", "POST"])
@login_required
@spa_required
def add_new_client_step2():
    spa_id = current_spa_id()
    selected_date = request.args.get("selected_date") or request.form.get("selected_date") or ""
            
    step1 = session.get("new_client_step1")
    if not step1:
        return redirect(url_for("add_new_client"))

    if request.method == "POST":
        session["new_client_step2"] = {
            "emergency_contact_name": request.form.get("emergency_contact_name", "").strip(),
            "emergency_contact_phone": request.form.get("emergency_contact_phone", "").strip(),
            "referrer_type": request.form.get("referrer_type", "").strip(),
            "referred_by": request.form.get("referred_by", "").strip(),
            "referrer_name": request.form.get("referrer_name", "").strip(),
            "referrer_business_name": request.form.get("referrer_business_name", "").strip(),
            "referrer_phone": request.form.get("referrer_phone", "").strip(),
            "referrer_email": request.form.get("referrer_email", "").strip(),
            "notes_one": request.form.get("notes_one", "").strip(),
            "notes_two": request.form.get("notes_two", "").strip(),
            "notes_three": request.form.get("notes_three", "").strip(),
            "active_client": request.form.get("active_client", "true")
        }

        action = request.form.get("action")

        if action == "back":
            return redirect(url_for("add_new_client", selected_date=selected_date))

        if action == "save":
            step1 = session.get("new_client_step1", {})
            step2 = session.get("new_client_step2", {})
            incoming_booking_data = session.get("incoming_booking_data", {})

            conn = get_db_connection()
            cur = conn.cursor()

            try:
                referred_by_value = None
                if step2.get("referrer_type") == "Client" and step2.get("referred_by"):
                    referred_by_value = step2.get("referred_by")

                cur.execute("""
                    INSERT INTO clients (
                        spa_id,
                        first_name,
                        last_name,
                        phone,
                        email,
                        birth_date,
                        address,
                        city,
                        state,
                        zip,
                        spa_location_id,
                        preferred_location_id,
                        client_status,
                        preferred_language,
                        ok_to_call,
                        ok_to_text,
                        ok_to_email,
                        preferred_contact_method,
                        emergency_contact_name,
                        emergency_contact_phone,
                        referred_by,
                        notes_one,
                        notes_two,
                        notes_three,
                        active_client
                    )
                    VALUES (
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s
                    )
                    RETURNING client_id
                """, (
                    spa_id,
                    step1.get("first_name", ""),
                    step1.get("last_name", ""),
                    step1.get("phone", ""),
                    step1.get("email", ""),
                    step1.get("birth_date") or None,
                    step1.get("address", ""),
                    step1.get("city", ""),
                    step1.get("state", ""),
                    step1.get("zip", ""),
                    step1.get("spa_location_id") or None,
                    step1.get("preferred_location_id") or None,
                    step1.get("client_status", "Current"),
                    step1.get("preferred_language") or None,
                    step1.get("ok_to_call", True),
                    step1.get("ok_to_text", True),
                    step1.get("ok_to_email", True),
                    step1.get("preferred_contact_method") or None,
                    step2.get("emergency_contact_name", ""),
                    step2.get("emergency_contact_phone", ""),
                    referred_by_value,
                    step2.get("notes_one", ""),
                    step2.get("notes_two", ""),
                    step2.get("notes_three", ""),
                    True if step2.get("active_client") == "true" else False
                ))

                new_client_id = cur.fetchone()[0]

                referrer_type = step2.get("referrer_type", "").strip()
                referred_by = step2.get("referred_by", "").strip()
                referrer_name = step2.get("referrer_name", "").strip()
                referrer_business_name = step2.get("referrer_business_name", "").strip()
                referrer_phone = step2.get("referrer_phone", "").strip()
                referrer_email = step2.get("referrer_email", "").strip()

                if referrer_type == "Client" and referred_by:
                    cur.execute("""
                        SELECT 1
                        FROM clients
                        WHERE client_id = %s
                          AND spa_id = %s
                    """, (referred_by, spa_id))

                    if cur.fetchone():
                        cur.execute("""
                            INSERT INTO referrals (
                                spa_id,
                                referred_client_id,
                                referrer_type,
                                referrer_client_id,
                                referral_date
                            )
                            VALUES (%s, %s, %s, %s, CURRENT_DATE)
                        """, (
                            spa_id,
                            new_client_id,
                            "Client",
                            int(referred_by)
                        ))

                elif referrer_type == "Non-Client" and referrer_name:
                    cur.execute("""
                        INSERT INTO referrals (
                            spa_id, 
                            referred_client_id,
                            referrer_type,
                            referrer_name,
                            referrer_business_name,
                            referrer_phone,
                            referrer_email,
                            referral_date  
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s, CURRENT_DATE)
                    """, (
                        spa_id,
                        new_client_id,
                        "Non-Client",
                        referrer_name,
                        referrer_business_name or None,
                        referrer_phone or None,
                        referrer_email or None
                    ))

                conn.commit()

            finally:
                cur.close()
                conn.close()

            flash("Client added successfully!", "success")

            session.pop("new_client_step1", None)
            session.pop("new_client_step2", None)

            if incoming_booking_data:
                session.pop("incoming_booking_data", None)
                return redirect(url_for(
                    "add_appointment",
                    client_id=new_client_id,
                    incoming_booking_id=incoming_booking_data.get("incoming_booking_id", ""),
                    appointment_date=incoming_booking_data.get("appointment_date", ""),
                    appointment_time=incoming_booking_data.get("appointment_time", ""),
                    service_name=incoming_booking_data.get("service_name", "")
                ))

            session.pop("incoming_booking_data", None)

            if selected_date:
                return redirect(url_for(
                    "add_appointment",
                    client_id=new_client_id,
                    selected_date=selected_date
                ))

            return redirect(url_for("client_history"))

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT client_id, first_name, last_name
        FROM clients
        WHERE spa_id = %s
          AND active_client = TRUE
        ORDER BY last_name, first_name
    """, (spa_id,))

    clients_for_referral = cur.fetchall()

    cur.close()
    conn.close()

    step2_data = session.get("new_client_step2", {})

    return render_template(
        "add_new_client_step2.html",
        step2_data=step2_data,
        selected_date=selected_date,
        clients_for_referral=clients_for_referral
    )






#  -----------------
#   EDIT CLIENT
#  4/28
#  ----------------


@app.route("/edit_client/<int:client_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_client(client_id):
    spa_id = current_spa_id()
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()

    client_filter = "WHERE c.client_id = %s"
    client_params = [client_id]

    if role != "master_admin":
        client_filter += " AND c.spa_id = %s"
        client_params.append(spa_id)

    cur.execute(f"""
        SELECT
            c.client_id,
            c.first_name,
            c.last_name,
            c.phone,
            c.email,
            c.birth_date,
            c.address,
            c.city,
            c.state,
            c.zip,
            c.spa_location_id,
            c.preferred_location_id,
            c.client_status,
            c.preferred_language,
            c.ok_to_call,
            c.ok_to_text,
            c.ok_to_email,
            c.preferred_contact_method,
            c.emergency_contact_name,
            c.emergency_contact_phone,
            c.referred_by,
            c.notes_one,
            c.notes_two,
            c.notes_three,
            c.active_client,
            c.spa_id
        FROM clients c
        {client_filter}
    """, tuple(client_params))

    client = cur.fetchone()

    if not client:
        cur.close()
        conn.close()
        flash("Client not found or not authorized.", "error")
        return redirect(url_for("client_history"))

    client_spa_id = client[25]

    # Load active locations for dropdowns
    cur.execute("""
        SELECT spa_location_id, location_name
        FROM spa_locations
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY location_name
    """, (client_spa_id,))

    locations = cur.fetchall()

    if request.method == "POST":
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        phone = request.form.get("phone", "").strip()
        email = request.form.get("email", "").strip()
        birth_date = request.form.get("birth_date") or None
        address = request.form.get("address", "").strip()
        city = request.form.get("city", "").strip()
        state = request.form.get("state", "").strip()
        zip_code = request.form.get("zip", "").strip()

        spa_location_id = request.form.get("spa_location_id") or None
        preferred_location_id = request.form.get("preferred_location_id") or None
        client_status = request.form.get("client_status", "Current").strip()
        preferred_language = request.form.get("preferred_language", "").strip() or None
        ok_to_call = "ok_to_call" in request.form
        ok_to_text = "ok_to_text" in request.form
        ok_to_email = "ok_to_email" in request.form
        preferred_contact_method = request.form.get("preferred_contact_method", "").strip() or None

        emergency_contact_name = request.form.get("emergency_contact_name", "").strip()
        emergency_contact_phone = request.form.get("emergency_contact_phone", "").strip()
        referred_by = request.form.get("referred_by", "").strip()
        notes_one = request.form.get("notes_one", "").strip()
        notes_two = request.form.get("notes_two", "").strip()
        notes_three = request.form.get("notes_three", "").strip()

        active_client = True if request.form.get("active_client") == "true" else False

        cur.execute("""
            UPDATE clients
            SET first_name = %s,
                last_name = %s,
                phone = %s,
                email = %s,
                birth_date = %s,
                address = %s,
                city = %s,
                state = %s,
                zip = %s,
                spa_location_id = %s,
                preferred_location_id = %s,
                client_status = %s,
                preferred_language = %s,
                ok_to_call = %s,
                ok_to_text = %s,
                ok_to_email = %s,
                preferred_contact_method = %s,
                emergency_contact_name = %s,
                emergency_contact_phone = %s,
                referred_by = %s,
                notes_one = %s,
                notes_two = %s,
                notes_three = %s,
                active_client = %s,
                updated_at = CURRENT_TIMESTAMP
            WHERE client_id = %s
              AND spa_id = %s
        """, (
            first_name,
            last_name,
            phone,
            email,
            birth_date,
            address,
            city,
            state,
            zip_code,
            spa_location_id,
            preferred_location_id,
            client_status,
            preferred_language,
            ok_to_call,
            ok_to_text,
            ok_to_email,
            preferred_contact_method,
            emergency_contact_name,
            emergency_contact_phone,
            referred_by,
            notes_one,
            notes_two,
            notes_three,
            active_client,
            client_id,
            client_spa_id
        ))

        conn.commit()
        cur.close()
        conn.close()

        flash("Client updated successfully!", "success")
        return redirect(url_for("client_history"))

    # --- CLIENT SUMMARY DATA ---

    cur.execute("""
        SELECT COUNT(*)
        FROM appointments
        WHERE client_id = %s
          AND spa_id = %s
          AND LOWER(status) = 'completed'
    """, (client_id, client_spa_id))
    total_visits = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(SUM(total_amount), 0)
        FROM income
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, client_spa_id))
    total_revenue = cur.fetchone()[0]

    cur.execute("""
        SELECT appointment_date
        FROM appointments
        WHERE client_id = %s
          AND spa_id = %s
          AND LOWER(status) = 'completed'
        ORDER BY appointment_date DESC
        LIMIT 1
    """, (client_id, client_spa_id))
    last_visit = cur.fetchone()

    cur.execute("""
        SELECT appointment_date, appointment_time
        FROM appointments
        WHERE client_id = %s
          AND spa_id = %s
          AND LOWER(status) = 'booked'
          AND appointment_date >= CURRENT_DATE
        ORDER BY appointment_date, appointment_time
        LIMIT 1
    """, (client_id, client_spa_id))
    next_appt = cur.fetchone()

    cur.execute("""
        SELECT
            appointment_id,
            appointment_date,
            appointment_time,
            duration_minutes,
            room_number,
            status,
            booking_channel,
            price_at_booking,
            notes
        FROM appointments
        WHERE client_id = %s
          AND spa_id = %s
        ORDER BY appointment_date DESC, appointment_time DESC
        LIMIT 10
    """, (client_id, client_spa_id))
    recent_appointments = cur.fetchall()

    cur.execute("""
        SELECT COALESCE(SUM(price_at_booking), 0)
        FROM appointments
        WHERE client_id = %s
          AND spa_id = %s
          AND LOWER(status) = 'completed'
    """, (client_id, client_spa_id))
    lifetime_value = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(AVG(price_at_booking), 0)
        FROM appointments
        WHERE client_id = %s
          AND spa_id = %s
          AND LOWER(status) = 'completed'
    """, (client_id, client_spa_id))
    average_ticket = cur.fetchone()[0]

    cur.execute("""
        SELECT COALESCE(SUM(amount), 0.00)
        FROM client_credit_transactions
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, client_spa_id))
    credit_balance = cur.fetchone()[0]

    cur.close()
    conn.close()

    return render_template(
        "edit_client.html",
        client=client,
        client_id=client_id,
        locations=locations,
        total_visits=total_visits,
        total_revenue=total_revenue,
        last_visit=last_visit,
        next_appt=next_appt,
        recent_appointments=recent_appointments,
        lifetime_value=lifetime_value,
        average_ticket=average_ticket,
        credit_balance=credit_balance
    )



#   ---------------------------
#
#    DEACTIVATE CLIENT
#
#
#   4/28
#   --------------------------



@app.route("/clients/deactivate/<int:client_id>", methods=["POST"])
@login_required
@spa_required
def deactivate_client(client_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE clients
        SET active_client = FALSE
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Client has been deactivated.", "success")
    return redirect(url_for("client_management", show_all=1))




#   ---------------------------
#           
#    INACTIVATE CLIENT
#           
#           
#   4/28
#   --------------------------


@app.route("/clients/inactive")
@login_required
@spa_required
def inactive_clients():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            client_id,
            first_name,
            last_name,
            phone,
            email,
            birth_date,
            client_status
        FROM clients
        WHERE spa_id = %s
          AND active_client = FALSE
        ORDER BY last_name, first_name
    """, (spa_id,))

    rows = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("inactive_clients.html", rows=rows)



#   ---------------------------
#    
#    RE - ACTIVATE CLIENT
#           
#   
#   4/28
#   --------------------------


@app.route("/clients/reactivate/<int:client_id>", methods=["POST"])
@login_required
@spa_required
def reactivate_client(client_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        UPDATE clients
        SET active_client = TRUE
        WHERE client_id = %s
          AND spa_id = %s
    """, (client_id, spa_id))

    conn.commit()
    cur.close()
    conn.close()

    flash("Client has been reactivated.", "success")
    return redirect(url_for("inactive_clients"))








#  -----------------------
#
#  DELETE CLIENT
# 4/28
#  this is not used
#  -----------------------


@app.route("/delete_client/<int:client_id>", methods=["POST"])
@login_required
@spa_required
def delete_client(client_id):
    spa_id = current_spa_id()
    role = session.get("role")

    conn = get_db_connection()
    cur = conn.cursor()

    filter_sql = "WHERE client_id = %s"
    params = [client_id]

    if role != "master_admin":
        filter_sql += " AND spa_id = %s"
        params.append(spa_id)

    cur.execute(f"""
        DELETE FROM clients
        {filter_sql}
    """, params)

    if cur.rowcount == 0:
        conn.rollback()
        cur.close()
        conn.close()
        flash("Client not found or not authorized.", "error")
        return redirect(url_for("client_history"))

    conn.commit()
    cur.close()
    conn.close()

    flash("Client deleted successfully!", "success")
    return redirect(url_for("client_history"))







#  -----------------
#   TIME ZONES
#       
#  ----------------





#  -----------------
#    TIME ZONES
# 4/28 good
#  ----------------



from datetime import datetime
from zoneinfo import ZoneInfo

def get_current_spa_timezone(spa_id=None):
    if not spa_id:
        spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()   
    
    cur.execute("""
        SELECT timezone_name
        FROM spas
        WHERE spa_id = %s
    """, (spa_id,))
    row = cur.fetchone()
    
    cur.close()
    conn.close()
    
    if row and row[0]:
        return row[0]
    
    return "America/Chicago"





def get_spa_now(spa_id=None):
    timezone_name = get_current_spa_timezone(spa_id)
    return datetime.now(ZoneInfo(timezone_name))


def get_utc_now():
    return datetime.now(ZoneInfo("UTC"))


#   ---------------------------------
#
#    ADMIN PAGE
#   4/28 good
#   --------------------------------



@app.route("/admin")
@login_required
@spa_required
def admin():
    spa_id = current_spa_id()

    current_timezone = get_current_spa_timezone(spa_id)
    utc_now = get_utc_now()
    spa_now = get_spa_now(spa_id)

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT dropdown_key, display_label
        FROM spa_dropdown_labels
        WHERE spa_id = %s
    """, (spa_id,))

    label_rows = cur.fetchall()
    dropdown_labels = {row[0]: row[1] for row in label_rows}

    cur.close()
    conn.close()
    
    return render_template(
        "admin.html",
        current_timezone=current_timezone,
        utc_now=utc_now,
        spa_now=spa_now,
        dropdown_labels=dropdown_labels
    )





#  --------------
#
#  SKIN TYPES   DROP DOWN
#
#   DROP DOWN
#  4/28 multi spa safe
#  -----------------------


@app.route("/skin_types", methods=["GET", "POST"])
@login_required
@spa_required
def skin_types():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        skin_type_name = request.form["skin_type_name"].strip()

        cur.execute("""
            INSERT INTO skin_types (
                spa_id,
                skin_type_name,
                is_active
            )
            VALUES (%s, %s)
        """, (spa_id, skin_type_name))

        conn.commit()
    
        cur.close()
        conn.close()
        return redirect(url_for("skin_types"))
        
    cur.execute("""
        SELECT skin_type_id, skin_type_name
        FROM skin_types
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY skin_type_name
    """, (spa_id,))

    skin_types_list = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("skin_types.html", skin_types=skin_types_list)












#  ---------------------
#  DELETE SKIN TYPE
#
#
#  4/28 multi safe
#  ---------------------

@app.route("/delete_skin_type/<int:skin_type_id>", methods=["POST"])
@login_required
@spa_required
def delete_skin_type(skin_type_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()
 
    cur.execute("""
        DELETE FROM skin_types
        WHERE skin_type_id = %s
          AND spa_id = %s
    """, (skin_type_id, spa_id))

    if cur.rowcount == 0:
        flash("Skin type not found or not authorized.", "error")
    else:
        flash("Skin type deleted.", "success")

    conn.commit()
    cur.close()
    conn.close()

    return redirect(url_for("skin_types"))







#  -------------------
#
#  FITZPATRICK DROP DOWN
# 4/28 multi safe
#  ------------------


@app.route("/fitzpatrick_types", methods=["GET", "POST"])
@login_required
@spa_required
def fitzpatrick_types():
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        fitzpatrick_level = request.form["fitzpatrick_level"].strip()
        description = request.form["description"].strip()
          
        cur.execute("""
            INSERT INTO fitzpatrick_types (
                spa_id,
                fitzpatrick_level,
                description
                is_active
            )
            VALUES (%s, %s, %s)
        """, (spa_id, fitzpatrick_level, description))
    
        conn.commit()
        cur.close()
        conn.close()
    
        flash("Fitzpatrick type added successfully.", "success")
        return redirect(url_for("fitzpatrick_types"))
    
    cur.execute("""
        SELECT fitzpatrick_id, fitzpatrick_level, description
        FROM fitzpatrick_types
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY fitzpatrick_level
    """, (spa_id,))

    fitzpatrick_types = cur.fetchall()

    cur.close()
    conn.close()

    return render_template("fitzpatrick_types.html", fitzpatrick_types=fitzpatrick_types)











#  ----------------------
#
#    DELETE FITZPATRICK TYPE
# 4/28 multi safe
#  ---------------------


@app.route("/delete_fitzpatrick_types/<int:fitzpatrick_id>", methods=["POST"])
@login_required
@spa_required
def delete_fitzpatrick_types(fitzpatrick_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        UPDATE fitzpatrick_types
        SET is_active = FALSE
        WHERE fitzpatrick_id = %s
          AND spa_id = %s
    """, (fitzpatrick_id, spa_id))

    if cur.rowcount == 0:
        flash("Fitzpatrick type not found or not authorized.", "error")
    else:
        flash("Fitzpatrick type deactivated successfully!", "success")
    
    conn.commit()
    cur.close()
    conn.close()
    
    return redirect(url_for("fitzpatrick_types"))







#  -----------------------
#
#    EDIT FITZPATRICK TYPE
#  4/28 multi safe
#  -----------------------


@app.route("/edit_fitzpatrick_types/<int:fitzpatrick_id>", methods=["GET", "POST"])
@login_required
@spa_required
def edit_fitzpatrick_types(fitzpatrick_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()   

    if request.method == "POST":
        fitzpatrick_level = request.form["fitzpatrick_level"].strip()
        description = request.form["description"].strip()
    
        cur.execute("""
            UPDATE fitzpatrick_types
            SET fitzpatrick_level = %s,
                description = %s
            WHERE fitzpatrick_id = %s
              AND spa_id = %s
        """, (fitzpatrick_level, description, fitzpatrick_id, spa_id))
    
        if cur.rowcount == 0:
            flash("Fitzpatrick type not found or not authorized.", "error")
        else:
            flash("Fitzpatrick type updated successfully.", "success")

        conn.commit()
        cur.close()
        conn.close()
        
        return redirect(url_for("fitzpatrick_types"))
            
    cur.execute("""
        SELECT fitzpatrick_id, fitzpatrick_level, description
        FROM fitzpatrick_types
        WHERE fitzpatrick_id = %s
          AND spa_id = %s
          AND is_active = TRUE
    """, (fitzpatrick_id, spa_id))

    fitzpatrick = cur.fetchone()
   
    cur.close()
    conn.close()

    if not fitzpatrick:
        flash("Fitzpatrick type not found or not authorized.", "error")
        return redirect(url_for("fitzpatrick_types"))
    
    return render_template(
        "edit_fitzpatrick_types.html",
        fitzpatrick=fitzpatrick
    )





#  -----------------------
#     	DROP DOWN
#  
#    REFERRAL SOURCES
#
#  4/28/26 multi safe
#  -----------------------

@app.route("/referral_sources", methods=["GET", "POST"])
@login_required
@spa_required
def referral_sources():
    spa_id = current_spa_id()
    conn = get_db_connection()
    cur = conn.cursor()

    if request.method == "POST":
        referral_source_name = request.form["referral_source_name"].strip()

        cur.execute("""
            INSERT INTO referral_sources (
                spa_id,
                referral_source_name,
                is_active
            )
            VALUES (%s, %s, TRUE)
        """, (spa_id, referral_source_name))

        conn.commit()
        cur.close()
        conn.close()

        flash("Referral source added.", "success")
        return redirect(url_for("referral_sources"))

    cur.execute("""
        SELECT referral_source_id, referral_source_name
        FROM referral_sources
        WHERE spa_id = %s
          AND is_active = TRUE
        ORDER BY referral_source_name
    """, (spa_id,))

    referral_sources_list = cur.fetchall()

    cur.close()
    conn.close()

    return render_template(
        "referral_sources.html",
        referral_sources=referral_sources_list
    )




#  --------------------
#
#   DELETE REFERRAL SOURCE
#  4/28/26  good
#  -------------------

@app.route("/delete_referral_source/<int:referral_source_id>", methods=["POST"])
@login_required
@spa_required
def delete_referral_source(referral_source_id):
    spa_id = current_spa_id()

    conn = get_db_connection()
    cur = conn.cursor()
        
    cur.execute("""
        UPDATE referral_sources
        SET is_active = FALSE
        WHERE referral_source_id = %s
          AND spa_id = %s
    """, (referral_source_id, spa_id))

    if cur.rowcount == 0:
        flash("Referral source not found or not authorized.", "error")
    else:
        flash("Referral source deactivated.", "success")

    conn.commit()
    cur.close()
    conn.close()   

    return redirect(url_for("referral_sources"))







#  -------------------
#   CLEARS ADD NEW FORM
#
#  good
#  ------------------


@app.route("/cancel_new_client")
@login_required
@spa_required
def cancel_new_client():
    session.pop("new_client_step1", None)
    session.pop("new_client_step2", None)
    return redirect(url_for("home"))
    

@app.route("/clear_new_client")
@login_required
@spa_required
def clear_new_client():  
    session.pop("new_client_step1", None)
    session.pop("new_client_step2", None)
    return redirect(url_for("add_new_client"))





#   ----------------------------
#  ------------------------------
#  ---------------------------------
#    END   END   END   END   END
#  --------------------------------

def scheduled_send_pending_reminders():
    print("Running scheduled reminder queue...", flush=True)

    conn = get_db_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT COUNT(*)
            FROM reminder_queue
            WHERE status = 'pending'
              AND scheduled_for <= NOW()
        """)
        count = cur.fetchone()[0]

        print(f"Pending reminders ready to send: {count}", flush=True)

    finally:
        cur.close()
        conn.close()


def scheduled_generate_birthdays():
    print("Running scheduled birthday generator...", flush=True)

    spa_id = 1

    created_count = generate_birthday_reminders(spa_id)

    print(f"Birthday reminders created: {created_count}", flush=True)



















#   ---------------------------
#
#
#
#    SCHEDULER
#
#
#
#   ----------------------------





def start_scheduler():
    scheduler = BackgroundScheduler()

    scheduler.add_job(
        scheduled_send_pending_reminders,
        "interval",
        minutes=1,
        id="send_pending_reminders_job",
        replace_existing=True
    )

    scheduler.add_job(
        scheduled_generate_birthdays,
        "cron",
        hour=8,
        minute=0,
        id="birthday_reminders_job",
        replace_existing=True
    )

    scheduler.add_job(
        lambda: poll_gmail_for_godaddy_bookings(1),
        "interval",
        minutes=5,
        id="poll_gmail_godaddy_bookings",
        replace_existing=True
    )


    scheduler.start()
    print("Scheduler started.", flush=True)



    log_scheduler("System logging initialized.")
#   -------------------
#
#  GO DADDY
#
#
#
#
#   -------------------


@app.route("/test-godaddy-post")
@login_required
@spa_required
def test_godaddy_post():
    with open("test_booking.txt", "r") as f:
        body = f.read()

    with app.test_client() as client:
        response = client.post(
            "/godaddy-booking-intake",
            data={"body": body}
        )

    return response.get_data(as_text=True)





@app.route("/godaddy-booking-intake", methods=["POST"])
def godaddy_booking_intake():

    secret = request.headers.get("X-Webhook-Secret")

    if secret != os.getenv("GODADDY_WEBHOOK_SECRET"):
        return {"error": "Unauthorized"}, 401

    body = request.form.get("body") or request.get_data(as_text=True)

    if not body:
        return {"error": "No booking body received"}, 400

    spa_id = 1

    result = import_godaddy_booking(body, spa_id)

    return result, 200



@app.route("/test-secure")
def test_godaddy_secure_post():
    with open("test_booking.txt", "r") as f:
        body = f.read()

    with app.test_client() as client:
        response = client.post(
            "/godaddy-booking-intake",
            data={"body": body},
            headers={
                "X-Webhook-Secret": os.getenv("GODADDY_WEBHOOK_SECRET")
            }
        )
    return response.get_data(as_text=True)




if os.environ.get("WERKZEUG_RUN_MAIN") == "true" or os.environ.get("RENDER"):
    start_scheduler()


if __name__ == "__main__":
    app.run(debug=True, port=5001)